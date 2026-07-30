from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from django.utils import timezone
from datetime import date, timedelta

from .models import Patient, RendezVous, Pathologie, TypeVisiteCurative
from .forms import (PatientForm, RendezVousForm, PathologieForm, TypeVisiteForm,
                    TypeVisiteCurativeForm)
from medecins.models import Medecin
from core.views import log_event
from core.utils import annees_avant
from gynecologie.models import TypeVisite


def _render_related_list(request, context):
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    template = 'patients/includes/related_list_body.html' if is_ajax else 'patients/related_list.html'
    return render(request, template, context)


@login_required
def patient_list(request):
    """Liste des patients : filtres cumulables, regroupements imbriqués, filtre et
    groupement personnalisés.

    Tout vient de core.listing, comme la liste des patientes de gynécologie qui
    porte le même modèle : cette vue ne fait que déclarer le jeu de départ et
    assembler le contexte. Elle avait auparavant sa propre mécanique — un seul
    niveau de groupes, des en-têtes posés sur les objets de la page, deux critères
    de filtre écrits en dur — et n'offrait ni condition personnalisée ni
    regroupement sur un champ non prévu.
    """
    from datetime import date as _date

    from core.listing import (Listing, appliquer_conditions, champs_pour_navigateur,
                              conditions_demandees, menu_filtres, menu_groupes,
                              paginer_groupes)
    from .patient_listing import (CHAMPS_RECHERCHE, champs_patients,
                                  construire_dimensions, dimensions_personnalisees,
                                  familles_patients)

    today   = _date.today()
    q       = request.GET.get('q', '').strip()
    groupes = request.GET.getlist('group')

    base_qs = Patient.objects.all()

    # En AJAX on ne renvoie que le fragment de liste : le titre « Patient N » n'en
    # fait pas partie, ce comptage est donc inutile.
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    # Champs établis une fois et partagés : ils servent au regroupement
    # personnalisé comme au constructeur de conditions.
    champs     = champs_patients()
    dims_perso = dimensions_personnalisees(champs)
    declarees  = construire_dimensions(today)
    listing = Listing(
        recherche=CHAMPS_RECHERCHE,
        familles=familles_patients(),
        dimensions=list(declarees.values()) + dims_perso,
        par_page=40,
        tri_defaut=('nom', 'prenoms'),
    )

    filtres = listing.filtres_demandes(request)
    qs = listing.appliquer_recherche(base_qs, q)
    qs = listing.appliquer_filtres(qs, filtres, {'aujourdhui': today})

    # Conditions personnalisées (champ + opérateur + valeur), validées contre les
    # champs découverts sur le modèle : une condition inconnue est ignorée.
    conditions = conditions_demandees(request, champs)
    mode_conditions = 'ou' if request.GET.get('cm') == 'ou' else 'et'
    qs = appliquer_conditions(qs, conditions, mode_conditions)
    qs = listing.trier(qs, groupes)

    # Avec un regroupement, on pagine les **groupes** et non les lignes : toutes
    # les lignes des groupes affichés sont chargées, si bien qu'un groupe visible
    # s'ouvre toujours et que déplier n'appelle jamais le serveur.
    toutes = dict(declarees)
    toutes.update({d.cle: d for d in dims_perso})
    dims = [toutes[g] for g in groupes if g in toutes]
    arbre = []
    if dims:
        arbre, page_obj, nb_groupes = paginer_groupes(qs, dims, request.GET.get('page'))
    else:
        nb_groupes = 0
        page_obj = Paginator(qs, listing.par_page).get_page(request.GET.get('page'))

    template = 'patients/includes/list_body.html' if is_ajax else 'patients/list.html'
    return render(request, template, {
        'page_obj':   page_obj,
        # `arbre` porte les groupes imbriqués ; vide sans regroupement, la liste
        # est alors rendue à plat depuis page_obj.
        'arbre':      arbre,
        'nb_groupes': nb_groupes,
        'stats':      {} if is_ajax else {'total': base_qs.count()},
        'q':          q,
        'filters':    filtres,
        'groups':     groupes,
        'filtre_pose':      bool(filtres),
        'selection_active': bool(filtres or groupes or q or conditions),
        # Menus générés depuis la déclaration : le gabarit ne fait que parcourir.
        'listing_filtres': menu_filtres(listing.familles, filtres),
        'listing_groupes': menu_groupes(list(declarees.values()) + dims_perso, groupes),
        'conditions':      conditions,
        'mode_conditions': mode_conditions,
        # L'entrée de menu doit survivre au rafraîchissement AJAX : elle dépend de
        # ce drapeau, pas des données JSON qui n'accompagnent que la page complète.
        'listing_filtre_perso': True,
        'listing_champs_json':  None if is_ajax else champs_pour_navigateur(champs),
        'breadcrumb': [{'title': 'Patients'}],
    })


# ── Export / Import des patients ────────────────────────────────────────────

_PATIENT_HDR = ['code_identifiant', 'nom', 'age', 'genre', 'mobile']


def _patient_row(p):
    return [p.code_patient, f'{p.nom} {p.prenoms}'.strip(), p.age_detail, p.sexe, p.telephone]


@login_required
def export_patients(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    from core.utils import csv_response
    import json as _json
    from django.http import HttpResponse

    fmt = request.GET.get('format', 'json')
    qs = Patient.objects.all()
    rows = [_patient_row(p) for p in qs]

    if fmt == 'csv':
        return csv_response('patients', _PATIENT_HDR, rows, delimiter=',')
    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io as _io
        wb = Workbook()
        ws = wb.active
        ws.title = 'Patients'
        fill = PatternFill(start_color='1F6E8C', end_color='1F6E8C', fill_type='solid')
        fnt = Font(color='FFFFFF', bold=True)
        ws.append(_PATIENT_HDR)
        for cell in ws[1]:
            cell.fill, cell.font = fill, fnt
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            ws.append(['' if v is None else v for v in row])
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 55)
        buf = _io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="patients.xlsx"'
        return resp

    data = [dict(zip(_PATIENT_HDR, r)) for r in rows]
    resp = HttpResponse(
        _json.dumps(data, ensure_ascii=False, indent=2, default=str),
        content_type='application/json',
    )
    resp['Content-Disposition'] = 'attachment; filename="patients.json"'
    return resp


@login_required
def patients_modele_excel(request):
    """Modèle Excel vierge pour l'import — mêmes colonnes que celles lues par import_patients."""
    if not request.user.is_superuser:
        raise PermissionDenied
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io as _io
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = 'Patients'
    fill = PatternFill(start_color='1F6E8C', end_color='1F6E8C', fill_type='solid')
    fnt = Font(color='FFFFFF', bold=True)
    ws.append(_PATIENT_HDR)
    for cell in ws[1]:
        cell.fill, cell.font = fill, fnt
        cell.alignment = Alignment(horizontal='center')
    ws.append(['', 'Koné Aminata', 34, 'F', '0708091011'])
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 55)
    buf = _io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="modele_patients.xlsx"'
    return resp


def _normalize_nom_patient(s):
    """Clé de comparaison insensible à la casse/accents/ponctuation, pour éviter les
    doublons patients à l'import (ex. « Koné » vs « KONE »)."""
    import re
    import unicodedata
    s = unicodedata.normalize('NFKD', str(s or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^A-Za-z0-9]+', '', s)
    return s.upper()


def _parse_age_to_date(v, today):
    """
    Convertit une valeur d'âge en date de naissance approximative. Accepte soit un
    nombre simple (ex. 34, '34'), soit le format composé « 2Ans4Mois1Jours »
    (ou tout sous-ensemble : « 30Ans », « 0Ans9Mois11Jours »…) utilisé par certains
    exports d'autres logiciels — dans ce dernier cas la date est bien plus précise
    (jusqu'au jour près) qu'un simple nombre d'années.
    """
    import re
    s = str(v or '').strip()
    if not s:
        return None

    m = re.fullmatch(r'\d+([.,]\d+)?', s)
    if m:
        try:
            years = int(float(s.replace(',', '.')))
        except ValueError:
            return None
        if years < 0:
            return None
        return annees_avant(today, years)

    m_y = re.search(r'(\d+)\s*Ans?', s, re.IGNORECASE)
    m_m = re.search(r'(\d+)\s*Mois', s, re.IGNORECASE)
    m_d = re.search(r'(\d+)\s*Jours?', s, re.IGNORECASE)
    if not (m_y or m_m or m_d):
        return None
    years = int(m_y.group(1)) if m_y else 0
    months = int(m_m.group(1)) if m_m else 0
    days = int(m_d.group(1)) if m_d else 0

    total_months = years * 12 + months
    year = today.year
    month = today.month - total_months
    while month <= 0:
        month += 12
        year -= 1
    day = today.day
    while True:
        try:
            base = date(year, month, day)
            break
        except ValueError:
            day -= 1
    return base - timedelta(days=days)


@login_required
def import_patients(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('patients:list')

    data, err = _parse_pathologie_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('patients:list')

    do_update = 'update' in request.POST
    created = updated = skipped = errors = 0
    today = timezone.now().date()

    for item in data:
        try:
            nom_complet = _s(item.get('nom', ''))
            if not nom_complet:
                errors += 1
                continue
            parts = nom_complet.split(' ', 1)
            nom = parts[0]
            prenoms = parts[1] if len(parts) > 1 else parts[0]

            date_naiss = _parse_age_to_date(item.get('age'), today)
            if not date_naiss:
                errors += 1
                continue

            sexe_raw = _s(item.get('genre', '')).lower()
            sexe = {
                'm': 'M', 'masculin': 'M', 'homme': 'M', 'male': 'M',
                'f': 'F', 'féminin': 'F', 'feminin': 'F', 'femme': 'F', 'female': 'F',
            }.get(sexe_raw, '')
            if sexe not in ('M', 'F'):
                errors += 1
                continue

            telephone = _s(item.get('mobile', ''))
            if not telephone:
                errors += 1
                continue

            defaults = {
                'nom': nom, 'prenoms': prenoms, 'sexe': sexe,
                'date_naissance': date_naiss, 'telephone': telephone,
            }

            code_identifiant = _s(item.get('code_identifiant', ''))
            existing = Patient.objects.filter(code_patient=code_identifiant).first() if code_identifiant else None
            matched_by_own_code = existing is not None

            if not existing:
                nom_norm = _normalize_nom_patient(nom)
                prenoms_norm = _normalize_nom_patient(prenoms)
                existing = next((
                    p for p in Patient.objects.filter(date_naissance=date_naiss)
                    if _normalize_nom_patient(p.nom) == nom_norm and _normalize_nom_patient(p.prenoms) == prenoms_norm
                ), None)

            if code_identifiant and not matched_by_own_code:
                # Ce n'est pas un de nos propres code_patient (ex. réimport de notre
                # export) : c'est l'identifiant d'un système externe, on le conserve
                # pour référence future sans l'afficher nulle part.
                defaults['ancien_identifiant'] = code_identifiant

            if existing:
                if do_update:
                    for k, v in defaults.items():
                        setattr(existing, k, v)
                    existing.save()
                    updated += 1
                else:
                    skipped += 1
            else:
                Patient.objects.create(**defaults)
                created += 1
        except Exception:
            errors += 1

    if errors:
        messages.warning(request, f'{created} créé(s), {updated} mis à jour, {skipped} ignoré(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} patient(s) importé(s), {updated} mis à jour, {skipped} ignoré(s).')
    return redirect('patients:list')


@login_required
def patient_detail(request, pk):
    patient = get_object_or_404(Patient, pk=pk)

    rdv_count = patient.rendez_vous.count()
    consultation_count = patient.consultations.count()
    facture_count = patient.factures.count()

    try:
        from consultations.models import Ordonnance
        ordonnance_count = Ordonnance.objects.filter(consultation__patient=patient).count()
    except Exception:
        ordonnance_count = 0

    try:
        from hospitalisation.models import Hospitalisation
        hospitalisation_count = Hospitalisation.objects.filter(patient=patient).count()
    except Exception:
        hospitalisation_count = 0

    try:
        from laboratoire.models import DemandeExamen, AnalyseLaboratoire
        demande_examens_count = DemandeExamen.objects.filter(patient=patient).count()
        resultat_examens_count = AnalyseLaboratoire.objects.filter(
            patient=patient, statut__in=['resultat', 'valide', 'envoye']
        ).count()
    except Exception:
        demande_examens_count = 0
        resultat_examens_count = 0

    # Navigation précédent/suivant dans la liste ordonnée
    ids = list(Patient.objects.order_by('-date_creation').values_list('pk', flat=True))
    try:
        idx = ids.index(pk)
        prev_pk = ids[idx - 1] if idx > 0 else None
        next_pk = ids[idx + 1] if idx < len(ids) - 1 else None
        position = idx + 1
    except ValueError:
        prev_pk = next_pk = None
        position = 1

    return render(request, 'patients/detail.html', {
        'patient': patient,
        'rdv_count': rdv_count,
        'consultation_count': consultation_count,
        'facture_count': facture_count,
        'ordonnance_count': ordonnance_count,
        'hospitalisation_count': hospitalisation_count,
        'demande_examens_count': demande_examens_count,
        'resultat_examens_count': resultat_examens_count,
        'total': len(ids),
        'position': position,
        'prev_pk': prev_pk,
        'next_pk': next_pk,
    })


@login_required
def patient_create(request):
    if request.method == 'POST':
        form = PatientForm(request.POST, request.FILES)
        if form.is_valid():
            patient = form.save()
            messages.success(request, f'Patient {patient.nom} {patient.prenoms} enregistré avec le code {patient.code_patient}.')
            return redirect('patients:list')
    else:
        form = PatientForm()
    return render(request, 'patients/form.html', {'form': form, 'titre': 'Nouveau patient', 'edit': False})


@login_required
def patient_edit(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        form = PatientForm(request.POST, request.FILES, instance=patient)
        if form.is_valid():
            form.save()
            messages.success(request, 'Dossier patient mis à jour.')
            return redirect('patients:detail', pk=patient.pk)
    else:
        form = PatientForm(instance=patient)
    return render(request, 'patients/form.html', {
        'form': form,
        'patient': patient,
        'titre': f'Modifier — {patient.nom} {patient.prenoms}',
        'edit': True,
    })


def _feuilles(noeuds):
    """Parcourt l'arbre de groupes et renvoie les nœuds qui portent des lignes."""
    for n in noeuds:
        if n['enfants']:
            yield from _feuilles(n['enfants'])
        else:
            yield n


def _rdv_listing(request, base_qs, template_page, rdv_url_name,
                 create_url=None, empty_sub=None, contexte_gyneco=False):
    """Filtrage, regroupement et pagination communs aux deux listes de rendez-vous.

    Les pages « Rendez-vous » (patients) et « Rendez-vous » (gynécologie) offrent
    les mêmes menus : seule leur requête de départ diffère. La logique vit dans
    patients/rdv_listing.py pour que les deux ne divergent pas.
    """
    from datetime import date
    from core.listing import (Listing, appliquer_conditions, champs_pour_navigateur,
                              conditions_demandees, menu_filtres, menu_groupes)
    from .rdv_listing import (FILTRES_PAR_DEFAUT, annoter_diagnostics, champs_rdv,
                              construire_dimensions, dimensions_menu,
                              dimensions_personnalisees, familles_rdv, libelle_periode,
                              trier_pour_groupes)

    today     = date.today()
    q         = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    groupes   = request.GET.getlist('group')

    # Déclaration de la liste : la brique en tire l'application des filtres et la
    # génération des menus. Les modules n'écrivent plus que cette déclaration.
    # Dimensions déclarées, complétées par celles générées depuis les champs des
    # formulaires : l'utilisateur peut regrouper sur n'importe lequel sans qu'on
    # l'ait prévu. La liste des champs est établie une fois et partagée entre le
    # regroupement et le constructeur de conditions.
    champs = champs_rdv()
    dims_perso = dimensions_personnalisees(champs)
    listing = Listing(
        recherche=('patient__nom', 'patient__prenoms', 'patient__code_patient'),
        familles=familles_rdv(contexte_gyneco),
        dimensions=list(construire_dimensions(today).values()) + dims_perso,
        par_page=25,
        filtres_defaut=FILTRES_PAR_DEFAUT,
        tri_defaut=('-date_heure',),
    )
    filtres = listing.filtres_demandes(request)   # défaut : la journée en cours

    qs = listing.appliquer_recherche(base_qs, q)
    qs = listing.appliquer_filtres(qs, filtres, {
        'user': request.user, 'aujourdhui': today,
        'date_from': date_from, 'date_to': date_to,
    })
    # Conditions personnalisées (champ + opérateur + valeur), validées contre la
    # liste des champs découverts : une condition inconnue est ignorée.
    conditions = conditions_demandees(request, champs)
    mode_conditions = 'ou' if request.GET.get('cm') == 'ou' else 'et'
    qs = appliquer_conditions(qs, conditions, mode_conditions)
    qs = trier_pour_groupes(qs, groupes, today)

    # Avec un regroupement actif, on pagine les **groupes** et non les lignes :
    # toutes les lignes des groupes affichés sont chargées, si bien qu'un groupe
    # visible s'ouvre toujours et que déplier n'appelle jamais le serveur.
    from core.listing import paginer_groupes
    declarees = dict(construire_dimensions(today))
    declarees.update({d.cle: d for d in dims_perso})
    dims = [declarees[g] for g in groupes if g in declarees]
    arbre = []
    if dims:
        arbre, page_obj, nb_groupes = paginer_groupes(qs, dims, request.GET.get('page'))
        annoter_diagnostics([o for n in _feuilles(arbre) for o in n['lignes']])
    else:
        nb_groupes = 0
        page_obj = Paginator(qs, 25).get_page(request.GET.get('page'))
        annoter_diagnostics(list(page_obj))

    # En AJAX on ne renvoie que les zones rafraîchies, pas la page entière.
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    # La période appliquée d'office n'est pas un filtre posé par l'utilisateur :
    # elle ne doit ni allumer le bouton « Filtres » ni faire apparaître
    # « Effacer », sinon on invite à effacer quelque chose qui revient aussitôt.
    # Elle reste visible sous forme d'une mention discrète à côté du titre.
    filtre_pose = not listing.est_selection_par_defaut(filtres) or bool(date_from or date_to)

    return render(request, 'includes/rdv_body.html' if is_ajax else template_page, {
        'page_obj':  page_obj,
        # `arbre` porte les groupes imbriqués ; vide sans regroupement, la liste
        # est alors rendue à plat depuis page_obj.
        'arbre':      arbre,
        'nb_groupes': nb_groupes,
        'today':     today.isoformat(),
        'q':         q,
        'filters':   filtres,
        'groups':    groupes,
        'date_from': date_from,
        'date_to':   date_to,
        'filtre_pose':      filtre_pose,
        'selection_active': filtre_pose or bool(groupes) or bool(q),
        'periode_libelle':  libelle_periode(filtres, date_from, date_to),
        # Menus générés depuis la déclaration : les gabarits n'ont plus qu'à
        # parcourir ces structures.
        'listing_filtres':  menu_filtres(listing.familles, filtres, date_from, date_to),
        'listing_groupes':  menu_groupes(dimensions_menu(contexte_gyneco, today) + dims_perso, groupes),
        'conditions':       conditions,
        'mode_conditions':  mode_conditions,
        # L'entrée de menu doit rester présente après un rafraîchissement AJAX :
        # elle est donc conditionnée à ce drapeau, et non aux données JSON — qui,
        # elles, n'accompagnent que la page complète.
        'listing_filtre_perso': True,
        'listing_champs_json': None if is_ajax else champs_pour_navigateur(champs),

        # Fournis par la vue (et non par un {% include with %}) pour que la page
        # complète et le fragment AJAX rendent exactement les mêmes contrôles.
        'subheader_title':       'Rendez-vous',
        'subheader_placeholder': 'Rechercher un patient…',
        'subheader_create_url':  create_url,
        'rdv_url_name':          rdv_url_name,
        'rdv_empty_sub':         empty_sub,
        # Les types de visite ne sont pas les mêmes des deux côtés : les CPN
        # configurés en gynécologie, le curatif pour les consultations.
        'contexte_gyneco':       contexte_gyneco,
        # Listes de configuration : le menu Filtres les affiche telles quelles,
        # donc un type ajouté en configuration devient filtrable aussitôt.
        'types_visite_cpn':      (TypeVisite.objects.filter(actif=True).order_by('nom')
                                  if contexte_gyneco else None),
        'types_visite_curative': (None if contexte_gyneco else
                                  TypeVisiteCurative.objects.filter(actif=True).order_by('nom')),
    })


@login_required
def rdv_global_list(request):
    base_qs = (RendezVous.objects
               .select_related('patient', 'medecin', 'departement', 'type_consultation',
                               'cur_type_visite', 'patient__assurance', 'medecin__employe')
               .prefetch_related('registre_curatif')
               .order_by('-date_heure'))
    from django.urls import reverse
    return _rdv_listing(request, base_qs, 'patients/rendez_vous.html',
                        rdv_url_name='patients:rdv_edit',
                        create_url=reverse('patients:rdv_create'),
                        empty_sub='Créez le premier rendez-vous.')


@login_required
def patient_info_json(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    return JsonResponse({'age': patient.age, 'telephone': patient.telephone})


@login_required
def patient_search_json(request):
    def _to_dict(p):
        return {
            'id': p.pk,
            'nom_complet': f"{p.nom} {p.prenoms}",
            'code': p.code_patient,
            'telephone': p.telephone or '',
            'age': p.age,
            'sexe_display': p.get_sexe_display(),
            'adresse': p.adresse or '',
            'assurance_nom': p.assurance.nom if p.assurance_id else '',
        }

    pk = request.GET.get('id', '').strip()
    if pk:
        try:
            p = Patient.objects.select_related('assurance').get(pk=int(pk))
            return JsonResponse({'results': [_to_dict(p)]})
        except (Patient.DoesNotExist, ValueError):
            return JsonResponse({'results': []})

    q = request.GET.get('q', '').strip()
    base_qs = Patient.objects.select_related('assurance').order_by('nom', 'prenoms')
    if not q:
        qs = base_qs[:20]
    elif len(q) < 2:
        return JsonResponse({'results': []})
    else:
        qs = base_qs.filter(
            Q(nom__icontains=q) | Q(prenoms__icontains=q) |
            Q(code_patient__icontains=q) | Q(telephone__icontains=q)
        )[:20]
    return JsonResponse({'results': [_to_dict(p) for p in qs]})


@login_required
def rdv_create(request):
    if request.method == 'POST':
        form = RendezVousForm(request.POST)
        patient_obj = None
        if form.is_valid():
            rdv = form.save(commit=False)
            code = request.POST.get('code_confirmation', '').strip()
            if code:
                rdv.code_confirmation = code
            rdv._skip_auto_log = True
            rdv.save()
            log_event(rdv, request.user, 'Rendez-vous créé.', type='system')
            messages.success(
                request,
                f'Rendez-vous créé pour {rdv.patient.nom} {rdv.patient.prenoms} '
                f'le {rdv.date_heure.strftime("%d/%m/%Y à %H:%M")}.'
            )
            action = request.POST.get('_action', '')
            if action == 'annuler':
                return redirect('patients:rdv_global')
            from django.urls import reverse
            return redirect(reverse('facture_create') + f'?patient={rdv.patient.pk}&rdv={rdv.pk}')
    else:
        initial = {'date_heure': timezone.now().strftime('%Y-%m-%dT%H:%M')}
        patient_pk = request.GET.get('patient')
        patient_obj = None
        if patient_pk:
            patient_obj = get_object_or_404(Patient, pk=patient_pk)
            initial['patient'] = patient_obj.pk
        form = RendezVousForm(initial=initial)
    return render(request, 'patients/rendez_vous_form.html', {
        'form':            form,
        'titre':           'Nouveau rendez-vous',
        'patient_prefill': patient_obj,
        'is_new':          True,
        'consultation':    None,
        'constante':       None,
        'pathologies':     Pathologie.objects.filter(actif=True, departement__code__in=('medg', 'MEDGEN')).order_by('nom'),
        # Types de visite curative, configurables depuis le menu Configurations
        # des rendez-vous (remplace trois valeurs autrefois écrites en dur).
        'types_visite_curative': TypeVisiteCurative.objects.filter(actif=True).order_by('nom'),
        'medecins':        Medecin.objects.filter(actif=True).select_related('employe').order_by('employe__nom', 'employe__prenoms'),
    })


@login_required
def rdv_edit(request, pk):
    rdv = get_object_or_404(RendezVous, pk=pk)

    try:
        from facturation.models import Facture
        facture_payee = Facture.objects.filter(
            Q(rendez_vous=rdv) | Q(patient=rdv.patient, statut='payee')
        ).exclude(statut='annulee').exists()
    except Exception:
        facture_payee = False

    # Consultation + constante liées à ce RDV
    consultation = None
    constante = None
    try:
        consultation = rdv.consultation
        try:
            constante = consultation.constantes
        except Exception:
            pass
    except Exception:
        pass

    if request.method == 'POST':
        action = request.POST.get('_action', '')

        if action == 'save_eval':
            # Sauvegarder le médecin sélectionné dans le modal
            medecin_pk = request.POST.get('eval_medecin', '').strip()
            if medecin_pk:
                try:
                    rdv.medecin = Medecin.objects.get(pk=medecin_pk)
                    rdv.save(update_fields=['medecin'])
                except Exception:
                    pass

            _eval_map = {
                'eval_poids': 'poids',
                'eval_taille': 'taille',
                'eval_temperature': 'temperature',
                'eval_tension_systolique': 'tension_systolique',
                'eval_tension_diastolique': 'tension_diastolique',
                'eval_tension_systolique_droite': 'tension_systolique_droite',
                'eval_tension_diastolique_droite': 'tension_diastolique_droite',
                'eval_pouls': 'pouls',
                'eval_frequence_respiratoire': 'frequence_respiratoire',
                'eval_saturation_oxygene': 'saturation_oxygene',
                'eval_glycemie': 'glycemie',
                'eval_albumine': 'albumine',
                'eval_perimetre_brachial': 'perimetre_brachial',
                'eval_niveau_douleur': 'niveau_douleur',
            }
            from consultations.models import Consultation as Consult, Constante as Const
            try:
                consult_obj = rdv.consultation
            except Exception:
                consult_obj = None
            if consult_obj is None:
                consult_obj = Consult.objects.create(
                    patient=rdv.patient,
                    medecin=rdv.medecin,
                    rendez_vous=rdv,
                    motif=rdv.motif or 'Évaluation clinique',
                    cree_par=request.user,
                )
            const_obj, _ = Const.objects.get_or_create(consultation=consult_obj)
            for post_key, model_field in _eval_map.items():
                val = request.POST.get(post_key, '').strip()
                if val != '':
                    setattr(const_obj, model_field, val)
            const_obj.save()
            messages.success(request, 'Évaluation enregistrée. Sélectionnez un médecin et cliquez sur « En Attente » pour continuer.')
            from django.urls import reverse
            return redirect(reverse('patients:rdv_edit', kwargs={'pk': rdv.pk}) + '?edit=1')

        if action == 'confirmer':
            if facture_payee:
                from django.utils import timezone as tz
                rdv.statut = 'confirme'
                rdv.date_confirme = tz.now()
                rdv._skip_auto_log = True
                rdv.save(update_fields=['statut', 'date_confirme'])
                log_event(rdv, request.user, 'État : Brouillon → Confirmer', type='statut')
                messages.success(request, 'Rendez-vous confirmé.')
                return redirect('patients:rdv_global')
            else:
                messages.error(request, 'Une facture est requise pour confirmer ce rendez-vous.')
                return redirect('patients:rdv_global')

        if action == 'en_attente':
            from django.utils import timezone as tz
            now = tz.now()
            rdv.statut = 'en_attente'
            rdv.date_en_attente = now
            if rdv.date_confirme:
                rdv.temps_constante_minutes = int((now - rdv.date_confirme).total_seconds() / 60)
            rdv.duree_minutes = rdv.temps_constante_minutes + rdv.temps_attente_minutes + rdv.temps_consultation_minutes
            update_fields = ['statut', 'date_en_attente', 'temps_constante_minutes', 'duree_minutes']
            medecin_pk = request.POST.get('medecin', '').strip()
            if medecin_pk:
                try:
                    rdv.medecin = Medecin.objects.get(pk=medecin_pk)
                    update_fields.append('medecin')
                except Exception:
                    pass
            rdv._skip_auto_log = True
            rdv.save(update_fields=update_fields)
            log_event(rdv, request.user, 'État : Confirmer → En Attente', type='statut')
            messages.success(request, 'Rendez-vous mis en attente de consultation.')
            from django.urls import reverse
            return redirect(reverse('patients:rdv_edit', kwargs={'pk': rdv.pk}))

        if action == 'en_consultation':
            from django.utils import timezone as tz
            now = tz.now()
            rdv.statut = 'en_consultation'
            rdv.date_en_consultation = now
            if rdv.date_en_attente:
                rdv.temps_attente_minutes = int((now - rdv.date_en_attente).total_seconds() / 60)
            rdv.duree_minutes = rdv.temps_constante_minutes + rdv.temps_attente_minutes + rdv.temps_consultation_minutes
            rdv._skip_auto_log = True
            rdv.save(update_fields=['statut', 'date_en_consultation', 'temps_attente_minutes', 'duree_minutes'])
            log_event(rdv, request.user, 'État : En Attente → En Consultation', type='statut')
            messages.success(request, 'Consultation démarrée.')
            from django.urls import reverse
            return redirect(reverse('patients:rdv_edit', kwargs={'pk': rdv.pk}))

        if action == 'terminer':
            from django.utils import timezone as tz
            now = tz.now()
            rdv.statut = 'termine'
            rdv.date_termine = now
            if rdv.date_en_consultation:
                rdv.temps_consultation_minutes = int((now - rdv.date_en_consultation).total_seconds() / 60)
            rdv.duree_minutes = rdv.temps_constante_minutes + rdv.temps_attente_minutes + rdv.temps_consultation_minutes
            rdv._skip_auto_log = True
            rdv.save(update_fields=['statut', 'date_termine', 'temps_consultation_minutes', 'duree_minutes'])
            log_event(rdv, request.user, 'État : En Consultation → Terminé', type='statut')
            messages.success(request, 'Consultation terminée.')
            return redirect('patients:rdv_global')

        if action == 'annuler':
            rdv.statut = 'annule'
            rdv._skip_auto_log = True
            rdv.save(update_fields=['statut'])
            log_event(rdv, request.user, 'Rendez-vous annulé.', type='statut')
            messages.success(request, 'Rendez-vous annulé.')
            return redirect('patients:rdv_global')

        form = RendezVousForm(request.POST, instance=rdv)
        if form.is_valid():
            rdv = form.save(commit=False)
            code = request.POST.get('code_confirmation', '').strip()
            if code:
                rdv.code_confirmation = code
            rdv._skip_auto_log = True
            rdv.save()
            log_event(rdv, request.user, 'Rendez-vous modifié.', type='modif')

            from patients.utils import save_registres
            save_registres(request, rdv)

            # Sauvegarder l'évaluation clinique si des champs sont remplis
            _eval_map = {
                'eval_poids': 'poids',
                'eval_taille': 'taille',
                'eval_temperature': 'temperature',
                'eval_tension_systolique': 'tension_systolique',
                'eval_tension_diastolique': 'tension_diastolique',
                'eval_tension_systolique_droite': 'tension_systolique_droite',
                'eval_tension_diastolique_droite': 'tension_diastolique_droite',
                'eval_pouls': 'pouls',
                'eval_frequence_respiratoire': 'frequence_respiratoire',
                'eval_saturation_oxygene': 'saturation_oxygene',
                'eval_glycemie': 'glycemie',
                'eval_albumine': 'albumine',
                'eval_perimetre_brachial': 'perimetre_brachial',
            }
            if any(request.POST.get(k, '').strip() for k in _eval_map):
                from consultations.models import Consultation as Consult, Constante as Const
                try:
                    consult_obj = rdv.consultation
                except Exception:
                    consult_obj = None
                if consult_obj is None:
                    consult_obj = Consult.objects.create(
                        patient=rdv.patient,
                        medecin=rdv.medecin,
                        rendez_vous=rdv,
                        motif=rdv.motif or 'Évaluation clinique',
                        cree_par=request.user,
                    )
                const_obj, _ = Const.objects.get_or_create(consultation=consult_obj)
                for post_key, model_field in _eval_map.items():
                    val = request.POST.get(post_key, '').strip()
                    if val != '':
                        setattr(const_obj, model_field, val)
                const_obj.save()

            messages.success(request, 'Rendez-vous modifié.')
            if action == 'créer une facture':
                from django.urls import reverse
                return redirect(reverse('facture_create') + f'?patient={rdv.patient.pk}&rdv={rdv.pk}')
            from django.urls import reverse
            return redirect(reverse('patients:rdv_edit', kwargs={'pk': rdv.pk}))
    else:
        form = RendezVousForm(instance=rdv)

    from patients.models import RegistreCPN, RegistreAccouchement, RegistrePostnatale, RegistreCuratif
    def _get_reg(Model):
        try:
            return Model.objects.get(rdv=rdv)
        except Model.DoesNotExist:
            return None

    return render(request, 'patients/rendez_vous_form.html', {
        'form':          form,
        'rdv':           rdv,
        'titre':         f'Rendez-vous — {rdv.patient.nom} {rdv.patient.prenoms}',
        'patient_prefill': rdv.patient,
        'facture_payee': facture_payee,
        'is_new':        False,
        'consultation':  consultation,
        'constante':     constante,
        'types_visite_curative': TypeVisiteCurative.objects.filter(actif=True).order_by('nom'),
        'pathologies':   Pathologie.objects.filter(actif=True, departement__code__in=('medg', 'MEDGEN')).order_by('nom'),
        'medecins':      Medecin.objects.filter(actif=True).select_related('employe').order_by('employe__nom', 'employe__prenoms'),
        'registre_cpn':          _get_reg(RegistreCPN),
        'registre_accouchement': _get_reg(RegistreAccouchement),
        'registre_postnatale':   _get_reg(RegistrePostnatale),
        'registre_curatif':      _get_reg(RegistreCuratif),
    })




# Les deux listes de gynécologie qui vivaient ici ont été retirées : le module a
# été refait dans core/views.py (`gynecologie_rdv` et `gynecologie_list`), qui en
# fait davantage — sélection élargie aux médecins de spécialité gynécologique,
# export Excel, et les menus générés depuis la déclaration de la liste. Leurs
# anciennes adresses renvoient désormais vers ces pages (cf. patients/urls.py) :
# ce fichier n'a plus à en garder une seconde version, qui divergeait déjà (ses
# menus Filtres et Regrouper par s'ouvraient sur du vide).


@login_required
def patient_rdv_list(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    items = patient.rendez_vous.select_related('medecin').order_by('-date_heure')
    return _render_related_list(request, {
        'patient': patient,
        'view_type': 'rdv',
        'titre': 'Rendez-vous',
        'items': items,
    })


@login_required
def patient_consultation_list(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    try:
        from consultations.models import Consultation
        items = Consultation.objects.filter(patient=patient).select_related('medecin').order_by('-date_heure')
    except Exception:
        items = []
    return _render_related_list(request, {
        'patient': patient,
        'view_type': 'consultation',
        'titre': 'Consultations',
        'items': items,
    })


@login_required
def patient_soin_list(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    try:
        from soins.models import Soin, ProcedureSoin
        from django.db.models import Prefetch
        items = Soin.objects.filter(patient=patient).prefetch_related(
            Prefetch(
                'procedures',
                queryset=ProcedureSoin.objects.select_related('infirmier', 'soin_type').order_by('date'),
                to_attr='procedures_list'
            )
        ).order_by('-date_heure')
    except Exception:
        items = []
    return _render_related_list(request, {
        'patient': patient,
        'view_type': 'soin',
        'titre': 'Soins infirmiers',
        'items': items,
    })


@login_required
def patient_ordonnance_list(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    try:
        from consultations.models import Ordonnance
        items = Ordonnance.objects.filter(
            consultation__patient=patient
        ).select_related('consultation').order_by('-date_emission')
    except Exception:
        items = []
    return _render_related_list(request, {
        'patient': patient,
        'view_type': 'ordonnance',
        'titre': 'Ordonnances',
        'items': items,
    })


@login_required
def patient_hospitalisation_list(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    try:
        from hospitalisation.models import Hospitalisation
        items = Hospitalisation.objects.filter(patient=patient).select_related(
            'medecin_traitant', 'chambre'
        ).order_by('-date_admission')
    except Exception:
        items = []
    return _render_related_list(request, {
        'patient': patient,
        'view_type': 'hospitalisation',
        'titre': 'Hospitalisations',
        'items': items,
    })


@login_required
def patient_demande_examens_list(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    try:
        from laboratoire.models import DemandeExamen
        items = DemandeExamen.objects.filter(patient=patient).prefetch_related('lignes').order_by('-date_creation')
    except Exception:
        items = []
    return _render_related_list(request, {
        'patient': patient,
        'view_type': 'demande_examens',
        'titre': "Demandes d'examens",
        'items': items,
    })


@login_required
def patient_resultat_examens_list(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    try:
        from laboratoire.models import AnalyseLaboratoire
        items = AnalyseLaboratoire.objects.filter(
            patient=patient, statut__in=['resultat', 'valide', 'envoye']
        ).order_by('-date_resultat')
    except Exception:
        items = []
    return _render_related_list(request, {
        'patient': patient,
        'view_type': 'resultat_examens',
        'titre': "Résultats d'examens de laboratoire",
        'items': items,
    })


@login_required
def ordonnance_create(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    from consultations.models import Consultation as Consult, Ordonnance, LigneOrdonnance

    consultation = None
    consultation_pk = request.GET.get('consultation') or request.POST.get('consultation_id')
    rdv_pk = request.GET.get('rdv') or request.POST.get('rdv_id')

    if consultation_pk:
        consultation = get_object_or_404(Consult, pk=consultation_pk)
    elif rdv_pk:
        rdv_obj = get_object_or_404(RendezVous, pk=rdv_pk)
        try:
            consultation = rdv_obj.consultation
        except Exception:
            consultation = Consult.objects.create(
                patient=patient,
                medecin=rdv_obj.medecin,
                rendez_vous=rdv_obj,
                motif=rdv_obj.motif or 'Consultation',
                cree_par=request.user,
            )

    if request.method == 'POST':
        if consultation is None:
            messages.error(request, "Impossible de créer une ordonnance sans consultation associée.")
            return redirect('patients:ordonnance_list', pk=pk)

        notes = request.POST.get('notes', '')
        date_expiration = request.POST.get('date_expiration') or None
        statut = request.POST.get('statut', 'emise')
        type_ordonnance = request.POST.get('type_ordonnance', 'interne')

        ordonnance = Ordonnance.objects.create(
            consultation=consultation,
            notes=notes,
            date_expiration=date_expiration,
            statut=statut,
            type_ordonnance=type_ordonnance,
        )

        medicaments = request.POST.getlist('medicament[]')
        medicaments_libres = request.POST.getlist('medicament_libre[]')
        posologies = request.POST.getlist('posologie[]')
        durees = request.POST.getlist('duree[]')
        quantites = request.POST.getlist('quantite[]')

        for i, posologie in enumerate(posologies):
            if not posologie.strip():
                continue
            med_id = medicaments[i] if i < len(medicaments) else ''
            med_libre = medicaments_libres[i] if i < len(medicaments_libres) else ''
            duree = durees[i] if i < len(durees) else ''
            quantite_val = quantites[i] if i < len(quantites) else '1'
            try:
                quantite = int(quantite_val)
            except (ValueError, TypeError):
                quantite = 1

            ligne = LigneOrdonnance(
                ordonnance=ordonnance,
                posologie=posologie,
                medicament_libre=med_libre,
                duree=duree,
                quantite=quantite,
            )
            if med_id:
                try:
                    ligne.medicament_id = int(med_id)
                except (ValueError, TypeError):
                    pass
            ligne.save()

        messages.success(request, f"Ordonnance {ordonnance.numero} créée avec succès.")
        return redirect('patients:ordonnance_list', pk=pk)

    try:
        from pharmacie.models import Medicament
        medicaments_dispo = list(Medicament.objects.filter(actif=True).values('pk', 'designation', 'dosage', 'forme'))
    except Exception:
        medicaments_dispo = []

    return render(request, 'pharmacie/ordonnance_create.html', {
        'patient': patient,
        'consultation': consultation,
        'medicaments_dispo': medicaments_dispo,
        'titre': 'Créer une ordonnance',
        'statuts': [('emise', 'Émise'), ('delivree', 'Délivrée'), ('partielle', 'Partielle'), ('expiree', 'Expirée')],
        'types': [('interne', 'Interne'), ('externe', 'Externe')],
    })


# ── Types de visite curative (configuration des rendez-vous) ────────────────
# Pendant de la configuration des types de visite gynécologiques (les CPN), mais
# rattaché au module Rendez-vous : ces types décrivent les consultations
# curatives (Consultant / Contrôle / Soins à l'origine).

@login_required
def typevisitecurative_list(request):
    qs = TypeVisiteCurative.objects.all()
    q  = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(code__icontains=q))

    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'patients/typevisitecurative_list.html', {
        'page_obj': page_obj,
        'q':        q,
        'total':    qs.count(),
    })


@login_required
def typevisitecurative_create(request):
    is_ajax = _is_ajax(request)
    if request.method == 'POST':
        form = TypeVisiteCurativeForm(request.POST)
        if form.is_valid():
            t = form.save()
            if is_ajax:
                return JsonResponse({'ok': True, 'message': f'Type de visite "{t.nom}" enregistré.'})
            messages.success(request, f'Type de visite "{t.nom}" enregistré.')
            return redirect('patients:typevisitecurative_list')
    else:
        form = TypeVisiteCurativeForm()
    template = ('patients/typevisitecurative_form_modal.html' if is_ajax
                else 'patients/typevisitecurative_form.html')
    return render(request, template, {
        'form': form, 'titre': 'Nouveau type de visite curative', 'edit': False,
    })


@login_required
def typevisitecurative_edit(request, pk):
    tvc = get_object_or_404(TypeVisiteCurative, pk=pk)
    is_ajax = _is_ajax(request)
    if request.method == 'POST':
        form = TypeVisiteCurativeForm(request.POST, instance=tvc)
        if form.is_valid():
            form.save()
            if is_ajax:
                return JsonResponse({'ok': True, 'message': 'Type de visite mis à jour.'})
            messages.success(request, 'Type de visite mis à jour.')
            return redirect('patients:typevisitecurative_list')
    else:
        form = TypeVisiteCurativeForm(instance=tvc)
    template = ('patients/typevisitecurative_form_modal.html' if is_ajax
                else 'patients/typevisitecurative_form.html')
    return render(request, template, {
        'form': form, 'titre': 'Modifier le type de visite', 'edit': True, 'object': tvc,
    })


@login_required
def typevisitecurative_delete(request, pk):
    tvc = get_object_or_404(TypeVisiteCurative, pk=pk)
    if request.method == 'POST':
        nom = tvc.nom
        # Les rendez-vous qui l'utilisent voient simplement leur type se vider
        # (SET_NULL) ; la valeur déjà écrite dans le registre curatif, elle, reste.
        tvc.delete()
        if _is_ajax(request):
            return JsonResponse({'ok': True, 'message': f'Type de visite "{nom}" supprimé.'})
        messages.success(request, f'Type de visite "{nom}" supprimé.')
    return redirect('patients:typevisitecurative_list')


@login_required
def pathologie_list(request):
    qs = Pathologie.objects.all()
    q  = request.GET.get('q', '').strip()

    if q:
        qs = qs.filter(nom__icontains=q)

    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'patients/pathologie_list.html', {
        'page_obj': page_obj,
        'q':        q,
        'total':    qs.count(),
    })


def _is_ajax(request):
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'


@login_required
def pathologie_create(request):
    is_ajax = _is_ajax(request)
    if request.method == 'POST':
        form = PathologieForm(request.POST)
        if form.is_valid():
            p = form.save()
            if is_ajax:
                return JsonResponse({'ok': True, 'message': f'Pathologie "{p.nom}" enregistrée.'})
            messages.success(request, f'Pathologie "{p.nom}" enregistrée.')
            return redirect('patients:pathologie_list')
    else:
        form = PathologieForm()
    template = 'patients/pathologie_form_modal.html' if is_ajax else 'patients/pathologie_form.html'
    return render(request, template, {
        'form': form, 'titre': 'Nouvelle pathologie', 'edit': False,
    })


@login_required
def pathologie_edit(request, pk):
    pathologie = get_object_or_404(Pathologie, pk=pk)
    is_ajax = _is_ajax(request)
    if request.method == 'POST':
        form = PathologieForm(request.POST, instance=pathologie)
        if form.is_valid():
            form.save()
            if is_ajax:
                return JsonResponse({'ok': True, 'message': 'Pathologie mise à jour.'})
            messages.success(request, 'Pathologie mise à jour.')
            return redirect('patients:pathologie_list')
    else:
        form = PathologieForm(instance=pathologie)
    template = 'patients/pathologie_form_modal.html' if is_ajax else 'patients/pathologie_form.html'
    return render(request, template, {
        'form': form, 'titre': 'Modifier la pathologie', 'edit': True, 'object': pathologie,
    })


@login_required
def pathologie_delete(request, pk):
    pathologie = get_object_or_404(Pathologie, pk=pk)
    if request.method == 'POST':
        nom = pathologie.nom
        pathologie.delete()
        if _is_ajax(request):
            return JsonResponse({'ok': True, 'message': f'Pathologie "{nom}" supprimée.'})
        messages.success(request, f'Pathologie "{nom}" supprimée.')
    return redirect('patients:pathologie_list')


# ── Export / Import des pathologies ─────────────────────────────────────────

_PATHOLOGIE_HDR = ['nom', 'categorie', 'departement', 'description', 'actif']


def _pathologie_row(p):
    return [p.nom, p.categorie, p.departement.code if p.departement_id else '', p.description, int(p.actif)]


@login_required
def export_pathologies(request):
    from core.utils import csv_response
    import json as _json
    from django.http import HttpResponse

    fmt = request.GET.get('format', 'json')
    qs = Pathologie.objects.select_related('departement').all()
    rows = [_pathologie_row(p) for p in qs]

    if fmt == 'csv':
        return csv_response('pathologies', _PATHOLOGIE_HDR, rows, delimiter=',')
    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io as _io
        wb = Workbook()
        ws = wb.active
        ws.title = 'Pathologies'
        fill = PatternFill(start_color='1F6E8C', end_color='1F6E8C', fill_type='solid')
        fnt = Font(color='FFFFFF', bold=True)
        ws.append(_PATHOLOGIE_HDR)
        for cell in ws[1]:
            cell.fill, cell.font = fill, fnt
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            ws.append(['' if v is None else v for v in row])
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 55)
        buf = _io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="pathologies.xlsx"'
        return resp

    data = [dict(zip(_PATHOLOGIE_HDR, r)) for r in rows]
    resp = HttpResponse(
        _json.dumps(data, ensure_ascii=False, indent=2, default=str),
        content_type='application/json',
    )
    resp['Content-Disposition'] = 'attachment; filename="pathologies.json"'
    return resp


def _parse_pathologie_upload(upload):
    import csv as _csv
    import io as _io
    import json as _json

    name = upload.name.lower()
    try:
        if name.endswith('.json'):
            return _json.loads(upload.read().decode('utf-8')), None
        if name.endswith('.csv'):
            text = upload.read().decode('utf-8-sig')
            reader = _csv.DictReader(_io.StringIO(text))
            return list(reader), None
        if name.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(upload.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return None, 'Fichier Excel vide.'
            hdrs = [str(h) if h is not None else '' for h in rows[0]]
            data = [dict(zip(hdrs, r)) for r in rows[1:] if any(v is not None for v in r)]
            return data, None
        return None, 'Format non supporté (.json, .csv ou .xlsx uniquement)'
    except Exception as e:
        return None, f'Erreur lecture fichier : {e}'


def _s(v):
    if v is None:
        return ''
    return str(v).replace('﻿', '').strip()


def _b(v):
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ('1', 'true', 'oui', 'yes')


@login_required
def import_pathologies(request):
    from medecins.models import Departement

    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('patients:pathologie_list')

    data, err = _parse_pathologie_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('patients:pathologie_list')

    do_update = 'update' in request.POST
    departements = {d.code: d for d in Departement.objects.all()}
    created = updated = skipped = errors = 0

    for item in data:
        try:
            nom = _s(item.get('nom', ''))
            if not nom:
                errors += 1
                continue
            departement_code = _s(item.get('departement', ''))
            defaults = {
                'categorie': _s(item.get('categorie', 'generale')) or 'generale',
                'departement': departements.get(departement_code),
                'description': _s(item.get('description', '')),
                'actif': _b(item.get('actif', True)),
            }
            obj, was_created = Pathologie.objects.get_or_create(nom=nom, defaults=defaults)
            if was_created:
                created += 1
            elif do_update:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                updated += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    if errors:
        messages.warning(request, f'{created} créée(s), {updated} mise(s) à jour, {skipped} ignorée(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} pathologie(s) importée(s), {updated} mise(s) à jour, {skipped} ignorée(s).')
    return redirect('patients:pathologie_list')


@login_required
def typevisite_list(request):
    qs = TypeVisite.objects.all()
    q  = request.GET.get('q', '').strip()

    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(code__icontains=q))

    paginator = Paginator(qs, 40)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'patients/typevisite_list.html', {
        'page_obj': page_obj,
        'q':        q,
        'total':    qs.count(),
    })


# Créer et modifier se font dans une modale, comme les pathologies et les types
# de visite curative : la liste reste à l'écran, on ne perd ni sa page ni sa
# recherche pour saisir deux champs. En AJAX la vue renvoie le seul fragment du
# formulaire, et du JSON quand l'enregistrement a réussi ; sans AJAX (lien ouvert
# dans un nouvel onglet, JavaScript indisponible) elle rend la page entière comme
# avant, si bien que les deux chemins restent praticables.

@login_required
def typevisite_create(request):
    is_ajax = _is_ajax(request)
    if request.method == 'POST':
        form = TypeVisiteForm(request.POST)
        if form.is_valid():
            tv = form.save()
            if is_ajax:
                return JsonResponse({'ok': True, 'message': f'Type de visite "{tv.nom}" enregistré.'})
            messages.success(request, f'Type de visite "{tv.nom}" enregistré.')
            return redirect('gynecologie_typevisite_list')
    else:
        form = TypeVisiteForm()
    template = ('patients/typevisite_form_modal.html' if is_ajax
                else 'patients/typevisite_form.html')
    return render(request, template, {
        'form': form, 'titre': 'Nouveau type de visite', 'edit': False,
    })


@login_required
def typevisite_edit(request, pk):
    tv = get_object_or_404(TypeVisite, pk=pk)
    is_ajax = _is_ajax(request)
    if request.method == 'POST':
        form = TypeVisiteForm(request.POST, instance=tv)
        if form.is_valid():
            form.save()
            if is_ajax:
                return JsonResponse({'ok': True, 'message': 'Type de visite mis à jour.'})
            messages.success(request, 'Type de visite mis à jour.')
            return redirect('gynecologie_typevisite_list')
    else:
        form = TypeVisiteForm(instance=tv)
    template = ('patients/typevisite_form_modal.html' if is_ajax
                else 'patients/typevisite_form.html')
    return render(request, template, {
        'form': form, 'titre': 'Modifier le type de visite', 'edit': True, 'object': tv,
    })


@login_required
def typevisite_delete(request, pk):
    tv = get_object_or_404(TypeVisite, pk=pk)
    if request.method == 'POST':
        nom = tv.nom
        tv.delete()
        if _is_ajax(request):
            return JsonResponse({'ok': True, 'message': f'Type de visite "{nom}" supprimé.'})
        messages.success(request, f'Type de visite "{nom}" supprimé.')
    return redirect('gynecologie_typevisite_list')


# ── Export / Import des types de visite ─────────────────────────────────────

_TYPEVISITE_HDR = ['nom', 'code', 'description', 'actif']


def _typevisite_row(tv):
    return [tv.nom, tv.code, tv.description, int(tv.actif)]


@login_required
def export_typevisite(request):
    import json as _json
    from django.http import HttpResponse

    fmt = request.GET.get('format', 'json')
    qs = TypeVisite.objects.all()
    rows = [_typevisite_row(tv) for tv in qs]

    if fmt == 'csv':
        from core.utils import csv_response
        return csv_response('types_visite', _TYPEVISITE_HDR, rows, delimiter=',')
    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io as _io
        wb = Workbook()
        ws = wb.active
        ws.title = 'Types de visite'
        fill = PatternFill(start_color='1F6E8C', end_color='1F6E8C', fill_type='solid')
        fnt = Font(color='FFFFFF', bold=True)
        ws.append(_TYPEVISITE_HDR)
        for cell in ws[1]:
            cell.fill, cell.font = fill, fnt
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            ws.append(['' if v is None else v for v in row])
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 55)
        buf = _io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="types_visite.xlsx"'
        return resp

    data = [dict(zip(_TYPEVISITE_HDR, r)) for r in rows]
    resp = HttpResponse(
        _json.dumps(data, ensure_ascii=False, indent=2, default=str),
        content_type='application/json',
    )
    resp['Content-Disposition'] = 'attachment; filename="types_visite.json"'
    return resp


@login_required
def import_typevisite(request):
    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('gynecologie_typevisite_list')

    data, err = _parse_pathologie_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('gynecologie_typevisite_list')

    do_update = 'update' in request.POST
    created = updated = skipped = errors = 0

    for item in data:
        try:
            nom = _s(item.get('nom', ''))
            code = _s(item.get('code', ''))
            if not nom or not code:
                errors += 1
                continue
            defaults = {
                'nom': nom,
                'description': _s(item.get('description', '')),
                'actif': _b(item.get('actif', True)),
            }
            obj, was_created = TypeVisite.objects.get_or_create(code__iexact=code, defaults={**defaults, 'code': code})
            if was_created:
                created += 1
            elif do_update:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                updated += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    if errors:
        messages.warning(request, f'{created} créé(s), {updated} mis à jour, {skipped} ignoré(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} type(s) de visite importé(s), {updated} mis à jour, {skipped} ignoré(s).')
    return redirect('gynecologie_typevisite_list')
