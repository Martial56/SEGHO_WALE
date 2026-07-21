from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme, urlencode
from django.urls import reverse
from django.db.models import F, Q
from datetime import timedelta

MAX_UNLOCK_ATTEMPTS = 5


def _safe_next_url(request, candidate):
    if candidate and url_has_allowed_host_and_scheme(
        candidate, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        return candidate
    return 'dashboard'


def _resolve_next_url(request, candidate):
    next_url = _safe_next_url(request, candidate)
    return reverse('dashboard') if next_url == 'dashboard' else next_url


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = _resolve_next_url(request, request.POST.get('next'))
            return redirect(f"{reverse('intro')}?{urlencode({'next': next_url})}")
        else:
            error = "Identifiant ou mot de passe incorrect."
    next_candidate = request.GET.get('next', '')
    return render(request, 'registration/login.html', {
        'error': error,
        'next': next_candidate if _safe_next_url(request, next_candidate) == next_candidate else '',
    })


@login_required(login_url='login')
def intro(request):
    return render(request, 'core/intro.html', {
        'next_url': _resolve_next_url(request, request.GET.get('next')),
    })


@require_POST
def logout_view(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
@require_POST
def lock_session(request):
    """Verrouille la session courante (appelé par le minuteur d'inactivité
    JS) sans détruire la session — voir core.middleware.SessionTimeoutMiddleware."""
    request.session['locked'] = True
    return JsonResponse({'ok': True})


@login_required(login_url='login')
def unlock_session(request):
    """Déverrouille la session après vérification du mot de passe de
    l'utilisateur déjà authentifié (pas de login()/logout() — la session et
    tout son contenu, y compris la page/le formulaire en cours côté client,
    restent intacts). Le nombre de tentatives est limité : au-delà, la session
    est entièrement détruite (retour à l'écran de connexion classique)."""
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    error = None
    if request.method == 'POST':
        password = request.POST.get('password', '')
        if request.user.check_password(password):
            request.session['locked'] = False
            request.session['last_activity'] = timezone.now().timestamp()
            request.session.pop('unlock_attempts', None)
            if is_ajax:
                return JsonResponse({'ok': True})
            return redirect(_safe_next_url(request, request.POST.get('next')))

        attempts = request.session.get('unlock_attempts', 0) + 1
        request.session['unlock_attempts'] = attempts
        if attempts >= MAX_UNLOCK_ATTEMPTS:
            logout(request)
            error = "Trop de tentatives incorrectes, vous avez été déconnecté."
            if is_ajax:
                return JsonResponse({'ok': False, 'logged_out': True, 'error': error}, status=403)
            messages.error(request, error)
            return redirect('login')

        error = f"Mot de passe incorrect ({MAX_UNLOCK_ATTEMPTS - attempts} tentative(s) restante(s))."
        if is_ajax:
            return JsonResponse({'ok': False, 'error': error}, status=400)

    next_candidate = request.GET.get('next', '')
    return render(request, 'registration/locked.html', {
        'error': error,
        'next': next_candidate if _safe_next_url(request, next_candidate) == next_candidate else '',
    })


@login_required(login_url='login')
def mon_compte(request):
    from django.contrib.auth import update_session_auth_hash
    from core.models import UserProfile
    user    = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)
    errors  = {}

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'profil':
            user.first_name = request.POST.get('first_name', '').strip()
            user.last_name  = request.POST.get('last_name', '').strip()
            user.email      = request.POST.get('email', '').strip()
            user.save()
            messages.success(request, 'Informations mises à jour.')
            return redirect('mon_compte')

        elif action == 'photo':
            photo = request.FILES.get('photo')
            if photo:
                if profile.photo:
                    profile.photo.delete(save=False)
                profile.photo = photo
                profile.save()
                messages.success(request, 'Photo de profil mise à jour.')
            return redirect('mon_compte')

        elif action == 'supprimer_photo':
            if profile.photo:
                profile.photo.delete(save=False)
                profile.photo = None
                profile.save()
                messages.success(request, 'Photo supprimée.')
            return redirect('mon_compte')

        elif action == 'session_timeout':
            try:
                minutes = int(request.POST.get('session_timeout_minutes', 30))
            except ValueError:
                minutes = 30
            profile.session_timeout_minutes = max(0, minutes)
            profile.save(update_fields=['session_timeout_minutes'])
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'ok': True, 'session_timeout_minutes': profile.session_timeout_minutes})
            messages.success(request, 'Délai de verrouillage mis à jour.')
            return redirect('mon_compte')

        elif action == 'password':
            current = request.POST.get('current_password', '')
            new_pw  = request.POST.get('new_password', '')
            confirm = request.POST.get('confirm_password', '')

            if not user.check_password(current):
                errors['password'] = 'Mot de passe actuel incorrect.'
            elif len(new_pw) < 8:
                errors['password'] = 'Le nouveau mot de passe doit contenir au moins 8 caractères.'
            elif new_pw != confirm:
                errors['password'] = 'Les deux mots de passe ne correspondent pas.'
            else:
                user.set_password(new_pw)
                user.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Mot de passe modifié avec succès.')
                return redirect('mon_compte')

    return render(request, 'utilisateur/mon_compte.html', {'errors': errors, 'profile': profile})


@login_required(login_url='login')
@require_POST
def accent_color_set(request):
    from core.models import UserProfile
    from core.utils import build_accent_css, is_valid_hex_color, DEFAULT_ACCENT_COLOR

    color = (request.POST.get('color') or '').strip()
    profile, _ = UserProfile.objects.get_or_create(user=request.user)

    if not color:
        profile.accent_color = None
        profile.save(update_fields=['accent_color'])
        return JsonResponse({'ok': True, 'color': None, 'default': DEFAULT_ACCENT_COLOR})

    if not is_valid_hex_color(color):
        return JsonResponse({'ok': False, 'error': 'Couleur invalide.'}, status=400)

    profile.accent_color = color
    profile.save(update_fields=['accent_color'])
    return JsonResponse({'ok': True, 'color': color, 'css': build_accent_css(color)})


def _get_dashboard_stats():
    from patients.models import Patient, RendezVous
    from consultations.models import Consultation
    from hospitalisation.models import Hospitalisation
    from pharmacie.models import Medicament
    from facturation.models import Facture
    from laboratoire.models import AnalyseLaboratoire
    from employer.models import Employe
    from soins.models import Soin

    today = timezone.now().date()
    return {
        'patients_total': Patient.objects.count(),
        'patients_today': Patient.objects.filter(date_creation__date=today).count(),
        'consultations_today': Consultation.objects.filter(date_heure__date=today).count(),
        'soins_aujourd_hui': Soin.objects.filter(date_creation__date=today).count(),
        'rdv_today': RendezVous.objects.filter(date_heure__date=today).count(),
        'hospitalisations': Hospitalisation.objects.filter(statut__in=['confirme', 'hospitalise']).count(),
        'analyses_pending': AnalyseLaboratoire.objects.filter(statut__in=['recu', 'en_analyse']).count(),
        'factures_impayees': Facture.objects.filter(statut='emise').count(),
        'medicaments_alerte': Medicament.objects.filter(stock_actuel__lte=F('stock_alerte')).count(),
        'employes_actifs': Employe.objects.filter(statut='actif').count(),
    }


@login_required(login_url='login')
def dashboard_stats_json(request):
    return JsonResponse(_get_dashboard_stats())


@login_required(login_url='login')
def dashboard(request):
    from patients.models import RendezVous
    from consultations.models import Consultation

    today = timezone.now().date()
    stats = _get_dashboard_stats()

    rdv_auj = RendezVous.objects.filter(
        date_heure__date=today,
        statut__in=['planifie', 'confirme', 'en_attente', 'en_consultation']
    ).select_related('patient', 'medecin').order_by('date_heure')[:8]

    last_cons = Consultation.objects.select_related(
        'patient', 'medecin'
    ).order_by('-date_heure')[:6]

    try:
        from modules_permissions.models import get_user_modules
        user_modules = get_user_modules(request.user)
        accessible_codes = set(user_modules.values_list('code', flat=True))
    except Exception:
        user_modules = []
        accessible_codes = set()

    # Plannings publiés non encore vus par cet utilisateur
    from planning.models import PlanningHebdomadaire, PlanningVu
    plannings_non_vus = PlanningHebdomadaire.objects.filter(
        publie=True
    ).exclude(
        vus_par__user=request.user
    ).count()

    response = render(request, 'core/dashboard.html', {
        'stats': stats,
        'rdv_auj': rdv_auj,
        'last_cons': last_cons,
        'today': today,
        'user': request.user,
        'user_modules': user_modules,
        'accessible_codes': accessible_codes,
        'plannings_non_vus': plannings_non_vus,
    })
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    response['Pragma'] = 'no-cache'
    return response


@login_required(login_url='login')
def patients_list(request):
    from patients.models import Patient
    from django.core.paginator import Paginator

    patients = Patient.objects.all().order_by('-date_creation')
    paginator = Paginator(patients, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    today = timezone.now().date()
    stats = {
        'total': Patient.objects.count(),
        'nouveaux_30j': Patient.objects.filter(date_creation__date__gte=today - timedelta(days=30)).count(),
    }
    breadcrumb = [{'title': 'Accueil', 'url': '/'}, {'title': 'Patients'}]
    return render(request, 'patients/list.html', {'page_obj': page_obj, 'stats': stats, 'breadcrumb': breadcrumb})


@login_required(login_url='login')
def medecins_list(request):
    from medecins.models import Medecin, Specialite
    from django.core.paginator import Paginator

    qs = Medecin.objects.select_related('specialite', 'employe').order_by('employe__nom')

    q          = request.GET.get('q', '').strip()
    specialite = request.GET.get('specialite', '')
    statut     = request.GET.get('statut', '')
    vue        = request.GET.get('vue', '')

    if q:
        qs = qs.filter(
            Q(employe__nom__icontains=q) | Q(employe__prenoms__icontains=q) | Q(employe__matricule__icontains=q)
        )
    if specialite:
        qs = qs.filter(specialite__pk=specialite)
    if statut == 'actif':
        qs = qs.filter(actif=True)
    elif statut == 'inactif':
        qs = qs.filter(actif=False)

    total       = Medecin.objects.count()
    actifs      = Medecin.objects.filter(actif=True).count()
    specialites = Specialite.objects.all().order_by('nom')

    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))

    # Groupement par spécialité pour la vue Kanban
    kanban_colonnes = []
    for sp in specialites:
        medecins_sp = [m for m in qs if m.specialite_id == sp.pk]
        if medecins_sp:
            kanban_colonnes.append({'titre': sp.nom, 'medecins': medecins_sp})
    sans_spec = [m for m in qs if m.specialite_id is None]
    if sans_spec:
        kanban_colonnes.append({'titre': 'Sans spécialité', 'medecins': sans_spec})

    return render(request, 'medecins/list.html', {
        'page_obj':         page_obj,
        'specialites':      specialites,
        'kanban_colonnes':  kanban_colonnes,
        'stats': {'total': total, 'actifs': actifs, 'specialites': specialites.count()},
        'q':                q,
        'specialite_filtre': specialite,
        'statut_filtre':    statut,
        'vue_active':       vue,
    })


@login_required(login_url='login')
def medecin_lookup_employe(request):
    """Recherche AJAX d'un employé par matricule — utilisée par le formulaire de
    création de médecin pour charger et afficher ses informations (lecture seule) :
    un médecin est avant tout un employé, créé une seule fois dans le module RH."""
    from employer.models import Employe

    matricule = request.GET.get('matricule', '').strip()
    if not matricule:
        return JsonResponse({'ok': False})

    employe = Employe.objects.filter(matricule=matricule).first()
    if not employe:
        return JsonResponse({'ok': False, 'message': 'Aucun employé trouvé avec ce matricule.'})
    if hasattr(employe, 'fiche_medecin'):
        return JsonResponse({'ok': False, 'message': f'{employe.nom_complet} est déjà enregistré(e) comme médecin.'})

    return JsonResponse({
        'ok': True,
        'nom': employe.nom,
        'prenoms': employe.prenoms,
        'telephone': employe.telephone or '—',
        'email': employe.email or '—',
        'service': employe.service.nom if employe.service else '—',
        'photo_url': employe.photo.url if employe.photo else '',
    })


@login_required(login_url='login')
def medecin_create(request):
    from medecins.models import Medecin, Specialite, Service, Departement
    from employer.models import Employe
    from django.contrib.auth.models import User

    specialites = Specialite.objects.order_by('nom')
    departements = Departement.objects.filter(actif=True).order_by('nom')
    services = Service.objects.filter(actif=True).order_by('nom')
    users_disponibles = User.objects.filter(medecin__isnull=True).order_by('last_name', 'first_name')
    errors = {}
    employe_trouve = None

    if request.method == 'POST':
        matricule = request.POST.get('matricule', '').strip()
        specialite_pk = request.POST.get('specialite', '')
        departement_pk = request.POST.get('departement', '')
        service_pk = request.POST.get('service', '')
        taux_honoraire = request.POST.get('taux_honoraire', '0').strip() or '0'
        actif = request.POST.get('actif') == 'on'
        user_pk = request.POST.get('user', '')

        if not matricule:
            errors['matricule'] = 'Le matricule employé est obligatoire.'
        else:
            employe_trouve = Employe.objects.filter(matricule=matricule).first()
            if not employe_trouve:
                errors['matricule'] = 'Aucun employé trouvé avec ce matricule.'
            elif hasattr(employe_trouve, 'fiche_medecin'):
                errors['matricule'] = f'{employe_trouve.nom_complet} est déjà enregistré(e) comme médecin.'

        if not specialite_pk:
            errors['specialite'] = 'La spécialité est obligatoire.'
        elif not Specialite.objects.filter(pk=specialite_pk).exists():
            errors['specialite'] = 'Spécialité invalide.'
        if not departement_pk:
            errors['departement'] = 'Le département est obligatoire.'
        elif not Departement.objects.filter(pk=departement_pk).exists():
            errors['departement'] = 'Département invalide.'

        if not errors:
            annee = timezone.now().year
            dernier_ord = Medecin.objects.filter(ordre_medecin__startswith=f'ORD{annee}').order_by('-ordre_medecin').first()
            if dernier_ord:
                try:
                    seq_ord = int(dernier_ord.ordre_medecin[-4:]) + 1
                except ValueError:
                    seq_ord = 1
            else:
                seq_ord = 1
            ordre_medecin = f'ORD{annee}{seq_ord:04d}'

            service_obj = Service.objects.filter(pk=service_pk).first() if service_pk else employe_trouve.service

            med = Medecin(employe=employe_trouve, ordre_medecin=ordre_medecin, actif=actif, service=service_obj)
            try:
                med.taux_honoraire = float(taux_honoraire)
            except ValueError:
                med.taux_honoraire = 0

            if specialite_pk:
                try:
                    med.specialite = Specialite.objects.get(pk=specialite_pk)
                except Specialite.DoesNotExist:
                    pass

            med.departement = Departement.objects.filter(pk=departement_pk).first() if departement_pk else None

            if user_pk:
                try:
                    med.user = User.objects.get(pk=user_pk)
                except User.DoesNotExist:
                    pass

            med.save()
            messages.success(request, f'Médecin {med} enregistré avec succès (matricule : {med.matricule}).')
            return redirect('medecins_list')

        post_data = request.POST
    else:
        post_data = None

    return render(request, 'medecins/form.html', {
        'mode': 'create',
        'specialites': specialites,
        'departements': departements,
        'services': services,
        'users_disponibles': users_disponibles,
        'errors': errors,
        'post': post_data,
        'employe_trouve': employe_trouve,
    })


@login_required(login_url='login')
def medecin_edit(request, pk):
    from medecins.models import Medecin, Specialite, Service, Departement
    from django.contrib.auth.models import User

    med = get_object_or_404(Medecin, pk=pk)
    specialites = Specialite.objects.order_by('nom')
    departements = Departement.objects.filter(actif=True).order_by('nom')
    services = Service.objects.filter(actif=True).order_by('nom')
    users_disponibles = User.objects.filter(
        Q(medecin__isnull=True) | Q(medecin=med)
    ).order_by('last_name', 'first_name')
    errors = {}

    if request.method == 'POST':
        specialite_pk = request.POST.get('specialite', '')
        departement_pk = request.POST.get('departement', '')
        service_pk = request.POST.get('service', '')
        taux_honoraire = request.POST.get('taux_honoraire', '0').strip() or '0'
        actif = request.POST.get('actif') == 'on'
        user_pk = request.POST.get('user', '')

        if not specialite_pk:
            errors['specialite'] = 'La spécialité est obligatoire.'
        elif not Specialite.objects.filter(pk=specialite_pk).exists():
            errors['specialite'] = 'Spécialité invalide.'
        if not departement_pk:
            errors['departement'] = 'Le département est obligatoire.'
        elif not Departement.objects.filter(pk=departement_pk).exists():
            errors['departement'] = 'Département invalide.'

        if not errors:
            med.actif = actif
            try:
                med.taux_honoraire = float(taux_honoraire)
            except ValueError:
                med.taux_honoraire = 0

            med.specialite = Specialite.objects.filter(pk=specialite_pk).first() if specialite_pk else None
            med.departement = Departement.objects.filter(pk=departement_pk).first() if departement_pk else None
            med.service = Service.objects.filter(pk=service_pk).first() if service_pk else None

            if user_pk:
                try:
                    med.user = User.objects.get(pk=user_pk)
                except User.DoesNotExist:
                    med.user = None
            else:
                med.user = None

            med.save()
            messages.success(request, f'Médecin {med} mis à jour avec succès.')
            return redirect('medecins_list')

    return render(request, 'medecins/form.html', {
        'mode': 'edit',
        'med': med,
        'specialites': specialites,
        'departements': departements,
        'services': services,
        'users_disponibles': users_disponibles,
        'errors': errors,
    })


@login_required(login_url='login')
def medecin_detail(request, pk):
    from medecins.models import Medecin
    from consultations.models import Consultation
    from patients.models import RendezVous

    med = get_object_or_404(Medecin, pk=pk)

    today = timezone.now().date()
    consultations_recentes = (
        Consultation.objects
        .filter(medecin=med)
        .select_related('patient')
        .order_by('-date_heure')[:10]
    )
    rdv_a_venir = (
        RendezVous.objects
        .filter(medecin=med, date_heure__date__gte=today, statut__in=['planifie', 'confirme', 'en_attente'])
        .select_related('patient')
        .order_by('date_heure')[:5]
    )
    stats = {
        'total_consultations': Consultation.objects.filter(medecin=med).count(),
        'consultations_mois':  Consultation.objects.filter(
            medecin=med,
            date_heure__month=today.month,
            date_heure__year=today.year,
        ).count(),
        'rdv_total':   RendezVous.objects.filter(medecin=med).count(),
        'rdv_a_venir': RendezVous.objects.filter(
            medecin=med, date_heure__date__gte=today,
            statut__in=['planifie', 'confirme', 'en_attente']
        ).count(),
    }
    return render(request, 'medecins/detail.html', {
        'med': med,
        'consultations_recentes': consultations_recentes,
        'rdv_a_venir': rdv_a_venir,
        'stats': stats,
        'today': today,
    })


@login_required(login_url='login')
def consultations_list(request):
    from consultations.models import Consultation
    from django.core.paginator import Paginator

    consultations = Consultation.objects.all().order_by('-date_heure')
    paginator = Paginator(consultations, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    today = timezone.now().date()
    stats = {
        'aujourdhui': Consultation.objects.filter(date_heure__date=today).count(),
        'ce_mois': Consultation.objects.filter(date_heure__month=today.month, date_heure__year=today.year).count(),
        'rdv_venir': 0,
        'diagnostic_rate': 0,
    }
    breadcrumb = [{'title': 'Accueil', 'url': '/'}, {'title': 'Consultations'}]
    return render(request, 'consultations/list.html', {'page_obj': page_obj, 'stats': stats, 'breadcrumb': breadcrumb})


@login_required(login_url='login')
def pharmacie_list(request):
    from pharmacie.views import pharmacie_accueil
    return pharmacie_accueil(request)


@login_required(login_url='login')
def laboratoire_list(request):
    from laboratoire.models import DemandeExamen
    from django.core.paginator import Paginator

    q      = request.GET.get('q', '').strip()
    statut = request.GET.get('statut', '')

    statut_choices = DemandeExamen.STATUT
    statut_label   = dict(statut_choices).get(statut, statut)

    demandes = DemandeExamen.objects.select_related(
        'patient', 'technicien', 'facture'
    ).prefetch_related('lignes').order_by('-date_creation')

    if q:
        demandes = demandes.filter(
            Q(numero__icontains=q) |
            Q(patient__nom__icontains=q) |
            Q(patient__prenoms__icontains=q) |
            Q(patient__code_patient__icontains=q) |
            Q(lignes__libelle__icontains=q)
        ).distinct()

    if statut:
        demandes = demandes.filter(statut=statut)

    now = timezone.now()
    stats = {
        'en_attente': DemandeExamen.objects.filter(statut__in=['brouillon', 'demande']).count(),
        'en_cours':   DemandeExamen.objects.filter(statut__in=['accepte', 'en_cours']).count(),
        'terminees':  DemandeExamen.objects.filter(statut='termine').count(),
        'ce_mois':    DemandeExamen.objects.filter(date_creation__month=now.month, date_creation__year=now.year).count(),
    }

    paginator   = Paginator(demandes, 25)
    page_number = request.GET.get('page')
    page_obj    = paginator.get_page(page_number)

    breadcrumb = [{'title': 'Accueil', 'url': '/'}, {'title': 'Laboratoire'}]
    return render(request, 'laboratoire/list.html', {
        'page_obj':       page_obj,
        'stats':          stats,
        'q':              q,
        'statut':         statut,
        'statut_label':   statut_label,
        'statut_choices': statut_choices,
        'total':          page_obj.paginator.count,
        'breadcrumb':     breadcrumb,
    })


@login_required(login_url='login')
def facturation_list(request):
    from facturation.models import Facture
    from django.core.paginator import Paginator

    q = request.GET.get('q', '').strip()
    factures = Facture.objects.all().order_by('-date_emission')
    if q:
        factures = factures.filter(
            Q(numero_facture__icontains=q) |
            Q(patient__nom__icontains=q) |
            Q(patient__prenoms__icontains=q)
        )
    total = factures.count()
    paginator = Paginator(factures, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    stats = {'montant_total': 0, 'montant_recu': 0, 'montant_attente': 0, 'taux_recouvrement': 0}
    breadcrumb = [{'title': 'Accueil', 'url': '/'}, {'title': 'Facturation'}]
    return render(request, 'facturation/list.html', {'page_obj': page_obj, 'stats': stats, 'q': q, 'total': total, 'breadcrumb': breadcrumb})


@login_required(login_url='login')
def caisse_list(request):
    from django.core.exceptions import PermissionDenied
    from caisse.models import SessionCaisse
    from django.core.paginator import Paginator
    from facturation.views import can_manage_paiement

    if not can_manage_paiement(request.user):
        raise PermissionDenied

    sessions = SessionCaisse.objects.all().order_by('-date_ouverture')
    paginator = Paginator(sessions, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    stats = {'solde_actuel': 0, 'entrees_jour': 0, 'sorties_jour': 0, 'transactions': 0}
    breadcrumb = [{'title': 'Accueil', 'url': '/'}, {'title': 'Caisse'}]
    return render(request, 'caisse/list.html', {'page_obj': page_obj, 'stats': stats, 'breadcrumb': breadcrumb})


@login_required(login_url='login')
def ressources_humaines_list(request):
    from employer.models import Employe
    from django.core.paginator import Paginator

    employes = Employe.objects.all().order_by('nom')
    paginator = Paginator(employes, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    stats = {'total_employes': Employe.objects.count(), 'employes_actifs': Employe.objects.filter(actif=True).count(), 'conges_attente': 0, 'taux_presence': 0}
    breadcrumb = [{'title': 'Accueil', 'url': '/'}, {'title': 'Ressources Humaines'}]
    return render(request, 'employer/list.html', {'page_obj': page_obj, 'stats': stats, 'breadcrumb': breadcrumb})


@login_required(login_url='login')
def facture_create(request):
    from facturation.models import Facture, LigneFacture, Acte, Paiement
    from facturation.forms import FactureForm
    from patients.models import Patient, RendezVous
    from caisse.models import Caisse
    from django.urls import reverse

    patient_pk = request.GET.get('patient') or request.POST.get('patient_id')
    patient = get_object_or_404(Patient, pk=patient_pk) if patient_pk else None
    actes = Acte.objects.filter(actif=True).order_by('categorie', 'libelle')
    caisses = Caisse.objects.filter(actif=True).order_by('nom')
    from services.models import Articleservice
    services = Articleservice.objects.all().order_by('nom')

    demande_pk = request.GET.get('demande') or request.POST.get('demande_id')
    demande_obj = None
    if demande_pk:
        from laboratoire.models import DemandeExamen
        try:
            demande_obj = DemandeExamen.objects.get(pk=demande_pk)
        except DemandeExamen.DoesNotExist:
            pass

    rdv_obj = None
    rdv_pk = request.GET.get('rdv') or request.POST.get('rdv_id')
    if rdv_pk:
        try:
            rdv_obj = RendezVous.objects.get(pk=rdv_pk)
        except RendezVous.DoesNotExist:
            pass

    ordonnance_pk = request.GET.get('ordonnance') or request.POST.get('ordonnance_id')
    ordonnance_obj = None
    if ordonnance_pk:
        from consultations.models import Ordonnance
        try:
            ordonnance_obj = Ordonnance.objects.prefetch_related('lignes__produit', 'lignes__medicament').get(pk=ordonnance_pk)
            if patient is None and ordonnance_obj.consultation:
                patient = ordonnance_obj.consultation.patient
            if patient is None and ordonnance_obj.patient_id:
                patient = ordonnance_obj.patient
        except Ordonnance.DoesNotExist:
            pass

    if ordonnance_obj:
        initial_type_facture = 'pharmacie'
        initial_ligne_libelle = ''
        initial_lignes = []
        for ligne in ordonnance_obj.lignes.select_related('produit', 'medicament').all():
            if ligne.produit:
                libelle = ligne.produit.nom
                prix = float(ligne.produit.prix_vente)
            elif ligne.medicament:
                libelle = ligne.medicament.designation
                prix = float(ligne.medicament.prix_vente)
            elif ligne.medicament_libre:
                libelle = ligne.medicament_libre
                prix = 0
            else:
                continue
            initial_lignes.append({'libelle': libelle, 'prix': prix, 'qte': int(ligne.quantite)})
    elif rdv_obj:
        initial_type_facture = 'consultation'
        initial_ligne_libelle = rdv_obj.type_consultation.nom if rdv_obj.type_consultation else ''
        initial_lignes = []
    elif demande_obj:
        initial_type_facture = 'laboratoire'
        initial_ligne_libelle = ''
        initial_lignes = [{'libelle': lg.libelle, 'prix': float(lg.prix), 'qte': 1} for lg in demande_obj.lignes.all()]
    else:
        initial_type_facture = ''
        initial_ligne_libelle = ''
        initial_lignes = []

    if request.method == 'POST':
        form = FactureForm(request.POST)
        if not patient:
            messages.error(request, 'Patient introuvable.')
            return redirect('facturation_list')
        if form.is_valid():
            facture = form.save(commit=False)
            facture.patient = patient
            facture.cree_par = request.user
            facture.rendez_vous = rdv_obj
            facture.save()

            total = 0
            i = 0
            while True:
                libelle = request.POST.get(f'ligne_libelle_{i}')
                if libelle is None:
                    break
                if libelle.strip():
                    try:
                        qte = float(request.POST.get(f'ligne_qte_{i}', 1) or 1)
                        prix = float(request.POST.get(f'ligne_prix_{i}', 0) or 0)
                        remise = float(request.POST.get(f'ligne_remise_{i}', 0) or 0)
                    except ValueError:
                        qte, prix, remise = 1, 0, 0
                    ligne = LigneFacture(
                        facture=facture,
                        libelle=libelle.strip(),
                        quantite=qte,
                        prix_unitaire=prix,
                        remise=remise,
                    )
                    acte_id = request.POST.get(f'ligne_acte_{i}')
                    if acte_id:
                        try:
                            ligne.acte_id = int(acte_id)
                        except ValueError:
                            pass
                    ligne.save()
                    total += qte * prix * (1 - remise / 100)
                i += 1

            facture.montant_total = total
            facture.save()

            pay_montant_raw = request.POST.get('pay_montant', '').strip()
            if pay_montant_raw:
                try:
                    pay_montant = float(pay_montant_raw)
                except ValueError:
                    pay_montant = 0
                if pay_montant > 0:
                    mode = request.POST.get('pay_mode', 'especes')
                    memo = request.POST.get('pay_memo', '')
                    compte = request.POST.get('pay_compte', '')
                    reference = compte if compte else memo
                    paiement = Paiement(
                        facture=facture,
                        montant=pay_montant,
                        mode_paiement=mode,
                        reference=reference,
                        notes=memo,
                        recu_par=request.user,
                    )
                    paiement.save()
                    facture.montant_paye = pay_montant
                    facture.statut = 'payee' if pay_montant >= total else 'partielle'
                    facture.save()

            if demande_obj:
                demande_obj.facture = facture
                demande_obj.save(update_fields=['facture'])

            if ordonnance_obj:
                facture.ordonnance = ordonnance_obj
                facture.save(update_fields=['ordonnance'])

            messages.success(request, f'Facture {facture.numero} créée avec succès.')
            if demande_obj:
                return redirect('laboratoire_detail', pk=demande_obj.pk)
            if ordonnance_pk:
                return redirect('ordonnance_detail', pk=ordonnance_pk)
            if rdv_pk:
                return redirect(reverse('patients:rdv_edit', kwargs={'pk': rdv_pk}))
            return redirect('facturation_list')
    else:
        initial = {'type_facture': initial_type_facture} if initial_type_facture else {}
        form = FactureForm(initial=initial)

    back_url = (
        reverse('ordonnance_detail', kwargs={'pk': ordonnance_pk}) if ordonnance_pk else
        f'/laboratoire/{demande_pk}/' if demande_pk else
        ''
    )

    return render(request, 'facturation/create_facture.html', {
        'form': form,
        'patient': patient,
        'actes': actes,
        'services': services,
        'caisses': caisses,
        'rdv': rdv_obj,
        'demande': demande_obj,
        'ordonnance': ordonnance_obj,
        'initial_ligne_libelle': initial_ligne_libelle,
        'initial_lignes': initial_lignes,
        'back_url': back_url,
        'breadcrumb': [
            {'title': 'Accueil', 'url': '/'},
            {'title': 'Facturation', 'url': '/facturation/'},
            {'title': 'Nouvelle facture'},
        ],
    })


@login_required(login_url='login')
def laboratoire_create(request):
    from laboratoire.models import DemandeExamen, LigneDemandeExamen, TypeExamen
    from patients.models import Patient
    from django.contrib.auth.models import User
    from django.utils.dateparse import parse_datetime

    patient_pk = request.GET.get('patient') or request.POST.get('patient_id')
    patient = get_object_or_404(Patient, pk=patient_pk) if patient_pk else None

    from medecins.models import Medecin
    medecins = Medecin.objects.filter(actif=True).select_related('employe', 'user').order_by('employe__nom')

    prefill_medecin = None
    medecin_pk = request.GET.get('medecin_id')
    if medecin_pk:
        try:
            prefill_medecin = int(medecin_pk)
        except (ValueError, TypeError):
            pass
    from services.models import Articleservice
    services_examens = Articleservice.objects.filter(
        categorie__code='EX', actif=True
    ).order_by('nom')

    if request.method == 'POST':
        if not patient:
            messages.error(request, 'Patient requis.')
            return redirect('laboratoire_list')

        action = request.POST.get('action', 'brouillon')
        demande = DemandeExamen(
            patient=patient,
            type_test=request.POST.get('type_test', ''),
            urgent=request.POST.get('urgent') == 'on',
            commentaire=request.POST.get('commentaire', '').strip(),
            statut='demande' if action == 'envoyer' else 'brouillon',
            cree_par=request.user,
        )
        med_id = request.POST.get('medecin_prescripteur')
        if med_id:
            try:
                demande.medecin_prescripteur_id = int(med_id)
            except ValueError:
                pass
        raw_date = request.POST.get('date_prelevement', '').strip()
        if raw_date:
            demande.date_prelevement = parse_datetime(raw_date)
        demande.save()

        total = 0
        i = 0
        while True:
            svc_nom = request.POST.get(f'ligne_examen_{i}')
            if svc_nom is None:
                break
            if svc_nom.strip():
                try:
                    prix = float(request.POST.get(f'ligne_prix_{i}', 0) or 0)
                except ValueError:
                    prix = 0
                LigneDemandeExamen.objects.create(
                    demande=demande,
                    libelle=svc_nom.strip(),
                    prix=prix,
                    instructions=request.POST.get(f'ligne_instructions_{i}', '').strip(),
                )
                total += prix
            i += 1

        demande.montant_total = total
        demande.save()
        messages.success(request, f'Demande {demande.numero} créée avec succès.')
        return redirect('laboratoire_detail', pk=demande.pk)

    return render(request, 'laboratoire/create_analyse.html', {
        'patient': patient,
        'services_examens': services_examens,
        'medecins': medecins,
        'prefill_medecin': prefill_medecin,
        'type_test_choices': DemandeExamen.TYPE_TEST,
        'breadcrumb': [
            {'title': 'Accueil', 'url': '/'},
            {'title': 'Laboratoire', 'url': '/laboratoire/'},
            {'title': 'Nouvelle demande'},
        ],
    })


@login_required(login_url='login')
def laboratoire_detail(request, pk):
    from laboratoire.models import DemandeExamen

    demande = get_object_or_404(DemandeExamen, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'envoyer_labo' and demande.facture and demande.statut == 'brouillon':
            from laboratoire.hprim.services import envoyer_demande
            try:
                echange = envoyer_demande(demande)
                if echange.statut == 'transmis':
                    messages.success(request, f'Demande {demande.numero} transmise au laboratoire.')
                elif echange.statut == 'en_attente':
                    messages.warning(request, f'Demande {demande.numero} enregistrée (FTP non configuré) : {echange.message_log}')
                else:
                    messages.error(request, f'Échec de l\'envoi : {echange.message_log}')
            except RuntimeError as exc:
                messages.error(request, str(exc))
        return redirect('laboratoire_detail', pk=pk)

    lignes = demande.lignes.select_related('type_examen').all()
    return render(request, 'laboratoire/detail_demande.html', {
        'demande': demande,
        'lignes': lignes,
        'facture_url': f'/facturation/nouvelle/?patient={demande.patient_id}&demande={demande.pk}&back=/laboratoire/{demande.pk}/',
        'breadcrumb': [
            {'title': 'Accueil', 'url': '/'},
            {'title': 'Laboratoire', 'url': '/laboratoire/'},
            {'title': demande.numero},
        ],
    })


@login_required(login_url='login')
def gynecologie_naissance_create(request):
    from patients.models import Naissance, Patient
    from medecins.models import Medecin

    patients = Patient.objects.all().order_by('nom')
    medecins = Medecin.objects.all().order_by('employe__nom')
    error = None

    if request.method == 'POST':
        from django.utils.dateparse import parse_datetime
        action = request.POST.get('action', 'brouillon')
        try:
            raw_date = request.POST.get('date_accouchement', '')
            date_acc = parse_datetime(raw_date) if raw_date else timezone.now()
            if date_acc is None:
                date_acc = timezone.now()
            n = Naissance(
                mere_id=request.POST.get('mere') or None,
                medecin_id=request.POST.get('medecin') or None,
                date_accouchement=date_acc,
                lieu_naissance=request.POST.get('lieu_naissance', ''),
                mode_accouchement=request.POST.get('mode_accouchement', 'voie_basse'),
                nom_enfant=request.POST.get('nom_enfant', ''),
                prenoms_enfant=request.POST.get('prenoms_enfant', ''),
                sexe_enfant=request.POST.get('sexe_enfant') or 'F',
                poids_naissance=request.POST.get('poids_naissance') or None,
                groupe_sanguin_enfant=request.POST.get('groupe_sanguin_enfant', ''),
                taux_hemoglobine=request.POST.get('taux_hemoglobine') or None,
                taille_naissance=request.POST.get('taille_naissance') or None,
                apgar_1min=request.POST.get('apgar_1min') or None,
                apgar_5min=request.POST.get('apgar_5min') or None,
                statut=request.POST.get('statut', 'vivant'),
                info_parents=request.POST.get('info_parents', ''),
                education_mere=request.POST.get('education_mere', ''),
                age_mere=request.POST.get('age_mere') or None,
                parite=int(request.POST.get('parite') or 0),
                nombre_garcons=int(request.POST.get('nombre_garcons') or 0),
                nombre_filles=int(request.POST.get('nombre_filles') or 0),
                remarques=request.POST.get('remarques', ''),
                statut_dossier='termine' if action == 'termine' else 'brouillon',
            )
            n.save()
            return redirect('gynecologie_naissances')
        except Exception as e:
            error = str(e)

    breadcrumb = [
        {'title': 'Accueil', 'url': '/'},
        {'title': 'Gynécologie', 'url': '/gynecologie/'},
        {'title': 'Registre des naissances', 'url': '/gynecologie/naissances/'},
    ]
    return render(request, 'gynecologie/registre_naissance_form.html', {
        'patients': patients,
        'medecins': medecins,
        'breadcrumb': breadcrumb,
        'error': error,
        'post': request.POST if request.method == 'POST' else {},
    })


@login_required(login_url='login')
def gynecologie_registre_naissance(request):
    from patients.models import Naissance
    from django.core.paginator import Paginator
    from django.db.models import Q

    export = request.GET.get('export', '')
    q = request.GET.get('q', '').strip()
    filter_type  = request.GET.get('filter_type', '')
    date_precise = request.GET.get('date_precise', '')
    date_debut   = request.GET.get('date_debut', '')
    date_fin     = request.GET.get('date_fin', '')
    filtre_statut = request.GET.get('filtre_statut', '')

    group_by = request.GET.get('group_by', '')

    naissances = Naissance.objects.select_related('mere', 'medecin', 'mere__assurance')

    if q:
        naissances = naissances.filter(
            Q(mere__nom__icontains=q) |
            Q(mere__prenoms__icontains=q) |
            Q(numero__icontains=q) |
            Q(nom_enfant__icontains=q)
        )

    if filter_type == 'mois':
        now = timezone.now()
        naissances = naissances.filter(
            date_accouchement__month=now.month,
            date_accouchement__year=now.year,
        )
    elif filter_type == 'date' and date_precise:
        naissances = naissances.filter(date_accouchement__date=date_precise)
    elif filter_type == 'plage':
        if date_debut:
            naissances = naissances.filter(date_accouchement__date__gte=date_debut)
        if date_fin:
            naissances = naissances.filter(date_accouchement__date__lte=date_fin)

    if filtre_statut in ('vivant', 'mort_ne'):
        naissances = naissances.filter(statut=filtre_statut)

    # Tri selon le regroupement
    ORDER_MAP = {
        'annee':         'date_accouchement',
        'trimestre':     'date_accouchement',
        'mois':          'date_accouchement',
        'semaine':       'date_accouchement',
        'jour':          'date_accouchement',
        'mere':          'mere__nom',
        'medecin':       'medecin__employe__nom',
        'mode':          'mode_accouchement',
        'statut':        'statut',
        'genre':         'sexe_enfant',
        'groupe_sanguin':'groupe_sanguin_enfant',
        'lieu':          'lieu_naissance',
        'parite':        'parite',
        'education':     'education_mere',
        'cree_le':       'date_creation',
    }
    order_field = ORDER_MAP.get(group_by, '-date_accouchement')
    naissances = naissances.order_by(order_field)

    # Préparer les données groupées
    from itertools import groupby as py_groupby
    from django.utils.formats import date_format

    def get_group_key(n):
        if group_by == 'annee':
            return str(n.date_accouchement.year)
        if group_by == 'trimestre':
            q = (n.date_accouchement.month - 1) // 3 + 1
            return f"T{q} {n.date_accouchement.year}"
        if group_by == 'mois':
            return n.date_accouchement.strftime('%B %Y').capitalize()
        if group_by == 'semaine':
            return f"Semaine {n.date_accouchement.strftime('%W')} – {n.date_accouchement.year}"
        if group_by == 'jour':
            return n.date_accouchement.strftime('%d/%m/%Y')
        if group_by == 'mere':
            return f"{n.mere.nom.upper()} {n.mere.prenoms}"
        if group_by == 'medecin':
            return str(n.medecin) if n.medecin else '— Sans médecin'
        if group_by == 'mode':
            return n.get_mode_accouchement_display()
        if group_by == 'statut':
            return n.get_statut_display()
        if group_by == 'genre':
            return 'Féminin' if n.sexe_enfant == 'F' else 'Masculin'
        if group_by == 'groupe_sanguin':
            return n.groupe_sanguin_enfant or '—'
        if group_by == 'lieu':
            return n.lieu_naissance or '—'
        if group_by == 'parite':
            return f"Parité {n.parite}"
        if group_by == 'education':
            return n.get_education_mere_display() or '—'
        if group_by == 'cree_le':
            return n.date_creation.strftime('%d/%m/%Y')
        return None

    naissances_list = list(naissances)

    # ── EXPORT EXCEL ──
    if export == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Registre des naissances'

        # En-têtes
        headers = [
            'Numéro', 'Mère', 'Date d\'accouchement', 'Médecin',
            'Mode d\'accouchement', 'Nom de l\'enfant', 'Prénoms de l\'enfant',
            'Sexe', 'Poids (g)', 'Taille (cm)', 'Apgar 1\'', 'Apgar 5\'',
            'Groupe sanguin', 'Lieu de naissance', 'Parité',
            'Garçons', 'Filles', 'Éducation mère', 'Statut', 'Remarques',
        ]
        header_fill = PatternFill('solid', fgColor='714B67')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        thin = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[1].height = 18

        # Données
        for row_idx, n in enumerate(naissances_list, 2):
            values = [
                n.numero,
                f"{n.mere.nom.upper()} {n.mere.prenoms}",
                n.date_accouchement.strftime('%d/%m/%Y %H:%M'),
                str(n.medecin) if n.medecin else '',
                n.get_mode_accouchement_display(),
                n.nom_enfant,
                n.prenoms_enfant,
                'Féminin' if n.sexe_enfant == 'F' else 'Masculin',
                float(n.poids_naissance) if n.poids_naissance else '',
                float(n.taille_naissance) if n.taille_naissance else '',
                n.apgar_1min if n.apgar_1min is not None else '',
                n.apgar_5min if n.apgar_5min is not None else '',
                n.groupe_sanguin_enfant,
                n.lieu_naissance,
                n.parite,
                n.nombre_garcons,
                n.nombre_filles,
                n.get_education_mere_display(),
                n.get_statut_display(),
                n.remarques,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='F9F4FC')

        # Largeurs auto
        col_widths = [14,28,18,22,18,18,20,10,10,10,9,9,12,18,8,8,8,16,10,30]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        filename = f"naissances_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    grouped_data = None
    if group_by and naissances_list:
        grouped_data = [(k, list(v)) for k, v in py_groupby(naissances_list, key=get_group_key)]

    paginator = Paginator(naissances_list, 80)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    breadcrumb = [
        {'title': 'Accueil', 'url': '/'},
        {'title': 'Gynécologie', 'url': '/gynecologie/'},
        {'title': 'Registre des naissances'},
    ]
    return render(request, 'gynecologie/registre_naissance.html', {
        'page_obj': page_obj,
        'grouped_data': grouped_data,
        'group_by': group_by,
        'breadcrumb': breadcrumb,
        'filter_type': filter_type,
        'date_precise': date_precise,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'filtre_statut': filtre_statut,
    })


def _rdv_gyn_qs():
    from patients.models import RendezVous
    from django.db.models import Q
    return RendezVous.objects.filter(
        Q(departement__code='GYN') | Q(medecin__specialite__nom__icontains='gyn')
    ).select_related('patient', 'medecin').order_by('-date_heure')


def _rdv_form_post(request, rdv):
    from django.utils.dateparse import parse_datetime
    action = request.POST.get('action', '')
    PIPELINE = ('planifie', 'confirme', 'en_attente', 'en_consultation', 'termine', 'annule', 'absent')
    if action in PIPELINE and rdv.pk:
        rdv.statut = action
        rdv.save(update_fields=['statut'])
        return None, True
    try:
        raw_date = request.POST.get('date_heure', '')
        date_rdv = parse_datetime(raw_date) if raw_date else rdv.date_heure
        if date_rdv is None:
            date_rdv = rdv.date_heure or timezone.now()
        raw_suivi = request.POST.get('date_suivi', '')
        date_suivi = parse_datetime(raw_suivi) if raw_suivi else None
        rdv.patient_id = request.POST.get('patient') or rdv.patient_id
        rdv.medecin_id = request.POST.get('medecin') or None
        rdv.docteur_jr_id = request.POST.get('docteur_jr') or None
        rdv.departement_id = request.POST.get('departement') or rdv.departement_id
        rdv.salle_consultation = request.POST.get('salle_consultation', '')
        rdv.date_heure = date_rdv
        rdv.date_suivi = date_suivi
        rdv.duree_minutes = int(request.POST.get('duree_minutes') or 30)
        rdv.type_rdv = request.POST.get('type_rdv', rdv.type_rdv)
        rdv.type_visite_cpn = request.POST.get('cpn_type_visite', '')
        rdv.niveau_urgence = request.POST.get('niveau_urgence', 'normal')
        rdv.motif = request.POST.get('motif', '')
        rdv.notes = request.POST.get('notes', '')
        rdv.maladies = request.POST.get('maladies', '')
        rdv.principales_plaintes = request.POST.get('principales_plaintes', '')
        rdv.antecedents_maladie = request.POST.get('antecedents_maladie', '')
        rdv.historique_passee = request.POST.get('historique_passee', '')
        rdv.rdv_exterieur = bool(request.POST.get('rdv_exterieur'))
        rdv.code_confirmation = request.POST.get('code_confirmation', '')
        rdv.temps_attente_minutes = int(request.POST.get('temps_attente_minutes') or 0)
        rdv.temps_consultation_minutes = int(request.POST.get('temps_consultation_minutes') or 0)
        if action in PIPELINE:
            rdv.statut = action
        rdv.save()
        return None, True
    except Exception as e:
        return str(e), False


@login_required(login_url='login')
def gynecologie_rdv_create(request):
    from patients.forms import RendezVousForm
    from medecins.models import Medecin

    medecins = Medecin.objects.all().order_by('employe__nom')

    if request.method == 'POST':
        form = RendezVousForm(request.POST)
        if form.is_valid():
            rdv = form.save(commit=False)
            code = request.POST.get('code_confirmation', '').strip()
            if code:
                rdv.code_confirmation = code
            rdv.save()
            from patients.utils import save_registres
            save_registres(request, rdv)
            action = request.POST.get('_action', '')
            if action == 'annuler':
                return redirect('gynecologie_rdv')
            from django.urls import reverse
            return redirect(reverse('facturation:create') + f'?patient={rdv.patient.pk}&rdv={rdv.pk}')
    else:
        from medecins.models import Departement
        gyn_departement = Departement.objects.filter(code='GYN').first()
        form = RendezVousForm(initial={
            'date_heure': timezone.now().strftime('%Y-%m-%dT%H:%M'),
            'departement': gyn_departement.pk if gyn_departement else None,
        })

    breadcrumb = [
        {'title': 'Gynécologie', 'url': '/gynecologie/'},
        {'title': 'Rendez-vous', 'url': '/gynecologie/rdv/'},
        {'title': 'Nouveau'},
    ]
    from patients.models import Pathologie
    from gynecologie.models import TypeVisite
    return render(request, 'gynecologie/rdv_form.html', {
        'form': form,
        'rdv': None,
        'is_new': True,
        'patient_prefill': None,
        'consultation': None,
        'constante': None,
        'facture_payee': False,
        'medecins': medecins,
        'pathologies': Pathologie.objects.filter(actif=True).order_by('nom'),
        'types_visite': TypeVisite.objects.filter(actif=True).order_by('nom'),
        'breadcrumb': breadcrumb,
        'registre_cpn': None,
        'registre_accouchement': None,
        'registre_postnatale': None,
        'registre_curatif': None,
    })


@login_required(login_url='login')
def gynecologie_rdv_detail(request, pk):
    from patients.forms import RendezVousForm
    from patients.models import RendezVous
    from medecins.models import Medecin

    rdv = get_object_or_404(RendezVous, pk=pk)
    medecins = Medecin.objects.all().order_by('employe__nom')

    try:
        from facturation.models import Facture
        facture_payee = Facture.objects.filter(
            Q(rendez_vous=rdv) | Q(patient=rdv.patient, statut='payee')
        ).exclude(statut='annulee').exists()
    except Exception as _e:
        import logging; logging.getLogger(__name__).warning('facture_payee check failed: %s', _e)
        facture_payee = False

    consultation = None
    constante = None
    try:
        consultation = rdv.consultation
        try:
            constante = consultation.constantes
        except Exception:
            pass
    except Exception:
        pass

    if request.method == 'POST':
        action = request.POST.get('_action', '')

        if action == 'save_eval':
            # Sauvegarder le médecin sélectionné dans le modal
            medecin_pk = request.POST.get('eval_medecin', '').strip()
            if medecin_pk:
                try:
                    from medecins.models import Medecin
                    rdv.medecin = Medecin.objects.get(pk=medecin_pk)
                    rdv.save(update_fields=['medecin'])
                except Exception:
                    pass

            _eval_map = {
                'eval_poids': 'poids', 'eval_taille': 'taille',
                'eval_temperature': 'temperature',
                'eval_tension_systolique': 'tension_systolique',
                'eval_tension_diastolique': 'tension_diastolique',
                'eval_tension_systolique_droite': 'tension_systolique_droite',
                'eval_tension_diastolique_droite': 'tension_diastolique_droite',
                'eval_pouls': 'pouls', 'eval_frequence_respiratoire': 'frequence_respiratoire',
                'eval_saturation_oxygene': 'saturation_oxygene', 'eval_glycemie': 'glycemie',
                'eval_albumine': 'albumine', 'eval_perimetre_brachial': 'perimetre_brachial',
                'eval_niveau_douleur': 'niveau_douleur',
            }
            from consultations.models import Consultation as Consult, Constante as Const
            try:
                consult_obj = rdv.consultation
            except Exception:
                consult_obj = None
            if consult_obj is None:
                consult_obj = Consult.objects.create(
                    patient=rdv.patient, medecin=rdv.medecin, rendez_vous=rdv,
                    motif=rdv.motif or 'Évaluation clinique', cree_par=request.user,
                )
            const_obj, _ = Const.objects.get_or_create(consultation=consult_obj)
            for post_key, model_field in _eval_map.items():
                val = request.POST.get(post_key, '').strip()
                if val != '':
                    setattr(const_obj, model_field, val)
            const_obj.save()
            messages.success(request, 'Évaluation enregistrée.')
            return redirect('gynecologie_rdv_detail', pk=rdv.pk)

        if action == 'confirmer':
            if facture_payee:
                rdv.statut = 'confirme'
                rdv.date_confirme = timezone.now()
                rdv._skip_auto_log = True
                rdv.save(update_fields=['statut', 'date_confirme'])
                log_event(rdv, request.user, 'État : Brouillon → Confirmer', type='statut')
                messages.success(request, 'Rendez-vous confirmé.')
            else:
                messages.error(request, 'Une facture est requise pour confirmer ce rendez-vous.')
            return redirect('gynecologie_rdv')

        if action == 'en_attente':
            now = timezone.now()
            rdv.statut = 'en_attente'
            rdv.date_en_attente = now
            if rdv.date_confirme:
                rdv.temps_constante_minutes = int((now - rdv.date_confirme).total_seconds() / 60)
            rdv._skip_auto_log = True
            rdv.save(update_fields=['statut', 'date_en_attente', 'temps_constante_minutes'])
            log_event(rdv, request.user, 'État : Confirmer → En Attente', type='statut')
            messages.success(request, 'Rendez-vous mis en attente.')
            return redirect('gynecologie_rdv_detail', pk=rdv.pk)

        if action == 'en_consultation':
            now = timezone.now()
            rdv.statut = 'en_consultation'
            rdv.date_en_consultation = now
            if rdv.date_en_attente:
                rdv.temps_attente_minutes = int((now - rdv.date_en_attente).total_seconds() / 60)
            rdv._skip_auto_log = True
            rdv.save(update_fields=['statut', 'date_en_consultation', 'temps_attente_minutes'])
            log_event(rdv, request.user, 'État : En Attente → En Consultation', type='statut')
            messages.success(request, 'Consultation démarrée.')
            return redirect('gynecologie_rdv_detail', pk=rdv.pk)

        if action == 'terminer':
            now = timezone.now()
            rdv.statut = 'termine'
            rdv.date_termine = now
            if rdv.date_en_consultation:
                rdv.temps_consultation_minutes = int((now - rdv.date_en_consultation).total_seconds() / 60)
            rdv._skip_auto_log = True
            rdv.save(update_fields=['statut', 'date_termine', 'temps_consultation_minutes'])
            log_event(rdv, request.user, 'État : En Consultation → Terminé', type='statut')
            messages.success(request, 'Consultation terminée.')
            return redirect('gynecologie_rdv')

        if action == 'annuler':
            rdv.statut = 'annule'
            rdv._skip_auto_log = True
            rdv.save(update_fields=['statut'])
            log_event(rdv, request.user, 'Rendez-vous annulé.', type='statut')
            messages.success(request, 'Rendez-vous annulé.')
            return redirect('gynecologie_rdv')

        # Sauvegarde normale du formulaire
        form = RendezVousForm(request.POST, instance=rdv)
        if form.is_valid():
            rdv = form.save(commit=False)
            code = request.POST.get('code_confirmation', '').strip()
            if code:
                rdv.code_confirmation = code
            from gynecologie.models import TypeVisite
            rdv.cpn_mode_entree = request.POST.get('cpn_mode_entree', '').strip()
            rdv.cpn_mode_entree_autre = request.POST.get('cpn_mode_entree_autre', '').strip()
            cpn_tv_pk = request.POST.get('cpn_type_visite', '').strip()
            if cpn_tv_pk:
                try:
                    rdv.cpn_type_visite = TypeVisite.objects.get(pk=int(cpn_tv_pk))
                except (ValueError, TypeVisite.DoesNotExist):
                    rdv.cpn_type_visite = None
            else:
                rdv.cpn_type_visite = None
            rdv.cur_mode_entree = request.POST.get('cur_mode_entree', '').strip()
            rdv.cur_mode_entree_autre = request.POST.get('cur_mode_entree_autre', '').strip()
            cur_tv_pk = request.POST.get('cur_type_visite', '').strip()
            if cur_tv_pk:
                try:
                    rdv.cur_type_visite = TypeVisite.objects.get(pk=int(cur_tv_pk))
                except (ValueError, TypeVisite.DoesNotExist):
                    rdv.cur_type_visite = None
            else:
                rdv.cur_type_visite = None
            rdv._skip_auto_log = True
            rdv.save()
            log_event(rdv, request.user, 'Rendez-vous modifié.', type='modif')
            from patients.utils import save_registres
            save_registres(request, rdv)
            messages.success(request, 'Rendez-vous modifié.')
            if action == 'créer une facture':
                from django.urls import reverse
                return redirect(reverse('facture_create') + f'?patient={rdv.patient.pk}&rdv={rdv.pk}')
            return redirect('gynecologie_rdv_detail', pk=rdv.pk)
    else:
        form = RendezVousForm(instance=rdv)

    # Navigation précédent/suivant dans la liste gynécologie
    qs_pks = list(_rdv_gyn_qs().values_list('pk', flat=True))
    total = len(qs_pks)
    try:
        pos = qs_pks.index(pk)
        nav_pos = pos + 1
        nav_prev = qs_pks[pos - 1] if pos > 0 else None
        nav_next = qs_pks[pos + 1] if pos < total - 1 else None
    except ValueError:
        nav_pos, nav_prev, nav_next = None, None, None

    breadcrumb = [
        {'title': 'Gynécologie', 'url': '/gynecologie/'},
        {'title': 'Rendez-vous', 'url': '/gynecologie/rdv/'},
        {'title': rdv.code_rdv or rdv.patient.code_patient},
    ]
    from patients.models import Pathologie, RegistreCPN, RegistreAccouchement, RegistrePostnatale, RegistreCuratif
    from gynecologie.models import TypeVisite
    def _get_reg(Model):
        try:
            return Model.objects.get(rdv=rdv)
        except Model.DoesNotExist:
            return None

    return render(request, 'gynecologie/rdv_form.html', {
        'form': form,
        'rdv': rdv,
        'is_new': False,
        'patient_prefill': rdv.patient,
        'facture_payee': facture_payee,
        'consultation': consultation,
        'constante': constante,
        'medecins': medecins,
        'pathologies': Pathologie.objects.filter(actif=True).order_by('nom'),
        'types_visite': TypeVisite.objects.filter(actif=True).order_by('nom'),
        'breadcrumb': breadcrumb,
        'nav_total': total,
        'nav_pos': nav_pos,
        'nav_prev': nav_prev,
        'nav_next': nav_next,
        'registre_cpn':          _get_reg(RegistreCPN),
        'registre_accouchement': _get_reg(RegistreAccouchement),
        'registre_postnatale':   _get_reg(RegistrePostnatale),
        'registre_curatif':      _get_reg(RegistreCuratif),
    })


@login_required(login_url='login')
def gynecologie_demarrer_consultation(request, pk):
    from patients.models import RendezVous
    from consultations.models import Consultation
    from django.core.exceptions import ObjectDoesNotExist

    rdv = get_object_or_404(RendezVous, pk=pk)

    # Si consultation existe déjà, retourner au RDV
    try:
        _ = rdv.consultation
        return redirect('gynecologie_rdv_detail', pk=rdv.pk)
    except ObjectDoesNotExist:
        pass

    Consultation.objects.create(
        patient=rdv.patient,
        medecin=rdv.medecin,
        rendez_vous=rdv,
        motif=rdv.motif or 'Consultation gynecologique',
        statut='en_cours',
        cree_par=request.user,
    )

    # Passer le RDV en consultation si pas encore terminé
    if rdv.statut in ('planifie', 'confirme', 'en_attente'):
        rdv.statut = 'en_consultation'
        rdv.save(update_fields=['statut'])

    return redirect('gynecologie_rdv_detail', pk=rdv.pk)


@login_required(login_url='login')
def gynecologie_rdv(request):
    from patients.models import RendezVous
    from django.core.paginator import Paginator
    from django.db.models import Q
    from datetime import date as _date

    q                 = request.GET.get('q', '').strip()
    filter_val        = request.GET.get('filter', '')
    group_val         = request.GET.get('group', '')
    date_from         = request.GET.get('date_from', '').strip()
    date_to           = request.GET.get('date_to', '').strip()
    type_rdv_val      = request.GET.get('type_rdv', '').strip()
    type_visite_cpn_val = request.GET.get('type_visite_cpn', '').strip()

    rdvs = RendezVous.objects.filter(
        Q(departement__code='GYN') |
        Q(medecin__specialite__nom__icontains='gyn')
    ).select_related('patient', 'medecin').order_by('-date_heure')

    if q:
        rdvs = rdvs.filter(
            Q(patient__nom__icontains=q) |
            Q(patient__prenoms__icontains=q) |
            Q(patient__code_patient__icontains=q)
        )

    # --- Filtre rapide (défaut : aujourd'hui sauf si filter=all ou autre filtre actif) ---
    no_active_filter = not filter_val and not type_rdv_val and not type_visite_cpn_val and not date_from and not date_to and not q
    if (filter_val == 'today' or no_active_filter) and not date_from and not date_to:
        rdvs = rdvs.filter(date_heure__date=_date.today())
    elif filter_val == 'all':
        pass  # tout afficher, aucun filtre date
    elif filter_val == 'mine':
        rdvs = rdvs.filter(medecin__user=request.user)
    elif filter_val == 'urgent':
        rdvs = rdvs.filter(niveau_urgence__in=['urgent', 'tres_urgent'])
    elif filter_val == 'urgence_medicale':
        rdvs = rdvs.filter(type_rdv='urgence')
    elif filter_val == 'consultation':
        rdvs = rdvs.filter(type_rdv='consultation')
    elif filter_val == 'suivi':
        rdvs = rdvs.filter(type_rdv='controle')
    elif filter_val == 'not_done':
        rdvs = rdvs.exclude(statut__in=['termine', 'annule', 'absent'])

    # --- Filtre type de visite (type_rdv) ---
    if type_rdv_val in ('consultation', 'controle', 'urgence', 'examen', 'vaccination'):
        rdvs = rdvs.filter(type_rdv=type_rdv_val)

    # --- Filtre rang CPN ---
    if type_visite_cpn_val in ('cpn1', 'cpn2', 'cpn3', 'cpn4', 'autre'):
        rdvs = rdvs.filter(type_visite_cpn=type_visite_cpn_val)

    # --- Plage de dates ---
    if date_from:
        try:
            rdvs = rdvs.filter(date_heure__date__gte=date_from)
        except (ValueError, TypeError):
            pass
    if date_to:
        try:
            rdvs = rdvs.filter(date_heure__date__lte=date_to)
        except (ValueError, TypeError):
            pass

    # --- Tri par colonne (priorité sur le tri par groupe) ---
    sort_val  = request.GET.get('sort', '').strip()
    order_val = request.GET.get('order', 'asc').strip()
    if order_val not in ('asc', 'desc'):
        order_val = 'asc'
    p = '-' if order_val == 'desc' else ''

    SORT_FIELDS = {
        'date':    ['date_heure'],
        'patient': ['patient__nom', 'patient__prenoms'],
        'medecin': ['medecin__employe__nom'],
        'statut':  ['statut'],
        'type':    ['type_rdv'],
        'age':     ['patient__date_naissance'],
    }
    if sort_val in SORT_FIELDS:
        rdvs = rdvs.order_by(*[p + f for f in SORT_FIELDS[sort_val]])
    # --- Tri selon regroupement (si pas de tri colonne) ---
    elif group_val in ('date_jour', 'date_semaine', 'date_mois', 'date_trimestre', 'date_annee'):
        rdvs = rdvs.order_by('date_heure')
    elif group_val == 'statut':
        rdvs = rdvs.order_by('statut', '-date_heure')
    elif group_val in ('medecin', 'referent'):
        rdvs = rdvs.order_by('medecin', '-date_heure')
    elif group_val in ('type_rdv', 'type_visite'):
        rdvs = rdvs.order_by('type_rdv', '-date_heure')
    elif group_val == 'patient':
        rdvs = rdvs.order_by('patient__nom', 'patient__prenoms')
    elif group_val == 'sexe':
        rdvs = rdvs.order_by('patient__sexe', '-date_heure')
    elif group_val == 'age':
        rdvs = rdvs.order_by('patient__date_naissance')
    elif group_val == 'type_visite_cpn':
        rdvs = rdvs.order_by('type_visite_cpn', '-date_heure')
    elif group_val == 'departement':
        rdvs = rdvs.order_by('departement', '-date_heure')
    elif group_val == 'assurance':
        rdvs = rdvs.order_by('patient__assurance__nom', '-date_heure')
    elif group_val == 'police':
        rdvs = rdvs.order_by('patient__numero_assurance', '-date_heure')

    # --- Compte par groupe pour les en-têtes visuels ---
    from django.db.models import Count as _Count
    import json as _json
    group_counts = {}
    if group_val == 'type_visite_cpn':
        rows = rdvs.values('type_visite_cpn').annotate(n=_Count('id'))
        group_counts = {r['type_visite_cpn'] or '': r['n'] for r in rows}
    elif group_val == 'statut':
        rows = rdvs.values('statut').annotate(n=_Count('id'))
        group_counts = {r['statut'] or '': r['n'] for r in rows}
    elif group_val in ('type_rdv', 'type_visite'):
        rows = rdvs.values('type_rdv').annotate(n=_Count('id'))
        group_counts = {r['type_rdv'] or '': r['n'] for r in rows}
    elif group_val == 'sexe':
        rows = rdvs.values('patient__sexe').annotate(n=_Count('id'))
        group_counts = {r['patient__sexe'] or '': r['n'] for r in rows}
    elif group_val == 'departement':
        rows = rdvs.values('departement').annotate(n=_Count('id'))
        group_counts = {r['departement'] or '': r['n'] for r in rows}
    elif group_val in ('medecin', 'referent'):
        rows = rdvs.values('medecin__employe__nom', 'medecin__employe__prenoms').annotate(n=_Count('id'))
        group_counts = {
            (f"Dr {r['medecin__employe__nom']} {r['medecin__employe__prenoms']}" if r['medecin__employe__nom'] else '—'): r['n']
            for r in rows
        }
    elif group_val == 'assurance':
        rows = rdvs.values('patient__assurance__nom').annotate(n=_Count('id'))
        group_counts = {r['patient__assurance__nom'] or '—': r['n'] for r in rows}
    elif group_val == 'police':
        rows = rdvs.values('patient__numero_assurance').annotate(n=_Count('id'))
        group_counts = {r['patient__numero_assurance'] or '—': r['n'] for r in rows}
    elif group_val == 'date_annee':
        rows = rdvs.values('date_heure__year').annotate(n=_Count('id'))
        group_counts = {str(r['date_heure__year']): r['n'] for r in rows}
    elif group_val == 'date_mois':
        rows = rdvs.values('date_heure__year', 'date_heure__month').annotate(n=_Count('id'))
        group_counts = {f"{r['date_heure__year']}-{r['date_heure__month']:02d}": r['n'] for r in rows}
    elif group_val == 'date_trimestre':
        rows = rdvs.values('date_heure__year', 'date_heure__month').annotate(n=_Count('id'))
        qt = {}
        for r in rows:
            key = f"{r['date_heure__year']}-T{(r['date_heure__month'] - 1) // 3 + 1}"
            qt[key] = qt.get(key, 0) + r['n']
        group_counts = qt
    elif group_val == 'date_semaine':
        from django.db.models.functions import ExtractWeek, ExtractIsoYear
        rows = rdvs.annotate(
            iso_yr=ExtractIsoYear('date_heure'), wk=ExtractWeek('date_heure')
        ).values('iso_yr', 'wk').annotate(n=_Count('id'))
        group_counts = {f"{r['iso_yr']}-{r['wk']}": r['n'] for r in rows}
    elif group_val == 'date_jour':
        rows = rdvs.values('date_heure__date').annotate(n=_Count('id'))
        group_counts = {str(r['date_heure__date']): r['n'] for r in rows}

    # --- Export Excel (respecte les filtres actifs) ---
    if request.GET.get('export') == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Rendez-vous'
        headers = ['Code RDV', 'Patient', 'Date', 'Médecin', 'Type', 'Motif', 'Âge', 'Genre', 'État']
        header_fill = PatternFill('solid', fgColor='00838F')
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        from datetime import date as _d2
        for ri, rdv in enumerate(rdvs, 2):
            age = ''
            if rdv.patient.date_naissance:
                t = _d2.today()
                dob = rdv.patient.date_naissance
                age = str(t.year - dob.year - ((t.month, t.day) < (dob.month, dob.day)))
            ws.append([
                rdv.code_rdv or rdv.patient.code_patient,
                f"{rdv.patient.nom.upper()} {rdv.patient.prenoms}",
                rdv.date_heure.strftime('%d/%m/%Y %H:%M'),
                str(rdv.medecin) if rdv.medecin else '',
                rdv.get_type_rdv_display(),
                rdv.motif[:100] if rdv.motif else '',
                age,
                rdv.patient.get_sexe_display() if rdv.patient.sexe else '',
                rdv.get_statut_display(),
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), max(len(str(c.value or '')) for c in col)) + 4
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="rendez-vous.xlsx"'
        wb.save(response)
        return response

    # --- Stats du jour (indépendant des filtres) ---
    today_base = RendezVous.objects.filter(
        Q(departement__code='GYN') | Q(medecin__specialite__nom__icontains='gyn'),
        date_heure__date=_date.today()
    )
    stat_rows = today_base.values('statut').annotate(n=_Count('id'))
    stats_today = {r['statut']: r['n'] for r in stat_rows}
    stats_today['total'] = sum(stats_today.values())

    paginator = Paginator(rdvs, 80)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    breadcrumb = [
        {'title': 'Accueil', 'url': '/'},
        {'title': 'Gynécologie', 'url': '/gynecologie/'},
        {'title': 'Rendez-vous'},
    ]
    return render(request, 'gynecologie/rdv.html', {
        'page_obj': page_obj,
        'breadcrumb': breadcrumb,
        'group_counts_json': _json.dumps(group_counts),
        'group_val': group_val,
        'filter_val': 'today' if no_active_filter else ('' if filter_val == 'all' else filter_val),
        'stats_today': stats_today,
        'sort_val': sort_val,
        'order_val': order_val,
        'statut_choices': [('planifie','Planifié'),('confirme','Confirmé'),('en_attente','En attente'),('en_consultation','En consultation'),('termine','Terminé'),('annule','Annulé'),('absent','Absent')],
    })


@login_required(login_url='login')
def gynecologie_rdv_set_statut(request, pk):
    from patients.models import RendezVous
    STATUTS = ('planifie', 'confirme', 'en_attente', 'en_consultation', 'termine', 'annule', 'absent')
    if request.method == 'POST':
        rdv = get_object_or_404(RendezVous, pk=pk)
        new_statut = request.POST.get('statut', '')
        if new_statut in STATUTS:
            rdv.statut = new_statut
            rdv.save(update_fields=['statut'])
    next_url = request.POST.get('next', '/gynecologie/rdv/')
    return redirect(next_url)


@login_required(login_url='login')
def gynecologie_rdv_bulk(request):
    from patients.models import RendezVous
    if request.method == 'POST':
        pks = request.POST.getlist('pks')
        action = request.POST.get('action', '')
        STATUTS = ('planifie', 'confirme', 'en_attente', 'en_consultation', 'termine', 'annule', 'absent')
        if action in STATUTS and pks:
            RendezVous.objects.filter(pk__in=pks).update(statut=action)
    next_url = request.POST.get('next', '/gynecologie/rdv/')
    return redirect(next_url)


@login_required(login_url='login')
def gynecologie_list(request):
    from patients.models import Patient
    from django.core.paginator import Paginator
    from datetime import date as _date

    q          = request.GET.get('q', '').strip()
    filter_val = request.GET.get('filter', '')
    group_val  = request.GET.get('group', '')

    patients = Patient.objects.filter(
        rendez_vous__departement__code='GYN'
    ).distinct().order_by('nom', 'prenoms')

    if q:
        patients = patients.filter(
            Q(nom__icontains=q) | Q(prenoms__icontains=q) | Q(code_patient__icontains=q) | Q(telephone__icontains=q)
        )

    # --- Filtres patients ---
    today = _date.today()
    if filter_val == 'femme':
        patients = patients.filter(sexe='F')
    elif filter_val == 'homme':
        patients = patients.filter(sexe='M')
    elif filter_val == 'mineur':
        cutoff = today.replace(year=today.year - 18)
        patients = patients.filter(date_naissance__gt=cutoff)
    elif filter_val == 'adulte':
        cutoff_18 = today.replace(year=today.year - 18)
        cutoff_60 = today.replace(year=today.year - 60)
        patients = patients.filter(date_naissance__lte=cutoff_18, date_naissance__gt=cutoff_60)
    elif filter_val == 'senior':
        cutoff = today.replace(year=today.year - 60)
        patients = patients.filter(date_naissance__lte=cutoff)

    # --- Tri selon regroupement ---
    if group_val == 'sexe':
        patients = patients.order_by('sexe', 'nom', 'prenoms')
    elif group_val == 'age':
        patients = patients.order_by('date_naissance')

    paginator = Paginator(patients, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    breadcrumb = [
        {'title': 'Gynécologie', 'url': '/gynecologie/'},
        {'title': 'Patients'},
    ]
    return render(request, 'gynecologie/list.html', {'page_obj': page_obj, 'breadcrumb': breadcrumb})


# ──────────────────────────────────────────────────────────────────
# Helpers de log d'activité générique (utilisés par hospitalisation,
# facturation, et tout module avec un LogActivite générique)
# ──────────────────────────────────────────────────────────────────
@login_required(login_url='login')
def post_note(request):
    if request.method != 'POST':
        return redirect('dashboard')
    from django.contrib.contenttypes.models import ContentType
    from core.models import LogActivite
    app_label  = request.POST.get('app_label', '')
    model_name = request.POST.get('model_name', '')
    object_id  = request.POST.get('object_id', '')
    note_text  = request.POST.get('note_text', '').strip()
    next_url   = request.POST.get('next', '/')
    if note_text and app_label and model_name and object_id:
        try:
            ct  = ContentType.objects.get(app_label=app_label, model=model_name)
            LogActivite.objects.create(
                content_type=ct,
                object_id=int(object_id),
                user=request.user,
                message=note_text,
                type='note',
            )
        except Exception:
            pass
    return redirect(next_url)


def log_event(instance, user, message, type='note'):
    from django.contrib.contenttypes.models import ContentType
    from core.models import LogActivite
    ct = ContentType.objects.get_for_model(instance)
    LogActivite.objects.create(
        content_type=ct,
        object_id=instance.pk,
        user=user if user and user.is_authenticated else None,
        message=message,
        type=type,
    )


def get_logs(instance):
    from django.contrib.contenttypes.models import ContentType
    from core.models import LogActivite
    ct = ContentType.objects.get_for_model(instance)
    return LogActivite.objects.filter(content_type=ct, object_id=instance.pk)
@login_required(login_url='login')
def kpi_dashboard(request):
    import json as _json
    from patients.models import Patient, RendezVous
    from consultations.models import Consultation
    from hospitalisation.models import Hospitalisation
    from pharmacie.models import Medicament
    from laboratoire.models import AnalyseLaboratoire
    from employer.models import Employe
    from soins.models import Soin
    from medecins.models import Medecin

    today = timezone.now().date()

    # Période sélectionnée
    period = request.GET.get('period', 'jour')
    if period not in ('jour', 'semaine', 'mois', 'tout'):
        period = 'jour'

    if period == 'jour':
        date_from = today
        period_label = f"Aujourd'hui — {today.strftime('%d/%m/%Y')}"
    elif period == 'semaine':
        date_from = today - timedelta(days=today.weekday())
        period_label = f"Cette semaine — du {date_from.strftime('%d/%m')} au {today.strftime('%d/%m/%Y')}"
    elif period == 'mois':
        date_from = today.replace(day=1)
        period_label = f"Ce mois — {today.strftime('%B %Y')}"
    else:  # tout
        from datetime import date as _date
        date_from = _date(2000, 1, 1)
        period_label = "Jusqu'à maintenant"

    date_to = today

    stats = {
        'patients_total': Patient.objects.count(),
        'patients_periode': Patient.objects.filter(date_creation__date__range=[date_from, date_to]).count(),
        'consultations_periode': Consultation.objects.filter(date_heure__date__range=[date_from, date_to]).count(),
        'rdv_periode': RendezVous.objects.filter(date_heure__date__range=[date_from, date_to]).count(),
        'hospitalisations': Hospitalisation.objects.filter(statut='hospitalise').count(),
        'analyses_pending': AnalyseLaboratoire.objects.filter(statut__in=['recu', 'en_analyse']).count(),
        'medicaments_alerte': Medicament.objects.filter(stock_actuel__lte=F('stock_alerte')).count(),
        'employes_actifs': Employe.objects.filter(statut='actif').count(),
        'soins_periode': Soin.objects.filter(date_creation__date__range=[date_from, date_to]).count(),
        'medecins_actifs': Medecin.objects.filter(actif=True).count(),
        'medecins_total': Medecin.objects.count(),
    }

    try:
        from facturation.models import Facture
        from django.db.models import Sum
        fq = Facture.objects.filter(statut__in=['brouillon', 'emise'])
        stats['factures_impayees_count'] = fq.count()
        m = fq.aggregate(t=Sum('montant_total'))['t'] or 0
        if m >= 1_000_000:
            stats['factures_montant_fmt'] = f"{m/1_000_000:.1f}M F"
        elif m >= 1_000:
            stats['factures_montant_fmt'] = f"{int(m/1_000)}k F"
        else:
            stats['factures_montant_fmt'] = f"{int(m)} F"
    except Exception:
        stats['factures_impayees_count'] = 0
        stats['factures_montant_fmt'] = "0 F"

    stats['patients_anniversaires'] = Patient.objects.filter(
        date_naissance__month=today.month, date_naissance__day=today.day
    ).count()
    stats['employes_anniversaires'] = Employe.objects.filter(
        date_naissance__month=today.month, date_naissance__day=today.day
    ).count()

    chart_labels, chart_patients, chart_rdv = [], [], []
    current_monday = today - timedelta(days=today.weekday())
    for i in range(7, -1, -1):
        w_start = current_monday - timedelta(weeks=i)
        w_end = w_start + timedelta(days=6)
        chart_labels.append(w_start.strftime('%d/%m'))
        chart_patients.append(Patient.objects.filter(date_creation__date__range=[w_start, w_end]).count())
        chart_rdv.append(RendezVous.objects.filter(date_heure__date__range=[w_start, w_end]).count())

    chart_data = _json.dumps({'labels': chart_labels, 'patients': chart_patients, 'rdv': chart_rdv})

    rdv_auj = RendezVous.objects.filter(
        date_heure__date__range=[date_from, date_to],
        statut__in=['planifie', 'confirme', 'en_attente', 'en_consultation']
    ).select_related('patient', 'medecin').order_by('date_heure')[:15]

    last_cons = Consultation.objects.select_related('patient', 'medecin').order_by('-date_heure')[:6]

    try:
        from modules_permissions.models import get_user_modules
        user_modules = get_user_modules(request.user)
        accessible_codes = set(user_modules.values_list('code', flat=True))
    except Exception:
        user_modules = []
        accessible_codes = set()

    return render(request, 'core/kpi_dashboard.html', {
        'stats': stats,
        'rdv_auj': rdv_auj,
        'last_cons': last_cons,
        'today': today,
        'user': request.user,
        'user_modules': user_modules,
        'accessible_codes': accessible_codes,
        'chart_data': chart_data,
        'period': period,
        'period_label': period_label,
    })


# ══════════════════════════════════════════════
#  MÉDECINS — Configuration : Spécialités
# ══════════════════════════════════════════════

def _specialite_form_class():
    from django import forms
    from medecins.models import Specialite
    class SpecialiteForm(forms.ModelForm):
        class Meta:
            model = Specialite
            fields = ['nom', 'code', 'description']
            widgets = {
                'nom':         forms.TextInput(attrs={'class': 'f-input', 'placeholder': 'Ex : Cardiologie'}),
                'code':        forms.TextInput(attrs={'class': 'f-input', 'placeholder': 'Ex : CARD'}),
                'description': forms.Textarea(attrs={'class': 'f-input', 'rows': 3, 'placeholder': 'Description optionnelle'}),
            }
    return SpecialiteForm


@login_required(login_url='login')
def medecins_specialites(request):
    from medecins.models import Specialite
    from django.core.paginator import Paginator

    q   = request.GET.get('q', '').strip()
    vue = request.GET.get('vue', 'liste')
    qs  = Specialite.objects.order_by('nom')
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(code__icontains=q))
    total = Specialite.objects.count()
    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'medecins/config/specialites_list.html', {
        'page_obj': page_obj, 'total': total,
        'total_filtre': qs.count(), 'q': q, 'vue': vue,
    })


@login_required(login_url='login')
def medecins_specialite_create(request):
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    Form = _specialite_form_class()
    form = Form(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        obj = form.save()
        if is_ajax:
            return JsonResponse({'ok': True, 'message': f'Spécialité « {obj} » créée.'})
        messages.success(request, 'Spécialité créée.')
        return redirect('medecins_specialites')
    template = 'medecins/config/specialite_form_modal.html' if is_ajax else 'medecins/config/specialite_form.html'
    return render(request, template, {
        'form': form, 'titre': 'Nouvelle spécialité',
    })


@login_required(login_url='login')
def medecins_specialite_edit(request, pk):
    from medecins.models import Specialite
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    obj  = get_object_or_404(Specialite, pk=pk)
    Form = _specialite_form_class()
    form = Form(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        if is_ajax:
            return JsonResponse({'ok': True, 'message': f'Spécialité « {obj} » mise à jour.'})
        messages.success(request, 'Spécialité mise à jour.')
        return redirect('medecins_specialites')
    template = 'medecins/config/specialite_form_modal.html' if is_ajax else 'medecins/config/specialite_form.html'
    return render(request, template, {
        'form': form, 'titre': f'Modifier – {obj.nom}', 'obj': obj,
    })


@login_required(login_url='login')
@require_POST
def medecins_specialite_bulk_delete(request):
    from medecins.models import Specialite
    ids = request.POST.getlist('ids[]')
    if ids:
        Specialite.objects.filter(pk__in=ids).delete()
    return JsonResponse({'ok': True})


_SPEC_HDR = ['code', 'nom', 'description']


def _spec_row(s):
    return [s.code, s.nom, s.description]


@login_required(login_url='login')
def medecins_export_specialites(request):
    from medecins.models import Specialite
    from services.views import _export_file
    fmt = request.GET.get('format', 'json')
    qs = Specialite.objects.order_by('nom')
    rows = [_spec_row(s) for s in qs]
    return _export_file(fmt, 'specialites', _SPEC_HDR, rows,
                        [dict(zip(_SPEC_HDR, r)) for r in rows])


@login_required(login_url='login')
@require_POST
def medecins_import_specialites(request):
    from medecins.models import Specialite
    from services.views import _parse_upload, _s

    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('medecins_specialites')

    data, err = _parse_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('medecins_specialites')

    do_update = 'update' in request.POST
    created = updated = skipped = errors = 0
    for item in data:
        try:
            code = _s(item.get('code', ''))
            if not code:
                errors += 1
                continue
            defaults = {
                'nom': _s(item.get('nom', code)),
                'description': _s(item.get('description', '')),
            }
            obj, was_created = Specialite.objects.get_or_create(code=code, defaults=defaults)
            if was_created:
                created += 1
            elif do_update:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                updated += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    if errors:
        messages.warning(request, f'{created} créée(s), {updated} mise(s) à jour, {skipped} ignorée(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} spécialité(s) importée(s), {updated} mise(s) à jour, {skipped} ignorée(s).')
    return redirect('medecins_specialites')


# ══════════════════════════════════════════════
#  MÉDECINS — Configuration : Départements
# ══════════════════════════════════════════════

def _departement_form_class():
    from django import forms
    from medecins.models import Departement
    class DepartementForm(forms.ModelForm):
        class Meta:
            model = Departement
            fields = ['nom', 'code', 'description', 'actif']
            widgets = {
                'nom':         forms.TextInput(attrs={'class': 'f-input', 'placeholder': 'Ex : Département Médical'}),
                'code':        forms.TextInput(attrs={'class': 'f-input', 'placeholder': 'Ex : DEPT-MED'}),
                'description': forms.Textarea(attrs={'class': 'f-input', 'rows': 3}),
            }
    return DepartementForm


@login_required(login_url='login')
def medecins_departements(request):
    from medecins.models import Departement
    from django.core.paginator import Paginator
    from django.db.models import Count

    q   = request.GET.get('q', '').strip()
    vue = request.GET.get('vue', 'liste')
    qs  = Departement.objects.annotate(nb_medecins=Count('medecins')).order_by('nom')
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(code__icontains=q))
    total = Departement.objects.count()
    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'medecins/config/departements_list.html', {
        'page_obj': page_obj, 'total': total,
        'total_filtre': qs.count(), 'q': q, 'vue': vue,
    })


@login_required(login_url='login')
def medecins_departement_create(request):
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    Form = _departement_form_class()
    form = Form(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        obj = form.save()
        if is_ajax:
            return JsonResponse({'ok': True, 'message': f'Département « {obj} » créé.'})
        messages.success(request, 'Département créé.')
        return redirect('medecins_departements')
    template = 'medecins/config/departement_form_modal.html' if is_ajax else 'medecins/config/departement_form.html'
    return render(request, template, {
        'form': form, 'titre': 'Nouveau département',
    })


@login_required(login_url='login')
def medecins_departement_edit(request, pk):
    from medecins.models import Departement
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    obj  = get_object_or_404(Departement, pk=pk)
    Form = _departement_form_class()
    form = Form(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        if is_ajax:
            return JsonResponse({'ok': True, 'message': f'Département « {obj} » mis à jour.'})
        messages.success(request, 'Département mis à jour.')
        return redirect('medecins_departements')
    template = 'medecins/config/departement_form_modal.html' if is_ajax else 'medecins/config/departement_form.html'
    return render(request, template, {
        'form': form, 'titre': f'Modifier – {obj.nom}', 'obj': obj,
    })


@login_required(login_url='login')
@require_POST
def medecins_departement_bulk_delete(request):
    from medecins.models import Departement
    ids = request.POST.getlist('ids[]')
    if ids:
        Departement.objects.filter(pk__in=ids).delete()
    return JsonResponse({'ok': True})


_DEPT_HDR = ['code', 'nom', 'description', 'actif']


def _dept_row(d):
    return [d.code, d.nom, d.description, int(d.actif)]


@login_required(login_url='login')
def medecins_export_departements(request):
    from medecins.models import Departement
    from services.views import _export_file
    fmt = request.GET.get('format', 'json')
    qs = Departement.objects.order_by('nom')
    rows = [_dept_row(d) for d in qs]
    return _export_file(fmt, 'departements', _DEPT_HDR, rows,
                        [dict(zip(_DEPT_HDR, r)) for r in rows])


@login_required(login_url='login')
@require_POST
def medecins_import_departements(request):
    from medecins.models import Departement
    from services.views import _parse_upload, _s, _b

    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('medecins_departements')

    data, err = _parse_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('medecins_departements')

    do_update = 'update' in request.POST
    created = updated = skipped = errors = 0
    for item in data:
        try:
            code = _s(item.get('code', ''))
            if not code:
                errors += 1
                continue
            defaults = {
                'nom': _s(item.get('nom', code)),
                'description': _s(item.get('description', '')),
                'actif': _b(item.get('actif', True)),
            }
            obj, was_created = Departement.objects.get_or_create(code=code, defaults=defaults)
            if was_created:
                created += 1
            elif do_update:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                updated += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    if errors:
        messages.warning(request, f'{created} créé(s), {updated} mis à jour, {skipped} ignoré(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} département(s) importé(s), {updated} mis à jour, {skipped} ignoré(s).')
    return redirect('medecins_departements')


### Le CRUD "Services" a été déplacé vers employer/views.py (menu Configuration du module Employé).
