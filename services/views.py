import csv
import io
import json

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import (
    Articleservice, CategorieArticle, FamilleArticle, CompagniePharma,
    LigneFournisseurArticle, ConditionnementArticle, VarianteAttributArticle, ReglePrix,
    UniteMesure, CategorieUniteMesure,
)
from .forms import CategorieArticleForm, CategorieUniteMesureForm, UniteMesureForm
from achats.models import Fournisseur
from django.contrib.auth.models import User


@login_required
def services_list(request):
    qs = Articleservice.objects.select_related('categorie', 'famille').all()

    # Filtres
    q = request.GET.get('q', '').strip()
    categorie_id = request.GET.get('categorie', '')
    type_produit = request.GET.get('type_produit', '')
    statut = request.GET.get('statut', '')
    filtre = request.GET.get('filtre', '')
    vue = request.GET.get('vue', 'liste')  # kanban ou liste

    if q:
        qs = qs.filter(
            Q(nom__icontains=q) |
            Q(reference_interne__icontains=q) |
            Q(code_barres__icontains=q)
        )
    if categorie_id:
        qs = qs.filter(categorie_id=categorie_id)
    if type_produit:
        qs = qs.filter(type_produit_hospitalier=type_produit)
    if statut == 'actif':
        qs = qs.filter(actif=True)
    elif statut == 'inactif':
        qs = qs.filter(actif=False)
    if filtre == 'services':
        qs = qs.filter(type_produit_hospitalier='service')
    elif filtre == 'articles':
        qs = qs.exclude(type_produit_hospitalier='service')
    elif filtre == 'peut_etre_vendu':
        qs = qs.filter(peut_etre_vendu=True)
    elif filtre == 'peut_etre_achete':
        qs = qs.filter(peut_etre_achete=True)
    elif filtre == 'favori':
        qs = qs.filter(favori=True)
    elif filtre == 'avertissement':
        qs = qs.filter(Q(avertissement_grossesse=True) | Q(avertissement_lactation=True))
    elif filtre == 'archive':
        qs = qs.filter(actif=False)

    total = qs.count()
    paginator = Paginator(qs, 24 if vue == 'kanban' else 40)
    page_obj = paginator.get_page(request.GET.get('page'))

    categories = CategorieArticle.objects.all()
    type_produit_choices = Articleservice.TYPE_PRODUIT_CHOICES

    return render(request, 'services/list.html', {
        'page_obj': page_obj,
        'categories': categories,
        'type_produit_choices': type_produit_choices,
        'q': q,
        'categorie_id': categorie_id,
        'type_produit': type_produit,
        'statut': statut,
        'filtre': filtre,
        'vue': vue,
        'total': total,
    })


@login_required
def service_form(request, pk=None):
    article = get_object_or_404(Articleservice, pk=pk) if pk else None
    is_new = article is None

    if request.method == 'POST':
        data = request.POST

        # Champs en-tête
        nom = data.get('nom', '').strip()
        if not nom:
            messages.error(request, "Le nom de l'article est obligatoire.")
            return redirect(request.path)

        reference_interne = data.get('reference_interne', '').strip().upper() or None
        if reference_interne:
            doublon = Articleservice.objects.filter(reference_interne=reference_interne)
            if not is_new:
                doublon = doublon.exclude(pk=article.pk)
            if doublon.exists():
                messages.error(request, f"La référence interne « {reference_interne} » est déjà utilisée par un autre article.")
                return redirect(request.path)

        if is_new:
            article = Articleservice(cree_par=request.user)

        # En-tête
        article.nom = nom
        article.reference_interne = reference_interne
        article.favori = 'favori' in data
        article.peut_etre_vendu = 'peut_etre_vendu' in data
        article.peut_etre_achete = 'peut_etre_achete' in data

        # Onglet 1 — Détails médicament
        article.forme = data.get('forme', '')
        article.voie_administration = data.get('voie_administration', '')
        article.dosage = data.get('dosage', '')
        article.dosage_unite = data.get('dosage_unite', '')
        article.quantite_prescription_manuelle = 'quantite_prescription_manuelle' in data
        article.frequence = data.get('frequence', '')
        article.composant_actif = data.get('composant_actif', '')
        article.effet_therapeutique = data.get('effet_therapeutique', '')
        article.effets_indesirables = data.get('effets_indesirables', '')
        compagnie_id = data.get('compagnie_pharmaceutique')
        article.compagnie_pharmaceutique_id = compagnie_id if compagnie_id else None
        article.code_produit = data.get('code_produit', '')
        article.url_produit = data.get('url_produit', '')
        article.nom_produit_fabricant = data.get('nom_produit_fabricant', '')
        article.avertissement_grossesse = 'avertissement_grossesse' in data
        article.avertissement_lactation = 'avertissement_lactation' in data
        article.indications = data.get('indications', '')
        article.remarques = data.get('remarques', '')

        # Onglet 2 — Information générale
        article.type_article = data.get('type_article', 'consommable')
        article.type_produit_hospitalier = data.get('type_produit_hospitalier', '')
        article.politique_facturation = data.get('politique_facturation', 'qtes_commandees')
        article.refacturer_depenses = data.get('refacturer_depenses', 'non')
        article.unite_mesure_id = data.get('unite_mesure') or None
        article.unite_achat_id = data.get('unite_achat') or None
        article.prix_vente = data.get('prix_vente') or 0
        article.taxes_vente = data.get('taxes_vente', '')
        article.cout = data.get('cout') or 0
        categorie_id = data.get('categorie')
        article.categorie_id = categorie_id if categorie_id else None
        article.code_barres = data.get('code_barres', '')
        famille_id = data.get('famille')
        article.famille_id = famille_id if famille_id else None
        article.notes_internes = data.get('notes_internes', '')

        # Onglet 4 — Vente
        article.description_vente = data.get('description_vente', '')

        # Onglet 5 — Achats
        article.taxes_fournisseur = data.get('taxes_fournisseur', '')
        article.politique_controle = data.get('politique_controle', 'qtes_recues')
        article.description_achat = data.get('description_achat', '')

        # Stock consommable / stockable
        article.quantite_stock = data.get('quantite_stock') or 0
        article.quantite_alerte = data.get('quantite_alerte') or 0

        # Onglet 6 — Stock
        responsable_id = data.get('responsable')
        article.responsable_id = responsable_id if responsable_id else None
        article.poids = data.get('poids') or 0
        article.volume = data.get('volume') or 0
        article.delai_livraison_client = data.get('delai_livraison_client') or 0
        article.description_reception = data.get('description_reception', '')
        article.description_livraison = data.get('description_livraison', '')
        article.description_transfert = data.get('description_transfert', '')

        # Onglet 7 — Comptabilité
        article.compte_revenus = data.get('compte_revenus', '')
        article.compte_charges = data.get('compte_charges', '')
        article.compte_ecart_prix = data.get('compte_ecart_prix', '')

        if not is_new:
            article.actif = 'actif' in data

        # Photo
        if 'photo' in request.FILES:
            article.photo = request.FILES['photo']

        article.save()
        messages.success(request, f"Article « {article.nom} » enregistré avec succès.")
        return redirect('services:detail', pk=article.pk)

    # GET — préparer le contexte
    fournisseurs = Fournisseur.objects.filter(actif=True).order_by('nom')
    categories = CategorieArticle.objects.all()
    cat_codes_json = json.dumps({str(c.pk): c.code for c in categories})
    familles = FamilleArticle.objects.all()
    compagnies = CompagniePharma.objects.all()
    users = User.objects.filter(is_active=True).order_by('last_name')
    unites_mesure = UniteMesure.objects.filter(actif=True).select_related('categorie').order_by('nom')

    lignes_fournisseurs = article.fournisseurs.select_related('fournisseur').all() if article else []
    lignes_conditionnements = article.conditionnements.all() if article else []
    lignes_variantes = article.variantes.all() if article else []
    regles = article.regles_prix.all() if article else []

    return render(request, 'services/form.html', {
        'article': article,
        'is_new': is_new,
        'fournisseurs': fournisseurs,
        'categories': categories,
        'familles': familles,
        'compagnies': compagnies,
        'users': users,
        'unites_mesure': unites_mesure,
        'lignes_fournisseurs': lignes_fournisseurs,
        'lignes_conditionnements': lignes_conditionnements,
        'lignes_variantes': lignes_variantes,
        'regles': regles,
        'forme_choices': Articleservice.FORME_CHOICES,
        'voie_choices': Articleservice.VOIE_CHOICES,
        'type_article_choices': Articleservice.TYPE_ARTICLE_CHOICES,
        'type_produit_choices': Articleservice.TYPE_PRODUIT_CHOICES,
        'politique_fact_choices': Articleservice.POLITIQUE_FACT_CHOICES,
        'refacturer_choices': Articleservice.REFACTURER_CHOICES,
        'politique_controle_choices': Articleservice.POLITIQUE_CONTROLE_CHOICES,
        'cat_codes_json': cat_codes_json,
    })


@login_required
def service_detail(request, pk):
    article = get_object_or_404(Articleservice, pk=pk)
    return render(request, 'services/detail.html', {
        'article': article,
    })


@login_required
def regles_prix(request, pk):
    article = get_object_or_404(Articleservice, pk=pk)
    regles = article.regles_prix.all()
    return render(request, 'services/regles_prix.html', {
        'article': article,
        'regles': regles,
    })


# ── Vues AJAX pour les lignes dynamiques ─────────────────────

@login_required
@require_POST
def service_bulk_delete(request):
    ids = request.POST.getlist('ids[]')
    if ids:
        count, _ = Articleservice.objects.filter(pk__in=ids).delete()
        return JsonResponse({'ok': True, 'count': count})
    return JsonResponse({'ok': False, 'error': 'Aucun élément sélectionné'}, status=400)


@login_required
@require_POST
def categorie_bulk_delete(request):
    ids = request.POST.getlist('ids[]')
    if ids:
        count, _ = CategorieArticle.objects.filter(pk__in=ids).delete()
        return JsonResponse({'ok': True, 'count': count})
    return JsonResponse({'ok': False}, status=400)


@login_required
@require_POST
def unite_bulk_delete(request):
    ids = request.POST.getlist('ids[]')
    if ids:
        count, _ = UniteMesure.objects.filter(pk__in=ids).delete()
        return JsonResponse({'ok': True, 'count': count})
    return JsonResponse({'ok': False}, status=400)


# ── Catégories de service ──────────────────────────────────────────────────

@login_required
def categories_list(request):
    q = request.GET.get('q', '').strip()
    vue = request.GET.get('vue', 'liste')
    qs = CategorieArticle.objects.all()
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(code__icontains=q))
    total_all = CategorieArticle.objects.count()
    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'services/categories/list.html', {
        'page_obj': page_obj,
        'q': q,
        'vue': vue,
        'total': total_all,
        'total_filtre': qs.count(),
    })


@login_required
def categorie_create(request):
    if request.method == 'POST':
        form = CategorieArticleForm(request.POST)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f'Catégorie « {obj.nom} » créée.')
            return redirect('services:categories')
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        form = CategorieArticleForm()
    return render(request, 'services/categories/form.html', {
        'form': form,
        'titre': 'Nouvelle catégorie de service',
        'edit': False,
    })


@login_required
def categorie_edit(request, pk):
    obj = get_object_or_404(CategorieArticle, pk=pk)
    if request.method == 'POST':
        form = CategorieArticleForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, f'Catégorie « {obj.nom} » mise à jour.')
            return redirect('services:categorie_detail', pk=obj.pk)
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        form = CategorieArticleForm(instance=obj)
    return render(request, 'services/categories/form.html', {
        'form': form,
        'obj': obj,
        'titre': f'Modifier — {obj.nom}',
        'edit': True,
    })


@login_required
def categorie_detail(request, pk):
    obj = get_object_or_404(CategorieArticle, pk=pk)
    return render(request, 'services/categories/detail.html', {'obj': obj})


@login_required
def categorie_delete(request, pk):
    obj = get_object_or_404(CategorieArticle, pk=pk)
    if request.method == 'POST':
        nom = obj.nom
        obj.delete()
        messages.success(request, f'Catégorie « {nom} » supprimée.')
    return redirect('services:categories')


# ── Unités de mesure ───────────────────────────────────────────────────────

@login_required
def unites_list(request):
    q = request.GET.get('q', '').strip()
    categorie_id = request.GET.get('categorie', '')
    vue = request.GET.get('vue', 'liste')
    qs = UniteMesure.objects.select_related('categorie').all()
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(code__icontains=q))
    if categorie_id:
        qs = qs.filter(categorie_id=categorie_id)
    total_all = UniteMesure.objects.count()
    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'services/unites/list.html', {
        'page_obj': page_obj,
        'categories_um': CategorieUniteMesure.objects.all(),
        'q': q,
        'categorie_id': categorie_id,
        'vue': vue,
        'total': total_all,
        'total_filtre': qs.count(),
    })


@login_required
def unite_create(request):
    if request.method == 'POST':
        form = UniteMesureForm(request.POST)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f'Unité « {obj.nom} » créée.')
            return redirect('services:unites')
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        form = UniteMesureForm()
    return render(request, 'services/unites/form.html', {
        'form': form,
        'titre': 'Nouvelle unité de mesure',
        'edit': False,
    })


@login_required
def unite_edit(request, pk):
    obj = get_object_or_404(UniteMesure, pk=pk)
    if request.method == 'POST':
        form = UniteMesureForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, f'Unité « {obj.nom} » mise à jour.')
            return redirect('services:unite_detail', pk=obj.pk)
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        form = UniteMesureForm(instance=obj)
    return render(request, 'services/unites/form.html', {
        'form': form,
        'obj': obj,
        'titre': f'Modifier — {obj.nom}',
        'edit': True,
    })


@login_required
def unite_detail(request, pk):
    obj = get_object_or_404(UniteMesure, pk=pk)
    return render(request, 'services/unites/detail.html', {'obj': obj})


@login_required
def unite_delete(request, pk):
    obj = get_object_or_404(UniteMesure, pk=pk)
    if request.method == 'POST':
        nom = obj.nom
        obj.delete()
        messages.success(request, f'Unité « {nom} » supprimée.')
    return redirect('services:unites')


@login_required
def categories_unites_list(request):
    q = request.GET.get('q', '').strip()
    vue = request.GET.get('vue', 'liste')
    qs = CategorieUniteMesure.objects.all()
    if q:
        qs = qs.filter(nom__icontains=q)
    total_all = CategorieUniteMesure.objects.count()
    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'services/unites/categories/list.html', {
        'page_obj': page_obj,
        'q': q,
        'vue': vue,
        'total': total_all,
        'total_filtre': qs.count(),
    })


@login_required
def categorie_unite_create(request):
    if request.method == 'POST':
        form = CategorieUniteMesureForm(request.POST)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f'Catégorie « {obj.nom} » créée.')
            return redirect('services:categories_unites')
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        form = CategorieUniteMesureForm()
    return render(request, 'services/unites/categories/form.html', {
        'form': form,
        'titre': "Nouvelle catégorie d'unité",
        'edit': False,
    })


@login_required
def categorie_unite_detail(request, pk):
    obj = get_object_or_404(CategorieUniteMesure, pk=pk)
    return render(request, 'services/unites/categories/detail.html', {'obj': obj})


@login_required
def categorie_unite_edit(request, pk):
    obj = get_object_or_404(CategorieUniteMesure, pk=pk)
    if request.method == 'POST':
        form = CategorieUniteMesureForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, f'Catégorie « {obj.nom} » mise à jour.')
            return redirect('services:categorie_unite_detail', pk=obj.pk)
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        form = CategorieUniteMesureForm(instance=obj)
    return render(request, 'services/unites/categories/form.html', {
        'form': form,
        'obj': obj,
        'titre': f'Modifier — {obj.nom}',
        'edit': True,
    })


@login_required
def categorie_unite_delete(request, pk):
    obj = get_object_or_404(CategorieUniteMesure, pk=pk)
    if request.method == 'POST':
        nom = obj.nom
        obj.delete()
        messages.success(request, f'Catégorie « {nom} » supprimée.')
    return redirect('services:categories_unites')


@login_required
@require_POST
def categorie_unite_bulk_delete(request):
    ids = request.POST.getlist('ids[]')
    if ids:
        count, _ = CategorieUniteMesure.objects.filter(pk__in=ids).delete()
        return JsonResponse({'ok': True, 'count': count})
    return JsonResponse({'ok': False}, status=400)


@login_required
@require_POST
def ajax_add_fournisseur(request):
    article_id = request.POST.get('article_id')
    article = get_object_or_404(Articleservice, pk=article_id)
    fournisseur_id = request.POST.get('fournisseur')
    if not fournisseur_id:
        return JsonResponse({'error': 'Fournisseur requis'}, status=400)
    unite_id = request.POST.get('unite_mesure') or None
    ligne = LigneFournisseurArticle.objects.create(
        article=article,
        fournisseur_id=fournisseur_id,
        nom_article_fournisseur=request.POST.get('nom_article_fournisseur', ''),
        reference_fournisseur=request.POST.get('reference_fournisseur', ''),
        quantite_min=request.POST.get('quantite_min') or 1,
        unite_mesure_id=unite_id,
        prix=request.POST.get('prix') or 0,
        delai_livraison=request.POST.get('delai_livraison') or 0,
    )
    unite_nom = ligne.unite_mesure.nom if ligne.unite_mesure else '—'
    return JsonResponse({'id': ligne.pk, 'fournisseur': ligne.fournisseur.nom, 'prix': str(ligne.prix), 'unite_mesure': unite_nom})


@login_required
@require_POST
def ajax_add_conditionnement(request):
    article_id = request.POST.get('article_id')
    article = get_object_or_404(Articleservice, pk=article_id)
    unite_id = request.POST.get('unite_mesure') or None
    ligne = ConditionnementArticle.objects.create(
        article=article,
        conditionnement=request.POST.get('conditionnement', ''),
        quantite=request.POST.get('quantite') or 1,
        unite_mesure_id=unite_id,
        pour_vente='pour_vente' in request.POST,
        pour_achat='pour_achat' in request.POST,
    )
    unite_nom = ligne.unite_mesure.nom if ligne.unite_mesure else '—'
    return JsonResponse({'id': ligne.pk, 'conditionnement': ligne.conditionnement, 'unite_mesure': unite_nom})


@login_required
@require_POST
def ajax_add_variante(request):
    article_id = request.POST.get('article_id')
    article = get_object_or_404(Articleservice, pk=article_id)
    ligne = VarianteAttributArticle.objects.create(
        article=article,
        caracteristique=request.POST.get('caracteristique', ''),
        valeurs=request.POST.get('valeurs', ''),
    )
    return JsonResponse({'id': ligne.pk, 'caracteristique': ligne.caracteristique})


@login_required
@require_POST
def ajax_add_regle_prix(request):
    article_id = request.POST.get('article_id')
    article = get_object_or_404(Articleservice, pk=article_id)
    ligne = ReglePrix.objects.create(
        article=article,
        liste_prix=request.POST.get('liste_prix', ''),
        applique_sur=request.POST.get('applique_sur', ''),
        quantite_min=request.POST.get('quantite_min') or 1,
        prix=request.POST.get('prix') or 0,
    )
    return JsonResponse({'id': ligne.pk, 'liste_prix': ligne.liste_prix})


@login_required
@require_POST
def ajax_delete_ligne(request, model, pk):
    model_map = {
        'fournisseur': LigneFournisseurArticle,
        'conditionnement': ConditionnementArticle,
        'variante': VarianteAttributArticle,
        'regle': ReglePrix,
    }
    Model = model_map.get(model)
    if not Model:
        return JsonResponse({'error': 'Modèle inconnu'}, status=400)
    obj = get_object_or_404(Model, pk=pk)
    obj.delete()
    return JsonResponse({'ok': True})


# ═══════════════════════════════════════════════════════════════════════════════
#  EXPORT / IMPORT
# ═══════════════════════════════════════════════════════════════════════════════

# ── Helpers partagés ────────────────────────────────────────────────────────

def _csv_dl(filename, headers, rows):
    resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    resp['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    w = csv.writer(resp)
    w.writerow(headers)
    for row in rows:
        w.writerow(['' if v is None else v for v in row])
    return resp


def _xlsx_dl(filename, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = filename[:31]
    fill = PatternFill(start_color='1F6E8C', end_color='1F6E8C', fill_type='solid')
    fnt  = Font(color='FFFFFF', bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill, cell.font = fill, fnt
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        ws.append(['' if v is None else str(v) for v in row])
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 55)
    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return resp


def _json_dl(filename, data):
    resp = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2, default=str),
        content_type='application/json',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}.json"'
    return resp


def _export_file(fmt, filename, headers, rows, json_data):
    if fmt == 'csv':
        return _csv_dl(filename, headers, rows)
    if fmt == 'xlsx':
        return _xlsx_dl(filename, headers, rows)
    return _json_dl(filename, json_data)


def _parse_upload(upload):
    name = upload.name.lower()
    try:
        if name.endswith('.json'):
            return json.loads(upload.read().decode('utf-8')), None
        if name.endswith('.csv'):
            text = upload.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            return list(reader), None
        if name.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(io.BytesIO(upload.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return None, 'Fichier Excel vide.'
            hdrs = [str(h) if h is not None else '' for h in rows[0]]
            data = [dict(zip(hdrs, r)) for r in rows[1:] if any(v is not None for v in r)]
            return data, None
        return None, 'Format non supporté (.json, .csv ou .xlsx uniquement)'
    except Exception as e:
        return None, f'Erreur lecture fichier : {e}'


def _s(v):
    return str(v).strip() if v is not None else ''


def _b(v):
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ('1', 'true', 'oui', 'yes')


_TYPE_ART_MAP = {
    'service': 'prestation', 'prestation': 'prestation',
    'consommable': 'consommable', 'stockable': 'stockable', 'autre': 'autre',
}

# ── Export Articles ──────────────────────────────────────────────────────────

_ART_HDR = [
    'reference_interne', 'nom', 'prix_vente', 'cout',
    'type_article', 'type_produit_hospitalier',
    'actif', 'peut_etre_vendu', 'peut_etre_achete',
    'categorie', 'unite_mesure', 'unite_achat', 'code_barres',
    'forme', 'voie_administration', 'dosage', 'dosage_unite',
    'composant_actif', 'effet_therapeutique', 'indications',
    'avertissement_grossesse', 'avertissement_lactation',
    'quantite_stock', 'quantite_alerte',
    'politique_facturation', 'refacturer_depenses', 'politique_controle',
    'notes_internes', 'compte_revenus', 'compte_charges', 'compte_ecart_prix',
]


def _art_row(a):
    return [
        a.reference_interne, a.nom, a.prix_vente, a.cout,
        a.type_article, a.type_produit_hospitalier,
        int(a.actif), int(a.peut_etre_vendu), int(a.peut_etre_achete),
        a.categorie.code if a.categorie else '',
        a.unite_mesure.code if a.unite_mesure else '',
        a.unite_achat.code if a.unite_achat else '',
        a.code_barres, a.forme, a.voie_administration,
        a.dosage, a.dosage_unite, a.composant_actif,
        a.effet_therapeutique, a.indications,
        int(a.avertissement_grossesse), int(a.avertissement_lactation),
        a.quantite_stock, a.quantite_alerte,
        a.politique_facturation, a.refacturer_depenses, a.politique_controle,
        a.notes_internes, a.compte_revenus, a.compte_charges, a.compte_ecart_prix,
    ]


@login_required
def export_articles(request):
    fmt = request.GET.get('format', 'json')
    qs  = Articleservice.objects.select_related('categorie', 'unite_mesure', 'unite_achat')
    rows = [_art_row(a) for a in qs]
    return _export_file(fmt, 'prestations', _ART_HDR, rows,
                        [dict(zip(_ART_HDR, r)) for r in rows])


# ── Export Catégories articles ───────────────────────────────────────────────

_CAT_HDR = [
    'code', 'nom', 'description', 'parent',
    'methode_cout', 'valorisation_inventaire', 'reservation_conditionnement',
    'bloquer_serie_lot', 'routes', 'strategie_enlevement',
    'sequence_code_barres', 'compte_revenus', 'compte_charges',
]


def _cat_row(c):
    return [
        c.code, c.nom, c.description,
        c.parent.code if c.parent else '',
        c.methode_cout, c.valorisation_inventaire, c.reservation_conditionnement,
        int(c.bloquer_serie_lot), c.routes, c.strategie_enlevement,
        c.sequence_code_barres, c.compte_revenus, c.compte_charges,
    ]


@login_required
def export_categories(request):
    fmt = request.GET.get('format', 'json')
    qs  = CategorieArticle.objects.select_related('parent')
    rows = [_cat_row(c) for c in qs]
    return _export_file(fmt, 'categories_articles', _CAT_HDR, rows,
                        [dict(zip(_CAT_HDR, r)) for r in rows])


# ── Export Unités de mesure ──────────────────────────────────────────────────

_UM_HDR = ['code', 'nom', 'categorie', 'type_unite', 'ratio', 'precision_arrondi', 'actif']


def _um_row(u):
    return [
        u.code, u.nom,
        u.categorie.nom if u.categorie else '',
        u.type_unite, float(u.ratio), float(u.precision_arrondi), int(u.actif),
    ]


@login_required
def export_unites(request):
    fmt = request.GET.get('format', 'json')
    qs  = UniteMesure.objects.select_related('categorie')
    rows = [_um_row(u) for u in qs]
    return _export_file(fmt, 'unites_mesure', _UM_HDR, rows,
                        [dict(zip(_UM_HDR, r)) for r in rows])


# ── Export Catégories unités ─────────────────────────────────────────────────

_CU_HDR = ['nom']


@login_required
def export_categories_unites(request):
    fmt = request.GET.get('format', 'json')
    qs  = CategorieUniteMesure.objects.all()
    rows = [[c.nom] for c in qs]
    return _export_file(fmt, 'categories_unites', _CU_HDR, rows,
                        [{'nom': c.nom} for c in qs])


# ── Import Articles ──────────────────────────────────────────────────────────

@login_required
@require_POST
def import_articles(request):
    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('services:list')

    data, err = _parse_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('services:list')

    do_update = 'update' in request.POST
    created = updated = skipped = errors = 0

    for item in data:
        try:
            ref = _s(item.get('reference_interne', '')).upper()
            nom = _s(item.get('nom', ''))
            if not ref and not nom:
                errors += 1
                continue

            cat = CategorieArticle.objects.filter(
                code=_s(item.get('categorie', ''))
            ).first() if item.get('categorie') else None

            um = UniteMesure.objects.filter(
                code=_s(item.get('unite_mesure', ''))
            ).first() if item.get('unite_mesure') else None

            ua = UniteMesure.objects.filter(
                code=_s(item.get('unite_achat', ''))
            ).first() if item.get('unite_achat') else None

            defaults = {
                'nom': nom,
                'prix_vente': item.get('prix_vente') or 0,
                'cout': item.get('cout') or 0,
                'type_article': _TYPE_ART_MAP.get(_s(item.get('type_article', '')), 'prestation'),
                'type_produit_hospitalier': _s(item.get('type_produit_hospitalier', '')),
                'actif': _b(item.get('actif', True)),
                'peut_etre_vendu': _b(item.get('peut_etre_vendu', True)),
                'peut_etre_achete': _b(item.get('peut_etre_achete', False)),
                'categorie': cat, 'unite_mesure': um, 'unite_achat': ua,
                'code_barres': _s(item.get('code_barres', '')),
                'forme': _s(item.get('forme', '')),
                'voie_administration': _s(item.get('voie_administration', '')),
                'dosage': _s(item.get('dosage', '')),
                'dosage_unite': _s(item.get('dosage_unite', '')),
                'composant_actif': _s(item.get('composant_actif', '')),
                'effet_therapeutique': _s(item.get('effet_therapeutique', '')),
                'indications': _s(item.get('indications', '')),
                'avertissement_grossesse': _b(item.get('avertissement_grossesse', False)),
                'avertissement_lactation': _b(item.get('avertissement_lactation', False)),
                'quantite_stock': int(item.get('quantite_stock') or 0),
                'quantite_alerte': int(item.get('quantite_alerte') or 0),
                'politique_facturation': _s(item.get('politique_facturation', 'qtes_commandees')),
                'refacturer_depenses': _s(item.get('refacturer_depenses', 'non')),
                'politique_controle': _s(item.get('politique_controle', 'qtes_recues')),
                'notes_internes': _s(item.get('notes_internes', '')),
                'compte_revenus': _s(item.get('compte_revenus', '')),
                'compte_charges': _s(item.get('compte_charges', '')),
                'compte_ecart_prix': _s(item.get('compte_ecart_prix', '')),
            }

            if ref:
                obj, was_created = Articleservice.objects.get_or_create(
                    reference_interne=ref, defaults=defaults)
            else:
                obj, was_created = Articleservice.objects.get_or_create(
                    nom=nom, defaults=defaults)

            if was_created:
                created += 1
            elif do_update:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                updated += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    if errors:
        messages.warning(request, f'{created} créé(s), {updated} mis à jour, {skipped} ignoré(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} prestation(s) importée(s), {updated} mise(s) à jour, {skipped} ignorée(s).')
    return redirect('services:list')


# ── Import Catégories articles ───────────────────────────────────────────────

@login_required
@require_POST
def import_categories(request):
    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('services:categories')

    data, err = _parse_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('services:categories')

    do_update = 'update' in request.POST
    id_to_code = {item.get('id'): item.get('code') for item in data if item.get('id')}
    created = updated = skipped = errors = 0

    for item in data:
        try:
            code = _s(item.get('code', ''))
            if not code:
                errors += 1
                continue
            defaults = {
                'nom': _s(item.get('nom', code)),
                'description': _s(item.get('description', '')),
                'methode_cout': _s(item.get('methode_cout', 'prix_standard')),
                'valorisation_inventaire': _s(item.get('valorisation_inventaire', 'manuelle')),
                'reservation_conditionnement': _s(item.get('reservation_conditionnement', 'partiels')),
                'bloquer_serie_lot': _b(item.get('bloquer_serie_lot', False)),
                'routes': _s(item.get('routes', '')),
                'strategie_enlevement': _s(item.get('strategie_enlevement', '')),
                'sequence_code_barres': _s(item.get('sequence_code_barres', '')),
                'compte_revenus': _s(item.get('compte_revenus', '')),
                'compte_charges': _s(item.get('compte_charges', '')),
                'parent': None,
            }
            obj, was_created = CategorieArticle.objects.get_or_create(code=code, defaults=defaults)
            if was_created:
                created += 1
            elif do_update:
                for k, v in defaults.items():
                    if k == 'parent':
                        continue
                    setattr(obj, k, v)
                obj.save()
                updated += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    # Passe 2 — résolution des parents
    for item in data:
        raw_parent = item.get('parent') or item.get('parent_id')
        if not raw_parent:
            continue
        try:
            code = _s(item.get('code', ''))
            parent_code = id_to_code.get(raw_parent) or id_to_code.get(int(raw_parent)) or _s(raw_parent)
            if parent_code:
                enfant = CategorieArticle.objects.filter(code=code).first()
                parent = CategorieArticle.objects.filter(code=parent_code).first()
                if enfant and parent and enfant.parent_id != parent.pk:
                    enfant.parent = parent
                    enfant.save()
        except Exception:
            pass

    if errors:
        messages.warning(request, f'{created} créée(s), {updated} mise(s) à jour, {skipped} ignorée(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} catégorie(s) importée(s), {updated} mise(s) à jour, {skipped} ignorée(s).')
    return redirect('services:categories')


# ── Import Unités de mesure ──────────────────────────────────────────────────

@login_required
@require_POST
def import_unites(request):
    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('services:unites')

    data, err = _parse_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('services:unites')

    do_update = 'update' in request.POST
    created = updated = skipped = errors = 0

    for item in data:
        try:
            code = _s(item.get('code', ''))
            if not code:
                errors += 1
                continue
            cat_nom = _s(item.get('categorie', ''))
            cat = None
            if cat_nom:
                cat, _ = CategorieUniteMesure.objects.get_or_create(nom=cat_nom)
            defaults = {
                'nom': _s(item.get('nom', code)),
                'categorie': cat,
                'type_unite': _s(item.get('type_unite', 'umrc')),
                'ratio': item.get('ratio') or 1,
                'precision_arrondi': item.get('precision_arrondi') or 0.01,
                'actif': _b(item.get('actif', True)),
            }
            obj, was_created = UniteMesure.objects.get_or_create(code=code, defaults=defaults)
            if was_created:
                created += 1
            elif do_update:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                updated += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    if errors:
        messages.warning(request, f'{created} créée(s), {updated} mise(s) à jour, {skipped} ignorée(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} unité(s) importée(s), {updated} mise(s) à jour, {skipped} ignorée(s).')
    return redirect('services:unites')


# ── Import Catégories unités ─────────────────────────────────────────────────

@login_required
@require_POST
def import_categories_unites(request):
    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('services:categories_unites')

    data, err = _parse_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('services:categories_unites')

    created = skipped = errors = 0
    for item in data:
        try:
            nom = _s(item.get('nom', ''))
            if not nom:
                errors += 1
                continue
            _, was_created = CategorieUniteMesure.objects.get_or_create(nom=nom)
            if was_created:
                created += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    if errors:
        messages.warning(request, f'{created} créée(s), {skipped} ignorée(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} catégorie(s) importée(s), {skipped} ignorée(s).')
    return redirect('services:categories_unites')
