"""
Catalogue des rapports du module.

Chaque rapport sait construire ses propres colonnes/lignes pour une période
donnée (periode_debut, periode_fin : objets date ; periode_debut peut être
None pour une recherche « jusqu'à une date ») ; le moteur d'export (voir
export.py) se charge ensuite de produire le fichier Excel/CSV.
"""

from django.core.exceptions import ObjectDoesNotExist


def _minutes(secondes):
    """Nombre brut de minutes (arrondi), pour une colonne de rapport triable/sommable."""
    return round(secondes / 60) if secondes is not None else 'Non applicable'


def _date_sortie(hosp):
    if hosp.heure_sortie:
        return hosp.heure_sortie.strftime('%d/%m/%Y %H:%M')
    return 'En cours' if hosp.statut == 'hospitalise' else 'Non applicable'


def _patients_inscrits(periode_debut, periode_fin):
    from patients.models import Patient
    qs = Patient.objects.all()
    if periode_debut:
        qs = qs.filter(date_creation__date__gte=periode_debut)
    qs = qs.filter(date_creation__date__lte=periode_fin).order_by('date_creation')
    columns = ['Code patient', 'Ville', "Date d'inscription", 'Nom', 'Prénoms', 'Sexe',
               'Date de naissance', 'Âge', 'Téléphone']
    rows = [[
        p.code_patient, p.ville, p.date_creation.strftime('%d/%m/%Y'),
        p.nom, p.prenoms, p.sexe, p.date_naissance.strftime('%d/%m/%Y'), p.age, p.telephone,
    ] for p in qs]
    return columns, rows


# ── Listing activité soins (Mise en Observation) ────────────────────────────
# Rapport riche : statut de facture + origine (directe/consultation médecine/
# gynéco) + tranches d'âge façon fiche papier, avec une feuille récapitulative
# âge × sexe.

def _facture_statut_hospitalisation(hosp):
    """Payé / Non payé / Partiellement payé / Annulée / Aucune facture."""
    from facturation.models import Facture
    factures = list(Facture.objects.filter(hospitalisation=hosp))
    if not factures:
        return 'Aucune facture'
    actives = [f for f in factures if f.statut != 'annulee']
    if not actives:
        return 'Annulée'
    payees = [f for f in actives if f.statut == 'payee']
    if len(payees) == len(actives):
        return 'Payé'
    if not payees:
        return 'Non payé'
    return 'Partiellement payé'


def _origine_hospitalisation(hosp):
    """Directe / Consultation (MDGN) / Consultation (GYN).

    Rattachement par département (module spécialisé 'gynecologie') ou, à
    défaut, par la spécialité du médecin de la consultation."""
    if not hosp.consultation_id:
        return 'Directe'
    consultation = hosp.consultation
    rdv = consultation.rendez_vous
    departement = rdv.departement if rdv else None
    medecin = consultation.medecin
    is_gyneco = False
    if departement and departement.code == 'GYN':
        is_gyneco = True
    elif medecin and medecin.specialite and 'gyn' in medecin.specialite.nom.lower():
        is_gyneco = True
    return 'Consultation (GYN)' if is_gyneco else 'Consultation (MDGN)'


AGE_BRACKETS_MO = ['0-11 mois', '1-4 ans', '5-9 ans', '10-14 ans',
                    '15-19 ans', '19-24 ans', '25-49 ans', '50 ans et plus']


def _age_bracket_mo(date_naissance, reference_date):
    """Reproduit les tranches de la fiche d'activité MO papier (y compris le
    chevauchement 15-19/19-24 présent sur le formulaire d'origine — on bascule
    à 20 ans dans la tranche '19-24 ans' pour ne pas compter deux fois)."""
    if not date_naissance or not reference_date:
        return None
    jours = (reference_date - date_naissance).days
    if jours < 0:
        return None
    mois = jours / 30.4368
    ans = jours / 365.25
    if mois < 12:
        return AGE_BRACKETS_MO[0]
    if ans < 5:
        return AGE_BRACKETS_MO[1]
    if ans < 10:
        return AGE_BRACKETS_MO[2]
    if ans < 15:
        return AGE_BRACKETS_MO[3]
    if ans < 20:
        return AGE_BRACKETS_MO[4]
    if ans < 25:
        return AGE_BRACKETS_MO[5]
    if ans < 50:
        return AGE_BRACKETS_MO[6]
    return AGE_BRACKETS_MO[7]


def _recap_mo_donnees(periode_debut, periode_fin):
    """Agrège les hospitalisations dont la facture a été créée ET payée,
    par tranche d'âge et sexe : nombre de MO et heures cumulées d'observation."""
    from hospitalisation.models import Hospitalisation
    qs = Hospitalisation.objects.select_related('patient')
    if periode_debut:
        qs = qs.filter(date_admission__date__gte=periode_debut)
    qs = qs.filter(date_admission__date__lte=periode_fin)

    nb = {b: {'F': 0, 'M': 0} for b in AGE_BRACKETS_MO}
    heures = {b: {'F': 0.0, 'M': 0.0} for b in AGE_BRACKETS_MO}

    for h in qs:
        if _facture_statut_hospitalisation(h) != 'Payé':
            continue
        patient = h.patient
        ref = h.date_admission.date() if h.date_admission else None
        bracket = _age_bracket_mo(patient.date_naissance, ref)
        sexe = patient.sexe
        if not bracket or sexe not in ('F', 'M'):
            continue
        nb[bracket][sexe] += 1
        if h.duree_observation is not None:
            heures[bracket][sexe] += h.duree_observation / 3600
    return nb, heures


def _populate_recap_mo_sheet(ws, periode_debut, periode_fin, genere_le, nb, heures):
    """Feuille récap façon fiche papier : tableau âge × sexe (Nombre de MO /
    Nombre de jours de MO en hrs)."""
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from .export import H_TITRE, H_ENTETE, H_STD, H_META, POLICE, _mois_texte, _periode_texte

    ws.title = 'Récapitulatif'

    n_cols = 1 + len(AGE_BRACKETS_MO) * 2 + 2
    gray_fill = PatternFill(start_color='FFE0E0E0', end_color='FFE0E0E0', fill_type='solid')
    bold = Font(name=POLICE, bold=True)
    regular = Font(name=POLICE)
    title_font = Font(name=POLICE, bold=True, size=14)
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='FFB0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value='RÉCAPITULATIF DES ACTIVITÉS DE SOINS MO')
    c.fill = gray_fill; c.font = title_font; c.alignment = center
    ws.row_dimensions[1].height = H_TITRE
    ws.row_dimensions[2].height = H_STD

    mid = n_cols // 2
    mois_txt = _mois_texte(periode_debut, periode_fin)
    genere_txt = genere_le.strftime('%d/%m/%Y %H:%M') if genere_le else ''
    for col in range(1, n_cols + 1):
        ws.cell(row=3, column=col).fill = gray_fill
    ws.cell(row=3, column=1, value="Mois :").font = bold
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=mid)
    ws.cell(row=3, column=2, value=mois_txt).font = regular
    ws.cell(row=3, column=mid + 1, value="Généré le :").font = bold
    ws.merge_cells(start_row=3, start_column=mid + 2, end_row=3, end_column=n_cols)
    ws.cell(row=3, column=mid + 2, value=genere_txt).font = regular
    ws.row_dimensions[3].height = H_META

    periode_txt = _periode_texte(periode_debut, periode_fin)
    for col in range(1, n_cols + 1):
        ws.cell(row=4, column=col).fill = gray_fill
    ws.cell(row=4, column=1, value="Période :").font = bold
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=n_cols)
    ws.cell(row=4, column=2, value=periode_txt).font = regular
    ws.row_dimensions[4].height = H_META
    ws.row_dimensions[5].height = H_STD

    # ── En-tête à deux niveaux (tranches d'âge, puis F/M) ──
    r1, r2 = 6, 7
    ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
    cell = ws.cell(row=r1, column=1, value='ACTIVITÉ')
    cell.fill = gray_fill; cell.font = bold; cell.alignment = center; cell.border = border

    col = 2
    for bracket in AGE_BRACKETS_MO:
        ws.merge_cells(start_row=r1, start_column=col, end_row=r1, end_column=col + 1)
        c1 = ws.cell(row=r1, column=col, value=bracket)
        c1.fill = gray_fill; c1.font = bold; c1.alignment = center; c1.border = border
        for j, sx in enumerate(('F', 'M')):
            c2 = ws.cell(row=r2, column=col + j, value=sx)
            c2.fill = gray_fill; c2.font = bold; c2.alignment = center; c2.border = border
        col += 2
    ws.merge_cells(start_row=r1, start_column=col, end_row=r1, end_column=col + 1)
    ct = ws.cell(row=r1, column=col, value='TOTAL')
    ct.fill = gray_fill; ct.font = bold; ct.alignment = center; ct.border = border
    for j, sx in enumerate(('F', 'M')):
        c2 = ws.cell(row=r2, column=col + j, value=sx)
        c2.fill = gray_fill; c2.font = bold; c2.alignment = center; c2.border = border
    ws.row_dimensions[r1].height = H_ENTETE
    ws.row_dimensions[r2].height = H_ENTETE

    # ── Lignes de données ──
    row_nb, row_h = r2 + 1, r2 + 2
    lbl_nb = ws.cell(row=row_nb, column=1, value='Nombre de MO')
    lbl_nb.font = bold; lbl_nb.border = border
    lbl_h = ws.cell(row=row_h, column=1, value='Nombre de jours de MO ( hrs)')
    lbl_h.font = bold; lbl_h.border = border

    col = 2
    total_nb = {'F': 0, 'M': 0}
    total_h = {'F': 0.0, 'M': 0.0}
    for bracket in AGE_BRACKETS_MO:
        for j, sx in enumerate(('F', 'M')):
            v_nb = nb[bracket][sx]
            v_h = round(heures[bracket][sx], 1)
            cell_nb = ws.cell(row=row_nb, column=col + j, value=v_nb or None)
            cell_nb.font = regular; cell_nb.alignment = center; cell_nb.border = border
            cell_h = ws.cell(row=row_h, column=col + j, value=v_h or None)
            cell_h.font = regular; cell_h.alignment = center; cell_h.border = border
            total_nb[sx] += v_nb
            total_h[sx] += v_h
        col += 2
    for j, sx in enumerate(('F', 'M')):
        cell_nb = ws.cell(row=row_nb, column=col + j, value=total_nb[sx] or None)
        cell_nb.font = bold; cell_nb.alignment = center; cell_nb.border = border
        cell_h = ws.cell(row=row_h, column=col + j, value=round(total_h[sx], 1) or None)
        cell_h.font = bold; cell_h.alignment = center; cell_h.border = border
    ws.row_dimensions[row_nb].height = H_STD * 1.6
    ws.row_dimensions[row_h].height = H_STD * 1.6

    ws.column_dimensions['A'].width = 26
    for i in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 7


def _hospitalisations_mo(periode_debut, periode_fin):
    from hospitalisation.models import Hospitalisation
    qs = Hospitalisation.objects.all()
    if periode_debut:
        qs = qs.filter(date_admission__date__gte=periode_debut)
    qs = qs.filter(date_admission__date__lte=periode_fin).select_related(
        'patient', 'consultation', 'consultation__rendez_vous__departement',
        'consultation__medecin__specialite',
    ).order_by('date_admission')
    columns = ['Code hospitalisation', 'Code patient', 'Nom', 'Prénoms', 'Âge', 'Sexe',
               "Date d'admission", 'Temps avant hospitalisation (min)',
               "Durée d'hospitalisation (min)", 'Date de sortie',
               'Facture', 'Origine', 'Statut']
    rows = [[
        h.numero, h.patient.code_patient, h.patient.nom, h.patient.prenoms, h.patient.age, h.patient.sexe,
        h.date_admission.strftime('%d/%m/%Y %H:%M') if h.date_admission else 'Non applicable',
        _minutes(h.temps_avant_hospitalisation),
        _minutes(h.duree_observation),
        _date_sortie(h),
        _facture_statut_hospitalisation(h),
        _origine_hospitalisation(h),
        h.get_statut_display(),
    ] for h in qs]
    return columns, rows


def build_listing_activite_mo_xlsx(periode_debut, periode_fin, genere_le):
    """Classeur à 2 feuilles pour le rapport MO : listing détaillé + récapitulatif
    (tableau âge/sexe), comme la fiche d'activité papier."""
    import io
    from openpyxl import Workbook
    from .export import populate_sheet

    wb = Workbook()
    columns, rows = _hospitalisations_mo(periode_debut, periode_fin)
    populate_sheet(wb.active, 'Listing activité soins (MO)', columns, rows, periode_debut, periode_fin, genere_le)

    ws_recap = wb.create_sheet()
    nb, heures = _recap_mo_donnees(periode_debut, periode_fin)
    _populate_recap_mo_sheet(ws_recap, periode_debut, periode_fin, genere_le, nb, heures)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Listing activité soins infirmiers (SN) ──────────────────────────────────
# Rapport détaillé des procédures de soins (soins.ProcedureSoin — le
# formulaire « Liste des soins »/« Procédure de soin ») : une ligne par soin
# effectué sur la période. La feuille récapitulative (comptage par type de
# soin, façon fiche papier) sera ajoutée dans un second temps.

def _facture_statut_procedure(proc):
    """Payé / Non payé / Partiellement payé / Annulée / Aucune facture."""
    if not proc.facture_id:
        return 'Aucune facture'
    facture = proc.facture
    if facture.statut == 'annulee':
        return 'Annulée'
    if facture.statut == 'payee':
        return 'Payé'
    if facture.montant_paye and facture.montant_paye > 0:
        return 'Partiellement payé'
    return 'Non payé'


def _origine_procedure_soin(proc):
    """Hospitalisation / Rendez-vous / Soins direct, selon le lien porté par la procédure
    (même logique que _origine_hospitalisation pour le rapport MO)."""
    if proc.soin_id and proc.soin.hospitalisation_id:
        return 'Hospitalisation'
    if proc.rendez_vous_id:
        return 'Rendez-vous'
    if proc.soin_id:
        return 'Soins direct'
    return 'Non renseigné'


def _listing_activite_soins_sn(periode_debut, periode_fin):
    from soins.models import ProcedureSoin
    qs = ProcedureSoin.objects.select_related(
        'patient', 'infirmier', 'soin_type', 'departement', 'facture',
        'soin__hospitalisation', 'rendez_vous',
    )
    if periode_debut:
        qs = qs.filter(date__date__gte=periode_debut)
    qs = qs.filter(date__date__lte=periode_fin).order_by('date')
    columns = ['Numéro', 'Code patient', 'Nom', 'Prénoms', 'Âge', 'Sexe',
               'Type de soin', 'Date', 'Infirmier', 'Département',
               'Prix', 'Facture', 'Origine', 'Statut']
    rows = [[
        proc.numero,
        proc.patient.code_patient, proc.patient.nom, proc.patient.prenoms,
        proc.patient.age, proc.patient.sexe,
        proc.soin_type.nom if proc.soin_type_id else 'Non renseigné',
        proc.date.strftime('%d/%m/%Y %H:%M') if proc.date else '',
        proc.infirmier.nom_complet if proc.infirmier_id else 'Non renseigné',
        proc.departement.nom if proc.departement_id else 'Non renseigné',
        float(proc.prix),
        _facture_statut_procedure(proc),
        _origine_procedure_soin(proc),
        proc.get_statut_display(),
    ] for proc in qs]
    return columns, rows


def _recap_soins_par_type(periode_debut, periode_fin):
    """Comptage des soins facturés ET payés, par type de soin distinct (pas de
    doublon : un type de soin = une ligne), pour la feuille récapitulative."""
    from soins.models import ProcedureSoin
    from django.db.models import Count

    qs = ProcedureSoin.objects.filter(facture__statut='payee', soin_type__isnull=False)
    if periode_debut:
        qs = qs.filter(date__date__gte=periode_debut)
    qs = qs.filter(date__date__lte=periode_fin)

    resultats = (
        qs.values('soin_type__nom')
        .annotate(n=Count('id'))
        .order_by('soin_type__nom')
    )
    columns = ['Type de soin', 'Nombre']
    rows = [[r['soin_type__nom'], r['n']] for r in resultats]
    return columns, rows


def build_listing_activite_soins_sn_xlsx(periode_debut, periode_fin, genere_le):
    """Classeur à 2 feuilles pour le rapport SN : listing détaillé + récapitulatif
    (comptage par type de soin, soins facturés et payés uniquement)."""
    import io
    from openpyxl import Workbook
    from .export import populate_sheet

    wb = Workbook()
    columns, rows = _listing_activite_soins_sn(periode_debut, periode_fin)
    populate_sheet(wb.active, 'Listing des soins infirmiers', columns, rows, periode_debut, periode_fin, genere_le)

    ws_recap = wb.create_sheet()
    columns_recap, rows_recap = _recap_soins_par_type(periode_debut, periode_fin)
    populate_sheet(ws_recap, 'Récapitulatif', columns_recap, rows_recap, periode_debut, periode_fin, genere_le)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _transactions_caisse(periode_debut, periode_fin):
    from caisse.models import TransactionCaisse
    qs = TransactionCaisse.objects.filter(
        date_transaction__date__range=[periode_debut, periode_fin]
    ).select_related('session__caisse', 'cree_par').order_by('date_transaction')
    columns = ['Numéro', 'Caisse', 'Type', 'Mode de paiement', 'Montant (FCFA)', 'Créé par', 'Date']
    rows = [[
        t.numero, str(t.session.caisse) if t.session_id else '—', t.get_type_transaction_display(),
        t.get_mode_paiement_display(), float(t.montant), str(t.cree_par) if t.cree_par else '—',
        t.date_transaction.strftime('%d/%m/%Y %H:%M'),
    ] for t in qs]
    return columns, rows


def _mouvements_stock_pharmacie(periode_debut, periode_fin):
    # Les mouvements de stock pharmacie réels passent tous par
    # MouvementPharmacie (lié à stock.Produit) — MouvementStock (lié à
    # l'ancien modèle Medicament) n'est écrit par aucun code actif du module
    # et ne contiendrait donc jamais de données pertinentes ici.
    from pharmacie.models import MouvementPharmacie
    qs = MouvementPharmacie.objects.filter(
        date__date__range=[periode_debut, periode_fin]
    ).select_related('produit', 'cree_par').order_by('date')
    columns = ['Pharmacie', 'Produit', 'Type', 'Quantité', 'Stock avant', 'Stock après', 'Référence', 'Créé par', 'Date']
    rows = [[
        m.get_pharmacie_display(), str(m.produit), m.get_type_display(), m.quantite,
        m.stock_avant, m.stock_apres, m.reference or '—',
        str(m.cree_par) if m.cree_par else '—',
        m.date.strftime('%d/%m/%Y %H:%M'),
    ] for m in qs]
    return columns, rows


def _employes_embauches(periode_debut, periode_fin):
    from employer.models import Employe
    qs = Employe.objects.filter(
        date_embauche__range=[periode_debut, periode_fin]
    ).select_related('service', 'fonction').order_by('date_embauche')
    columns = ['Matricule', 'Nom', 'Prénoms', 'Service', 'Fonction', "Date d'embauche", 'Statut']
    rows = [[
        e.matricule, e.nom, e.prenoms, str(e.service) if e.service else '—',
        str(e.fonction) if e.fonction else '—', e.date_embauche.strftime('%d/%m/%Y'), e.get_statut_display(),
    ] for e in qs]
    return columns, rows


def _conges(periode_debut, periode_fin):
    from employer.models import Conge
    qs = Conge.objects.filter(
        date_debut__range=[periode_debut, periode_fin]
    ).select_related('employe').order_by('date_debut')
    columns = ['Employé', 'Type de congé', 'Du', 'Au', 'Durée (jours)', 'Statut']
    rows = [[
        str(c.employe), c.get_type_conge_display(), c.date_debut.strftime('%d/%m/%Y'),
        c.date_fin.strftime('%d/%m/%Y'), c.duree, c.get_statut_display(),
    ] for c in qs]
    return columns, rows


def _presences(periode_debut, periode_fin):
    from employer.models import Presence
    qs = Presence.objects.filter(
        date__range=[periode_debut, periode_fin]
    ).select_related('employe').order_by('date')
    columns = ['Employé', 'Date', 'Arrivée matin', 'Départ matin', 'Arrivée soir', 'Départ soir',
               'Présent', 'Durée totale (min)']
    rows = [[
        str(p.employe), p.date.strftime('%d/%m/%Y'),
        p.heure_arrivee_matin.strftime('%H:%M') if p.heure_arrivee_matin else '—',
        p.heure_depart_matin.strftime('%H:%M') if p.heure_depart_matin else '—',
        p.heure_arrivee_soir.strftime('%H:%M') if p.heure_arrivee_soir else '—',
        p.heure_depart_soir.strftime('%H:%M') if p.heure_depart_soir else '—',
        'Oui' if p.present else 'Non', p.duree_totale if p.duree_totale is not None else '—',
    ] for p in qs]
    return columns, rows


def _facturation(periode_debut, periode_fin):
    from facturation.models import Facture
    qs = Facture.objects.select_related('patient')
    if periode_debut:
        qs = qs.filter(date_emission__date__gte=periode_debut)
    qs = qs.filter(date_emission__date__lte=periode_fin).order_by('date_emission')
    columns = ['N° Facture', 'Code patient', 'Nom', 'Prénoms', 'Type', 'Date',
               'Total (FCFA)', 'Payé (FCFA)', 'Reste (FCFA)', 'Statut']
    rows = [[
        f.numero, f.patient.code_patient, f.patient.nom, f.patient.prenoms,
        f.get_type_facture_display(), f.date_emission.strftime('%d/%m/%Y'),
        float(f.montant_total), float(f.montant_paye), float(f.solde_restant),
        f.get_statut_display(),
    ] for f in qs]
    return columns, rows


def _diagnostic_retenu_rdv(rdv):
    """Diagnostic retenu / autres pathologies associées (registre curatif, sinon
    champ libre « maladies »), même logique que la liste des rendez-vous."""
    diag = ''
    try:
        diag = rdv.registre_curatif.diagnostic_display
    except ObjectDoesNotExist:
        diag = ''
    return diag or rdv.maladies or 'Non renseigné'


RDV_COLUMNS = ['Code RDV', 'Code patient', 'Nom', 'Prénoms', 'Sexe', 'Âge', 'Date et heure',
               'Médecin', 'Type de consultation', 'Diagnostics retenus', 'Statut']


def _rdv_row(rdv):
    return [
        rdv.code_rdv, rdv.patient.code_patient, rdv.patient.nom, rdv.patient.prenoms,
        rdv.patient.sexe, rdv.patient.age, rdv.date_heure.strftime('%d/%m/%Y %H:%M'),
        str(rdv.medecin) if rdv.medecin_id else 'Non renseigné',
        rdv.type_consultation.nom if rdv.type_consultation_id else 'Non renseigné',
        _diagnostic_retenu_rdv(rdv),
        rdv.get_statut_display(),
    ]


def _rendez_vous_periode_qs(periode_debut, periode_fin):
    from patients.models import RendezVous
    qs = RendezVous.objects.select_related('patient', 'medecin', 'type_consultation').prefetch_related('registre_curatif')
    if periode_debut:
        qs = qs.filter(date_heure__date__gte=periode_debut)
    return qs.filter(date_heure__date__lte=periode_fin).order_by('date_heure')


def _rendez_vous(periode_debut, periode_fin):
    qs = _rendez_vous_periode_qs(periode_debut, periode_fin)
    return RDV_COLUMNS, [_rdv_row(rdv) for rdv in qs]


def _rendez_vous_sans_doublon(periode_debut, periode_fin):
    """Un seul rendez-vous par patient sur la période (le plus récent),
    pour un listing patients sans doublon."""
    qs = _rendez_vous_periode_qs(periode_debut, periode_fin)
    par_patient = {}
    for rdv in qs:
        par_patient[rdv.patient_id] = rdv
    rdvs = sorted(par_patient.values(), key=lambda r: r.date_heure)
    return RDV_COLUMNS, [_rdv_row(rdv) for rdv in rdvs]


def _examens_labo(periode_debut, periode_fin):
    from laboratoire.models import DemandeExamen
    qs = DemandeExamen.objects.select_related('patient', 'medecin_prescripteur').prefetch_related('lignes')
    if periode_debut:
        qs = qs.filter(date_creation__date__gte=periode_debut)
    qs = qs.filter(date_creation__date__lte=periode_fin).order_by('date_creation')
    columns = ['Numéro', 'Code patient', 'Nom', 'Prénoms', 'Sexe', 'Âge', 'Médecin prescripteur',
               'Tests demandés', 'Urgent', 'Date de création', 'Montant (FCFA)', 'Statut']
    rows = [[
        d.numero, d.patient.code_patient, d.patient.nom, d.patient.prenoms, d.patient.sexe, d.patient.age,
        str(d.medecin_prescripteur) if d.medecin_prescripteur_id else 'Non renseigné',
        ', '.join(l.libelle for l in d.lignes.all()) or 'Non renseigné',
        'Oui' if d.urgent else 'Non',
        d.date_creation.strftime('%d/%m/%Y %H:%M'),
        float(d.montant_total),
        d.get_statut_display(),
    ] for d in qs]
    return columns, rows


def _ordonnances_prescrites(periode_debut, periode_fin):
    from consultations.models import Ordonnance
    qs = Ordonnance.objects.select_related(
        'consultation__patient', 'patient', 'medecin', 'consultation__medecin'
    ).prefetch_related('lignes')
    if periode_debut:
        qs = qs.filter(date_emission__date__gte=periode_debut)
    qs = qs.filter(date_emission__date__lte=periode_fin).order_by('date_emission')
    columns = ['Numéro', 'Code patient', 'Nom', 'Prénoms', 'Médecin prescripteur', 'Médicaments',
               'Type', "Date d'émission", "Date d'expiration", 'Statut']
    rows = []
    for o in qs:
        patient = o.consultation.patient if o.consultation_id and o.consultation.patient_id else o.patient
        medecin = o.medecin if o.medecin_id else (o.consultation.medecin if o.consultation_id else None)
        medicaments = ', '.join(
            l.medicament_libre or (str(l.produit) if l.produit_id else None) or (str(l.medicament) if l.medicament_id else 'Non renseigné')
            for l in o.lignes.all()
        ) or 'Non renseigné'
        rows.append([
            o.numero,
            patient.code_patient if patient else 'Non renseigné',
            patient.nom if patient else 'Non renseigné',
            patient.prenoms if patient else '',
            str(medecin) if medecin else 'Non renseigné',
            medicaments,
            o.get_type_ordonnance_display(),
            o.date_emission.strftime('%d/%m/%Y %H:%M'),
            o.date_expiration.strftime('%d/%m/%Y') if o.date_expiration else 'Non applicable',
            o.get_statut_display(),
        ])
    return columns, rows


CDIP_PROPOSE_LABELS = {'oui': 'Oui', 'na': 'NA'}
CDIP_REALISE_LABELS = {'oui': 'Oui', 'non': 'Non', 'na': 'NA'}

CDIP_COLUMNS = ['Code patient', 'Nom', 'Prénoms', 'Sexe', 'Âge', 'Date de consultation',
                'CDIP proposé', 'CDIP réalisé']


def _cdip_row(reg):
    rdv, patient = reg.rdv, reg.rdv.patient
    return [
        patient.code_patient, patient.nom, patient.prenoms, patient.sexe, patient.age,
        rdv.date_heure.strftime('%d/%m/%Y %H:%M'),
        CDIP_PROPOSE_LABELS.get(reg.donnees.get('cur_cdip_propose'), 'Non renseigné'),
        CDIP_REALISE_LABELS.get(reg.donnees.get('cur_cdip_realise'), 'Non renseigné'),
    ]


def _cdip_med_generale(periode_debut, periode_fin):
    from patients.models import RegistreCuratif
    qs = RegistreCuratif.objects.filter(rdv__departement__code='medg').select_related('rdv', 'rdv__patient')
    if periode_debut:
        qs = qs.filter(rdv__date_heure__date__gte=periode_debut)
    qs = qs.filter(rdv__date_heure__date__lte=periode_fin).order_by('rdv__date_heure')
    return CDIP_COLUMNS, [_cdip_row(reg) for reg in qs]


def _cdip_gynecologie(periode_debut, periode_fin):
    from django.db.models import Q
    from patients.models import RegistreCuratif
    qs = RegistreCuratif.objects.filter(
        Q(rdv__departement__code='GYN') | Q(rdv__medecin__specialite__nom__icontains='gyn')
    ).select_related('rdv', 'rdv__patient')
    if periode_debut:
        qs = qs.filter(rdv__date_heure__date__gte=periode_debut)
    qs = qs.filter(rdv__date_heure__date__lte=periode_fin).order_by('rdv__date_heure')
    return CDIP_COLUMNS, [_cdip_row(reg) for reg in qs]


VIH_STATUT_ACCUEIL_LABELS = {'inconnu': 'Inconnu', 'negatif': 'Négatif', 'positif': 'Positif'}
VIH_PROPOSITION_LABELS = {'propose': 'Proposé', 'accepte': 'Accepté', 'refuse': 'Refusé'}
VIH_RESULTAT_LABELS = {'negatif': 'Négatif', 'positif': 'Positif', 'na': 'NA'}


def _depistage_vih_cpn(periode_debut, periode_fin):
    from patients.models import RegistreCPN
    qs = RegistreCPN.objects.select_related('rdv', 'rdv__patient')
    if periode_debut:
        qs = qs.filter(rdv__date_heure__date__gte=periode_debut)
    qs = qs.filter(rdv__date_heure__date__lte=periode_fin).order_by('rdv__date_heure')
    columns = ['Code patient', 'Nom', 'Prénoms', 'Âge', 'Date de consultation',
               "Statut VIH à l'accueil", 'Proposition de test VIH', 'Résultat du test VIH']
    rows = []
    for reg in qs:
        rdv, patient, d = reg.rdv, reg.rdv.patient, reg.donnees
        rows.append([
            patient.code_patient, patient.nom, patient.prenoms, patient.age,
            rdv.date_heure.strftime('%d/%m/%Y %H:%M'),
            VIH_STATUT_ACCUEIL_LABELS.get(d.get('cpn_statut_vih_accueil'), 'Non renseigné'),
            VIH_PROPOSITION_LABELS.get(d.get('cpn_proposition_test_vih'), 'Non renseigné'),
            VIH_RESULTAT_LABELS.get(d.get('cpn_resultat_vih'), 'Non renseigné'),
        ])
    return columns, rows


REPORT_CATALOGUE = [
    {
        'nom': 'Patients',
        'icone': 'bi-people-fill',
        'rapports': [
            {'slug': 'patients_inscrits', 'nom': 'Patients inscrits', 'icone': 'bi-person-plus-fill',
             'description': 'Liste des patients enregistrés sur la période.', 'fn': _patients_inscrits},
        ],
    },
    {
        'nom': 'Rendez-vous',
        'icone': 'bi-calendar-check-fill',
        'rapports': [
            {'slug': 'rendez_vous', 'nom': 'Rendez-vous', 'icone': 'bi-calendar-check-fill',
             'description': 'Liste des rendez-vous pris sur la période.', 'fn': _rendez_vous},
            {'slug': 'rendez_vous_sans_doublon', 'nom': 'Rendez-vous sans doublon', 'icone': 'bi-calendar2-check',
             'description': 'Liste des rendez-vous sur la période, un seul par patient (le plus récent).',
             'fn': _rendez_vous_sans_doublon},
        ],
    },
    {
        'nom': 'Consultations',
        'icone': 'bi-clipboard2-pulse-fill',
        'rapports': [
            {'slug': 'cdip_med_generale', 'nom': 'CDIP proposé et réalisé - Consultation générale',
             'icone': 'bi-clipboard2-pulse-fill',
             'description': "CDIP proposé et CDIP réalisé pour les consultations du département médecine générale.",
             'fn': _cdip_med_generale},
            {'slug': 'cdip_gynecologie', 'nom': 'CDIP proposé et réalisé - Consultation gynécologie',
             'icone': 'bi-gender-female',
             'description': "CDIP proposé et CDIP réalisé pour les consultations du département gynécologie.",
             'fn': _cdip_gynecologie},
            {'slug': 'depistage_vih_cpn', 'nom': 'Dépistage VIH en CPN', 'icone': 'bi-virus2',
             'description': "Statut VIH à l'accueil, proposition et résultat du test VIH pour les consultations prénatales.",
             'fn': _depistage_vih_cpn},
        ],
    },
    {
        'nom': 'Hospitalisation',
        'icone': 'bi-hospital-fill',
        'rapports': [
            {'slug': 'listing_activite_soins_mo', 'nom': 'Listing des mises en observation', 'icone': 'bi-hospital-fill',
             'description': "Hospitalisations (mise en observation) demandées sur la période.",
             'fn': _hospitalisations_mo, 'build_xlsx_fn': build_listing_activite_mo_xlsx},
        ],
    },
    {
        'nom': 'Soins infirmiers',
        'icone': 'bi-bandaid-fill',
        'rapports': [
            {'slug': 'listing_activite_soins_sn', 'nom': 'Listing des soins infirmiers', 'icone': 'bi-bandaid-fill',
             'description': "Soins et procédures infirmiers effectués sur la période.",
             'fn': _listing_activite_soins_sn, 'build_xlsx_fn': build_listing_activite_soins_sn_xlsx},
        ],
    },
    {
        'nom': 'Laboratoire',
        'icone': 'bi-eyedropper',
        'rapports': [
            {'slug': 'examens_labo', 'nom': 'Examens labo', 'icone': 'bi-eyedropper',
             'description': "Demandes d'examens de laboratoire créées sur la période.", 'fn': _examens_labo},
        ],
    },
    {
        'nom': 'Pharmacie',
        'icone': 'bi-file-earmark-medical-fill',
        'rapports': [
            {'slug': 'ordonnances_prescrites', 'nom': 'Ordonnances prescrites', 'icone': 'bi-file-earmark-medical-fill',
             'description': 'Ordonnances émises sur la période.', 'fn': _ordonnances_prescrites},
        ],
    },
    {
        'nom': 'Facturation',
        'icone': 'bi-receipt-cutoff',
        'rapports': [
            {'slug': 'facturation', 'nom': 'Facturation', 'icone': 'bi-receipt-cutoff',
             'description': 'Liste des factures émises sur la période.', 'fn': _facturation},
        ],
    },
]

REPORTS_BY_SLUG = {
    rapport['slug']: {**rapport, 'categorie': categorie['nom']}
    for categorie in REPORT_CATALOGUE
    for rapport in categorie['rapports']
}
