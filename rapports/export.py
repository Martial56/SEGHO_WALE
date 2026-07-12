import csv
import io

from django.http import HttpResponse

CM_TO_PT = 72 / 2.54  # 1 cm en points

H_TITRE  = round(1.7  * CM_TO_PT, 2)
H_ENTETE = round(1.3  * CM_TO_PT, 2)
H_STD    = round(0.55 * CM_TO_PT, 2)
H_META   = round(0.7  * CM_TO_PT, 2)
POLICE   = 'Times New Roman'


def _periode_texte(periode_debut, periode_fin):
    if not periode_fin:
        return ''
    f = periode_fin.strftime('%d/%m/%Y')
    if not periode_debut:
        return f"jusqu'au {f}"
    d = periode_debut.strftime('%d/%m/%Y')
    return d if periode_debut == periode_fin else f"du {d} au {f}"


MOIS_FR = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin', 'juillet',
           'août', 'septembre', 'octobre', 'novembre', 'décembre']


def _mois_texte(periode_debut, periode_fin):
    if not periode_fin:
        return ''
    if not periode_debut:
        return MOIS_FR[periode_fin.month - 1].capitalize()
    mois = []
    y, m = periode_debut.year, periode_debut.month
    while (y, m) <= (periode_fin.year, periode_fin.month):
        label = MOIS_FR[m - 1].capitalize()
        if label not in mois:
            mois.append(label)
        m += 1
        if m > 12:
            m = 1
            y += 1
    return ', '.join(mois)


def populate_sheet(ws, titre, columns, rows, periode_debut=None, periode_fin=None, genere_le=None):
    """Remplit une feuille déjà créée avec le style standard des listings
    (bandeau titre, bloc Mois/Généré le/Période, en-têtes fond gris/gras).
    Réutilisable pour construire un classeur à plusieurs feuilles."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.title = titre[:31] or 'Rapport'

    n_cols = max(len(columns), 1)
    gray_fill = PatternFill(start_color='FFE0E0E0', end_color='FFE0E0E0', fill_type='solid')
    bold = Font(name=POLICE, bold=True)
    regular = Font(name=POLICE)
    title_font = Font(name=POLICE, bold=True, size=14)
    center = Alignment(horizontal='center', vertical='center')

    # Ligne 1 : bandeau titre plein largeur
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=(titre or '').upper())
    c.fill = gray_fill; c.font = title_font; c.alignment = center
    ws.row_dimensions[1].height = H_TITRE

    ws.row_dimensions[2].height = H_STD

    mid = max(n_cols // 2, 1)

    # Ligne 3 : Mois | Généré le — fond gris sur toute la ligne
    mois_txt = _mois_texte(periode_debut, periode_fin)
    genere_txt = genere_le.strftime('%d/%m/%Y %H:%M') if genere_le else ''
    for col in range(1, n_cols + 1):
        ws.cell(row=3, column=col).fill = gray_fill
    ws.cell(row=3, column=1, value="Mois :").font = bold
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=mid)
    ws.cell(row=3, column=2, value=mois_txt).font = regular
    if n_cols > mid:
        ws.cell(row=3, column=mid + 1, value="Généré le :").font = bold
        ws.merge_cells(start_row=3, start_column=mid + 2, end_row=3, end_column=n_cols)
        ws.cell(row=3, column=mid + 2, value=genere_txt).font = regular
    ws.row_dimensions[3].height = H_META

    # Ligne 4 : Période — fond gris sur toute la ligne
    periode_txt = _periode_texte(periode_debut, periode_fin)
    for col in range(1, n_cols + 1):
        ws.cell(row=4, column=col).fill = gray_fill
    ws.cell(row=4, column=1, value="Période :").font = bold
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=n_cols)
    ws.cell(row=4, column=2, value=periode_txt).font = regular
    ws.row_dimensions[4].height = H_META

    ws.row_dimensions[5].height = H_STD

    header_row = 6
    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.fill = gray_fill
        cell.font = bold
        cell.alignment = center
    ws.row_dimensions[header_row].height = H_ENTETE

    for r_idx, row in enumerate(rows, start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=value)
            cell.font = regular
            cell.alignment = center
        ws.row_dimensions[r_idx].height = H_STD

    widths = [len(str(c)) for c in columns]
    for row in rows:
        for i, value in enumerate(row):
            widths[i] = max(widths[i], len(str(value)) if value is not None else 0)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 12), 45)

    ws.freeze_panes = f'A{header_row + 1}'
    return header_row


def build_xlsx(titre, columns, rows, periode_debut=None, periode_fin=None, genere_le=None):
    """Construit un classeur Excel à une seule feuille et retourne son contenu binaire (bytes)."""
    from openpyxl import Workbook
    wb = Workbook()
    populate_sheet(wb.active, titre, columns, rows, periode_debut, periode_fin, genere_le)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def xlsx_response(filename, titre, columns, rows, periode_debut=None, periode_fin=None, genere_le=None):
    content = build_xlsx(titre, columns, rows, periode_debut, periode_fin, genere_le)
    response = HttpResponse(
        content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response, content


def build_csv(columns, rows):
    """Construit le contenu CSV (bytes, BOM UTF-8 pour Excel)."""
    buffer = io.StringIO()
    buffer.write('﻿')
    writer = csv.writer(buffer, delimiter=';')
    writer.writerow(columns)
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode('utf-8')


def csv_response(filename, columns, rows):
    content = build_csv(columns, rows)
    response = HttpResponse(content, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    return response, content
