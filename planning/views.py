from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db import models, transaction
from django.db.models import Q
from datetime import date, timedelta
from collections import defaultdict
from types import SimpleNamespace
import re as _re
import calendar as cal_module
import logging

logger = logging.getLogger(__name__)

from .models import (Bureau, PlageHoraire, PlanningHebdomadaire,
                     Affectation, PlanningVu, PlanningModification, PlanningConfig,
                     PlanningGabarit, GabaritAffectation, LignePermanence,
                     MedecinSignataire, FONCTION_SIGNATAIRE_CHOICES)
from medecins.models import Medecin

JOURS_LABELS = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']

PLANNING_WRITE_GROUPS          = {'Médecin Chef', 'Médecin Chef Adjoint', 'Administrateur', 'Directeur'}
PLANNING_DELETE_PUBLISHED_GROUPS = {'Médecin Chef', 'Médecin Chef Adjoint', 'Administrateur'}

MOIS_NOMS = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
             'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']


def can_manage_planning(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    return user.groups.filter(name__in=PLANNING_WRITE_GROUPS).exists()


def can_delete_published(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    return user.groups.filter(name__in=PLANNING_DELETE_PUBLISHED_GROUPS).exists()


def _medecins_json(medecins_qs):
    return [
        {
            'label': f"Dr {m.nom} {m.prenoms}",
            'spec':  m.specialite.nom if m.specialite else '',
        }
        for m in medecins_qs
    ]


def _base_ctx(request):
    """Context variables shared by all planning views (nav alertes + permissions)."""
    can_manage = can_manage_planning(request.user)
    alertes = []
    if can_manage:
        today = date.today()
        monday_now = today - timedelta(days=today.weekday())
        for i in range(3, -1, -1):
            monday = monday_now - timedelta(weeks=i)
            has_pub = PlanningHebdomadaire.objects.filter(
                semaine_debut=monday, publie=True
            ).exists()
            if not has_pub:
                alertes.append({
                    'label': f"Semaine du {monday.strftime('%d/%m/%Y')} "
                             f"au {(monday + timedelta(days=5)).strftime('%d/%m/%Y')}",
                    'monday': monday,
                })
    notifs_count = 0
    if request.user.is_authenticated:
        viewed_ids = PlanningVu.objects.filter(
            user=request.user
        ).values_list('planning_id', flat=True)
        notifs_count = PlanningHebdomadaire.objects.filter(
            publie=True
        ).exclude(pk__in=viewed_ids).count()
    return {
        'planning_alertes': alertes,
        'pl_can_manage':    can_manage,
        'pl_notifs_count':  notifs_count,
    }


def _conges_semaine(semaine_debut, semaine_fin):
    """Retourne la liste des employés en congé sur la semaine donnée."""
    try:
        from employer.models import Conge
    except ImportError:
        logger.warning("Module employer introuvable — congés non chargés dans le planning.")
        return []
    try:
        conges = Conge.objects.filter(
            date_debut__lte=semaine_fin,
            date_fin__gte=semaine_debut,
            statut__in=['approuve', 'en_cours', 'valide_service'],
        ).select_related('employe')
        result = []
        seen = set()
        for c in conges:
            e = c.employe
            if e.pk not in seen:
                seen.add(e.pk)
                result.append({
                    'key':     e.nom.lower(),
                    'display': f"{e.nom} {e.prenoms}",
                    'type':    c.get_type_conge_display(),
                })
        return result
    except Exception as exc:
        logger.error("Erreur lecture congés pour planning (%s → %s) : %s", semaine_debut, semaine_fin, exc)
        return []


def _conges_conflicts(planning, absents):
    """Affectations dont le personnel correspond à un employé en congé cette semaine —
    même heuristique de correspondance que le surlignage JS de modifier.html
    (nom du congé inclus dans le libellé de l'affectation, insensible à la casse).
    Contrairement à `absents`, qui n'est utilisé que côté client sur la page d'édition,
    ce signal doit aussi être visible sur un planning déjà publié (`planning_detail`),
    où aucune alerte n'existait auparavant."""
    if not absents:
        return []
    conflicts = []
    affectations = (
        planning.affectations.exclude(personnel='')
        .select_related('plage__bureau')
    )
    for aff in affectations:
        for name in split_names(aff.personnel):
            name_low = name.lower()
            match = next((a for a in absents if a['key'] in name_low), None)
            if match:
                conflicts.append({
                    'jour':      JOURS_LABELS[aff.jour],
                    'bureau':    aff.plage.bureau.nom,
                    'plage':     aff.plage.code,
                    'personnel': name,
                    'type':      match['type'],
                })
    return conflicts


def build_grid_rows(bureaux, aff_map):
    rows = []
    for bureau in bureaux:
        plages = list(bureau.plages.all())
        for i, plage in enumerate(plages):
            cells = []
            for j in range(6):
                aff = aff_map.get((plage.pk, j))
                cells.append({
                    'input_name': f'cell_{plage.pk}_{j}',
                    'value': aff.personnel if aff else '',
                    'note': aff.note if aff else '',
                })
            rows.append({
                'bureau': bureau,
                'plage': plage,
                'is_first': i == 0,
                'bureau_rowspan': len(plages),
                'cells': cells,
            })
    return rows


def get_bureaux():
    return Bureau.objects.filter(actif=True).prefetch_related('plages')


def split_names(val):
    """Découpe 'Dr X / Dr Y' en ['Dr X', 'Dr Y']."""
    return [n.strip() for n in _re.split(r'[/,]', val) if n.strip()]


def validate_planning(posted, bureaux):
    """Retourne une liste d'erreurs bloquantes selon les règles métier."""
    errors = []
    slot_map = defaultdict(list)  # (jour, code) → [(name_lower, bureau_nom, name_orig)]

    for bureau in bureaux:
        for plage in bureau.plages.all():
            for j in range(6):
                val = posted.get(f'cell_{plage.pk}_{j}', '').strip()
                if not val:
                    continue
                names = split_names(val)
                if len(names) > 2:
                    errors.append(
                        f"{bureau.nom} / {plage.code} / {JOURS_LABELS[j]} : "
                        f"maximum 2 médecins par créneau ({len(names)} saisis)."
                    )
                for name in names:
                    slot_map[(j, plage.code)].append((name.lower(), bureau.nom, name))

    for (jour, code), entries in slot_map.items():
        name_bureaux = defaultdict(list)
        for name_lower, bureau_nom, name_orig in entries:
            name_bureaux[name_lower].append((bureau_nom, name_orig))
        for name_lower, bureau_list in name_bureaux.items():
            if len(bureau_list) > 1:
                bureaux_str = ', '.join(b[0] for b in bureau_list)
                errors.append(
                    f"{bureau_list[0][1]} est affecté(e) dans {len(bureau_list)} bureaux "
                    f"le {JOURS_LABELS[jour]} ({code}) : {bureaux_str}."
                )
    return errors


def _posted_from_affectations(affectations):
    """Reconstruit un dict façon POST ({cell_<plage>_<jour>: personnel}) à partir
    d'un queryset/liste d'Affectation ou GabaritAffectation — pour pouvoir repasser
    par validate_planning() lors d'une application de gabarit ou d'une duplication,
    qui sinon contournaient entièrement la détection de conflits."""
    return {f'cell_{a.plage_id}_{a.jour}': a.personnel for a in affectations}


def build_grid_rows_from_posted(bureaux, posted):
    """Reconstruit les lignes du tableau à partir des valeurs POST."""
    rows = []
    for bureau in bureaux:
        plages = list(bureau.plages.all())
        for i, plage in enumerate(plages):
            cells = []
            for j in range(6):
                key = f'cell_{plage.pk}_{j}'
                note_key = f'note_{plage.pk}_{j}'
                cells.append({
                    'input_name': key,
                    'value': posted.get(key, ''),
                    'note': posted.get(note_key, ''),
                })
            rows.append({
                'bureau': bureau, 'plage': plage,
                'is_first': i == 0, 'bureau_rowspan': len(plages),
                'cells': cells,
            })
    return rows


# ── Liste ──────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def planning_list(request):
    plannings = PlanningHebdomadaire.objects.select_related('cree_par').all()
    if not can_manage_planning(request.user):
        # Les brouillons (non publiés) restent internes à la hiérarchie tant
        # qu'ils ne sont pas publiés — un utilisateur sans droit de gestion ne
        # doit voir que les plannings déjà publiés.
        plannings = plannings.filter(publie=True)

    annee   = request.GET.get('annee',   '').strip()
    mois    = request.GET.get('mois',    '').strip()
    semaine = request.GET.get('semaine', '').strip()

    if annee:
        plannings = plannings.filter(semaine_debut__year=annee)
    if mois:
        plannings = plannings.filter(semaine_debut__month=mois)
    if semaine:
        plannings = plannings.filter(semaine_debut__week=semaine)

    annees_dispo = sorted(
        set(PlanningHebdomadaire.objects.values_list('semaine_debut__year', flat=True)),
        reverse=True,
    )

    MOIS_LABELS = [
        (1,'Janvier'),(2,'Février'),(3,'Mars'),(4,'Avril'),
        (5,'Mai'),(6,'Juin'),(7,'Juillet'),(8,'Août'),
        (9,'Septembre'),(10,'Octobre'),(11,'Novembre'),(12,'Décembre'),
    ]

    return render(request, 'planning/liste.html', {
        'plannings':            plannings,
        'annees_dispo':         annees_dispo,
        'mois_labels':          MOIS_LABELS,
        'annee':                annee,
        'mois':                 mois,
        'semaine':              semaine,
        'can_manage':           can_manage_planning(request.user),
        'can_delete_published': can_delete_published(request.user),
        **_base_ctx(request),
    })


# ── Nouveau ────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def planning_nouveau(request):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        date_str      = request.POST.get('semaine_debut', '').strip()
        signataire_pk = request.POST.get('signataire', '').strip()
        signataire    = get_object_or_404(MedecinSignataire, pk=signataire_pk) if signataire_pk else None
        try:
            d = date.fromisoformat(date_str)
        except (ValueError, TypeError):
            d = date.today()
        monday = d - timedelta(days=d.weekday())
        planning, created = PlanningHebdomadaire.objects.get_or_create(
            semaine_debut=monday,
            defaults={'cree_par': request.user, 'signataire': signataire}
        )
        if not created and signataire and signataire.pk != planning.signataire_id:
            if planning.publie:
                # Un planning publié ne doit être modifiable (y compris son signataire)
                # que via planning_modifier, qui refuse justement les plannings publiés —
                # laisser passer ici contournait silencieusement cette règle, sans trace d'audit.
                messages.warning(
                    request,
                    f"Un planning existe déjà pour cette semaine et est publié — "
                    f"son signataire n'a pas été modifié."
                )
            else:
                ancien = planning.signataire
                planning.signataire = signataire
                planning.save(update_fields=['signataire'])
                PlanningModification.objects.create(
                    planning=planning,
                    modifie_par=request.user,
                    resume=f"Signataire: «{ancien.nom if ancien else '—'}»→«{signataire.nom}» (via création).",
                )
        return redirect('planning_modifier', pk=planning.pk)
    today  = date.today()
    monday = today - timedelta(days=today.weekday())
    return render(request, 'planning/nouveau.html', {
        'default_date': monday.isoformat(),
        'medecins_signataires': MedecinSignataire.objects.filter(actif=True),
        'default_signataire_id': PlanningConfig.get().medecin_defaut_id,
        **_base_ctx(request),
    })


# ── Détail ─────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def planning_detail(request, pk):
    planning = get_object_or_404(PlanningHebdomadaire, pk=pk)
    if not planning.publie and not can_manage_planning(request.user):
        raise PermissionDenied
    bureaux  = get_bureaux()
    aff_map  = {(a.plage_id, a.jour): a for a in planning.affectations.all()}
    rows     = build_grid_rows(bureaux, aff_map)

    if planning.publie:
        PlanningVu.objects.get_or_create(user=request.user, planning=planning)

    prev_planning = PlanningHebdomadaire.objects.filter(
        semaine_debut__lt=planning.semaine_debut
    ).order_by('-semaine_debut').first()

    next_planning = PlanningHebdomadaire.objects.filter(
        semaine_debut__gt=planning.semaine_debut
    ).order_by('semaine_debut').first()

    confirm_publish = request.GET.get('confirm_publish') == '1'
    pub_empty_days  = [d for d in request.GET.get('empty_days', '').split(',') if d] if confirm_publish else []

    perm_map  = {p.jour: p.personnel for p in planning.permanences.all()}
    perm_list = [perm_map.get(j, '') for j in range(6)]

    absents          = _conges_semaine(planning.semaine_debut, planning.semaine_fin)
    conflits_conges  = _conges_conflicts(planning, absents)

    return render(request, 'planning/hebdomadaire.html', {
        'planning':        planning,
        'rows':            rows,
        'jours_labels':    JOURS_LABELS,
        'perm_list':       perm_list,
        'can_manage':      can_manage_planning(request.user),
        'can_delete_pub':  can_delete_published(request.user),
        'prev_planning':   prev_planning,
        'next_planning':   next_planning,
        'modifications':   planning.modifications.select_related('modifie_par')[:10],
        'confirm_publish': confirm_publish,
        'pub_empty_days':  pub_empty_days,
        'conflits_conges': conflits_conges,
        'fonction_signataire': dict(FONCTION_SIGNATAIRE_CHOICES).get(
            PlanningConfig.get().fonction_signataire, ''
        ),
        **_base_ctx(request),
    })


# ── Modifier ───────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def planning_modifier(request, pk):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    planning = get_object_or_404(PlanningHebdomadaire, pk=pk)
    if planning.publie:
        messages.error(request, 'Ce planning est publié et ne peut plus être modifié.')
        return redirect('planning_detail', pk=pk)
    bureaux  = get_bureaux()
    medecins = Medecin.objects.filter(actif=True).select_related('specialite', 'employe').order_by('employe__nom')
    absents  = _conges_semaine(planning.semaine_debut, planning.semaine_fin)

    if request.method == 'POST':
        posted = {
            f'cell_{plage.pk}_{j}': request.POST.get(f'cell_{plage.pk}_{j}', '').strip()
            for bureau in bureaux for plage in bureau.plages.all() for j in range(6)
        }

        errors = validate_planning(posted, bureaux)
        if errors:
            for err in errors:
                messages.error(request, err)
            rows = build_grid_rows_from_posted(bureaux, posted)
            return render(request, 'planning/modifier.html', {
                'planning': planning, 'rows': rows,
                'jours_labels': JOURS_LABELS, 'medecins': medecins,
                'medecins_json': _medecins_json(medecins),
                'absents_json':  absents,
                **_base_ctx(request),
            })

        # Un seul aller-retour DB pour charger l'existant, puis diff en mémoire —
        # remplace ~200-300 requêtes (1 SELECT + 1-2 write par cellule) par
        # quelques requêtes en tout (1 SELECT, 1 bulk_update, 1 bulk_create, 1 delete).
        old_aff_map = {(a.plage_id, a.jour): a for a in planning.affectations.all()}

        changes = []
        to_create = []
        to_update = []
        to_delete_ids = []
        for bureau in bureaux:
            for plage in bureau.plages.all():
                for j in range(6):
                    val      = posted[f'cell_{plage.pk}_{j}']
                    note_val = request.POST.get(f'note_{plage.pk}_{j}', '').strip()
                    old_aff  = old_aff_map.get((plage.pk, j))
                    old_val  = old_aff.personnel if old_aff else ''
                    if val != old_val:
                        changes.append(
                            f"{bureau.nom}/{plage.code}/{JOURS_LABELS[j]}: «{old_val}»→«{val}»"
                        )
                    if val or note_val:
                        if old_aff:
                            old_aff.personnel = val
                            old_aff.note = note_val
                            to_update.append(old_aff)
                        else:
                            to_create.append(Affectation(
                                planning=planning, plage=plage, jour=j,
                                personnel=val, note=note_val,
                            ))
                    elif old_aff:
                        to_delete_ids.append(old_aff.pk)

        signataire_pk = request.POST.get('signataire', '').strip()
        signataire    = get_object_or_404(MedecinSignataire, pk=signataire_pk) if signataire_pk else None
        if signataire_pk != str(planning.signataire_id or ''):
            ancien_nom = planning.signataire.nom if planning.signataire else '—'
            nouveau_nom = signataire.nom if signataire else '—'
            changes.append(f"Signataire: «{ancien_nom}»→«{nouveau_nom}»")
        planning.signataire = signataire

        with transaction.atomic():
            planning.save()
            if to_create:
                Affectation.objects.bulk_create(to_create)
            if to_update:
                Affectation.objects.bulk_update(to_update, ['personnel', 'note'])
            if to_delete_ids:
                Affectation.objects.filter(pk__in=to_delete_ids).delete()

            if changes:
                PlanningModification.objects.create(
                    planning=planning,
                    modifie_par=request.user,
                    resume='; '.join(changes[:30]),
                )

        # Permanence
        for j in range(6):
            val = request.POST.get(f'perm_{j}', '').strip()
            if val:
                LignePermanence.objects.update_or_create(
                    planning=planning, jour=j,
                    defaults={'personnel': val}
                )
            else:
                LignePermanence.objects.filter(planning=planning, jour=j).delete()

        messages.success(request, 'Planning enregistré avec succès.')
        return redirect('planning_list')

    aff_map   = {(a.plage_id, a.jour): a for a in planning.affectations.all()}
    rows      = build_grid_rows(bureaux, aff_map)
    gabarits  = PlanningGabarit.objects.all()
    perm_map  = {p.jour: p.personnel for p in planning.permanences.all()}
    perm_list = [perm_map.get(j, '') for j in range(6)]
    return render(request, 'planning/modifier.html', {
        'planning':       planning,
        'rows':           rows,
        'jours_labels':   JOURS_LABELS,
        'perm_list':      perm_list,
        'medecins':       medecins,
        'medecins_json':  _medecins_json(medecins),
        'medecins_signataires': MedecinSignataire.objects.filter(actif=True),
        'absents_json':   absents,
        'gabarits':       gabarits,
        **_base_ctx(request),
    })


# ── Dupliquer ──────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def planning_dupliquer(request, pk):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    source = get_object_or_404(PlanningHebdomadaire, pk=pk)
    if request.method == 'POST':
        date_str = request.POST.get('semaine_debut', '').strip()
        try:
            d = date.fromisoformat(date_str)
        except (ValueError, TypeError):
            messages.error(request, 'Date invalide.')
            return redirect('planning_detail', pk=pk)
        monday = d - timedelta(days=d.weekday())
        if monday == source.semaine_debut:
            messages.error(request, 'Choisissez une semaine différente de la source.')
            return redirect('planning_detail', pk=pk)

        bureaux = get_bureaux()
        posted  = _posted_from_affectations(source.affectations.all())
        errors  = validate_planning(posted, bureaux)
        if errors:
            for err in errors:
                messages.error(request, err)
            messages.error(
                request,
                'Le planning source contient des conflits (ci-dessus) — duplication annulée. '
                'Corrigez le planning source avant de réessayer.'
            )
            return redirect('planning_detail', pk=pk)

        new_pl, created = PlanningHebdomadaire.objects.get_or_create(
            semaine_debut=monday,
            defaults={'cree_par': request.user, 'signataire': source.signataire}
        )
        if created:
            with transaction.atomic():
                Affectation.objects.bulk_create([
                    Affectation(
                        planning=new_pl, plage=aff.plage,
                        jour=aff.jour, personnel=aff.personnel, note=aff.note,
                    )
                    for aff in source.affectations.all()
                ])
                LignePermanence.objects.bulk_create([
                    LignePermanence(planning=new_pl, jour=perm.jour, personnel=perm.personnel)
                    for perm in source.permanences.all()
                ])
                PlanningModification.objects.create(
                    planning=new_pl,
                    modifie_par=request.user,
                    resume=f'Dupliqué depuis la semaine du {source.semaine_debut.strftime("%d/%m/%Y")}.',
                )
            messages.success(request, f'Planning dupliqué pour la semaine du {monday.strftime("%d/%m/%Y")}.')
            return redirect('planning_modifier', pk=new_pl.pk)
        else:
            messages.warning(request, f'Un planning existe déjà pour la semaine du {monday.strftime("%d/%m/%Y")}.')
            return redirect('planning_detail', pk=new_pl.pk)
    return redirect('planning_detail', pk=pk)


# ── Publier / Dépublier ────────────────────────────────────────────────────────

@login_required(login_url='login')
def planning_publier(request, pk):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    if request.method != 'POST':
        return redirect('planning_detail', pk=pk)

    planning = get_object_or_404(PlanningHebdomadaire, pk=pk)

    # Dépublier (toggle off)
    if planning.publie:
        planning.publie = False
        planning.save()
        # Efface les "vus" existants : si le contenu change avant republication,
        # les utilisateurs doivent être re-notifiés au lieu que le planning reste
        # marqué comme déjà consulté indéfiniment.
        planning.vus_par.all().delete()
        PlanningModification.objects.create(
            planning=planning, modifie_par=request.user, resume='Planning dépublié.',
        )
        return redirect('planning_detail', pk=pk)

    # ── Règle 1 : refus si planning entièrement vide ──────────────────────────
    filled = planning.affectations.exclude(personnel='').count()
    if filled == 0:
        messages.error(
            request,
            'Impossible de publier un planning entièrement vide. '
            'Veuillez d\'abord saisir les affectations.'
        )
        return redirect('planning_detail', pk=pk)

    # ── Règle 2 : jours ouvrables non fériés sans aucune affectation ──────────
    try:
        from employer.models import JourFerie
        feries = set(
            JourFerie.objects.filter(
                date__range=(planning.semaine_debut, planning.semaine_fin)
            ).values_list('date', flat=True)
        )
    except ImportError:
        feries = set()
    except Exception as exc:
        logger.error("Erreur lecture jours fériés : %s", exc)
        feries = set()

    empty_days = []
    for j in range(6):
        day_date = planning.semaine_debut + timedelta(days=j)
        if day_date in feries:
            continue
        if not planning.affectations.exclude(personnel='').filter(jour=j).exists():
            empty_days.append(JOURS_LABELS[j])

    justification = request.POST.get('justification', '').strip()

    if empty_days and not justification:
        from django.urls import reverse as _reverse
        url = _reverse('planning_detail', args=[pk])
        days_param = ','.join(empty_days)
        return redirect(f'{url}?confirm_publish=1&empty_days={days_param}')

    # ── Publication ───────────────────────────────────────────────────────────
    planning.publie = True
    planning.save()
    resume = 'Planning publié.'
    if justification:
        resume += f' Justification ({", ".join(empty_days)}) : {justification}'
    PlanningModification.objects.create(
        planning=planning, modifie_par=request.user, resume=resume,
    )
    count = _send_publication_email(planning, request.user)
    if count > 0:
        messages.success(request, f'Planning publié. {count} notification(s) envoyée(s).')
    else:
        messages.success(request, 'Planning publié avec succès.')
    return redirect('planning_detail', pk=pk)


# ── Supprimer ──────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def planning_supprimer(request, pk):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        planning = get_object_or_404(PlanningHebdomadaire, pk=pk)
        if planning.publie and not can_delete_published(request.user):
            raise PermissionDenied
        planning.delete()
        messages.success(request, 'Planning supprimé.')
    return redirect('planning_list')


# ── Vue mensuelle ──────────────────────────────────────────────────────────────

@login_required(login_url='login')
def planning_mensuel(request):
    today = date.today()
    try:
        annee = int(request.GET.get('annee', today.year))
        mois  = int(request.GET.get('mois',  today.month))
    except (ValueError, TypeError):
        annee, mois = today.year, today.month
    annee = max(2020, min(annee, 2099))
    mois  = max(1,    min(mois,  12))

    _, days_in_month = cal_module.monthrange(annee, mois)
    first_day = date(annee, mois, 1)
    last_day  = date(annee, mois, days_in_month)

    plannings = PlanningHebdomadaire.objects.filter(
        semaine_debut__gte=first_day - timedelta(days=6),
        semaine_debut__lte=last_day,
    ).order_by('semaine_debut')

    date_to_planning = {}
    for pl in plannings:
        for j in range(6):
            d = pl.semaine_debut + timedelta(days=j)
            if d.month == mois:
                date_to_planning[d] = pl

    weeks_raw  = cal_module.monthcalendar(annee, mois)
    weeks_data = []
    for week in weeks_raw:
        cells = []
        for day_num in week:
            if day_num == 0:
                cells.append({'day_num': 0, 'is_other': True, 'planning': None, 'is_today': False})
            else:
                d  = date(annee, mois, day_num)
                cells.append({
                    'day_num':  day_num,
                    'is_other': False,
                    'planning': date_to_planning.get(d),
                    'is_today': d == today,
                })
        weeks_data.append(cells)

    prev_annee = annee if mois > 1 else annee - 1
    prev_mois  = mois - 1 if mois > 1 else 12
    next_annee = annee if mois < 12 else annee + 1
    next_mois  = mois + 1 if mois < 12 else 1

    return render(request, 'planning/mensuel.html', {
        'annee':      annee,
        'mois':       mois,
        'mois_nom':   MOIS_NOMS[mois],
        'weeks_data': weeks_data,
        'today':      today,
        'prev_annee': prev_annee,
        'prev_mois':  prev_mois,
        'next_annee': next_annee,
        'next_mois':  next_mois,
        'can_manage': can_manage_planning(request.user),
        **_base_ctx(request),
    })


# ── Vue par médecin / personnel ────────────────────────────────────────────────

@login_required(login_url='login')
def planning_par_medecin(request):
    nom = request.GET.get('nom', '').strip()
    affectations = []
    if nom:
        affectations = (
            Affectation.objects
            .filter(personnel__icontains=nom)
            .select_related('planning', 'plage', 'plage__bureau')
            .order_by('-planning__semaine_debut', 'jour', 'plage__bureau__ordre')
        )
    personnel_list = (
        Affectation.objects
        .exclude(personnel='')
        .values_list('personnel', flat=True)
        .distinct()
        .order_by('personnel')
    )
    return render(request, 'planning/par_medecin.html', {
        'nom':            nom,
        'affectations':   affectations,
        'personnel_list': personnel_list,
        'jours_labels':   JOURS_LABELS,
        **_base_ctx(request),
    })


# ── Autocomplétion JSON ────────────────────────────────────────────────────────

@login_required(login_url='login')
def planning_medecins_json(request):
    q  = request.GET.get('q', '').strip()
    qs = Medecin.objects.filter(actif=True).select_related('specialite', 'employe')
    if q:
        qs = qs.filter(Q(employe__nom__icontains=q) | Q(employe__prenoms__icontains=q))
    data = [
        {
            'label':      f"Dr {m.nom} {m.prenoms}",
            'specialite': m.specialite.nom if m.specialite else '',
        }
        for m in qs[:20]
    ]
    return JsonResponse({'results': data})


# ── Email de publication ───────────────────────────────────────────────────────

def _send_publication_email(planning, published_by):
    from django.core.mail import send_mass_mail
    from django.contrib.auth.models import User as AuthUser
    from django.db.models import Q as DQ
    # Collect emails: all active médecins + all users in write groups
    med_emails = set(Medecin.objects.filter(actif=True).exclude(employe__email='').values_list('employe__email', flat=True))
    user_emails = set(AuthUser.objects.filter(
        is_active=True, email__gt=''
    ).filter(
        DQ(is_superuser=True) | DQ(groups__name__in=PLANNING_WRITE_GROUPS)
    ).values_list('email', flat=True))
    all_emails = list(med_emails | user_emails)
    if not all_emails:
        return 0
    subject = f"Planning publié — semaine du {planning.semaine_debut.strftime('%d/%m/%Y')}"
    body = (
        f"Bonjour,\n\n"
        f"Le planning de la semaine du {planning.semaine_debut.strftime('%d/%m/%Y')} "
        f"au {planning.semaine_fin.strftime('%d/%m/%Y')} a été publié"
        f" par {published_by.get_full_name() or published_by.username}.\n\n"
        f"Consultez-le sur l'application Centre Médico-Social WALÉ.\n\n"
        f"— Système de gestion SEGHO-WALÉ"
    )
    from django.conf import settings as django_settings
    from_email = getattr(django_settings, 'DEFAULT_FROM_EMAIL', 'planning@cms-wale.ci')
    msgs = [(subject, body, from_email, [e]) for e in all_emails]
    try:
        sent = send_mass_mail(msgs, fail_silently=False)
        return sent
    except Exception as exc:
        logger.error("Erreur envoi emails planning %s : %s", planning.pk, exc)
        return 0


# ── Semaine en cours / Tableau de bord ───────────────────────────────────────

@login_required(login_url='login')
def planning_courant(request):
    today       = date.today()
    monday      = today - timedelta(days=today.weekday())
    next_monday = monday + timedelta(weeks=1)
    prev_monday = monday - timedelta(weeks=1)

    planning_courant_ = PlanningHebdomadaire.objects.filter(semaine_debut=monday).first()
    planning_prochain = PlanningHebdomadaire.objects.filter(semaine_debut=next_monday).first()
    planning_precedent = PlanningHebdomadaire.objects.filter(semaine_debut=prev_monday).first()

    # 6 derniers plannings (hors semaine courante)
    recents = (
        PlanningHebdomadaire.objects
        .exclude(semaine_debut=monday)
        .order_by('-semaine_debut')[:6]
    )

    # Stats rapides
    active_bureaux = list(Bureau.objects.filter(actif=True).prefetch_related('plages'))
    total_cells_week = sum(len(b.plages.all()) * 6 for b in active_bureaux)
    filled_courant = 0
    if planning_courant_:
        filled_courant = planning_courant_.affectations.exclude(personnel='').count()
    pct_courant = round(filled_courant * 100 / total_cells_week) if total_cells_week else 0

    total_published = PlanningHebdomadaire.objects.filter(publie=True).count()
    total_plannings = PlanningHebdomadaire.objects.count()

    return render(request, 'planning/dashboard.html', {
        'today':               today,
        'monday':              monday,
        'next_monday':         next_monday,
        'prev_monday':         prev_monday,
        'planning_courant':    planning_courant_,
        'planning_prochain':   planning_prochain,
        'planning_precedent':  planning_precedent,
        'recents':             recents,
        'total_cells_week':    total_cells_week,
        'filled_courant':      filled_courant,
        'pct_courant':         pct_courant,
        'total_published':     total_published,
        'total_plannings':     total_plannings,
        **_base_ctx(request),
    })


# ── Gestion des bureaux et plages ─────────────────────────────────────────────

@login_required(login_url='login')
def planning_bureaux(request):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    bureaux = Bureau.objects.prefetch_related('plages').order_by('ordre')
    return render(request, 'planning/bureaux.html', {
        'bureaux':    bureaux,
        'can_manage': True,
        **_base_ctx(request),
    })


@login_required(login_url='login')
def planning_bureau_save(request):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        pk  = request.POST.get('pk', '').strip()
        nom = request.POST.get('nom', '').strip()
        actif = request.POST.get('actif') == '1'
        if not nom:
            messages.error(request, 'Le nom du bureau est obligatoire.')
            return redirect('planning_bureaux')
        if pk:
            bureau = get_object_or_404(Bureau, pk=pk)
            bureau.nom   = nom
            bureau.actif = actif
            bureau.save()
            messages.success(request, f'Bureau « {nom} » mis à jour.')
        else:
            max_ordre = Bureau.objects.aggregate(m=models.Max('ordre'))['m'] or 0
            Bureau.objects.create(nom=nom, actif=actif, ordre=max_ordre + 1)
            messages.success(request, f'Bureau « {nom} » créé.')
    return redirect('planning_bureaux')


@login_required(login_url='login')
def planning_bureau_delete(request, pk):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        bureau = get_object_or_404(Bureau, pk=pk)
        has_aff = Affectation.objects.filter(plage__bureau=bureau).exists()
        if has_aff:
            messages.error(
                request,
                f'Impossible de supprimer « {bureau.nom} » : des affectations existent pour ce bureau.'
            )
        else:
            nom = bureau.nom
            bureau.delete()
            messages.success(request, f'Bureau « {nom} » supprimé.')
    return redirect('planning_bureaux')


@login_required(login_url='login')
def planning_bureau_ordre(request, pk):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        bureau    = get_object_or_404(Bureau, pk=pk)
        direction = request.POST.get('direction', '')
        bureaux   = list(Bureau.objects.order_by('ordre'))
        idx       = next((i for i, b in enumerate(bureaux) if b.pk == bureau.pk), None)
        if idx is None:
            return redirect('planning_bureaux')
        if direction == 'up' and idx > 0:
            bureaux[idx], bureaux[idx - 1] = bureaux[idx - 1], bureaux[idx]
        elif direction == 'down' and idx < len(bureaux) - 1:
            bureaux[idx], bureaux[idx + 1] = bureaux[idx + 1], bureaux[idx]
        for i, b in enumerate(bureaux):
            if b.ordre != i:
                b.ordre = i
                b.save(update_fields=['ordre'])
    return redirect('planning_bureaux')


@login_required(login_url='login')
def planning_plage_save(request):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        pk        = request.POST.get('pk', '').strip()
        bureau_pk = request.POST.get('bureau_pk', '').strip()
        code      = request.POST.get('code', '').strip()
        if not code:
            messages.error(request, 'Le code de la plage est obligatoire.')
            return redirect('planning_bureaux')
        if pk:
            plage      = get_object_or_404(PlageHoraire, pk=pk)
            plage.code = code
            plage.save()
            messages.success(request, f'Plage « {code} » mise à jour.')
        else:
            bureau    = get_object_or_404(Bureau, pk=bureau_pk)
            max_ordre = bureau.plages.aggregate(m=models.Max('ordre'))['m'] or 0
            PlageHoraire.objects.create(bureau=bureau, code=code, ordre=max_ordre + 1)
            messages.success(request, f'Plage « {code} » créée.')
    return redirect('planning_bureaux')


@login_required(login_url='login')
def planning_plage_delete(request, pk):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        plage   = get_object_or_404(PlageHoraire, pk=pk)
        has_aff = Affectation.objects.filter(plage=plage).exists()
        if has_aff:
            messages.error(
                request,
                f'Impossible de supprimer la plage « {plage.code} » : des affectations existent.'
            )
        else:
            code = plage.code
            plage.delete()
            messages.success(request, f'Plage « {code} » supprimée.')
    return redirect('planning_bureaux')


# ── Configuration (signataire) ────────────────────────────────────────────────

@login_required(login_url='login')
def planning_configuration(request):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    config = PlanningConfig.get()
    if request.method == 'POST' and request.POST.get('form') == 'signataire':
        fonction = request.POST.get('fonction_signataire', '').strip()
        medecin_pk = request.POST.get('medecin_defaut', '').strip()
        if fonction not in dict(FONCTION_SIGNATAIRE_CHOICES):
            messages.error(request, 'Fonction de signataire invalide.')
        else:
            config.fonction_signataire = fonction
            config.medecin_defaut = (
                get_object_or_404(MedecinSignataire, pk=medecin_pk) if medecin_pk else None
            )
            config.save(update_fields=['fonction_signataire', 'medecin_defaut'])
            messages.success(request, 'Configuration du signataire enregistrée.')
        return redirect('planning_configuration')
    return render(request, 'planning/configuration.html', {
        'config': config,
        'fonction_choices': FONCTION_SIGNATAIRE_CHOICES,
        'medecins': MedecinSignataire.objects.all(),
        'can_manage': True,
        **_base_ctx(request),
    })


@login_required(login_url='login')
def planning_medecin_save(request):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        pk    = request.POST.get('pk', '').strip()
        nom   = request.POST.get('nom', '').strip()
        actif = request.POST.get('actif') == '1'
        if not nom:
            messages.error(request, 'Le nom du médecin signataire est obligatoire.')
            return redirect('planning_configuration')
        if pk:
            medecin = get_object_or_404(MedecinSignataire, pk=pk)
            medecin.nom   = nom
            medecin.actif = actif
            medecin.save()
            messages.success(request, f'Médecin signataire « {nom} » mis à jour.')
        else:
            max_ordre = MedecinSignataire.objects.aggregate(m=models.Max('ordre'))['m'] or 0
            MedecinSignataire.objects.create(nom=nom, actif=actif, ordre=max_ordre + 1)
            messages.success(request, f'Médecin signataire « {nom} » ajouté.')
    return redirect('planning_configuration')


@login_required(login_url='login')
def planning_medecin_delete(request, pk):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        medecin = get_object_or_404(MedecinSignataire, pk=pk)
        if medecin.plannings.exists():
            messages.error(
                request,
                f'Impossible de supprimer « {medecin.nom} » : il est lié à au moins un planning. '
                f'Vous pouvez le désactiver à la place.'
            )
        else:
            nom = medecin.nom
            medecin.delete()
            messages.success(request, f'Médecin signataire « {nom} » supprimé.')
    return redirect('planning_configuration')


# ── Statistiques de couverture ─────────────────────────────────────────────────

@login_required(login_url='login')
def planning_stats(request):
    from collections import defaultdict as _defaultdict

    plannings_qs = (
        PlanningHebdomadaire.objects
        .prefetch_related('affectations')
        .order_by('-semaine_debut')[:12]
    )
    plannings_list = list(plannings_qs)

    active_bureaux = list(Bureau.objects.filter(actif=True).prefetch_related('plages'))
    # Indépendant de `pl` — calculé une seule fois au lieu de 12 fois (et len() sur
    # une liste déjà préchargée plutôt que .count(), qui ignore toujours le cache prefetch).
    total_cells = sum(len(bureau.plages.all()) * 6 for bureau in active_bureaux)

    stats = []
    for pl in plannings_list:
        filled = sum(1 for a in pl.affectations.all() if a.personnel)
        pct    = round(filled * 100 / total_cells) if total_cells else 0
        stats.append({
            'planning':    pl,
            'total_cells': total_cells,
            'filled':      filled,
            'pct':         pct,
        })

    # Monthly averages
    monthly_raw = _defaultdict(list)
    for s in stats:
        key = (s['planning'].semaine_debut.year, s['planning'].semaine_debut.month)
        monthly_raw[key].append(s['pct'])

    monthly_stats = []
    for (year, month), pcts in sorted(monthly_raw.items(), reverse=True):
        monthly_stats.append({
            'year':     year,
            'month':    month,
            'mois_nom': MOIS_NOMS[month],
            'avg_pct':  round(sum(pcts) / len(pcts)),
            'count':    len(pcts),
        })

    avg_global = round(sum(s['pct'] for s in stats) / len(stats)) if stats else 0
    best  = max(stats, key=lambda s: s['pct']) if stats else None
    worst = min(stats, key=lambda s: s['pct']) if stats else None

    # ── Stats par bureau ──────────────────────────────────────────────────────
    plage_to_bureau = {}
    for bureau in active_bureaux:
        for plage in bureau.plages.all():
            plage_to_bureau[plage.pk] = bureau.pk

    bureau_filled = _defaultdict(int)
    for pl in plannings_list:
        for aff in pl.affectations.all():
            if aff.personnel:
                bpk = plage_to_bureau.get(aff.plage_id)
                if bpk:
                    bureau_filled[bpk] += 1

    bureau_stats = []
    for bureau in active_bureaux:
        total = len(bureau.plages.all()) * 6 * len(plannings_list)
        filled_count = bureau_filled[bureau.pk]
        pct = round(filled_count * 100 / total) if total else 0
        bureau_stats.append({
            'bureau': bureau,
            'total':  total,
            'filled': filled_count,
            'pct':    pct,
        })
    bureau_stats.sort(key=lambda x: x['pct'], reverse=True)

    # ── Top 10 médecins / personnel ───────────────────────────────────────────
    medecin_counts = _defaultdict(int)
    for pl in plannings_list:
        for aff in pl.affectations.all():
            if aff.personnel:
                for name in split_names(aff.personnel):
                    medecin_counts[name.strip()] += 1
    top_medecins = sorted(medecin_counts.items(), key=lambda x: x[1], reverse=True)[:10]

    return render(request, 'planning/stats.html', {
        'stats':           stats,
        'monthly_stats':   monthly_stats,
        'avg_global':      avg_global,
        'best':            best,
        'worst':           worst,
        'total_plannings': len(plannings_list),
        'bureau_stats':    bureau_stats,
        'top_medecins':    top_medecins,
        'can_manage':      can_manage_planning(request.user),
        **_base_ctx(request),
    })


# ── Gabarits ───────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def planning_gabarit_sauvegarder(request, pk):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    planning = get_object_or_404(PlanningHebdomadaire, pk=pk)
    if request.method == 'POST':
        nom = request.POST.get('nom', '').strip()
        if not nom:
            messages.error(request, 'Le nom du gabarit est obligatoire.')
            return redirect('planning_detail', pk=pk)
        gabarit = PlanningGabarit.objects.create(nom=nom, cree_par=request.user)
        GabaritAffectation.objects.bulk_create([
            GabaritAffectation(
                gabarit=gabarit, plage=aff.plage, jour=aff.jour,
                personnel=aff.personnel, note=aff.note,
            )
            for aff in planning.affectations.exclude(personnel='')
        ])
        messages.success(request, f'Gabarit « {nom} » enregistré.')
    return redirect('planning_detail', pk=pk)


@login_required(login_url='login')
def planning_gabarit_appliquer(request, pk):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    planning = get_object_or_404(PlanningHebdomadaire, pk=pk)
    if planning.publie:
        messages.error(request, "Impossible d'appliquer un gabarit sur un planning publié.")
        return redirect('planning_detail', pk=pk)
    if request.method == 'POST':
        gabarit_pk = request.POST.get('gabarit_pk', '').strip()
        gabarit = get_object_or_404(PlanningGabarit, pk=gabarit_pk)

        bureaux = get_bureaux()
        posted  = _posted_from_affectations(gabarit.affectations.all())
        errors  = validate_planning(posted, bureaux)
        if errors:
            for err in errors:
                messages.error(request, err)
            messages.error(
                request,
                f"Le gabarit « {gabarit.nom} » contient des conflits (ci-dessus) — "
                f"il n'a pas été appliqué. Corrigez le gabarit avant de réessayer."
            )
            return redirect('planning_modifier', pk=pk)

        try:
            with transaction.atomic():
                planning.affectations.all().delete()
                Affectation.objects.bulk_create([
                    Affectation(
                        planning=planning, plage=ga.plage, jour=ga.jour,
                        personnel=ga.personnel, note=ga.note,
                    )
                    for ga in gabarit.affectations.all()
                ])
                PlanningModification.objects.create(
                    planning=planning,
                    modifie_par=request.user,
                    resume=f'Gabarit « {gabarit.nom} » appliqué (id={gabarit.pk}).',
                )
            messages.success(request, f'Gabarit « {gabarit.nom} » appliqué.')
        except Exception as exc:
            logger.error("Erreur application gabarit %s sur planning %s : %s", gabarit.pk, pk, exc)
            messages.error(request, "Erreur lors de l'application du gabarit. Le planning n'a pas été modifié.")
    return redirect('planning_modifier', pk=pk)


@login_required(login_url='login')
def planning_gabarit_supprimer(request, gabarit_pk):
    if not can_manage_planning(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        gabarit = get_object_or_404(PlanningGabarit, pk=gabarit_pk)
        nom = gabarit.nom
        gabarit.delete()
        messages.success(request, f'Gabarit « {nom} » supprimé.')
    return redirect('planning_bureaux')


# ── Export Excel ──────────────────────────────────────────────────────────────

@login_required(login_url='login')
def planning_export_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    planning = get_object_or_404(PlanningHebdomadaire, pk=pk)
    if not planning.publie and not can_manage_planning(request.user):
        raise PermissionDenied
    bureaux  = get_bureaux()
    aff_map  = {(a.plage_id, a.jour): a for a in planning.affectations.all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sem. {planning.semaine_debut.strftime('%d-%m-%Y')}"

    hdr_fill = PatternFill(start_color='445F35', end_color='445F35', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    bur_fill = PatternFill(start_color='DEECD4', end_color='DEECD4', fill_type='solid')
    thin     = Side(style='thin', color='CCCCCC')
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)

    num_cols = 8
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title = ws.cell(row=1, column=1)
    title.value = (
        f"Centre Médico-Social WALÉ — Planning semaine du "
        f"{planning.semaine_debut.strftime('%d/%m/%Y')} au {planning.semaine_fin.strftime('%d/%m/%Y')}"
    )
    title.font      = Font(bold=True, size=12, color='2B3E22')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    headers = ['BUREAU', 'PLAGE', 'LUNDI', 'MARDI', 'MERCREDI', 'JEUDI', 'VENDREDI', 'SAMEDI']
    for c, h in enumerate(headers, 1):
        cell            = ws.cell(row=2, column=c, value=h)
        cell.fill       = hdr_fill
        cell.font       = hdr_font
        cell.alignment  = Alignment(horizontal='center')
        cell.border     = brd

    row_num = 3
    for bureau in bureaux:
        plages    = list(bureau.plages.all())
        start_row = row_num
        for plage in plages:
            bcell           = ws.cell(row=row_num, column=1, value=bureau.nom)
            bcell.font      = Font(bold=True, size=10)
            bcell.fill      = bur_fill
            bcell.alignment = Alignment(horizontal='center', vertical='center')
            bcell.border    = brd
            pcell           = ws.cell(row=row_num, column=2, value=plage.code)
            pcell.alignment = Alignment(horizontal='center')
            pcell.border    = brd
            for j in range(6):
                aff          = aff_map.get((plage.pk, j))
                cell         = ws.cell(row=row_num, column=j + 3,
                                       value=aff.personnel if aff else '')
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                cell.border  = brd
            row_num += 1
        if len(plages) > 1:
            ws.merge_cells(
                start_row=start_row, start_column=1,
                end_row=row_num - 1, end_column=1,
            )
            ws.cell(row=start_row, column=1).alignment = Alignment(
                horizontal='center', vertical='center'
            )

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 22

    if planning.signataire:
        sig_row = row_num + 2
        fonction_label = dict(FONCTION_SIGNATAIRE_CHOICES).get(
            PlanningConfig.get().fonction_signataire, ''
        )
        ws.cell(row=sig_row, column=7, value=f'{fonction_label} :').font = Font(
            italic=True, color='888888'
        )
        ws.cell(row=sig_row, column=8, value=planning.signataire.nom).font = Font(
            bold=True, color='2B3E22'
        )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname = planning.semaine_debut.strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="planning_{fname}.xlsx"'
    wb.save(response)
    return response
