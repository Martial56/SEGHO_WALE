from datetime import date, time, timedelta, datetime
import calendar
import json

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.db import transaction
from django.utils import timezone

from employer.models import Employe, Presence, JourFerie, Conge
from employer.views import can_manage_rh

_MOIS_FR = {
    1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
    5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
    9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre',
}
_JOURS_FR = {0: 'Lundi', 1: 'Mardi', 2: 'Mercredi', 3: 'Jeudi',
             4: 'Vendredi', 5: 'Samedi', 6: 'Dimanche'}


def _employes_actifs():
    return (Employe.objects.filter(statut='actif')
            .select_related('service', 'fonction')
            .order_by('nom', 'prenoms'))


def _parse_time(val):
    if not val or not val.strip():
        return None
    try:
        from datetime import time
        h, m = val.strip().split(':')
        return time(int(h), int(m))
    except (ValueError, IndexError):
        return None


def can_unlock_registre(user):
    """Seuls les administrateurs peuvent rouvrir un registre verrouillé —
    la RH qui l'a verrouillé (même par erreur) ne peut pas revenir en arrière."""
    return user.is_authenticated and (
        user.is_superuser or user.groups.filter(name='Administrateur').exists()
    )


def _nav_month(year, month):
    if month == 1:
        py, pm = year - 1, 12
    else:
        py, pm = year, month - 1
    if month == 12:
        ny, nm = year + 1, 1
    else:
        ny, nm = year, month + 1
    return py, pm, ny, nm


_LABELS_CRENEAU = {
    'heure_arrivee_matin': 'arrivée matin',
    'heure_depart_matin':  'départ matin',
    'heure_arrivee_soir':  'arrivée soir',
    'heure_depart_soir':   'départ soir',
}


def _creneaux_attendus(permanence, jour_date):
    """Champs heure_* attendus pour un employé présent ce jour."""
    if permanence:
        return ['heure_arrivee_matin', 'heure_depart_soir']
    if jour_date.weekday() == 5:  # samedi : pas de session soir
        return ['heure_arrivee_matin', 'heure_depart_matin']
    return ['heure_arrivee_matin', 'heure_depart_matin', 'heure_arrivee_soir', 'heure_depart_soir']


def _creneaux_manquants(obj):
    """Champs heure_* attendus mais non renseignés, pour un employé présent.
    Liste vide si absent (rien n'est attendu) ou si tout est renseigné."""
    if not obj.present:
        return []
    attendus = _creneaux_attendus(obj.permanence, obj.date)
    return [champ for champ in attendus if getattr(obj, champ) is None]


# ── Registre quotidien ────────────────────────────────────────────────────────

@login_required(login_url='login')
def presence_registre(request):
    if not can_manage_rh(request.user):
        raise PermissionDenied

    date_str = request.GET.get('date', '')
    try:
        selected_date = date.fromisoformat(date_str) if date_str else date.today()
    except ValueError:
        selected_date = date.today()

    from presence.models import RegistreVerrou
    verrou = RegistreVerrou.objects.select_related('verrouille_par').filter(date=selected_date).first()

    # Filtre service
    from medecins.models import Service
    services   = Service.objects.order_by('nom')
    service_id = request.GET.get('service', '')
    service_sel = None
    employes = _employes_actifs()
    if service_id:
        try:
            service_sel = Service.objects.get(pk=service_id)
            employes = employes.filter(service=service_sel)
        except Service.DoesNotExist:
            pass

    if request.method == 'POST':
        if verrou:
            # Journée déjà verrouillée : aucune modification possible, même
            # pour la RH qui l'a verrouillée — seul un administrateur peut
            # la rouvrir (voir presence_deverrouiller).
            messages.error(request, f"Le registre du {selected_date:%d/%m/%Y} est verrouillé, aucune modification n'est possible.")
            url = f"{request.path}?date={selected_date.isoformat()}"
            if service_id:
                url += f"&service={service_id}"
            return redirect(url)

        # IMPORTANT : itérer sur `employes` (le queryset déjà filtré par
        # service ci-dessus), PAS sur _employes_actifs() — le formulaire ne
        # soumet des champs que pour ce sous-ensemble ; boucler sur tous les
        # employés actifs écrasait silencieusement présence/permanence/heures
        # de tous les employés des AUTRES services à chaque enregistrement.
        employes_post = list(employes)
        emp_ids = [e.pk for e in employes_post]
        presences_map = {p.employe_id: p for p in
                         Presence.objects.filter(date=selected_date, employe_id__in=emp_ids)}
        to_create, to_update = [], []
        for emp in employes_post:
            obj = presences_map.get(emp.pk)
            is_new = obj is None
            present = (request.POST.get(f'present_{emp.pk}', '0') == '1')
            if is_new:
                obj = Presence(employe=emp, date=selected_date, present=present)
            deja_pointe = obj.am_in_locked or obj.am_out_locked or obj.pm_in_locked or obj.pm_out_locked
            # Un employé qui a pointé au kiosk est forcément présent : on ne
            # laisse pas le registre le repasser en absent après coup.
            obj.present    = True if deja_pointe else present
            obj.permanence = (request.POST.get(f'perm_{emp.pk}', '0') == '1')
            # Respecter le verrouillage kiosk : ne pas écraser les heures pointées.
            # Un absent n'a aucune heure valide à saisir — la valeur soumise
            # est ignorée (défense en profondeur : les champs sont déjà
            # désactivés côté formulaire, mais on ne fait pas confiance au client).
            if obj.present:
                if not obj.am_in_locked:
                    obj.heure_arrivee_matin = _parse_time(request.POST.get(f'h_am_{emp.pk}'))
                if not obj.am_out_locked:
                    obj.heure_depart_matin  = _parse_time(request.POST.get(f'h_dm_{emp.pk}'))
                if not obj.pm_in_locked:
                    obj.heure_arrivee_soir  = _parse_time(request.POST.get(f'h_as_{emp.pk}'))
                if not obj.pm_out_locked:
                    obj.heure_depart_soir   = _parse_time(request.POST.get(f'h_ds_{emp.pk}'))
            else:
                if not obj.am_in_locked:
                    obj.heure_arrivee_matin = None
                if not obj.am_out_locked:
                    obj.heure_depart_matin = None
                if not obj.pm_in_locked:
                    obj.heure_arrivee_soir = None
                if not obj.pm_out_locked:
                    obj.heure_depart_soir = None
            obj.remarques = request.POST.get(f'rem_{emp.pk}', '').strip()
            if not obj.present:
                obj.motif_absence = request.POST.get(f'motif_{emp.pk}', '').strip()
            else:
                obj.motif_absence = ''
            obj.modifie_par = request.user
            obj.modifie_le  = timezone.now()
            (to_create if is_new else to_update).append(obj)

        # Vérification des créneaux manquants — recalculée côté serveur,
        # sans se fier à la confirmation JS (qui n'est qu'une aide à l'usage).
        incomplets = []
        for obj in to_create + to_update:
            manquants = _creneaux_manquants(obj)
            if manquants:
                incomplets.append((obj.employe, manquants))

        if incomplets and request.POST.get('confirmer_verrouillage') != '1':
            noms = ', '.join(
                f"{emp.nom_complet} ({', '.join(_LABELS_CRENEAU[m] for m in manquants)})"
                for emp, manquants in incomplets
            )
            messages.warning(
                request,
                f"Créneaux non pointés pour : {noms}. Cliquez à nouveau sur "
                "« Enregistrer » pour confirmer malgré tout — le registre du "
                "jour sera alors verrouillé définitivement."
            )
            url = f"{request.path}?date={selected_date.isoformat()}"
            if service_id:
                url += f"&service={service_id}"
            return redirect(url)

        with transaction.atomic():
            if to_create:
                Presence.objects.bulk_create(to_create)
            if to_update:
                Presence.objects.bulk_update(to_update, [
                    'present', 'permanence', 'heure_arrivee_matin', 'heure_depart_matin',
                    'heure_arrivee_soir', 'heure_depart_soir', 'remarques', 'motif_absence',
                    'modifie_par', 'modifie_le',
                ])
            RegistreVerrou.objects.get_or_create(
                date=selected_date, defaults={'verrouille_par': request.user})
        messages.success(request, f"Registre du {selected_date:%d/%m/%Y} enregistré et verrouillé.")
        url = f"{request.path}?date={selected_date.isoformat()}"
        if service_id:
            url += f"&service={service_id}"
        return redirect(url)

    employes = list(employes)
    emp_ids  = [e.pk for e in employes]
    presences_map = {p.employe_id: p for p in
                     Presence.objects.filter(date=selected_date, employe_id__in=emp_ids)}
    conges_map = {c.employe_id: c for c in Conge.objects.filter(
        employe_id__in=emp_ids,
        date_debut__lte=selected_date,
        date_fin__gte=selected_date,
        statut__in=['approuve', 'en_cours'],
    )}

    # Permanences planifiées pour ce jour
    from presence.models import AffectationPermanence
    permanences_planifiees = set(
        AffectationPermanence.objects.filter(
            date=selected_date, employe_id__in=emp_ids
        ).values_list('employe_id', flat=True)
    )

    # Récupérer les heures de permanence des plannings correspondants — un
    # planning distinct par type (personnel / médecins), chacun avec son
    # propre créneau horaire.
    from presence.models import PlanningPermanence
    from medecins.models import Medecin
    lundi = selected_date - timedelta(days=selected_date.weekday())
    plannings_semaine = {
        p.type_permanence: p
        for p in PlanningPermanence.objects.filter(semaine_du=lundi)
    }
    medecin_emp_ids = set(
        Medecin.objects.filter(employe_id__in=emp_ids).values_list('employe_id', flat=True)
    )

    lignes = []
    for i, emp in enumerate(employes, 1):
        p     = presences_map.get(emp.pk)
        conge = conges_map.get(emp.pk)
        if p:
            present = p.present
        elif conge:
            present = False
        else:
            present = True
        # Pré-cocher permanence si planifiée et pas encore saisie
        perm_planifiee = emp.pk in permanences_planifiees
        if p is None and perm_planifiee:
            perm_active = True
        elif p is not None:
            perm_active = p.permanence
        else:
            perm_active = False
        present_locked = bool(p and (p.am_in_locked or p.am_out_locked or p.pm_in_locked or p.pm_out_locked))
        type_perm = 'medecins' if emp.pk in medecin_emp_ids else 'personnel'
        lignes.append({
            'num': i, 'emp': emp, 'p': p,
            'present': present, 'conge': conge,
            'present_locked': present_locked,
            'perm_active': perm_active,
            'perm_planifiee': perm_planifiee,
            'arrivee_manquante': _arrivee_manquante(p),
            'planning_perm': plannings_semaine.get(type_perm),
        })

    # Jour férié ?
    jour_ferie = JourFerie.objects.filter(date=selected_date).first()

    # Navigation jours ouvrés
    prev_d = selected_date - timedelta(days=1)
    next_d = selected_date + timedelta(days=1)
    while prev_d.weekday() >= 5 or JourFerie.objects.filter(date=prev_d).exists():
        prev_d -= timedelta(days=1)
    while next_d.weekday() >= 5 or JourFerie.objects.filter(date=next_d).exists():
        next_d += timedelta(days=1)

    nb_presents = sum(1 for l in lignes if l['present'])

    return render(request, 'presence/registre.html', {
        'selected_date':     selected_date,
        'selected_date_iso': selected_date.isoformat(),
        'jour_fr':           _JOURS_FR[selected_date.weekday()],
        'mois_fr':           _MOIS_FR[selected_date.month],
        'prev_date':         prev_d,
        'next_date':         next_d,
        'lignes':            lignes,
        'services':          services,
        'service_sel':       service_sel,
        'nb_presents':       nb_presents,
        'nb_absents':        len(lignes) - nb_presents,
        'jour_ferie':        jour_ferie,
        'today':             date.today(),
        'can_manage':        True,
        'verrou':            verrou,
        'can_unlock':        can_unlock_registre(request.user),
    })


@login_required(login_url='login')
@require_POST
def presence_deverrouiller(request):
    """Rouvre le registre d'un jour verrouillé — réservé aux administrateurs."""
    if not can_unlock_registre(request.user):
        raise PermissionDenied

    from presence.models import RegistreVerrou
    date_str = request.POST.get('date', '')
    try:
        selected_date = date.fromisoformat(date_str)
    except ValueError:
        raise PermissionDenied

    RegistreVerrou.objects.filter(date=selected_date).delete()
    messages.success(request, f"Registre du {selected_date:%d/%m/%Y} déverrouillé.")

    from django.urls import reverse
    url = reverse('presence_registre') + f"?date={selected_date.isoformat()}"
    service_id = request.POST.get('service', '')
    if service_id:
        url += f"&service={service_id}"
    return redirect(url)


# ── Récapitulatif mensuel (tableau des jours) ─────────────────────────────────

@login_required(login_url='login')
def presence_recap_mensuel(request):
    if not can_manage_rh(request.user):
        raise PermissionDenied

    today = date.today()
    try:
        year  = int(request.GET.get('year',  today.year))
        month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        year, month = today.year, today.month

    last_day = calendar.monthrange(year, month)[1]
    date_embauches = list(
        Employe.objects.filter(statut='actif').values_list('date_embauche', flat=True)
    )
    emp_count = len(date_embauches)

    par_jour = {}
    for p in Presence.objects.filter(date__year=year, date__month=month).values('date', 'present'):
        entry = par_jour.setdefault(p['date'], {'p': 0, 'a': 0})
        entry['p' if p['present'] else 'a'] += 1

    jours = []
    for day in range(1, last_day + 1):
        d = date(year, month, day)
        if d.weekday() >= 5:
            continue
        # Dénominateur = employés déjà embauchés à cette date (pas encore
        # embauché ≠ absent).
        emp_count_jour = sum(1 for de in date_embauches if de <= d)
        counts = par_jour.get(d, {'p': 0, 'a': 0})
        nb_p, nb_a = counts['p'], counts['a']
        jours.append({
            'date':    d,
            'jour_fr': _JOURS_FR[d.weekday()],
            'nb_p':    nb_p,
            'nb_a':    nb_a,
            'nb_ns':   max(0, emp_count_jour - nb_p - nb_a),
            'taux':    round(nb_p / emp_count_jour * 100) if emp_count_jour else 0,
            'futur':   d > today,
        })

    py, pm, ny, nm = _nav_month(year, month)
    return render(request, 'presence/recap_mensuel.html', {
        'jours': jours, 'year': year, 'month': month, 'mois_fr': _MOIS_FR[month],
        'emp_count': emp_count, 'today': today,
        'prev_year': py, 'prev_month': pm, 'next_year': ny, 'next_month': nm,
        'can_manage': True,
    })


# ── Historique d'un employé ───────────────────────────────────────────────────

@login_required(login_url='login')
def presence_employe(request, pk):
    if not can_manage_rh(request.user):
        raise PermissionDenied

    employe = get_object_or_404(Employe.objects.select_related('service', 'fonction'), pk=pk)
    today = date.today()
    try:
        year  = int(request.GET.get('year',  today.year))
        month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        year, month = today.year, today.month

    last_day = calendar.monthrange(year, month)[1]
    presences_map = {p.date: p for p in
                     Presence.objects.filter(employe=employe, date__year=year, date__month=month)}

    # Fetch all approved/en_cours congés overlapping this month in one query
    conges_mois = list(Conge.objects.filter(
        employe=employe,
        statut__in=['approuve', 'en_cours'],
        date_debut__lte=date(year, month, last_day),
        date_fin__gte=date(year, month, 1),
    ))

    def _conge_pour(d):
        for c in conges_mois:
            if c.date_debut <= d <= c.date_fin:
                return c
        return None

    jours = []
    for day in range(1, last_day + 1):
        d = date(year, month, day)
        jours.append({
            'date':    d,
            'jour_fr': _JOURS_FR[d.weekday()],
            'weekend': d.weekday() >= 5,
            'p':       presences_map.get(d),
            'conge':   _conge_pour(d) if d.weekday() < 5 else None,
        })

    jouvres  = sum(1 for j in jours if not j['weekend'])
    nb_p     = sum(1 for j in jours if j['p'] and j['p'].present and not j['weekend'])
    nb_a     = sum(1 for j in jours if j['p'] and not j['p'].present and not j['weekend'])
    nb_retards = sum(
        1 for j in jours
        if j['p'] and j['p'].present and not j['weekend']
        and (j['p'].retard_matin_min > 0 or j['p'].retard_soir_min > 0)
    )
    total_min = sum(
        (j['p'].duree_totale or 0) for j in jours
        if j['p'] and j['p'].present and not j['weekend']
    )
    total_h = total_min // 60
    total_m = total_min % 60

    py, pm, ny, nm = _nav_month(year, month)
    return render(request, 'presence/employe.html', {
        'employe': employe, 'jours': jours,
        'year': year, 'month': month, 'mois_fr': _MOIS_FR[month],
        'jouvres': jouvres, 'nb_p': nb_p, 'nb_a': nb_a,
        'nb_ns':    max(0, jouvres - nb_p - nb_a),
        'nb_retards': nb_retards,
        'taux':     round(nb_p / jouvres * 100, 1) if jouvres else 0,
        'total_h':  total_h,
        'total_m':  total_m,
        'prev_year': py, 'prev_month': pm, 'next_year': ny, 'next_month': nm,
        'can_manage': True,
    })


# ── Statistiques ──────────────────────────────────────────────────────────────

@login_required(login_url='login')
def presence_stats(request):
    if not can_manage_rh(request.user):
        raise PermissionDenied

    today = date.today()
    try:
        year  = int(request.GET.get('year',  today.year))
        month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        year, month = today.year, today.month

    last_day  = calendar.monthrange(year, month)[1]
    jours_ouvres_mois = [date(year, month, d) for d in range(1, last_day + 1)
                         if date(year, month, d).weekday() < 5]
    jouvres = len(jours_ouvres_mois)

    from medecins.models import Service
    services   = Service.objects.order_by('nom')
    service_id = request.GET.get('service', '')
    service_sel = None
    if service_id:
        try:
            service_sel = Service.objects.get(pk=service_id)
        except Service.DoesNotExist:
            pass

    # Un seul aller-retour DB pour tous les employés actifs et toutes les
    # présences du mois (au lieu de 2 requêtes par employé + 2 par service).
    tous_actifs = list(_employes_actifs())
    emp_ids = [e.pk for e in tous_actifs]
    par_emp = {}
    for p in Presence.objects.filter(employe_id__in=emp_ids, date__year=year, date__month=month) \
                              .values('employe_id', 'present'):
        entry = par_emp.setdefault(p['employe_id'], {'p': 0, 'a': 0})
        entry['p' if p['present'] else 'a'] += 1

    def _jouvres_emp(emp):
        # Jours ouvrés du mois à compter de l'embauche : ne pas compter comme
        # "absent" un employé pas encore embauché.
        return sum(1 for d in jours_ouvres_mois if d >= emp.date_embauche)

    stats = []
    stats_svc_map = {}
    for emp in tous_actifs:
        jouvres_emp = _jouvres_emp(emp)
        counts = par_emp.get(emp.pk, {'p': 0, 'a': 0})
        nb_p, nb_a = counts['p'], counts['a']

        if emp.service_id:
            svc_entry = stats_svc_map.setdefault(
                emp.service_id, {'service': emp.service, 'nb_emps': 0, 'total_p': 0, 'total_poss': 0})
            svc_entry['nb_emps']     += 1
            svc_entry['total_p']    += nb_p
            svc_entry['total_poss'] += jouvres_emp

        if service_sel and emp.service_id != service_sel.pk:
            continue
        stats.append({
            'emp': emp,
            'nb_p': nb_p,
            'nb_a': nb_a,
            'nb_ns': max(0, jouvres_emp - nb_p - nb_a),
            'taux': round(nb_p / jouvres_emp * 100, 1) if jouvres_emp else 0,
        })
    stats.sort(key=lambda x: x['taux'])
    paginator  = Paginator(stats, 30)
    page_stats = paginator.get_page(request.GET.get('page'))

    stats_svc = sorted((
        {
            'service': v['service'],
            'nb_emps': v['nb_emps'],
            'taux': round(v['total_p'] / v['total_poss'] * 100, 1) if v['total_poss'] else 0,
        }
        for v in stats_svc_map.values()
    ), key=lambda x: x['service'].nom)

    py, pm, ny, nm = _nav_month(year, month)
    return render(request, 'presence/stats.html', {
        'stats': page_stats, 'stats_svc': stats_svc,
        'page_obj': page_stats,
        'year': year, 'month': month, 'mois_fr': _MOIS_FR[month],
        'jouvres': jouvres, 'services': services, 'service_sel': service_sel,
        'prev_year': py, 'prev_month': pm, 'next_year': ny, 'next_month': nm,
        'can_manage': True,
    })


# ── Rapport de présence ──────────────────────────────────────────────────────

@login_required(login_url='login')
def presence_rapport(request):
    if not can_manage_rh(request.user):
        raise PermissionDenied

    today = date.today()
    try:
        year  = int(request.GET.get('year',  today.year))
        month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        year, month = today.year, today.month

    last_day = calendar.monthrange(year, month)[1]
    jouvres_list = [date(year, month, d) for d in range(1, last_day + 1)
                    if date(year, month, d).weekday() < 5]
    nb_jouvres = len(jouvres_list)

    from medecins.models import Service
    services   = Service.objects.order_by('nom')
    service_id = request.GET.get('service', '')
    service_sel = None
    employes = _employes_actifs()
    if service_id:
        try:
            service_sel = Service.objects.get(pk=service_id)
            employes = employes.filter(service=service_sel)
        except Service.DoesNotExist:
            pass

    employes = list(employes)
    emp_ids  = [e.pk for e in employes]
    presences_par_emp = {}
    for p in Presence.objects.filter(employe_id__in=emp_ids, date__year=year, date__month=month):
        presences_par_emp.setdefault(p.employe_id, {})[p.date] = p

    stats = []
    retards_all = []

    for emp in employes:
        presences = presences_par_emp.get(emp.pk, {})
        # Jours ouvrés à compter de l'embauche : ne pas compter comme
        # "absent" un employé pas encore embauché sur ce mois.
        nb_jouvres_emp = sum(1 for d in jouvres_list if d >= emp.date_embauche)
        nb_p = sum(1 for p in presences.values() if p.present)
        nb_a = sum(1 for p in presences.values() if not p.present)

        nb_retards = 0
        min_retards = 0
        for p in presences.values():
            if not p.present:
                continue
            rm = p.retard_matin_min
            rs = p.retard_soir_min
            if rm > 0:
                nb_retards += 1
                min_retards += rm
                retards_all.append({
                    'emp':          emp,
                    'date':         p.date,
                    'jour_fr':      _JOURS_FR[p.date.weekday()],
                    'type':         'Matin',
                    'heure_arrivee': p.heure_arrivee_matin,
                    'minutes':      rm,
                    'ref':          '08:00',
                })
            if rs > 0:
                nb_retards += 1
                min_retards += rs
                retards_all.append({
                    'emp':          emp,
                    'date':         p.date,
                    'jour_fr':      _JOURS_FR[p.date.weekday()],
                    'type':         'Soir',
                    'heure_arrivee': p.heure_arrivee_soir,
                    'minutes':      rs,
                    'ref':          '15:00',
                })

        # Heures totales travaillées
        total_min = sum(
            (p.duree_totale or 0) for p in presences.values() if p.present
        )
        total_h = total_min // 60
        total_m = total_min % 60

        # Absences avec motif
        absences_motif = [
            {'date': p.date, 'motif': p.motif_absence}
            for p in presences.values()
            if not p.present and p.motif_absence
        ]

        stats.append({
            'emp':           emp,
            'nb_p':          nb_p,
            'nb_a':          nb_a,
            'nb_ns':         max(0, nb_jouvres_emp - nb_p - nb_a),
            'nb_retards':    nb_retards,
            'min_retards':   min_retards,
            'taux':          round(nb_p / nb_jouvres_emp * 100, 1) if nb_jouvres_emp else 0,
            'total_h':       total_h,
            'total_m':       total_m,
            'absences_motif': absences_motif,
        })

    retards_all.sort(key=lambda x: (x['date'], x['emp'].nom))

    # Absences par motif (agrégation globale)
    motifs_count = {}
    for s in stats:
        for ab in s['absences_motif']:
            m = ab['motif'].strip() or 'Non précisé'
            motifs_count[m] = motifs_count.get(m, 0) + 1
    total_motifs = sum(motifs_count.values())
    motifs_list = [
        {'motif': m, 'nb': nb, 'pct': round(nb / total_motifs * 100, 1) if total_motifs else 0}
        for m, nb in sorted(motifs_count.items(), key=lambda x: -x[1])
    ]

    taux_moyen = round(sum(s['taux'] for s in stats) / len(stats), 1) if stats else 0

    # Pagination du récap
    paginator   = Paginator(stats, 30)
    page_stats  = paginator.get_page(request.GET.get('page'))

    py, pm, ny, nm = _nav_month(year, month)
    return render(request, 'presence/rapport.html', {
        'stats':        page_stats,
        'stats_all':    stats,
        'page_obj':     page_stats,
        'retards_all':  retards_all,
        'motifs_list':  motifs_list,
        'year':         year,
        'month':        month,
        'mois_fr':      _MOIS_FR[month],
        'nb_jouvres':   nb_jouvres,
        'nb_employes':  len(stats),
        'nb_retards':   len(retards_all),
        'taux_moyen':   taux_moyen,
        'services':     services,
        'service_sel':  service_sel,
        'prev_year':    py,
        'prev_month':   pm,
        'next_year':    ny,
        'next_month':   nm,
        'today':        today,
        'can_manage':   True,
    })


# ── Export Excel rapport ─────────────────────────────────────────────────────

@login_required(login_url='login')
def presence_rapport_export(request):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    today = date.today()
    try:
        year  = int(request.GET.get('year',  today.year))
        month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        year, month = today.year, today.month

    last_day = calendar.monthrange(year, month)[1]
    jouvres_list = [date(year, month, d) for d in range(1, last_day + 1)
                    if date(year, month, d).weekday() < 5]
    nb_jouvres = len(jouvres_list)

    from medecins.models import Service
    service_id  = request.GET.get('service', '')
    employes    = _employes_actifs()
    if service_id:
        try:
            employes = employes.filter(service=Service.objects.get(pk=service_id))
        except Service.DoesNotExist:
            pass
    employes = list(employes)
    emp_ids  = [e.pk for e in employes]
    presences_par_emp = {}
    for p in Presence.objects.filter(employe_id__in=emp_ids, date__year=year, date__month=month):
        presences_par_emp.setdefault(p.employe_id, {})[p.date] = p

    wb = openpyxl.Workbook()
    # ── Feuille récapitulatif ──
    ws = wb.active
    ws.title = f"Présences {_MOIS_FR[month]} {year}"

    header_fill = PatternFill('solid', fgColor='3A5924')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'Nom & Prénoms', 'Service', 'Présences', 'Absences',
               'Non saisi', 'Retards (nb)', 'Retards (min)', 'Heures trav.', 'Taux %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row = 2
    for i, emp in enumerate(employes, 1):
        presences = presences_par_emp.get(emp.pk, {})
        # Jours ouvrés à compter de l'embauche : ne pas compter comme
        # "absent" un employé pas encore embauché sur ce mois.
        nb_jouvres_emp = sum(1 for d in jouvres_list if d >= emp.date_embauche)
        nb_p = sum(1 for p in presences.values() if p.present)
        nb_a = sum(1 for p in presences.values() if not p.present)
        nb_r = 0; min_r = 0
        for p in presences.values():
            if p.present:
                rm = p.retard_matin_min; rs = p.retard_soir_min
                if rm > 0: nb_r += 1; min_r += rm
                if rs > 0: nb_r += 1; min_r += rs
        tot_min = sum((p.duree_totale or 0) for p in presences.values() if p.present)
        heures = f"{tot_min // 60}h{tot_min % 60:02d}"
        taux   = round(nb_p / nb_jouvres_emp * 100, 1) if nb_jouvres_emp else 0
        data = [i, emp.nom_complet, emp.service.nom if emp.service else '—',
                nb_p, nb_a, max(0, nb_jouvres_emp - nb_p - nb_a), nb_r, min_r, heures, taux]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col not in (2, 3) else 'left')
            if row % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='EDF6E8')
        row += 1

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            (len(str(c.value)) if c.value else 0) for c in col) + 4

    # ── Feuille retards ──
    ws2 = wb.create_sheet("Retards")
    h2 = ['#', 'Employé', 'Service', 'Date', 'Session', 'Heure limite', 'Heure arrivée', 'Retard (min)']
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = border

    row2 = 2; idx = 1
    for emp in employes:
        presences = [p for p in presences_par_emp.get(emp.pk, {}).values() if p.present]
        for p in sorted(presences, key=lambda x: x.date):
            for typ, heure, ref, min_r in [
                ('Matin', p.heure_arrivee_matin, '08:00', p.retard_matin_min),
                ('Soir',  p.heure_arrivee_soir,  '15:00', p.retard_soir_min),
            ]:
                if min_r > 0:
                    data2 = [idx, emp.nom_complet, emp.service.nom if emp.service else '—',
                             p.date.strftime('%d/%m/%Y'), typ, ref,
                             heure.strftime('%H:%M') if heure else '—', min_r]
                    for col, val in enumerate(data2, 1):
                        cell = ws2.cell(row=row2, column=col, value=val)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center' if col not in (2, 3) else 'left')
                        if row2 % 2 == 0:
                            cell.fill = PatternFill('solid', fgColor='FFF3E0')
                    row2 += 1; idx += 1

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = max(
            (len(str(c.value)) if c.value else 0) for c in col) + 4

    fname = f"presence_{_MOIS_FR[month].lower()}_{year}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


# ── Paramètres permanence ────────────────────────────────────────────────────

@login_required(login_url='login')
def presence_parametres(request, type_permanence='personnel'):
    from django.db import IntegrityError
    from presence.models import PlanningPermanence, AffectationPermanence
    from medecins.models import Service

    if type_permanence == 'medecins':
        from medecins.views import can_manage_medecins
        if not can_manage_medecins(request.user):
            raise PermissionDenied
    else:
        if not can_manage_rh(request.user):
            raise PermissionDenied

    today = date.today()

    # Semaine à afficher (par défaut : semaine prochaine)
    days_to_monday = (7 - today.weekday()) % 7 or 7
    default_semaine = today + timedelta(days=days_to_monday)

    semaine_str = request.GET.get('semaine', default_semaine.isoformat())
    try:
        semaine = date.fromisoformat(semaine_str)
        semaine -= timedelta(days=semaine.weekday())  # force lundi
    except ValueError:
        semaine = default_semaine

    # Jours de la semaine (lundi → samedi)
    jours = [semaine + timedelta(days=i) for i in range(6)]

    # Planning de la semaine (créé si inexistant) — un par (semaine, type) :
    # personnel et médecins ont des créneaux indépendants. Créneau par défaut
    # différent selon le type (modifiable ensuite via le formulaire).
    if type_permanence == 'medecins':
        defaults_horaire = {'heure_debut': time(12, 0), 'heure_fin': time(15, 0)}
    else:
        defaults_horaire = {'heure_debut': time(8, 0), 'heure_fin': time(15, 0)}
    planning, _ = PlanningPermanence.objects.get_or_create(
        semaine_du=semaine, type_permanence=type_permanence,
        defaults=defaults_horaire,
    )

    # Filtre service — non pertinent pour les médecins : la permanence
    # concerne l'ensemble des médecins, pas un découpage par service.
    services   = Service.objects.order_by('nom')
    service_id = request.GET.get('service', '') if type_permanence != 'medecins' else ''
    employes   = _employes_actifs()
    if type_permanence == 'medecins':
        employes = employes.filter(fiche_medecin__isnull=False)
    if service_id:
        try:
            employes = employes.filter(service=Service.objects.get(pk=service_id))
        except Service.DoesNotExist:
            pass

    # Le créneau est figé dès qu'au moins un employé (personnel ou médecin)
    # est affecté à ce planning pour un jour de la semaine — plus personne, RH
    # ou administrateur compris, ne peut ensuite modifier heure_debut/heure_fin.
    # Un planning sans aucune affectation reste librement modifiable.
    heures_verrouillees = planning.heures_verrouillees

    if request.method == 'POST':
        if heures_verrouillees:
            # Créneau déjà validé : on ignore toute valeur soumise pour les
            # heures, elles ne bougent plus.
            h_debut = planning.heure_debut
            h_fin   = planning.heure_fin
        else:
            # Heures — repli sur le créneau par défaut du type si non fourni/invalide
            h_debut = _parse_time(request.POST.get('heure_debut', '')) or defaults_horaire['heure_debut']
            h_fin   = _parse_time(request.POST.get('heure_fin',   '')) or defaults_horaire['heure_fin']

        # IMPORTANT : itérer sur `employes` (déjà filtré par service/type
        # ci-dessus), PAS sur _employes_actifs() — le formulaire ne soumet
        # des cases que pour ce sous-ensemble ; boucler sur tous les employés
        # actifs écrasait silencieusement la planification des AUTRES
        # services à chaque enregistrement (même bug que presence_registre).
        employes_post = list(employes)
        nouvelles_affectations = [
            AffectationPermanence(planning=planning, employe=emp, date=jour)
            for emp in employes_post
            for jour in jours
            if request.POST.get(f'perm_{emp.pk}_{jour.isoformat()}') == '1'
        ]

        try:
            with transaction.atomic():
                planning.heure_debut = h_debut
                planning.heure_fin   = h_fin
                planning.cree_par    = request.user
                planning.save()

                # Supprimer les anciennes affectations de cette semaine et
                # recréer selon les cases cochées, en une seule requête groupée.
                AffectationPermanence.objects.filter(planning=planning).delete()
                if nouvelles_affectations:
                    AffectationPermanence.objects.bulk_create(nouvelles_affectations)
        except IntegrityError:
            # Un employé ne peut être en permanence que d'un seul type par
            # jour (unique_together employe+date) — il est déjà affecté à
            # l'autre planning (personnel/médecins) ce jour-là.
            messages.error(request,
                "Enregistrement impossible : un ou plusieurs employés sont déjà "
                "affectés à une autre permanence (personnel/médecins) sur l'un de ces jours.")
            url = f"{request.path}?semaine={semaine.isoformat()}"
            if service_id:
                url += f"&service={service_id}"
            return redirect(url)

        messages.success(request,
            f"Planning permanence semaine du {semaine:%d/%m/%Y} enregistré.")
        url = f"{request.path}?semaine={semaine.isoformat()}"
        if service_id:
            url += f"&service={service_id}"
        return redirect(url)

    # Affectations existantes → set (emp_pk, date)
    affectations = {
        (a.employe_id, a.date)
        for a in AffectationPermanence.objects.filter(planning=planning)
    }

    # Grille : chaque ligne = un employé, chaque colonne = un jour
    grid = [
        {
            'emp':  emp,
            'jours': [(j, (emp.pk, j) in affectations) for j in jours],
        }
        for emp in employes
    ]

    prev_sem = semaine - timedelta(weeks=1)
    next_sem = semaine + timedelta(weeks=1)

    return render(request, 'presence/parametres.html', {
        'planning':            planning,
        'type_permanence':     type_permanence,
        'heures_verrouillees': heures_verrouillees,
        'semaine':             semaine,
        'jours':               jours,
        'grid':                grid,
        'services':            services,
        'service_id':          service_id,
        'prev_sem':            prev_sem,
        'next_sem':            next_sem,
        'today':               today,
        'can_manage':          True,
    })


# ── Kiosque de pointage ───────────────────────────────────────────────────────

_LABELS_ACTION = {
    'arrivee_matin': 'Arrivée matin',
    'depart_matin':  'Départ matin',
    'arrivee_soir':  'Arrivée soir',
    'depart_soir':   'Départ soir',
    # Permanence (horaire continu 8h–15h)
    'arrivee':       'Arrivée (permanence)',
    'depart':        'Départ (permanence)',
}


def _conge_du_jour(emp, d):
    """Retourne le congé approuvé/en cours de l'employé pour la date d, ou None."""
    return Conge.objects.filter(
        employe=emp,
        date_debut__lte=d,
        date_fin__gte=d,
        statut__in=['approuve', 'en_cours'],
    ).first()


def _est_jour_ouvre(d):
    """Retourne (True, None) si ouvré, (False, raison) sinon."""
    if d.weekday() == 6:
        return False, 'Dimanche — repos'
    ferie = JourFerie.objects.filter(date=d).first()
    if ferie:
        return False, f'Jour férié — {ferie.description}'
    return True, None


_MIDI = time(12, 0)


def _est_en_permanence(emp, d, presence_obj=None):
    """Est-ce que l'employé est en permanence ce jour — soit déjà coché dans
    le registre (Presence.permanence), soit simplement planifié via
    Paramètres/permanences sans que la RH n'ait encore ouvert le registre de
    ce jour précis."""
    if presence_obj is None:
        presence_obj = Presence.objects.filter(employe=emp, date=d).first()
    if presence_obj and presence_obj.permanence:
        return True
    from presence.models import AffectationPermanence
    return AffectationPermanence.objects.filter(employe=emp, date=d).exists()


def _arrivee_manquante(presence_obj):
    """Une permanence dont l'arrivée n'a jamais été pointée — signal à
    afficher à la RH dans le registre/l'historique, pas un champ stocké."""
    return bool(presence_obj and presence_obj.permanence and not presence_obj.heure_arrivee_matin)


def _detecter_action(emp, d):
    """Détermine la prochaine action à enregistrer selon ce qui est déjà saisi.

    Permanence (8h–15h) : seulement 2 actions — 'arrivee' puis 'depart'.
    Normal : arrivee_matin → depart_matin → arrivee_soir → depart_soir.
    Après 12h00 : si matin non enregistré, on saute à arrivee_soir.
    Samedi : pas de session soir.
    """
    try:
        p = Presence.objects.get(employe=emp, date=d)
    except Presence.DoesNotExist:
        p = None

    # Sans ce contrôle (via _est_en_permanence), le kiosk traitait l'employé
    # selon le flux normal à 4 pointages tant que personne n'avait pensé à
    # ouvrir le registre avant son arrivée.
    en_permanence = _est_en_permanence(emp, d, presence_obj=p)
    if en_permanence:
        if not (p and p.heure_arrivee_matin):
            # Arrivée ratée : après midi, on ne propose plus de pointer
            # l'arrivée (elle est considérée manquée) — seul le départ en fin
            # de journée reste possible. La RH régularise ensuite manuellement
            # dans le registre (voir `_arrivee_manquante`).
            if datetime.now().time() >= _MIDI:
                return 'depart'
            return 'arrivee'
        if not p.heure_depart_soir:
            return 'depart'
        return 'complet'

    # ── Horaire normal ──
    apres_midi = datetime.now().time() >= _MIDI
    a_soir = d.weekday() < 5
    matin_saute = apres_midi and (p is None or not p.heure_arrivee_matin)

    if not matin_saute:
        if p is None or not p.heure_arrivee_matin:
            return 'arrivee_matin'
        if not p.heure_depart_matin:
            return 'depart_matin'

    if not a_soir:
        return 'complet'
    if p is None or not p.heure_arrivee_soir:
        return 'arrivee_soir'
    if not p.heure_depart_soir:
        return 'depart_soir'
    return 'complet'


_MATIN_LIMITE = time(6, 0)


def _date_reference(emp, now=None):
    """Détermine à quel jour rattacher un pointage à l'instant `now`.

    Si une session soir d'hier est restée ouverte (arrivée pointée, jamais de
    départ) et qu'il est tôt le matin, on rattache le pointage à HIER plutôt
    qu'à aujourd'hui — sinon une garde chevauchant minuit se retrouvait
    réattribuée au jour suivant comme une nouvelle "arrivée matin", laissant
    la session de la veille définitivement incomplète."""
    now = now or datetime.now()
    today = now.date()
    if now.time() < _MATIN_LIMITE:
        hier = today - timedelta(days=1)
        p_hier = Presence.objects.filter(employe=emp, date=hier).first()
        if p_hier and p_hier.heure_arrivee_soir and not p_hier.heure_depart_soir:
            return hier
    return today


def presence_pointage(request):
    """Page kiosque publique — l'employé pointe lui-même."""
    today = date.today()
    now   = datetime.now()
    ouvre, raison_ferme = _est_jour_ouvre(today)
    return render(request, 'presence/pointage.html', {
        'today':        today,
        'ouvre':        ouvre,
        'raison_ferme': raison_ferme,
        'jour_fr':      _JOURS_FR[today.weekday()],
        'mois_fr':      _MOIS_FR[today.month],
        'heure_init':   now.strftime('%H:%M'),
        # Lien "Accès administration" — visible seulement si un RH/Admin/
        # Directeur/Médecin Chef est déjà connecté sur ce poste (can_manage_rh
        # couvre exactement ces rôles), pas au grand public du kiosk.
        'can_access_admin': request.user.is_authenticated and can_manage_rh(request.user),
    })


@require_POST
def presence_chercher(request):
    """AJAX — recherche d'un employé par matricule, QR code ou identifiant biométrique."""
    try:
        data = json.loads(request.body)
        query = data.get('matricule', '').strip().upper()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'ok': False, 'erreur': 'Données invalides'})

    if not query:
        return JsonResponse({'ok': False, 'erreur': 'Identifiant requis'})

    from django.db.models import Q as _Q2
    try:
        emp = Employe.objects.select_related('service', 'fonction').get(
            _Q2(matricule=query) | _Q2(biometric_id=query),
            statut='actif'
        )
    except Employe.DoesNotExist:
        return JsonResponse({'ok': False, 'erreur': 'Matricule introuvable ou employé inactif'})

    today = _date_reference(emp)
    conge = _conge_du_jour(emp, today)
    if conge:
        return JsonResponse({
            'ok':    False,
            'conge': True,
            'erreur': f"En congé — {conge.get_type_conge_display()} (jusqu'au {conge.date_fin:%d/%m/%Y})",
        })

    p_today       = Presence.objects.filter(employe=emp, date=today).first()
    action        = _detecter_action(emp, today)
    en_permanence = _est_en_permanence(emp, today, presence_obj=p_today)
    # Arrivée jamais pointée mais on lui propose quand même "Départ" (règle
    # après-midi ci-dessus) : à distinguer du cas normal pour prévenir
    # l'employé — la RH devra régulariser l'arrivée dans le registre.
    arrivee_manquee = bool(en_permanence and action == 'depart' and not (p_today and p_today.heure_arrivee_matin))

    return JsonResponse({
        'ok': True,
        'emp': {
            'pk':         emp.pk,
            'nom':        emp.nom_complet,
            'matricule':  emp.matricule,
            'service':    emp.service.nom if emp.service else '—',
            'fonction':   emp.fonction.nom if emp.fonction else '',
            'photo':      emp.photo.url if emp.photo else None,
        },
        'action':           action,
        'action_label':     _LABELS_ACTION.get(action, ''),
        'complet':          action == 'complet',
        'en_permanence':    en_permanence,
        'arrivee_manquee':  arrivee_manquee,
    })


@require_POST
def presence_pointer(request):
    """AJAX — enregistre le pointage avec l'heure courante."""
    try:
        data   = json.loads(request.body)
        emp_pk = int(data.get('emp_pk'))
        action = data.get('action', '').strip()
    except (json.JSONDecodeError, AttributeError, TypeError, ValueError):
        return JsonResponse({'ok': False, 'erreur': 'Données invalides'})

    if action not in _LABELS_ACTION:
        return JsonResponse({'ok': False, 'erreur': 'Action inconnue'})


    try:
        emp = Employe.objects.get(pk=emp_pk, statut='actif')
    except Employe.DoesNotExist:
        return JsonResponse({'ok': False, 'erreur': 'Employé introuvable'})

    today = _date_reference(emp)
    ouvre, raison = _est_jour_ouvre(today)
    if not ouvre:
        return JsonResponse({'ok': False, 'erreur': raison})

    # Recontrôlé ici (pas seulement dans presence_chercher) : un client qui
    # sauterait l'étape de recherche pourrait sinon pointer "présent" un
    # employé en congé approuvé sans jamais passer par ce contrôle.
    conge = _conge_du_jour(emp, today)
    if conge:
        return JsonResponse({
            'ok': False,
            'erreur': f"En congé — {conge.get_type_conge_display()} (jusqu'au {conge.date_fin:%d/%m/%Y})",
        })

    from django.db import transaction
    with transaction.atomic():
        # select_for_update verrouille la ligne pour la durée de la
        # transaction : sérialise deux pointages quasi simultanés (double-tap
        # réseau) sur le même employé/jour au lieu de laisser les deux lire
        # le même état "current" avant que l'un des deux n'ait sauvegardé.
        p, _created = Presence.objects.select_for_update().get_or_create(
            employe=emp, date=today, defaults={'present': True}
        )

        # Vérification cohérence — refaite sous verrou, sur l'état le plus
        # frais possible.
        current = _detecter_action(emp, today)
        if current == 'complet':
            return JsonResponse({'ok': False, 'erreur': 'Journée déjà complète'})
        if current != action:
            return JsonResponse({'ok': False, 'erreur': 'Veuillez actualiser et recommencer'})

        now = datetime.now().time().replace(second=0, microsecond=0)
        p.present = True

        if action == 'arrivee_matin':
            p.heure_arrivee_matin = now
            p.am_in_locked = True
        elif action == 'depart_matin':
            p.heure_depart_matin = now
            p.am_out_locked = True
        elif action == 'arrivee_soir':
            p.heure_arrivee_soir = now
            p.pm_in_locked = True
        elif action == 'depart_soir':
            p.heure_depart_soir = now
            p.pm_out_locked = True
        # Permanence — sans ce marquage, le registre (qui ne se fie au
        # planning que tant qu'aucune Presence n'existe) retombait sur
        # `p.permanence=False` dès le premier pointage kiosk du jour.
        elif action == 'arrivee':
            p.permanence = True
            p.heure_arrivee_matin = now
            p.am_in_locked = True
        elif action == 'depart':
            p.permanence = True
            p.heure_depart_soir = now
            p.pm_out_locked = True

        p.save()

    # Prochaine action
    next_action = _detecter_action(emp, today)
    return JsonResponse({
        'ok':          True,
        'heure':       now.strftime('%H:%M'),
        'action':      action,
        'label':       _LABELS_ACTION[action],
        'next_action': next_action,
        'complet':     next_action == 'complet',
    })
