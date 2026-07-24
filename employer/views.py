from datetime import date, timedelta
import io
import re

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse

from django.utils import timezone
from .models import (Employe, Fonction, Grade, TypeContrat, Nationalite,
                     DocumentEmploye, InfoSupplementaire, AlerteContrat,
                     AlerteDocument, HistoriqueEmploye, Conge, Presence,
                     JourFerie, CredentialBiometrique, DOCS_OBLIGATOIRES)
from medecins.models import Service

RH_MANAGE_GROUPS = {'Médecin Chef', 'Médecin Chef Adjoint', 'Administrateur', 'Directeur', 'RH'}


def _generate_service_code(nom):
    """Code unique pour un Service créé à la volée (import Excel) — Service.code est unique."""
    import re
    base = re.sub(r'[^A-Z0-9]', '', nom.upper())[:16] or 'SVC'
    code = base
    n = 1
    while Service.objects.filter(code=code).exists():
        n += 1
        code = f'{base[:16 - len(str(n))]}{n}'
    return code


def _upper_no_accent(s):
    """Majuscules sans accent — imposé sur le champ « nom » des formulaires de
    saisie Grade/Fonction/Type de contrat (Configuration)."""
    import unicodedata
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return s.upper()


def _normalize_nom(s):
    """Clé de comparaison insensible à la casse/accents/ponctuation — évite de créer un
    doublon « DIRECTEUR GENERAL » quand « Directeur Général » existe déjà (import Excel)."""
    import re
    import unicodedata
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^A-Za-z0-9]+', '', s)
    return s.upper()


def _find_or_create_ref(model, nom, extra_create=None):
    """Cherche un objet (Service/Fonction/Grade/TypeContrat) par nom en ignorant
    casse/accents/ponctuation avant d'en créer un nouveau — utilisé par l'import Excel."""
    cible = _normalize_nom(nom)
    for obj in model.objects.all():
        if _normalize_nom(obj.nom) == cible:
            return obj
    kwargs = {'nom': nom}
    if extra_create:
        kwargs.update(extra_create)
    return model.objects.create(**kwargs)

LABELS_DOCS = {'cni': "Carte d'identité", 'contrat': 'Contrat signé', 'diplome': 'Diplôme'}

def _empty_form():
    d = {
        'nom': '', 'prenoms': '', 'sexe': 'M', 'date_naissance': '',
        'lieu_naissance': '', 'nationalite': '', 'situation_matrimoniale': '',
        'nombre_enfants': '0', 'telephone': '', 'telephone2': '', 'email': '',
        'adresse': '', 'service': '', 'fonction': '', 'grade': '', 'type_contrat': '',
        'date_embauche': '', 'date_fin_contrat': '', 'salaire_base': '', 'statut': 'actif',
        'notes': '',
    }
    defaut = Nationalite.objects.filter(nom__iexact='Ivoirienne').first()
    if defaut:
        d['nationalite'] = str(defaut.pk)
    return d


def _employe_to_form(e):
    fmt = lambda d: d.strftime('%Y-%m-%d') if d else ''
    return {
        'nom':                    e.nom or '',
        'prenoms':                e.prenoms or '',
        'sexe':                   e.sexe or 'M',
        'date_naissance':         fmt(e.date_naissance),
        'lieu_naissance':         e.lieu_naissance or '',
        'nationalite':            str(e.nationalite_id) if e.nationalite_id else '',
        'situation_matrimoniale': e.situation_matrimoniale or '',
        'nombre_enfants':         str(e.nombre_enfants) if e.nombre_enfants is not None else '0',
        'telephone':              e.telephone or '',
        'telephone2':             e.telephone2 or '',
        'email':                  e.email or '',
        'adresse':                e.adresse or '',
        'service':                str(e.service_id)      if e.service_id      else '',
        'fonction':               str(e.fonction_id)     if e.fonction_id     else '',
        'grade':                  str(e.grade_id)        if e.grade_id        else '',
        'type_contrat':           str(e.type_contrat_id) if e.type_contrat_id else '',
        'date_embauche':          fmt(e.date_embauche),
        'date_fin_contrat':       fmt(e.date_fin_contrat),
        'salaire_base':           str(int(e.salaire_base)) if e.salaire_base else '',
        'statut':                 e.statut or 'actif',
        'notes':                  e.notes or '',
    }


def can_manage_rh(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    return user.groups.filter(name__in=RH_MANAGE_GROUPS).exists()


def _rh_selects(user):
    ordre_cat = ['direction', 'medical', 'paramedical', 'communautaire', 'support']
    labels_cat = dict(Fonction.CATEGORIE_CHOICES)
    fonctions_qs = Fonction.objects.all()
    groupes = []
    for cat in ordre_cat:
        items = [f for f in fonctions_qs if f.categorie == cat]
        if items:
            groupes.append({'label': labels_cat.get(cat, cat), 'items': items})
    # Fonctions sans catégorie
    autres = [f for f in fonctions_qs if f.categorie not in ordre_cat]
    if autres:
        groupes.append({'label': 'Autres', 'items': autres})
    return {
        'fonctions':         fonctions_qs,
        'fonctions_groupes': groupes,
        'grades':            Grade.objects.all(),
        'types_contrat':     TypeContrat.objects.all(),
        'nationalites':      Nationalite.objects.all(),
        'services':          Service.objects.filter(actif=True).order_by('nom'),
        'can_manage':        can_manage_rh(user),
    }


def _enregistrer_historique(employe, type_changement, ancienne, nouvelle, note, user):
    if ancienne != nouvelle:
        HistoriqueEmploye.objects.create(
            employe=employe,
            type_changement=type_changement,
            ancienne_valeur=str(ancienne),
            nouvelle_valeur=str(nouvelle),
            note=note,
            fait_par=user,
        )


# â"€â"€ Dashboard â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def rh_dashboard(request):
    today = date.today()

    total    = Employe.objects.count()
    actifs   = Employe.objects.filter(statut='actif').count()
    suspendus = Employe.objects.filter(statut='suspendu').count()
    quittes  = Employe.objects.filter(statut='quitte').count()

    # Répartition par type de contrat
    par_contrat = (
        Employe.objects.filter(statut='actif')
        .values('type_contrat__nom')
        .annotate(n=Count('id'))
        .order_by('-n')
    )

    # Répartition par service
    par_service = (
        Employe.objects.filter(statut='actif')
        .values('service__nom')
        .annotate(n=Count('id'))
        .order_by('-n')
    )

    # Répartition par nationalité
    par_nationalite = (
        Employe.objects.filter(statut='actif')
        .values('nationalite__nom')
        .annotate(n=Count('id'))
        .order_by('-n')
    )

    # Contrats expirant dans 90 jours
    fin_90j = (
        Employe.objects.filter(
            statut='actif',
            date_fin_contrat__gte=today,
            date_fin_contrat__lte=today + timedelta(days=90),
        )
        .select_related('service', 'type_contrat')
        .order_by('date_fin_contrat')
    )

    # Pyramide des âges (tranches)
    tranches = [
        ('<25 ans',  0,  24),
        ('25-34',   25,  34),
        ('35-44',   35,  44),
        ('45-54',   45,  54),
        ('55+',     55, 200),
    ]
    pyramide = []
    for label, amin, amax in tranches:
        n = sum(
            1 for e in Employe.objects.filter(statut='actif', date_naissance__isnull=False)
            if amin <= e.age <= amax
        )
        pyramide.append({'label': label, 'n': n})
    # Employés actifs sans date de naissance renseignée — affiché à part pour
    # que le total de la pyramide corresponde bien à l'effectif actif réel
    # (sinon ils disparaissent silencieusement du graphique).
    nb_age_non_renseigne = Employe.objects.filter(statut='actif', date_naissance__isnull=True).count()
    if nb_age_non_renseigne:
        pyramide.append({'label': 'Non renseigné', 'n': nb_age_non_renseigne})

    # Docs manquants
    employes_docs_manquants = [
        e for e in Employe.objects.filter(statut='actif').select_related('service').prefetch_related('documents')
        if e.docs_manquants
    ]
    _alertes_svc = {}
    for _e in employes_docs_manquants:
        _svc = _e.service.nom if _e.service else 'Non affecté'
        if _svc not in _alertes_svc:
            _alertes_svc[_svc] = {'service': _svc, 'employes': 0, 'docs': 0}
        _alertes_svc[_svc]['employes'] += 1
        _alertes_svc[_svc]['docs'] += len(_e.docs_manquants)
    alertes_docs_par_service = sorted(_alertes_svc.values(), key=lambda x: -x['docs'])

    docs_paginator = Paginator(employes_docs_manquants, 5)
    docs_page_obj = docs_paginator.get_page(request.GET.get('docs_page'))

    # ── Turnover ─────────────────────────────────────────────────────────────
    annee = today.year
    entrees_annee = Employe.objects.filter(date_embauche__year=annee).count()
    departs_annee = Employe.objects.filter(statut='quitte', date_depart__year=annee).count()
    # Taux de turnover (formule simplifiée BTP/RH) : départs / effectif actif * 100
    taux_turnover = round(departs_annee / actifs * 100, 1) if actifs > 0 else 0
    # Ancienneté moyenne + tranches
    anciennete_moy = None
    emp_avec_date = list(Employe.objects.filter(statut='actif', date_embauche__isnull=False))
    if emp_avec_date:
        total_annees = sum(e.anciennete['annees'] for e in emp_avec_date)
        anciennete_moy = round(total_annees / len(emp_avec_date), 1)

    tranches_anc = [
        ('< 1 an',     0,   0),
        ('1 – 3 ans',  1,   2),
        ('3 – 5 ans',  3,   4),
        ('5 – 10 ans', 5,   9),
        ('10 – 15 ans',10,  14),
        ('15 ans +',  15, 999),
    ]
    anciennete_tranches = [
        {'label': label, 'n': sum(1 for e in emp_avec_date if amin <= e.anciennete['annees'] <= amax)}
        for label, amin, amax in tranches_anc
    ]

    # Anniversaires du mois en cours
    MOIS_FR = ['janvier','février','mars','avril','mai','juin',
               'juillet','août','septembre','octobre','novembre','décembre']
    mois_label = MOIS_FR[today.month - 1]
    anniversaires = (
        Employe.objects
        .filter(statut='actif', date_naissance__isnull=False, date_naissance__month=today.month)
        .select_related('service')
        .order_by('date_naissance__day')
    )
    anniversaires = [
        {
            'employe': e,
            'jour': e.date_naissance.day,
            'age': today.year - e.date_naissance.year,
            'aujourd_hui': e.date_naissance.day == today.day,
        }
        for e in anniversaires
    ]

    # ── Répartition Homme / Femme ─────────────────────────────────────────────
    nb_hommes  = Employe.objects.filter(statut='actif', sexe='M').count()
    nb_femmes  = Employe.objects.filter(statut='actif', sexe='F').count()
    nb_sexe_non_renseigne = actifs - nb_hommes - nb_femmes
    genre_data = {
        'hommes': nb_hommes, 'femmes': nb_femmes,
        'non_renseigne': nb_sexe_non_renseigne,
        'total': actifs,
    }

    # ── Nouveaux employés du mois ─────────────────────────────────────────────
    nouveaux_mois = list(
        Employe.objects
        .filter(date_embauche__year=today.year, date_embauche__month=today.month)
        .select_related('service', 'fonction')
        .order_by('date_embauche')
    )

    # ── Congés en cours ───────────────────────────────────────────────────────
    conges_en_cours = list(
        Conge.objects
        .filter(statut='en_cours')
        .select_related('employe', 'employe__service')
        .order_by('date_fin')
    )

    # ── Employés en période d'essai (embauchés depuis < 90 jours) ────────────
    periode_essai = list(
        Employe.objects
        .filter(statut='actif', date_embauche__gte=today - timedelta(days=90))
        .select_related('service', 'type_contrat')
        .order_by('date_embauche')
    )

    # ── Taux de présence du mois ──────────────────────────────────────────────
    presences_mois  = Presence.objects.filter(date__year=today.year, date__month=today.month)
    total_presences = presences_mois.count()
    presents_count  = presences_mois.filter(present=True).count()
    taux_presence   = round(presents_count / total_presences * 100, 1) if total_presences > 0 else None


    return render(request, 'employer/dashboard.html', {
        'total': total, 'actifs': actifs, 'suspendus': suspendus, 'quittes': quittes,
        'par_contrat': list(par_contrat),
        'par_service': list(par_service),
        'par_nationalite': list(par_nationalite),
        'fin_90j': fin_90j,
        'pyramide': pyramide,
        'employes_docs_manquants': employes_docs_manquants,
        'docs_page_obj': docs_page_obj,
        'annee':          annee,
        'entrees_annee':  entrees_annee,
        'departs_annee':  departs_annee,
        'taux_turnover':  taux_turnover,
        'anciennete_moy': anciennete_moy,
        'anciennete_tranches': anciennete_tranches,
        'anniversaires': anniversaires,
        'mois_label': mois_label,
        'genre_data':      genre_data,
        'nouveaux_mois':   nouveaux_mois,
        'conges_en_cours': conges_en_cours,
        'periode_essai':   periode_essai,
        'taux_presence':   taux_presence,
        'total_presences': total_presences,
        'presents_count':  presents_count,
        'alertes_docs_par_service': alertes_docs_par_service,
        'can_manage': can_manage_rh(request.user),
    })


# ── Identification employé → médecin ──────────────────────────────────────────

@login_required(login_url='login')
@require_POST
def employe_supprimer(request, pk):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    employe = get_object_or_404(Employe, pk=pk)
    nom = employe.nom_complet

    # Medecin.employe est en CASCADE : supprimer l'employé supprimerait sa
    # fiche médecin, qui déconnecterait à son tour (SET_NULL) tout son
    # historique clinique (consultations, ordonnances, rendez-vous,
    # hospitalisations). On refuse plutôt que de perdre cette traçabilité
    # silencieusement — il faut d'abord retirer/désactiver la fiche médecin.
    if hasattr(employe, 'fiche_medecin'):
        messages.error(
            request,
            f"Impossible de supprimer {nom} : cette personne a une fiche médecin "
            f"active (historique de consultations/ordonnances/rendez-vous). "
            f"Désactivez d'abord sa fiche médecin depuis le module Médecins."
        )
        return redirect('ressources_humaines_list')

    employe.delete()
    messages.success(request, f"L'employé {nom} a été supprimé.")
    return redirect('ressources_humaines_list')


@login_required(login_url='login')
def employe_vers_medecin(request, pk):
    from medecins.models import Medecin

    if not can_manage_rh(request.user):
        raise PermissionDenied

    employe = get_object_or_404(Employe, pk=pk)

    # Déjà lié : rester sur la fiche employé
    if hasattr(employe, 'fiche_medecin'):
        return redirect('rh_detail', pk=employe.pk)

    med = Medecin(
        actif=(employe.statut == 'actif'),
        employe=employe,
        service=employe.service,
    )
    med.save()

    messages.success(
        request,
        f"Fiche médecin créée pour {employe.nom_complet} ({employe.matricule}). "
        f"Complétez la spécialité, le taux honoraire et le numéro d'ordre depuis le module Médecins."
    )
    return redirect('rh_detail', pk=employe.pk)


# ── Liste ─────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def employe_list(request):
    qs = Employe.objects.select_related('service', 'fonction', 'grade', 'type_contrat')

    q         = request.GET.get('q', '').strip()
    f_service = request.GET.get('service', '').strip()
    f_statut  = request.GET.get('statut', '').strip()

    if q:
        qs = qs.filter(
            Q(nom__icontains=q) | Q(prenoms__icontains=q) | Q(matricule__icontains=q)
        )
    if f_service:
        qs = qs.filter(service_id=f_service)
    if f_statut:
        qs = qs.filter(statut=f_statut)

    paginator = Paginator(qs, 10)
    page_obj  = paginator.get_page(request.GET.get('page'))

    stats = {
        'total':     Employe.objects.count(),
        'actifs':    Employe.objects.filter(statut='actif').count(),
        'suspendus': Employe.objects.filter(statut='suspendu').count(),
        'quittes':   Employe.objects.filter(statut='quitte').count(),
    }

    return render(request, 'employer/list.html', {
        'page_obj':  page_obj,
        'stats':     stats,
        'services':  Service.objects.filter(actif=True).order_by('nom'),
        'statuts':   Employe.STATUT_CHOICES,
        'q':         q,
        'f_service': f_service,
        'f_statut':  f_statut,
        'can_manage': can_manage_rh(request.user),
    })


# â"€â"€ Détail â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def employe_detail(request, pk):
    employe = get_object_or_404(
        Employe.objects.select_related('service', 'fonction', 'grade', 'type_contrat')
                       .prefetch_related('documents', 'historique__fait_par'),
        pk=pk
    )
    from .models import TYPE_DOC_CHOICES
    docs_manquants = [
        {'code': d, 'label': LABELS_DOCS.get(d, d)} for d in employe.docs_manquants
    ]
    conges = Conge.objects.filter(employe=employe).order_by('-date_demande')[:10]
    return render(request, 'employer/detail.html', {
        'employe':          employe,
        'docs':             employe.documents.select_related('ajoute_par').all(),
        'infos_supp':       employe.infos_supp.all(),
        'type_doc_choices': TYPE_DOC_CHOICES,
        'docs_manquants':   docs_manquants,
        'historique':       employe.historique.all()[:30],
        'conges':           conges,
        'can_manage':       can_manage_rh(request.user),
    })


# â"€â"€ Créer â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def employe_nouveau(request):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    ctx = _rh_selects(request.user)
    form_data = _empty_form()
    form_errors = {}
    if request.method == 'POST':
        employe, form_errors = _save_employe(request, None)
        if employe:
            HistoriqueEmploye.objects.create(
                employe=employe, type_changement='creation',
                nouvelle_valeur=employe.nom_complet, fait_par=request.user,
            )
            messages.success(request, f'Employé {employe.nom_complet} créé avec succès.')
            return redirect('rh_detail', pk=employe.pk)
        messages.error(request, 'Veuillez corriger les erreurs signalées ci-dessous.')
        form_data = request.POST
    return render(request, 'employer/form.html', {
        **ctx, 'employe': None, 'form_data': form_data, 'form_errors': form_errors,
    })


# â"€â"€ Modifier â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def employe_modifier(request, pk):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    employe = get_object_or_404(Employe, pk=pk)
    ctx = _rh_selects(request.user)
    form_data = _employe_to_form(employe)
    form_errors = {}
    if request.method == 'POST':
        # Snapshot avant modif
        old_statut  = employe.statut
        old_salaire = str(int(employe.salaire_base)) if employe.salaire_base else '0'
        old_service = str(employe.service) if employe.service else ''

        updated, form_errors = _save_employe(request, employe)
        if updated:
            _enregistrer_historique(updated, 'statut', old_statut, updated.statut,
                                    '', request.user)
            new_salaire = str(int(updated.salaire_base)) if updated.salaire_base else '0'
            _enregistrer_historique(updated, 'salaire', old_salaire + ' FCFA',
                                    new_salaire + ' FCFA', '', request.user)
            new_service = str(updated.service) if updated.service else ''
            _enregistrer_historique(updated, 'service', old_service, new_service,
                                    '', request.user)
            messages.success(request, 'Dossier mis à jour avec succès.')
            return redirect('rh_detail', pk=employe.pk)
        messages.error(request, 'Veuillez corriger les erreurs signalées ci-dessous.')
        form_data = request.POST
    return render(request, 'employer/form.html', {
        **ctx, 'employe': employe, 'form_data': form_data, 'form_errors': form_errors,
    })


def _save_employe(request, employe):
    """Retourne (employe, errors) — employe est None et errors non-vide si la
    validation échoue ; errors est un dict {nom_du_champ: message} affiché
    directement sous le champ concerné dans le template."""
    p = request.POST
    nom           = p.get('nom', '').strip()
    prenoms       = p.get('prenoms', '').strip()
    date_embauche = p.get('date_embauche', '').strip()

    errors = {}
    if not nom:
        errors['nom'] = 'Le nom est obligatoire.'
    if not prenoms:
        errors['prenoms'] = 'Les prénoms sont obligatoires.'
    if not date_embauche:
        errors['date_embauche'] = "La date d'embauche est obligatoire."
    if not p.get('sexe'):
        errors['sexe'] = 'Le sexe est obligatoire.'
    if not p.get('date_naissance'):
        errors['date_naissance'] = 'La date de naissance est obligatoire.'
    if not p.get('nationalite'):
        errors['nationalite'] = 'La nationalité est obligatoire.'
    telephone = p.get('telephone', '').strip()
    if not telephone:
        errors['telephone'] = 'Le téléphone principal est obligatoire.'
    elif len(re.sub(r'\D', '', telephone)) != 10:
        errors['telephone'] = 'Le téléphone doit contenir exactement 10 chiffres.'
    telephone2 = p.get('telephone2', '').strip()
    if telephone2 and len(re.sub(r'\D', '', telephone2)) != 10:
        errors['telephone2'] = 'Le téléphone doit contenir exactement 10 chiffres.'
    if not p.get('fonction'):
        errors['fonction'] = 'La fonction est obligatoire.'
    if not p.get('type_contrat'):
        errors['type_contrat'] = 'Le type de contrat est obligatoire.'
    if errors:
        return None, errors

    nom_norm = _normalize_nom(nom)
    prenoms_norm = _normalize_nom(prenoms)
    qs_dup = Employe.objects.all()
    if employe is not None and employe.pk:
        qs_dup = qs_dup.exclude(pk=employe.pk)
    for autre in qs_dup:
        if _normalize_nom(autre.nom) == nom_norm and _normalize_nom(autre.prenoms) == prenoms_norm:
            msg = f'Un employé nommé « {nom} {prenoms} » existe déjà (matricule {autre.matricule}).'
            return None, {'nom': msg, 'prenoms': msg}

    if employe is None:
        employe = Employe()

    employe.nom     = nom
    employe.prenoms = prenoms
    employe.sexe    = p.get('sexe', '')
    employe.date_naissance         = p.get('date_naissance') or None
    employe.lieu_naissance         = p.get('lieu_naissance', '').strip()
    employe.situation_matrimoniale = p.get('situation_matrimoniale', '')
    employe.nombre_enfants         = int(p.get('nombre_enfants') or 0)
    employe.telephone  = p.get('telephone', '').strip()
    employe.telephone2 = p.get('telephone2', '').strip()
    employe.email      = p.get('email', '').strip()
    employe.adresse    = p.get('adresse', '').strip()

    def fk_or_none(model, pk_str):
        try:
            return model.objects.get(pk=int(pk_str)) if pk_str else None
        except (model.DoesNotExist, ValueError, TypeError):
            return None

    employe.service      = fk_or_none(Service, p.get('service'))
    employe.fonction     = fk_or_none(Fonction, p.get('fonction'))
    employe.grade        = fk_or_none(Grade, p.get('grade'))
    employe.type_contrat = fk_or_none(TypeContrat, p.get('type_contrat'))
    employe.nationalite  = fk_or_none(Nationalite, p.get('nationalite'))
    employe.date_embauche    = p.get('date_embauche') or None
    employe.date_fin_contrat = p.get('date_fin_contrat') or None
    employe.salaire_base     = p.get('salaire_base') or 0
    employe.statut           = p.get('statut', 'actif')
    employe.notes            = p.get('notes', '').strip()

    if 'photo' in request.FILES:
        if employe.photo:
            employe.photo.delete(save=False)
        employe.photo = request.FILES['photo']
    elif p.get('photo_supprimer') == '1' and employe.photo:
        employe.photo.delete(save=False)
        employe.photo = None

    employe.save()
    return employe, {}


# â"€â"€ Renouvellement de contrat â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def employe_renouveler(request, pk):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    employe = get_object_or_404(Employe, pk=pk)
    if request.method == 'POST':
        nouvelle_date = request.POST.get('date_fin_contrat', '').strip()
        if not nouvelle_date:
            messages.error(request, 'La nouvelle date de fin de contrat est obligatoire.')
            return redirect('rh_detail', pk=pk)
        ancienne = employe.date_fin_contrat.strftime('%d/%m/%Y') if employe.date_fin_contrat else 'â€"'
        employe.date_fin_contrat = nouvelle_date
        employe.save()
        # Réinitialiser les alertes
        AlerteContrat.objects.filter(employe=employe).delete()
        # Historique
        HistoriqueEmploye.objects.create(
            employe=employe,
            type_changement='contrat',
            ancienne_valeur=ancienne,
            nouvelle_valeur=date.fromisoformat(nouvelle_date).strftime('%d/%m/%Y'),
            fait_par=request.user,
        )
        messages.success(request, 'Contrat renouvelé avec succès.')
    return redirect('rh_detail', pk=pk)


# â"€â"€ Export Excel â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def employe_export_excel(request):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    qs = Employe.objects.select_related('service', 'fonction', 'grade', 'type_contrat')
    q         = request.GET.get('q', '').strip()
    f_service = request.GET.get('service', '').strip()
    f_statut  = request.GET.get('statut', '').strip()
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(prenoms__icontains=q) | Q(matricule__icontains=q))
    if f_service:
        qs = qs.filter(service_id=f_service)
    if f_statut:
        qs = qs.filter(statut=f_statut)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employés"

    header_fill = PatternFill('solid', fgColor='4A7236')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),  bottom=Side(style='thin', color='D0D0D0'),
    )

    headers = [
        'Matricule', 'Nom', 'Prénoms', 'Sexe', 'Date naissance', 'Nationalité',
        'Téléphone', 'Email', 'Service', 'Fonction', 'Grade', 'Type contrat',
        'Date embauche', 'Date fin contrat', 'Ancienneté', 'Salaire base', 'Statut',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
    ws.row_dimensions[1].height = 20

    alt_fill = PatternFill('solid', fgColor='EDF6E8')
    for row_n, emp in enumerate(qs, 2):
        fill = alt_fill if row_n % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        anc = emp.anciennete
        row = [
            emp.matricule, emp.nom, emp.prenoms,
            'M' if emp.sexe == 'M' else 'F',
            emp.date_naissance.strftime('%d/%m/%Y') if emp.date_naissance else '',
            emp.nationalite.nom if emp.nationalite else '',
            emp.telephone, emp.email,
            emp.service.nom if emp.service else '',
            emp.fonction.nom if emp.fonction else '',
            emp.grade.nom if emp.grade else '',
            emp.type_contrat.nom if emp.type_contrat else '',
            emp.date_embauche.strftime('%d/%m/%Y') if emp.date_embauche else '',
            emp.date_fin_contrat.strftime('%d/%m/%Y') if emp.date_fin_contrat else '',
            anc['label'],
            int(emp.salaire_base) if emp.salaire_base else 0,
            emp.get_statut_display(),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_n, column=col, value=val)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(vertical='center')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True
        ws.column_dimensions[get_column_letter(col)].width = max(12, 18)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="employes.xlsx"'
    return resp


# â"€â"€ Export PDF (fiche individuelle) â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def employe_qrcode(request, pk):
    """Retourne le QR code de l'employé en PNG (contenu = matricule)."""
    if not can_manage_rh(request.user):
        raise PermissionDenied
    employe = get_object_or_404(Employe, pk=pk)
    import qrcode, io, base64
    qr = qrcode.QRCode(version=1, box_size=8, border=2,
                       error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(employe.matricule)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return HttpResponse(buf.getvalue(), content_type='image/png')


@login_required(login_url='login')
def employe_badge(request, pk):
    """Page badge imprimable avec photo + QR + infos."""
    if not can_manage_rh(request.user):
        raise PermissionDenied
    employe = get_object_or_404(
        Employe.objects.select_related('service', 'fonction'), pk=pk
    )
    import qrcode, io, base64
    qr = qrcode.QRCode(version=1, box_size=6, border=2,
                       error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(employe.matricule)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#1b2e13', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()
    from datetime import date as _date
    return render(request, 'employer/badge.html', {
        'employe': employe,
        'qr_b64':  qr_b64,
        'today':   _date.today(),
    })


@login_required(login_url='login')
@require_POST
def employe_biometric_save(request, pk):
    """Enregistre l'identifiant biométrique de l'employé."""
    if not can_manage_rh(request.user):
        raise PermissionDenied
    employe = get_object_or_404(Employe, pk=pk)
    biometric_id = request.POST.get('biometric_id', '').strip()
    employe.biometric_id = biometric_id
    employe.save(update_fields=['biometric_id'])
    messages.success(request, f"Identifiant biométrique mis à jour pour {employe.nom_complet}.")
    return redirect('rh_detail', pk=pk)


@login_required(login_url='login')
def employe_fiche_pdf(request, pk):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    employe = get_object_or_404(
        Employe.objects.select_related('service', 'fonction', 'grade', 'type_contrat')
                       .prefetch_related('documents'),
        pk=pk
    )
    return render(request, 'employer/fiche_print.html', {
        'employe': employe,
        'today': date.today(),
    })



# ── Registre unique du personnel ──────────────────────────────────────────────

@login_required(login_url='login')
def rh_registre(request):
    qs = Employe.objects.select_related('service', 'fonction', 'type_contrat')\
                        .order_by('date_embauche', 'nom')
    statut = request.GET.get('statut', '')
    if statut:
        qs = qs.filter(statut=statut)
    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'employer/registre.html', {
        'employes':      page_obj,
        'page_obj':      page_obj,
        'statut_filtre': statut,
        'can_manage':    can_manage_rh(request.user),
        'today':         date.today(),
    })


@login_required(login_url='login')
def rh_registre_export(request):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    qs = Employe.objects.select_related('service', 'fonction', 'type_contrat')\
                        .order_by('date_embauche', 'nom')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registre du Personnel'

    vert  = '1b4332'
    blanc = 'FFFFFF'
    gris  = 'F2F7EF'

    entetes = [
        'No', 'Matricule', 'Nom & Prenoms', 'Sexe', 'Date de naissance',
        'Lieu de naissance', 'Nationalite', 'Fonction', 'Service',
        'Type de contrat', "Date d'embauche", 'Date fin contrat', 'Statut', 'Observations',
    ]
    largeurs = [5, 14, 28, 8, 16, 20, 14, 22, 18, 18, 16, 16, 12, 28]

    # Titre
    ws.merge_cells('A1:N1')
    titre = ws['A1']
    titre.value = 'REGISTRE UNIQUE DU PERSONNEL - Centre Medico-Social WALE'
    titre.font = Font(bold=True, size=13, color=blanc)
    titre.fill = PatternFill('solid', fgColor=vert)
    titre.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # En-têtes
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, (label, larg) in enumerate(zip(entetes, largeurs), 1):
        cell = ws.cell(row=2, column=col, value=label)
        cell.font = Font(bold=True, size=10, color=blanc)
        cell.fill = PatternFill('solid', fgColor='345726')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = larg
    ws.row_dimensions[2].height = 32

    fmt_date = lambda d: d.strftime('%d/%m/%Y') if d else '-'
    statuts = {'actif': 'Actif', 'suspendu': 'Suspendu', 'quitte': 'Quitte'}

    for i, emp in enumerate(qs, 1):
        row = i + 2
        fill = PatternFill('solid', fgColor=gris) if i % 2 == 0 else None
        vals = [
            i,
            emp.matricule,
            emp.nom + ' ' + emp.prenoms,
            emp.get_sexe_display() if emp.sexe else '-',
            fmt_date(emp.date_naissance),
            emp.lieu_naissance or '-',
            emp.nationalite.nom if emp.nationalite else '-',
            emp.fonction.nom if emp.fonction else '-',
            emp.service.nom if emp.service else '-',
            emp.type_contrat.nom if emp.type_contrat else '-',
            fmt_date(emp.date_embauche),
            fmt_date(emp.date_fin_contrat),
            statuts.get(emp.statut, emp.statut),
            emp.notes or '',
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border
            if fill:
                cell.fill = fill
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = 'A3'

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = "attachment; filename='registre_personnel.xlsx'"
    wb.save(resp)
    return resp


# ── Import Excel ──────────────────────────────────────────────────────────────

@login_required(login_url='login')
def employe_import(request):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        fichier = request.FILES.get('fichier')
        if not fichier:
            messages.error(request, 'Veuillez sélectionner un fichier Excel.')
            return redirect('rh_import')
        try:
            import openpyxl
            wb = openpyxl.load_workbook(fichier, data_only=True)
            ws = wb.active
            created = 0
            errors = []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row or not row[0]:
                    continue
                try:
                    nom     = str(row[0]).strip() if row[0] else ''
                    prenoms = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    if not nom or not prenoms:
                        errors.append(f'Ligne {i} : nom ou prénoms manquants.')
                        continue
                    date_emb_raw = row[2] if len(row) > 2 else None
                    if hasattr(date_emb_raw, 'date'):
                        date_emb = date_emb_raw.date()
                    elif date_emb_raw:
                        date_emb = date.fromisoformat(str(date_emb_raw)[:10])
                    else:
                        errors.append(f'Ligne {i} : date d\'embauche manquante.')
                        continue

                    service_nom  = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                    fonction_nom = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                    grade_nom    = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                    contrat_nom  = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                    statut_val   = str(row[7]).strip().lower() if len(row) > 7 and row[7] else 'actif'
                    matricule_val = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                    sexe_raw = str(row[9]).strip().lower() if len(row) > 9 and row[9] else ''
                    sexe_val = {
                        'm': 'M', 'masculin': 'M', 'homme': 'M',
                        'f': 'F', 'féminin': 'F', 'feminin': 'F', 'femme': 'F',
                    }.get(sexe_raw, '')

                    if matricule_val and Employe.objects.filter(matricule=matricule_val).exists():
                        errors.append(f'Ligne {i} : le matricule « {matricule_val} » est déjà utilisé.')
                        continue

                    nom_norm = _normalize_nom(nom)
                    prenoms_norm = _normalize_nom(prenoms)
                    doublon = any(
                        _normalize_nom(e.nom) == nom_norm and _normalize_nom(e.prenoms) == prenoms_norm
                        for e in Employe.objects.all()
                    )
                    if doublon:
                        errors.append(f'Ligne {i} : un employé nommé « {nom} {prenoms} » existe déjà.')
                        continue

                    service = None
                    if service_nom:
                        service = _find_or_create_ref(
                            Service, service_nom,
                            extra_create={'code': _generate_service_code(service_nom)},
                        )

                    fonction = None
                    if fonction_nom:
                        fonction = _find_or_create_ref(Fonction, fonction_nom)

                    grade = None
                    if grade_nom:
                        grade = _find_or_create_ref(Grade, grade_nom)

                    contrat = None
                    if contrat_nom:
                        contrat = _find_or_create_ref(TypeContrat, contrat_nom)

                    nationalite = _find_or_create_ref(Nationalite, 'Ivoirienne')

                    emp = Employe(
                        nom=nom, prenoms=prenoms, date_embauche=date_emb, sexe=sexe_val,
                        service=service, fonction=fonction, grade=grade, type_contrat=contrat,
                        nationalite=nationalite,
                        statut=statut_val if statut_val in ('actif', 'suspendu', 'quitte') else 'actif',
                    )
                    if matricule_val:
                        emp.matricule = matricule_val
                    emp.save()
                    HistoriqueEmploye.objects.create(
                        employe=emp, type_changement='creation',
                        nouvelle_valeur='Import Excel', fait_par=request.user,
                    )
                    created += 1
                except Exception as e:
                    errors.append(f'Ligne {i} : {e}')

            if created:
                messages.success(request, f'{created} employé(s) importé(s) avec succès.')
            for err in errors[:10]:
                messages.warning(request, err)
        except Exception as e:
            messages.error(request, f'Erreur lecture fichier : {e}')
        return redirect('ressources_humaines_list')

    return render(request, 'employer/import.html', {
        'can_manage': can_manage_rh(request.user),
    })


@login_required(login_url='login')
def employe_import_modele(request):
    """Modèle Excel vierge pour l'import — mêmes colonnes que celles lues
    par employe_import (dans le même ordre)."""
    if not can_manage_rh(request.user):
        raise PermissionDenied
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employés"

    header_fill = PatternFill('solid', fgColor='4A7236')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),  bottom=Side(style='thin', color='D0D0D0'),
    )

    headers = ['Nom', 'Prénoms', "Date d'embauche", 'Service', 'Fonction', 'Grade', 'Type de contrat', 'Statut', 'Matricule', 'Sexe']
    exemple = ['KONÉ', 'Aminata', '2024-01-15', 'Maternité', 'Sage-femme', 'Agent 2ème échelon', 'CDI', 'actif', '', 'F']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
    ws.row_dimensions[1].height = 20

    for col, val in enumerate(exemple, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.border = thin
        cell.font = Font(italic=True, color='888888')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="modele_import_employes.xlsx"'
    return resp


# â"€â"€ Organigramme â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def rh_organigramme(request):
    return redirect('rh_annuaire')


# â"€â"€ Documents â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def employe_doc_upload(request, pk):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    employe = get_object_or_404(Employe, pk=pk)
    if request.method == 'POST':
        titre   = request.POST.get('titre', '').strip()
        fichier = request.FILES.get('fichier')
        if not titre or not fichier:
            messages.error(request, 'Le titre et le fichier sont obligatoires.')
        else:
            date_exp = request.POST.get('date_expiration', '').strip() or None
            DocumentEmploye.objects.create(
                employe         = employe,
                type_document   = request.POST.get('type_document', 'autre'),
                titre           = titre,
                fichier         = fichier,
                date_expiration = date_exp,
                notes           = request.POST.get('doc_notes', '').strip(),
                ajoute_par      = request.user,
            )
            messages.success(request, f'Document « {titre} » ajouté.')
    return redirect('rh_detail', pk=pk)


@login_required(login_url='login')
def employe_doc_delete(request, pk, doc_pk):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    doc = get_object_or_404(DocumentEmploye, pk=doc_pk, employe_id=pk)
    if request.method == 'POST':
        titre = doc.titre
        if doc.fichier:
            doc.fichier.delete(save=False)
        doc.delete()
        messages.success(request, f'Document « {titre} » supprimé.')
    return redirect('rh_detail', pk=pk)


# â"€â"€ Informations supplémentaires â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def employe_info_save(request, pk):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    employe = get_object_or_404(Employe, pk=pk)
    if request.method == 'POST':
        cle    = request.POST.get('cle', '').strip()
        valeur = request.POST.get('valeur', '').strip()
        if cle and valeur:
            InfoSupplementaire.objects.create(employe=employe, cle=cle, valeur=valeur)
            messages.success(request, f'Information « {cle} » ajoutée.')
        else:
            messages.error(request, 'Le champ et la valeur sont obligatoires.')
    return redirect('rh_detail', pk=pk)


@login_required(login_url='login')
def employe_info_delete(request, pk, info_pk):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    info = get_object_or_404(InfoSupplementaire, pk=info_pk, employe_id=pk)
    if request.method == 'POST':
        cle = info.cle
        info.delete()
        messages.success(request, f'Information « {cle} » supprimée.')
    return redirect('rh_detail', pk=pk)



# â"€â"€ Annuaire â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def rh_annuaire(request):
    qs = Employe.objects.filter(statut='actif').select_related('service', 'fonction')
    q         = request.GET.get('q', '').strip()
    f_service = request.GET.get('service', '').strip()
    if q:
        qs = qs.filter(
            Q(nom__icontains=q) | Q(prenoms__icontains=q) |
            Q(fonction__nom__icontains=q)
        )
    if f_service:
        qs = qs.filter(service_id=f_service)
    qs = qs.order_by('service__nom', 'nom')
    paginator = Paginator(qs, 24)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'employer/annuaire.html', {
        'employes':   page_obj,
        'page_obj':   page_obj,
        'services':   Service.objects.filter(actif=True).order_by('nom'),
        'q':          q,
        'f_service':  f_service,
        'can_manage': can_manage_rh(request.user),
    })


# â"€â"€ Alertes contrat â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def alerte_marquer_lue(request, alerte_id):
    if request.method == 'POST':
        AlerteContrat.objects.filter(pk=alerte_id).update(lue=True)
    return JsonResponse({'ok': True})


@login_required(login_url='login')
def alertes_tout_lire(request):
    if request.method == 'POST':
        AlerteContrat.objects.filter(lue=False).update(lue=True)
        AlerteDocument.objects.filter(lue=False).update(lue=True)
    return JsonResponse({'ok': True})


# â"€â"€ Alertes document â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@login_required(login_url='login')
def alerte_doc_lue(request, alerte_id):
    if request.method == 'POST':
        AlerteDocument.objects.filter(pk=alerte_id).update(lue=True)


# ── Présence — historique par employé ─────────────────────────────────────────

_MOIS_FR_LONG = ['Janvier','Février','Mars','Avril','Mai','Juin',
                 'Juillet','Août','Septembre','Octobre','Novembre','Décembre']


@login_required(login_url='login')
def presence_historique(request, employe_pk):
    import calendar
    from .models import JourFerie

    employe = get_object_or_404(Employe, pk=employe_pk)
    today = date.today()

    annee = int(request.GET.get('annee', today.year))
    mois  = int(request.GET.get('mois',  today.month))
    mois  = max(1, min(12, mois))

    prev_mois  = 12 if mois == 1  else mois - 1
    prev_annee = annee - 1 if mois == 1  else annee
    next_mois  = 1  if mois == 12 else mois + 1
    next_annee = annee + 1 if mois == 12 else annee

    _, nb_jours = calendar.monthrange(annee, mois)
    debut_mois  = date(annee, mois, 1)
    fin_mois    = date(annee, mois, nb_jours)

    presences_dict = {
        p.date: p
        for p in Presence.objects.filter(employe=employe, date__year=annee, date__month=mois)
    }

    conges_approuves = list(Conge.objects.filter(
        employe=employe,
        statut__in=['approuve', 'en_cours', 'termine'],
        date_debut__lte=fin_mois,
        date_fin__gte=debut_mois,
    ))
    conge_jours = {}
    for cg in conges_approuves:
        d = cg.date_debut
        while d <= cg.date_fin:
            if d.year == annee and d.month == mois:
                conge_jours[d] = cg
            d += timedelta(days=1)

    jours_feries = set(
        JourFerie.objects.filter(date__year=annee, date__month=mois).values_list('date', flat=True)
    )

    jours = []
    for day_num in range(1, nb_jours + 1):
        d = date(annee, mois, day_num)
        weekend  = d.weekday() >= 5
        ferie    = d in jours_feries
        presence = presences_dict.get(d)
        conge    = conge_jours.get(d)

        if ferie:
            statut = 'ferie'
        elif weekend:
            statut = 'weekend'
        elif presence:
            statut = 'present' if presence.present else ('conge' if conge else 'absent')
        elif conge:
            statut = 'conge'
        else:
            statut = 'non_saisi'

        jours.append({
            'date': d, 'day': day_num,
            'weekday': ['Lun','Mar','Mer','Jeu','Ven','Sam','Dim'][d.weekday()],
            'weekend': weekend, 'ferie': ferie,
            'presence': presence, 'conge': conge, 'statut': statut,
        })

    nb_ouvrables   = sum(1 for j in jours if not j['weekend'] and not j['ferie'])
    nb_presents    = sum(1 for j in jours if j['statut'] == 'present')
    nb_absents     = sum(1 for j in jours if j['statut'] == 'absent')
    nb_conge       = sum(1 for j in jours if j['statut'] == 'conge')
    nb_non_saisi   = sum(1 for j in jours if j['statut'] == 'non_saisi')
    nb_retards     = 0
    total_retard_min = 0
    total_min      = 0
    for j in jours:
        p = j['presence']
        if p and p.present:
            retard = (p.retard_matin_min or 0) + (p.retard_soir_min or 0)
            if retard > 0:
                nb_retards += 1
                total_retard_min += retard
            if p.duree_totale:
                total_min += p.duree_totale

    return render(request, 'employer/presence/historique.html', {
        'employe':        employe,
        'annee':          annee, 'mois': mois,
        'mois_nom':       _MOIS_FR_LONG[mois - 1],
        'jours':          jours,
        'prev_mois':      prev_mois, 'prev_annee': prev_annee,
        'next_mois':      next_mois, 'next_annee': next_annee,
        'nb_ouvrables':   nb_ouvrables,
        'nb_presents':    nb_presents,
        'nb_absents':     nb_absents,
        'nb_conge':       nb_conge,
        'nb_non_saisi':   nb_non_saisi,
        'total_heures':   total_min // 60,
        'total_min_reste': total_min % 60,
        'nb_retards':     nb_retards,
        'total_retard_min': total_retard_min,
        'can_manage':     can_manage_rh(request.user),
    })


# ── Présence — rapport mensuel ────────────────────────────────────────────────

@login_required(login_url='login')
def presence_rapport_mensuel(request):
    import calendar
    from .models import JourFerie

    today = date.today()
    annee = int(request.GET.get('annee', today.year))
    mois  = int(request.GET.get('mois',  today.month))
    mois  = max(1, min(12, mois))

    prev_mois  = 12 if mois == 1  else mois - 1
    prev_annee = annee - 1 if mois == 1  else annee
    next_mois  = 1  if mois == 12 else mois + 1
    next_annee = annee + 1 if mois == 12 else annee

    _, nb_jours = calendar.monthrange(annee, mois)
    debut_mois  = date(annee, mois, 1)
    fin_mois    = date(annee, mois, nb_jours)

    jours_feries = set(
        JourFerie.objects.filter(date__year=annee, date__month=mois).values_list('date', flat=True)
    )
    nb_ouvrables = sum(
        1 for d in range(1, nb_jours + 1)
        if date(annee, mois, d).weekday() < 5 and date(annee, mois, d) not in jours_feries
    )

    service_id  = request.GET.get('service') or None
    employe_id  = request.GET.get('employe') or None
    employes_qs = Employe.objects.filter(statut='actif').select_related('service', 'fonction')
    if service_id:
        employes_qs = employes_qs.filter(service_id=service_id)
    employe_filtre = None
    if employe_id:
        employes_qs = employes_qs.filter(pk=employe_id)
        employe_filtre = employes_qs.first()

    presences_mois = list(
        Presence.objects.filter(date__year=annee, date__month=mois).select_related('employe')
    )
    conges_mois = list(Conge.objects.filter(
        statut__in=['approuve', 'en_cours', 'termine'],
        date_debut__lte=fin_mois,
        date_fin__gte=debut_mois,
    ))

    conge_par_employe    = {}
    for cg in conges_mois:
        conge_par_employe.setdefault(cg.employe_id, []).append(cg)

    presence_par_employe = {}
    for p in presences_mois:
        presence_par_employe.setdefault(p.employe_id, []).append(p)

    employes_stats = []
    for emp in employes_qs:
        presences = presence_par_employe.get(emp.pk, [])
        conges    = conge_par_employe.get(emp.pk, [])

        conge_jours_set = set()
        for cg in conges:
            d = cg.date_debut
            while d <= cg.date_fin:
                if d.year == annee and d.month == mois:
                    conge_jours_set.add(d)
                d += timedelta(days=1)

        nb_presents        = sum(1 for p in presences if p.present)
        nb_absents_non_just = 0
        nb_absents_conge    = 0
        presences_dates     = {p.date for p in presences}

        for p in presences:
            if not p.present:
                if p.date in conge_jours_set:
                    nb_absents_conge += 1
                else:
                    nb_absents_non_just += 1

        for d in conge_jours_set:
            if d not in presences_dates and d.weekday() < 5 and d not in jours_feries:
                nb_absents_conge += 1

        nb_retards = 0
        total_retard_min = 0
        total_min = 0
        for p in presences:
            if p.present:
                retard = (p.retard_matin_min or 0) + (p.retard_soir_min or 0)
                if retard > 0:
                    nb_retards += 1
                    total_retard_min += retard
                if p.duree_totale:
                    total_min += p.duree_totale

        employes_stats.append({
            'employe':            emp,
            'nb_presents':        nb_presents,
            'nb_absents_non_just': nb_absents_non_just,
            'nb_absents_conge':   nb_absents_conge,
            'nb_retards':         nb_retards,
            'total_retard_min':   total_retard_min,
            'total_heures':       total_min // 60,
            'total_min_reste':    total_min % 60,
            'taux': round(nb_presents / nb_ouvrables * 100) if nb_ouvrables > 0 else None,
        })

    employes_stats.sort(key=lambda x: (
        x['employe'].service.nom if x['employe'].service else '',
        x['employe'].nom
    ))

    services = Service.objects.filter(employes__statut='actif').distinct().order_by('nom')

    try:
        page_size = int(request.GET.get('page_size', 15))
    except (TypeError, ValueError):
        page_size = 15
    page_size = max(5, min(50, page_size))
    paginator = Paginator(employes_stats, page_size)
    page_obj  = paginator.get_page(request.GET.get('page'))

    return render(request, 'employer/presence/rapport_mensuel.html', {
        'employes_stats':  employes_stats,
        'page_obj':        page_obj,
        'page_size':       page_size,
        'annee':           annee, 'mois': mois,
        'mois_nom':        _MOIS_FR_LONG[mois - 1],
        'nb_ouvrables':    nb_ouvrables,
        'prev_mois':       prev_mois, 'prev_annee': prev_annee,
        'next_mois':       next_mois, 'next_annee': next_annee,
        'services':        services,
        'service_id':      service_id,
        'employe_filtre':  employe_filtre,
        'total_presents':  sum(e['nb_presents'] for e in employes_stats),
        'total_absents':   sum(e['nb_absents_non_just'] for e in employes_stats),
        'total_conge':     sum(e['nb_absents_conge'] for e in employes_stats),
        'total_retards':   sum(e['nb_retards'] for e in employes_stats),
        'can_manage':      can_manage_rh(request.user),
        'annees_dispo':    list(range(today.year - 2, today.year + 1)),
    })


# ─────────────────────────────────────────────────────────────────
#  Import biométrique (pointeuse ZKTeco)
# ─────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def presence_import_bio(request):
    if not can_manage_rh(request.user):
        raise PermissionDenied

    # ── Étape 3 : confirmation → création en base ──────────────────
    if request.method == 'POST' and request.POST.get('action') == 'confirmer':
        import json
        from datetime import time as _time

        try:
            records = json.loads(request.POST.get('data_json', '[]'))
        except (json.JSONDecodeError, ValueError):
            messages.error(request, "Données invalides, recommencez l'import.")
            return redirect('rh_presence_import')

        employe_map = {
            str(e.biometric_id): e
            for e in Employe.objects.filter(statut='actif')
                .exclude(biometric_id__isnull=True).exclude(biometric_id='')
        }

        def _pt(s):
            if not s:
                return None
            h, m = s.split(':')
            return _time(int(h), int(m))

        created = updated = erreurs = 0
        for r in records:
            emp = employe_map.get(str(r.get('bio_id', '')))
            if not emp:
                erreurs += 1
                continue
            try:
                _, was_created = Presence.objects.update_or_create(
                    employe=emp,
                    date=r['date'],
                    defaults={
                        'present': True,
                        'heure_arrivee_matin': _pt(r.get('am_in')),
                        'heure_depart_matin':  _pt(r.get('am_out')),
                        'heure_arrivee_soir':  _pt(r.get('pm_in')),
                        'heure_depart_soir':   _pt(r.get('pm_out')),
                        'modifie_par': request.user,
                        'modifie_le':  timezone.now(),
                    }
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except Exception:
                erreurs += 1

        messages.success(
            request,
            f"Import terminé : {created} présence(s) créée(s), {updated} mise(s) à jour"
            + (f", {erreurs} ignorée(s)." if erreurs else ".")
        )
        return redirect('rh_presence_rapport')

    # ── Étape 2 : parsing du fichier → prévisualisation ────────────
    if request.method == 'POST' and request.FILES.get('fichier'):
        import re, json
        from datetime import datetime as _dt, time as _time
        from collections import defaultdict

        try:
            contenu = request.FILES['fichier'].read().decode('utf-8', errors='replace')
        except Exception as e:
            messages.error(request, f"Erreur de lecture du fichier : {e}")
            return redirect('rh_presence_import')

        punches = []
        for line in contenu.splitlines():
            line = line.strip()
            if not line:
                continue
            if re.match(r'(?i)(no\b|userid|user.id|^#|^date)', line):
                continue

            # Normalise séparateurs multiples → tab
            normed = re.sub(r'\t|[ ]{2,}', '\t', line)
            parts = [p.strip() for p in normed.split('\t') if p.strip()]
            if len(parts) < 2:
                continue

            bio_id = parts[0]
            if not re.match(r'^\d+$', bio_id):
                continue

            datetime_str = None

            # Cas A : parts[1] = "YYYY-MM-DD HH:MM:SS" ou "YYYY-MM-DD HH:MM"
            m = re.match(r'(\d{4}[-/]\d{2}[-/]\d{2})[T ](\d{2}:\d{2}(?::\d{2})?)', parts[1])
            if m:
                datetime_str = m.group(1).replace('/', '-') + ' ' + m.group(2)

            # Cas B : parts[1] = date, parts[2] = heure
            if not datetime_str and len(parts) >= 3:
                dm = re.match(r'(\d{4}[-/]\d{2}[-/]\d{2})$', parts[1])
                tm = re.match(r'(\d{2}:\d{2}(?::\d{2})?)', parts[2])
                if dm and tm:
                    datetime_str = dm.group(1).replace('/', '-') + ' ' + tm.group(1)

            if not datetime_str:
                continue

            try:
                fmt = '%Y-%m-%d %H:%M:%S' if datetime_str.count(':') == 2 else '%Y-%m-%d %H:%M'
                dt = _dt.strptime(datetime_str, fmt)
                punches.append((bio_id, dt.date(), dt.time()))
            except ValueError:
                continue

        if not punches:
            messages.error(request, "Aucune ligne valide trouvée. Vérifiez le format du fichier.")
            return redirect('rh_presence_import')

        # Grouper par (bio_id, date) et trier les heures
        groups = defaultdict(list)
        for bio_id, d, t in punches:
            groups[(bio_id, d)].append(t)
        for k in groups:
            groups[k].sort()

        employe_map = {
            str(e.biometric_id): e
            for e in Employe.objects.filter(statut='actif')
                .exclude(biometric_id__isnull=True).exclude(biometric_id='')
        }

        preview = []
        inconnus = set()

        for (bio_id, d), times in sorted(groups.items(), key=lambda x: (x[0][0], x[0][1])):
            emp = employe_map.get(bio_id)
            if not emp:
                inconnus.add(bio_id)
                continue

            t_cut = _time(13, 0)
            matin = [t for t in times if t < t_cut]
            soir  = [t for t in times if t >= t_cut]

            am_in  = matin[0]         if matin               else None
            am_out = matin[-1]        if len(matin) >= 2     else None
            pm_in  = soir[0]          if soir                else None
            pm_out = soir[-1]         if len(soir) >= 2      else None

            # Permanence style (2 pointages seulement, tous en matin)
            if len(times) == 2 and not soir:
                am_in = times[0]; am_out = None
                pm_in = None;     pm_out = times[1]

            def fmt_t(t):
                return t.strftime('%H:%M') if t else ''

            preview.append({
                'bio_id':       bio_id,
                'employe':      emp.nom_complet,
                'date':         str(d),
                'nb_pointages': len(times),
                'am_in':        fmt_t(am_in),
                'am_out':       fmt_t(am_out),
                'pm_in':        fmt_t(pm_in),
                'pm_out':       fmt_t(pm_out),
            })

        data_json = json.dumps([
            {k: v for k, v in r.items() if k not in ('employe', 'nb_pointages')}
            for r in preview
        ])

        return render(request, 'employer/presence/import_bio.html', {
            'step':        'preview',
            'preview':     preview,
            'data_json':   data_json,
            'nb_total':    len(preview),
            'nb_inconnus': len(inconnus),
            'inconnus':    sorted(inconnus),
            'can_manage':  True,
        })

    # ── Étape 1 : formulaire d'upload ──────────────────────────────
    employes_bio = Employe.objects.filter(
        statut='actif', biometric_id__isnull=False
    ).exclude(biometric_id='').order_by('nom')

    return render(request, 'employer/presence/import_bio.html', {
        'step':       'upload',
        'employes_bio': employes_bio,
        'nb_mapped':  employes_bio.count(),
        'can_manage': can_manage_rh(request.user),
    })


# ─────────────────────────────────────────────────────────────────
#  Kiosk de pointage (self-service employé)
# ─────────────────────────────────────────────────────────────────

def presence_kiosk(request):
    """Page kiosk accessible sans connexion RH."""
    return render(request, 'employer/presence/kiosk.html', {
        'can_enroll_bio': request.user.is_authenticated and can_manage_rh(request.user),
    })


def presence_pointer(request):
    """
    AJAX POST : enregistre un pointage à l'heure courante.
    Body (JSON ou form) : { "matricule": "EMP001" }
    Retourne JSON : { ok, employe, slot, heure, photo_url, message }

    Délègue la détection de l'action (_detecter_action), le contrôle jour
    ouvré (_est_jour_ouvre) et le contrôle congé (_conge_du_jour) à
    presence.views — ce kiosk et celui du module Présence écrivent sur le
    même modèle Presence ; avoir deux logiques différentes désynchronisait
    les règles (celle-ci ignorait congés/jours fériés/permanence).
    """
    import json as _json
    from django.db import transaction
    from presence.views import (
        _detecter_action, _est_jour_ouvre, _conge_du_jour, _LABELS_ACTION, _date_reference,
        _est_en_permanence,
    )

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'message': 'Méthode non autorisée.'}, status=405)

    # Lire le matricule depuis JSON ou form
    try:
        body = _json.loads(request.body)
        matricule = str(body.get('matricule', '')).strip().upper()
    except (ValueError, AttributeError):
        matricule = request.POST.get('matricule', '').strip().upper()

    if not matricule:
        return JsonResponse({'ok': False, 'message': 'Matricule requis.'})

    emp = Employe.objects.filter(matricule=matricule, statut='actif').first()
    if not emp:
        return JsonResponse({'ok': False, 'message': 'Matricule non reconnu ou employé inactif.'})

    heure = timezone.localtime().time().replace(second=0, microsecond=0)
    today = _date_reference(emp)

    ouvre, raison = _est_jour_ouvre(today)
    if not ouvre:
        return JsonResponse({'ok': False, 'message': raison})

    conge = _conge_du_jour(emp, today)
    if conge:
        return JsonResponse({
            'ok': False,
            'message': f"En congé — {conge.get_type_conge_display()} (jusqu'au {conge.date_fin:%d/%m/%Y})",
        })

    with transaction.atomic():
        # select_for_update : sérialise deux pointages quasi simultanés sur le
        # même employé/jour (même verrou que presence.views.presence_pointer).
        presence, _created = Presence.objects.select_for_update().get_or_create(
            employe=emp, date=today,
            defaults={'present': True}
        )

        slot = _detecter_action(emp, today)
        if slot == 'complet':
            return JsonResponse({'ok': False, 'message': 'Journée déjà complète.'})
        slot_label = _LABELS_ACTION.get(slot, slot)
        # Après midi, l'arrivée en permanence n'est plus proposée (elle est
        # considérée manquée) : slot='depart' est alors renvoyé directement
        # sans qu'une arrivée n'ait jamais été pointée.
        arrivee_manquee = (
            slot == 'depart'
            and _est_en_permanence(emp, today, presence_obj=presence)
            and not presence.heure_arrivee_matin
        )

        champ_heure, champ_locked = {
            'arrivee_matin': ('heure_arrivee_matin', 'am_in_locked'),
            'depart_matin':  ('heure_depart_matin',  'am_out_locked'),
            'arrivee_soir':  ('heure_arrivee_soir',  'pm_in_locked'),
            'depart_soir':   ('heure_depart_soir',   'pm_out_locked'),
            'arrivee':       ('heure_arrivee_matin', 'am_in_locked'),
            'depart':        ('heure_depart_soir',   'pm_out_locked'),
        }[slot]
        setattr(presence, champ_heure, heure)
        setattr(presence, champ_locked, True)
        presence.present    = True
        # Sans ce marquage, le registre (qui ne se fie au planning que tant
        # qu'aucune Presence n'existe) retombait sur permanence=False dès le
        # premier pointage kiosk du jour.
        if slot in ('arrivee', 'depart'):
            presence.permanence = True
        presence.modifie_le = timezone.now()
        presence.save()

    photo_url = emp.photo.url if emp.photo else None

    message = f'Pointage enregistré — {slot_label} : {heure:%H:%M}'
    if arrivee_manquee:
        message += " — Arrivée non pointée avant midi, à régulariser par la RH."

    return JsonResponse({
        'ok':              True,
        'employe':         emp.nom_complet,
        'service':         emp.service.nom if emp.service else '',
        'slot':            slot,
        'slot_label':      slot_label,
        'heure':           heure.strftime('%H:%M'),
        'photo_url':       photo_url,
        'arrivee_manquee': arrivee_manquee,
        'message':         message,
    })


# ─────────────────────────────────────────────────────────────────
#  WebAuthn — empreinte digitale (pointage biométrique)
# ─────────────────────────────────────────────────────────────────

def _wba_rp(request):
    from django.conf import settings
    host    = request.get_host()
    domain  = host.split(':')[0]
    scheme  = 'https' if request.is_secure() else 'http'
    rp_id   = getattr(settings, 'WEBAUTHN_RP_ID',   domain)
    rp_name = getattr(settings, 'WEBAUTHN_RP_NAME',  'Centre Médico-Social WALÉ')
    origin  = getattr(settings, 'WEBAUTHN_ORIGIN',   f'{scheme}://{host}')
    return rp_id, rp_name, origin


@require_POST
def bio_enroll_options(request):
    import json as _j
    # L'enrôlement biométrique lie durablement une empreinte à une identité
    # de pointage — contrairement au pointage quotidien, ce n'est PAS une
    # opération en libre-service depuis le kiosk public : elle doit être
    # supervisée par un membre RH connecté (sinon n'importe qui peut lier
    # sa propre empreinte au matricule d'un collègue).
    if not (request.user.is_authenticated and can_manage_rh(request.user)):
        return JsonResponse({'ok': False, 'message': 'Réservé au personnel RH connecté.'}, status=403)
    from webauthn import generate_registration_options, options_to_json
    from webauthn.helpers import base64url_to_bytes, bytes_to_base64url
    from webauthn.helpers.structs import (
        AuthenticatorSelectionCriteria, AuthenticatorAttachment,
        ResidentKeyRequirement, UserVerificationRequirement,
        PublicKeyCredentialDescriptor, PublicKeyCredentialType,
    )
    from webauthn.helpers.cose import COSEAlgorithmIdentifier

    try:
        body = _j.loads(request.body)
    except (ValueError, AttributeError):
        body = request.POST

    matricule = str(body.get('matricule', '')).strip().upper()
    if not matricule:
        return JsonResponse({'ok': False, 'message': 'Matricule requis.'}, status=400)

    emp = Employe.objects.filter(matricule=matricule, statut='actif').first()
    if not emp:
        return JsonResponse({'ok': False, 'message': 'Employé introuvable.'}, status=404)

    rp_id, rp_name, _ = _wba_rp(request)

    exclude = [
        PublicKeyCredentialDescriptor(
            type=PublicKeyCredentialType.PUBLIC_KEY,
            id=base64url_to_bytes(c.credential_id),
        )
        for c in emp.credentials_bio.all()
    ]

    opts = generate_registration_options(
        rp_id=rp_id,
        rp_name=rp_name,
        user_id=str(emp.pk).encode(),
        user_name=emp.matricule,
        user_display_name=emp.nom_complet,
        authenticator_selection=AuthenticatorSelectionCriteria(
            authenticator_attachment=AuthenticatorAttachment.PLATFORM,
            resident_key=ResidentKeyRequirement.REQUIRED,
            user_verification=UserVerificationRequirement.REQUIRED,
        ),
        supported_pub_key_algs=[COSEAlgorithmIdentifier.ECDSA_SHA_256],
        exclude_credentials=exclude,
    )

    request.session['wba_reg_challenge'] = bytes_to_base64url(opts.challenge)
    request.session['wba_reg_emp_pk']    = emp.pk

    return HttpResponse(options_to_json(opts), content_type='application/json')


@require_POST
def bio_enroll_verify(request):
    import json as _j
    if not (request.user.is_authenticated and can_manage_rh(request.user)):
        return JsonResponse({'ok': False, 'message': 'Réservé au personnel RH connecté.'}, status=403)
    from webauthn import verify_registration_response
    from webauthn.helpers import base64url_to_bytes, bytes_to_base64url
    from webauthn.helpers.structs import (
        RegistrationCredential, AuthenticatorAttestationResponse,
    )

    challenge_b64 = request.session.get('wba_reg_challenge')
    emp_pk        = request.session.get('wba_reg_emp_pk')
    if not challenge_b64 or not emp_pk:
        return JsonResponse({'ok': False, 'message': 'Session expirée, recommencez.'}, status=400)

    emp = Employe.objects.filter(pk=emp_pk, statut='actif').first()
    if not emp:
        return JsonResponse({'ok': False, 'message': 'Employé introuvable.'}, status=404)

    try:
        body = _j.loads(request.body)
    except (ValueError, AttributeError):
        return JsonResponse({'ok': False, 'message': 'Données invalides.'}, status=400)

    rp_id, _, origin = _wba_rp(request)

    try:
        resp = body.get('response', {})
        credential = RegistrationCredential(
            id=body['id'],
            raw_id=base64url_to_bytes(body['rawId']),
            response=AuthenticatorAttestationResponse(
                client_data_json=base64url_to_bytes(resp['clientDataJSON']),
                attestation_object=base64url_to_bytes(resp['attestationObject']),
            ),
            type=body['type'],
        )
        verified = verify_registration_response(
            credential=credential,
            expected_challenge=base64url_to_bytes(challenge_b64),
            expected_rp_id=rp_id,
            expected_origin=origin,
            require_user_verification=True,
        )
    except Exception as e:
        return JsonResponse({'ok': False, 'message': f'Vérification échouée : {e}'}, status=400)

    CredentialBiometrique.objects.update_or_create(
        credential_id=bytes_to_base64url(verified.credential_id),
        defaults={
            'employe':     emp,
            'public_key':  bytes_to_base64url(verified.credential_public_key),
            'sign_count':  verified.sign_count,
            'aaguid':      str(verified.aaguid) if verified.aaguid else '',
            'device_name': body.get('device_name', ''),
        }
    )

    request.session.pop('wba_reg_challenge', None)
    request.session.pop('wba_reg_emp_pk', None)

    return JsonResponse({'ok': True, 'message': f'Empreinte enregistrée pour {emp.nom_complet}.'})


def bio_auth_options(request):
    from webauthn import generate_authentication_options, options_to_json
    from webauthn.helpers import bytes_to_base64url
    from webauthn.helpers.structs import UserVerificationRequirement

    rp_id, _, _ = _wba_rp(request)

    opts = generate_authentication_options(
        rp_id=rp_id,
        user_verification=UserVerificationRequirement.REQUIRED,
    )

    request.session['wba_auth_challenge'] = bytes_to_base64url(opts.challenge)

    return HttpResponse(options_to_json(opts), content_type='application/json')


@require_POST
def bio_auth_verify(request):
    import json as _j
    from datetime import time as _time
    from webauthn import verify_authentication_response
    from webauthn.helpers import base64url_to_bytes, bytes_to_base64url
    from webauthn.helpers.structs import (
        AuthenticationCredential, AuthenticatorAssertionResponse,
    )

    challenge_b64 = request.session.get('wba_auth_challenge')
    if not challenge_b64:
        return JsonResponse({'ok': False, 'message': 'Session expirée, réessayez.'}, status=400)

    try:
        body = _j.loads(request.body)
    except (ValueError, AttributeError):
        return JsonResponse({'ok': False, 'message': 'Données invalides.'}, status=400)

    rp_id, _, origin = _wba_rp(request)

    cred_id  = body.get('id', '')
    cred_obj = CredentialBiometrique.objects.select_related('employe').filter(
        credential_id=cred_id
    ).first()
    if not cred_obj:
        return JsonResponse({'ok': False, 'message': 'Empreinte non reconnue — badge non enregistré sur ce kiosk.'})

    try:
        resp       = body.get('response', {})
        user_h_raw = resp.get('userHandle')
        credential = AuthenticationCredential(
            id=body['id'],
            raw_id=base64url_to_bytes(body['rawId']),
            response=AuthenticatorAssertionResponse(
                client_data_json=base64url_to_bytes(resp['clientDataJSON']),
                authenticator_data=base64url_to_bytes(resp['authenticatorData']),
                signature=base64url_to_bytes(resp['signature']),
                user_handle=base64url_to_bytes(user_h_raw) if user_h_raw else None,
            ),
            type=body['type'],
        )
        verified = verify_authentication_response(
            credential=credential,
            expected_challenge=base64url_to_bytes(challenge_b64),
            expected_rp_id=rp_id,
            expected_origin=origin,
            credential_public_key=base64url_to_bytes(cred_obj.public_key),
            credential_current_sign_count=cred_obj.sign_count,
            require_user_verification=True,
        )
    except Exception as e:
        return JsonResponse({'ok': False, 'message': f'Authentification échouée : {e}'})

    cred_obj.sign_count = verified.new_sign_count
    cred_obj.last_used  = timezone.now()
    cred_obj.save(update_fields=['sign_count', 'last_used'])
    request.session.pop('wba_auth_challenge', None)

    emp   = cred_obj.employe
    now   = timezone.localtime()
    today = now.date()
    heure = now.time().replace(second=0, microsecond=0)
    t_cut = _time(13, 0)

    presence, _ = Presence.objects.get_or_create(
        employe=emp, date=today,
        defaults={'present': True}
    )

    if heure < t_cut:
        if not presence.heure_arrivee_matin:
            slot, slot_label = 'arrivee_matin', 'Arrivée matin'
            presence.heure_arrivee_matin = heure
            presence.am_in_locked = True
        elif not presence.heure_depart_matin:
            slot, slot_label = 'depart_matin', 'Départ matin'
            presence.heure_depart_matin = heure
            presence.am_out_locked = True
        else:
            return JsonResponse({
                'ok': False,
                'message': f'Session matin déjà complète ({presence.heure_arrivee_matin:%H:%M} → {presence.heure_depart_matin:%H:%M}).',
            })
    else:
        if not presence.heure_arrivee_soir:
            slot, slot_label = 'arrivee_soir', 'Arrivée soir'
            presence.heure_arrivee_soir = heure
            presence.pm_in_locked = True
        elif not presence.heure_depart_soir:
            slot, slot_label = 'depart_soir', 'Départ soir'
            presence.heure_depart_soir = heure
            presence.pm_out_locked = True
        else:
            return JsonResponse({
                'ok': False,
                'message': f'Session soir déjà complète ({presence.heure_arrivee_soir:%H:%M} → {presence.heure_depart_soir:%H:%M}).',
            })

    presence.present    = True
    presence.modifie_le = timezone.now()
    presence.save()

    photo_url = emp.photo.url if emp.photo else None
    return JsonResponse({
        'ok':         True,
        'employe':    emp.nom_complet,
        'service':    emp.service.nom if emp.service else '',
        'slot':       slot,
        'slot_label': slot_label,
        'heure':      heure.strftime('%H:%M'),
        'photo_url':  photo_url,
        'message':    f'Pointage enregistré — {slot_label} : {heure:%H:%M}',
    })


# ── Configuration : Grades / Fonctions / Types de contrat ──────────────────
# Trois tables de référence utilisées par le formulaire employé (menus
# déroulants Grade/Fonction/Type de contrat, voir _rh_selects ci-dessus).
# CRUD générique paramétré par "slug" plutôt que 3 vues quasi-identiques.

RH_CONFIG_MODELS = {
    'grade': {
        'model': Grade, 'fields': ['nom', 'code', 'description'],
        'label': 'Grade', 'label_plural': 'Grades', 'icone': 'bi-award',
        'sort_fields': ['nom', 'code'],
    },
    'fonction': {
        'model': Fonction, 'fields': ['nom', 'code', 'categorie', 'description'],
        'label': 'Fonction', 'label_plural': 'Fonctions', 'icone': 'bi-briefcase',
        'sort_fields': ['nom', 'code', 'categorie'],
    },
    'type_contrat': {
        'model': TypeContrat, 'fields': ['nom', 'code', 'droit_au_conge', 'description'],
        'label': 'Type de contrat', 'label_plural': 'Types de contrat', 'icone': 'bi-file-earmark-text',
        'sort_fields': ['nom', 'code', 'droit_au_conge'],
    },
    'nationalite': {
        'model': Nationalite, 'fields': ['nom'],
        'label': 'Nationalité', 'label_plural': 'Nationalités', 'icone': 'bi-flag',
        'sort_fields': ['nom'],
    },
}


def _rh_config_or_404(slug):
    from django.http import Http404
    cfg = RH_CONFIG_MODELS.get(slug)
    if not cfg:
        raise Http404("Configuration inconnue.")
    return cfg


def _rh_config_form_class(cfg):
    from django import forms
    from django.forms.widgets import CheckboxInput

    class ConfigForm(forms.ModelForm):
        class Meta:
            model = cfg['model']
            fields = cfg['fields']

        def clean_nom(self):
            nom = self.cleaned_data.get('nom', '')
            nom = _upper_no_accent(nom) if nom else nom
            if cfg['model'] is Fonction and nom:
                qs = Fonction.objects.filter(nom=nom)
                if self.instance.pk:
                    qs = qs.exclude(pk=self.instance.pk)
                if qs.exists():
                    raise forms.ValidationError('Cette fonction existe déjà.')
            return nom

        if cfg['model'] is Fonction:
            def clean_code(self):
                import re
                base = re.sub(r'[^A-Z0-9]', '', self.cleaned_data.get('nom', '')[:3]) or 'XXX'
                code = f'FON-{base}'
                qs = Fonction.objects.filter(code=code)
                if self.instance.pk:
                    qs = qs.exclude(pk=self.instance.pk)
                n = 2
                original = code
                while qs.exists():
                    code = f'{original}{n}'
                    n += 1
                    qs = Fonction.objects.filter(code=code)
                    if self.instance.pk:
                        qs = qs.exclude(pk=self.instance.pk)
                return code

        if cfg['model'] is TypeContrat:
            def clean_code(self):
                if self.instance.pk and self.instance.code:
                    return self.instance.code
                import re
                max_n = 0
                qs = TypeContrat.objects.all()
                if self.instance.pk:
                    qs = qs.exclude(pk=self.instance.pk)
                for c in qs.values_list('code', flat=True):
                    m = re.match(r'^CON-(\d{3})$', c or '')
                    if m:
                        max_n = max(max_n, int(m.group(1)))
                return f'CON-{max_n + 1:03d}'

    if cfg['model'] is Fonction:
        ConfigForm.base_fields['code'].required = False
        ConfigForm.base_fields['code'].widget.attrs['readonly'] = True
        ConfigForm.base_fields['code'].help_text = 'Généré automatiquement : FON- + 3 premières lettres du nom.'

    if cfg['model'] is TypeContrat:
        ConfigForm.base_fields['code'].required = False
        ConfigForm.base_fields['code'].widget.attrs['readonly'] = True
        ConfigForm.base_fields['code'].help_text = 'Généré automatiquement : CON- + numéro d\'ordre sur 3 chiffres.'

    for name in cfg['fields']:
        widget = ConfigForm.base_fields[name].widget
        if not isinstance(widget, CheckboxInput):
            widget.attrs.setdefault('class', 'rh-input')
    return ConfigForm


@login_required(login_url='login')
def rh_config_list(request, slug):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    cfg = _rh_config_or_404(slug)
    q = request.GET.get('q', '').strip()
    qs = cfg['model'].objects.all()
    if q:
        qs = qs.filter(nom__icontains=q)

    sort = request.GET.get('sort', 'nom')
    if sort not in cfg['sort_fields']:
        sort = 'nom'
    direction = request.GET.get('dir', 'asc')
    if direction not in ('asc', 'desc'):
        direction = 'asc'
    order = sort if direction == 'asc' else f'-{sort}'

    paginator = Paginator(qs.order_by(order), 12)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'employer/config/list.html', {
        'slug': slug, 'cfg': cfg, 'page_obj': page_obj, 'q': q,
        'sort': sort, 'dir': direction,
        'can_manage': True,
    })


def _rh_config_is_ajax(request):
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'


@login_required(login_url='login')
def rh_config_create(request, slug):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    cfg = _rh_config_or_404(slug)
    Form = _rh_config_form_class(cfg)
    is_ajax = _rh_config_is_ajax(request)
    if request.method == 'POST':
        form = Form(request.POST)
        if form.is_valid():
            obj = form.save()
            if is_ajax:
                return JsonResponse({'ok': True, 'message': f"{cfg['label']} « {obj} » créé(e)."})
            messages.success(request, f"{cfg['label']} « {obj} » créé(e).")
            return redirect('rh_config_list', slug=slug)
    else:
        form = Form()
    template = 'employer/config/form_modal.html' if is_ajax else 'employer/config/form.html'
    return render(request, template, {
        'slug': slug, 'cfg': cfg, 'form': form,
        'titre': f"Nouveau — {cfg['label']}", 'edit': False,
        'can_manage': True,
    })


@login_required(login_url='login')
def rh_config_edit(request, slug, pk):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    cfg = _rh_config_or_404(slug)
    obj = get_object_or_404(cfg['model'], pk=pk)
    Form = _rh_config_form_class(cfg)
    is_ajax = _rh_config_is_ajax(request)
    if request.method == 'POST':
        form = Form(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            if is_ajax:
                return JsonResponse({'ok': True, 'message': f"{cfg['label']} « {obj} » mis à jour."})
            messages.success(request, f"{cfg['label']} « {obj} » mis à jour.")
            return redirect('rh_config_list', slug=slug)
    else:
        form = Form(instance=obj)
    template = 'employer/config/form_modal.html' if is_ajax else 'employer/config/form.html'
    return render(request, template, {
        'slug': slug, 'cfg': cfg, 'form': form, 'obj': obj,
        'titre': f"Modifier — {obj}", 'edit': True,
        'can_manage': True,
    })


@login_required(login_url='login')
@require_POST
def rh_config_delete(request, slug, pk):
    if not can_manage_rh(request.user):
        raise PermissionDenied
    cfg = _rh_config_or_404(slug)
    obj = get_object_or_404(cfg['model'], pk=pk)
    nom = str(obj)
    obj.delete()
    messages.success(request, f"{cfg['label']} « {nom} » supprimé(e).")
    return redirect('rh_config_list', slug=slug)


# ══════════════════════════════════════════════
#  EMPLOYÉS — Configuration : Services
# ══════════════════════════════════════════════

def _service_form_class():
    from django import forms
    from medecins.models import Medecin
    class ServiceForm(forms.ModelForm):
        chef_service = forms.ModelChoiceField(
            queryset=Medecin.objects.filter(actif=True).select_related('employe').order_by('employe__nom', 'employe__prenoms'),
            required=False,
            empty_label='— Aucun chef de service —',
        )

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            # Le chef de service actuel doit rester proposé même s'il est
            # devenu inactif entre-temps — sinon modifier n'importe quel
            # autre champ du service (nom, description…) efface silencieusement
            # cette référence, faute d'être dans la liste déroulante.
            actuel_id = self.instance.chef_service_id if self.instance else None
            if actuel_id and not self.fields['chef_service'].queryset.filter(pk=actuel_id).exists():
                self.fields['chef_service'].queryset = (
                    Medecin.objects.filter(Q(actif=True) | Q(pk=actuel_id))
                    .select_related('employe').order_by('employe__nom', 'employe__prenoms')
                )

        class Meta:
            model = Service
            fields = ['nom', 'code', 'description', 'chef_service', 'actif']
            widgets = {
                'nom':         forms.TextInput(attrs={'class': 'f-input', 'placeholder': 'Ex : Médecine interne'}),
                'code':        forms.TextInput(attrs={'class': 'f-input', 'placeholder': 'Ex : MED-INT'}),
                'description': forms.Textarea(attrs={'class': 'f-input', 'rows': 3}),
            }
    return ServiceForm


@login_required(login_url='login')
def employer_services(request):
    q   = request.GET.get('q', '').strip()
    vue = request.GET.get('vue', 'liste')
    qs  = Service.objects.annotate(nb_employes=Count('employes')).order_by('nom')
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(code__icontains=q))
    total = Service.objects.count()
    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'employer/config/services_list.html', {
        'page_obj': page_obj, 'total': total,
        'total_filtre': qs.count(), 'q': q, 'vue': vue,
        'can_manage': can_manage_rh(request.user),
    })


@login_required(login_url='login')
def employer_service_detail(request, pk):
    obj = get_object_or_404(Service, pk=pk)
    pks = list(Service.objects.order_by('nom').values_list('pk', flat=True))
    pos = pks.index(pk) if pk in pks else None
    employes = obj.employes.order_by('nom', 'prenoms')
    return render(request, 'employer/config/service_detail.html', {
        'obj':      obj,
        'employes': employes,
        'prev_pk': pks[pos - 1] if pos and pos > 0 else None,
        'next_pk': pks[pos + 1] if pos is not None and pos < len(pks) - 1 else None,
        'can_manage': can_manage_rh(request.user),
    })


@login_required(login_url='login')
def employer_service_create(request):
    Form = _service_form_class()
    form = Form(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Service créé.')
        return redirect('employer_services')
    return render(request, 'employer/config/service_form.html', {
        'form': form, 'titre': 'Nouveau service',
        'can_manage': can_manage_rh(request.user),
    })


@login_required(login_url='login')
def employer_service_edit(request, pk):
    obj  = get_object_or_404(Service, pk=pk)
    Form = _service_form_class()
    form = Form(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Service mis à jour.')
        return redirect('employer_service_detail', pk=pk)
    return render(request, 'employer/config/service_form.html', {
        'form': form, 'titre': f'Modifier – {obj.nom}', 'obj': obj,
        'can_manage': can_manage_rh(request.user),
    })


@login_required(login_url='login')
@require_POST
def employer_service_bulk_delete(request):
    ids = request.POST.getlist('ids[]')
    if ids:
        Service.objects.filter(pk__in=ids).delete()
    return JsonResponse({'ok': True})


_SVC_HDR = ['code', 'nom', 'description', 'chef_service', 'actif']


def _svc_row(s):
    return [
        s.code, s.nom, s.description,
        s.chef_service.matricule if s.chef_service else '',
        int(s.actif),
    ]


@login_required(login_url='login')
def employer_export_services(request):
    from services.views import _export_file
    fmt = request.GET.get('format', 'json')
    qs = Service.objects.select_related('chef_service').order_by('nom')
    rows = [_svc_row(s) for s in qs]
    return _export_file(fmt, 'services', _SVC_HDR, rows,
                        [dict(zip(_SVC_HDR, r)) for r in rows])


@login_required(login_url='login')
@require_POST
def employer_import_services(request):
    from medecins.models import Medecin
    from services.views import _parse_upload, _s, _b, _fk_warning

    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('employer_services')

    data, err = _parse_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('employer_services')

    do_update = 'update' in request.POST
    created = updated = skipped = errors = 0
    chef_manquants = set()
    for item in data:
        try:
            code = _s(item.get('code', ''))
            if not code:
                errors += 1
                continue
            chef_matricule = _s(item.get('chef_service', ''))
            chef_service = Medecin.objects.filter(employe__matricule=chef_matricule).first() if chef_matricule else None
            if chef_matricule and not chef_service:
                chef_manquants.add(chef_matricule)

            defaults = {
                'nom': _s(item.get('nom', code)),
                'description': _s(item.get('description', '')),
                'chef_service': chef_service,
                'actif': _b(item.get('actif', True)),
            }
            obj, was_created = Service.objects.get_or_create(code=code, defaults=defaults)
            if was_created:
                created += 1
            elif do_update:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                updated += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    fk_msg = _fk_warning([('Médecin(s) chef de service', chef_manquants)])
    if errors or fk_msg:
        messages.warning(request, f'{created} créé(s), {updated} mis à jour, {skipped} ignoré(s)'
                          + (f', {errors} erreur(s)' if errors else '') + '.' + fk_msg)
    else:
        messages.success(request, f'{created} service(s) importé(s), {updated} mis à jour, {skipped} ignoré(s).')
    return redirect('employer_services')
