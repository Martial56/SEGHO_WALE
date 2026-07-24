from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.db.models import Q, Sum, Count, F
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from .models import Produit, CategorieStock, LotProduit, MouvementStock, CommandeStock, LigneCommande, Inventaire, LigneInventaire, DemandePharmacie, LigneDemande, PHARMACIES, FicheBesoins, LigneFicheBesoins, UniteMesure
from achats.models import Fournisseur

STOCK_MANAGE_GROUPS = {'Gestionnaire stock', 'Pharmacien', 'Administrateur', 'Directeur'}


def can_manage_stock(user):
    return user.is_superuser or user.is_staff or user.groups.filter(name__in=STOCK_MANAGE_GROUPS).exists()


@login_required(login_url='login')
def stock_dashboard(request):
    import json as _json
    today = timezone.now().date()

    ok      = Produit.objects.filter(actif=True, stock_actuel__gt=F('stock_alerte')).count()
    alertes = Produit.objects.filter(actif=True, stock_actuel__gt=0, stock_actuel__lte=F('stock_alerte')).count()
    ruptures = Produit.objects.filter(actif=True, stock_actuel__lte=0).count()

    stats = {
        'total_produits':     Produit.objects.filter(actif=True).count(),
        'medicaments':        Produit.objects.filter(type='medicament', actif=True).count(),
        'consommables':       Produit.objects.filter(type='consommable', actif=True).count(),
        'equipements':        Produit.objects.filter(type='equipement', actif=True).count(),
        'ruptures':           ruptures,
        'alertes':            alertes,
        'commandes_en_cours': CommandeStock.objects.filter(statut__in=['brouillon', 'envoye', 'partiel']).count(),
        'lots_perimes':       LotProduit.objects.filter(date_peremption__lt=today, quantite_actuelle__gt=0).count(),
    }

    # ── Graphe 1 : Répartition statut stock (donut) ──
    chart_statut = {
        'labels': ['Stock OK', 'En alerte', 'Rupture'],
        'data':   [ok, alertes, ruptures],
        'colors': ['#1a237e', '#f57c00', '#d32f2f'],
    }

    # ── Graphe 2 : Entrées vs Livraisons sur 7 jours (ligne) ──
    jours_labels, entrees_data, sorties_data = [], [], []
    for i in range(6, -1, -1):
        d = today - timezone.timedelta(days=i)
        jours_labels.append(d.strftime('%d/%m'))
        entrees_data.append(
            int(MouvementStock.objects.filter(type='entree', date__date=d).aggregate(
                t=Sum('quantite'))['t'] or 0)
        )
        sorties_data.append(
            int(MouvementStock.objects.filter(type='livraison', date__date=d).aggregate(
                t=Sum('quantite'))['t'] or 0)
        )
    chart_mouvements = {
        'labels':  jours_labels,
        'entrees': entrees_data,
        'sorties': sorties_data,
    }

    # ── Graphe 3 : Top 6 produits livrés (30 jours) ──
    from collections import defaultdict
    debut_mois = today - timezone.timedelta(days=30)
    mvts = MouvementStock.objects.filter(
        type='livraison', date__date__gte=debut_mois
    ).select_related('produit')
    conso = defaultdict(float)
    for mv in mvts:
        conso[mv.produit.nom[:20]] += float(mv.quantite)
    top_items = sorted(conso.items(), key=lambda x: x[1], reverse=True)[:6]
    chart_top = {
        'labels': [t[0] for t in top_items],
        'data':   [t[1] for t in top_items],
    }

    # ── Graphe 4 : Répartition par type (bar) — réutilise les comptages déjà
    # faits dans `stats`, au lieu de refaire les 3 mêmes requêtes.
    chart_types = {
        'labels': ['Médicaments', 'Consommables', 'Équipements'],
        'data':   [stats['medicaments'], stats['consommables'], stats['equipements']],
        'colors': ['#4a6741', '#1a237e', '#4527a0'],
    }

    # Fenêtre plus large que l'affichage réel (5/page) : cartes paginées côté
    # client dans le template, il faut donc plus de lignes que ce qu'une
    # seule page en montre.
    produits_alerte = Produit.objects.filter(
        actif=True, stock_actuel__lte=F('stock_alerte')
    ).order_by('stock_actuel')[:30]

    lots_a_surveiller = LotProduit.objects.filter(
        quantite_actuelle__gt=0,
        date_peremption__lte=today + timezone.timedelta(days=90)
    ).select_related('produit').order_by('date_peremption')[:30]

    # Fenêtre plus large que l'affichage réel (5/page) : les deux cartes sont
    # paginées côté client dans le template, il faut donc plus de lignes que
    # ce qu'une seule page en montre.
    derniers_mouvements = MouvementStock.objects.select_related('produit', 'produit__unite_mesure').order_by('-date')[:30]

    commandes_recentes = CommandeStock.objects.select_related('fournisseur').exclude(
        statut='annule'
    ).order_by('-date_creation')[:20]

    # Point 1 — Alertes critiques (ruptures, sous seuil minimum, lots périmant dans 30j)
    produits_rupture      = Produit.objects.filter(actif=True, stock_actuel__lte=0).order_by('nom')[:5]
    produits_sous_minimum = Produit.objects.filter(
        actif=True, stock_actuel__gt=0, stock_actuel__lte=F('stock_minimum')
    ).order_by('stock_actuel')[:5]
    lots_expiration_30 = LotProduit.objects.filter(
        quantite_actuelle__gt=0,
        date_peremption__gt=today,
        date_peremption__lte=today + timezone.timedelta(days=30),
    ).select_related('produit').order_by('date_peremption')[:5]

    # Points 3 & 6 — À commander + KPIs (une seule requête produits partagée)
    produits_all = list(Produit.objects.filter(actif=True).select_related('unite_mesure'))
    a_commander_dash = sorted(
        [{'produit': p, 'qte': p.qte_a_commander, 'stock': p.stock_actuel, 'minimum': p.stock_minimum}
         for p in produits_all if p.qte_a_commander > 0],
        key=lambda x: x['qte'], reverse=True
    )[:30]
    kpi_valeur_stock     = round(sum(float(p.stock_actuel) * float(p.prix_achat) for p in produits_all))
    total_livr_kpi       = float(MouvementStock.objects.filter(type='livraison').aggregate(t=Sum('quantite'))['t'] or 0)
    stock_total_kpi      = sum(float(p.stock_actuel) for p in produits_all)
    kpi_taux_rotation    = round(total_livr_kpi / stock_total_kpi, 2) if stock_total_kpi > 0 else 0
    _td = float(LigneDemande.objects.aggregate(t=Sum('quantite_demandee'))['t'] or 0)
    _ta = float(LigneDemande.objects.aggregate(t=Sum('quantite_approuvee'))['t'] or 0)
    kpi_taux_service     = round(_ta / _td * 100, 1) if _td > 0 else 0
    lots_obsoletes_count = LotProduit.objects.filter(
        quantite_actuelle__gt=0,
        date_reception__lt=today - timezone.timedelta(days=180),
    ).count()

    return render(request, 'stock/dashboard.html', {
        'stats':               stats,
        'produits_alerte':     produits_alerte,
        'lots_a_surveiller':   lots_a_surveiller,
        'derniers_mouvements': derniers_mouvements,
        'commandes_recentes':  commandes_recentes,
        'today':               today,
        'chart_statut':        chart_statut,
        'chart_mouvements':    chart_mouvements,
        'chart_top':           chart_top,
        'chart_types':         chart_types,
        # Point 1
        'produits_rupture':      produits_rupture,
        'produits_sous_minimum': produits_sous_minimum,
        'lots_expiration_30':    lots_expiration_30,
        # Point 3
        'a_commander_dash':      a_commander_dash,
        # Point 6
        'kpi_valeur_stock':      kpi_valeur_stock,
        'kpi_taux_rotation':     kpi_taux_rotation,
        'kpi_taux_service':      kpi_taux_service,
        'lots_obsoletes_count':  lots_obsoletes_count,
    })


@login_required(login_url='login')
def produits_list(request):
    qs = Produit.objects.select_related('categorie', 'unite_mesure').prefetch_related('lots').filter(actif=True)
    type_filtre      = request.GET.get('type', '')
    statut_filtre    = request.GET.get('statut', '')
    categorie_filtre = request.GET.get('categorie', '')
    q = request.GET.get('q', '').strip()

    if type_filtre:
        qs = qs.filter(type=type_filtre)
    if categorie_filtre:
        qs = qs.filter(categorie__pk=categorie_filtre)
    if statut_filtre == 'rupture':
        qs = qs.filter(stock_actuel__lte=0)
    elif statut_filtre == 'alerte':
        qs = qs.filter(stock_actuel__gt=0, stock_actuel__lte=F('stock_alerte'))
    elif statut_filtre == 'ok':
        qs = qs.filter(stock_actuel__gt=F('stock_alerte'))
    if q:
        qs = qs.filter(
            Q(nom__icontains=q) | Q(code__icontains=q) |
            Q(dci__icontains=q) | Q(lots__numero_lot__icontains=q)
        ).distinct()

    qs = qs.order_by('type', 'nom')
    paginator = Paginator(qs, 30)
    page_obj  = paginator.get_page(request.GET.get('page'))

    categories = CategorieStock.objects.filter(actif=True).order_by('nom')
    stats = {
        'total':      Produit.objects.filter(actif=True).count(),
        'ruptures':   Produit.objects.filter(actif=True, stock_actuel__lte=0).count(),
        'alertes':    Produit.objects.filter(actif=True, stock_actuel__gt=0, stock_actuel__lte=F('stock_alerte')).count(),
    }

    return render(request, 'stock/produits/list.html', {
        'page_obj':      page_obj,
        'categories':    categories,
        'type_filtre':      type_filtre,
        'statut_filtre':    statut_filtre,
        'categorie_filtre': categorie_filtre,
        'q':                q,
        'stats':         stats,
    })


@login_required(login_url='login')
def produit_detail(request, pk):
    produit = get_object_or_404(Produit.objects.select_related('categorie', 'unite_mesure', 'fournisseur_principal'), pk=pk)
    lots = produit.lots.order_by('date_peremption')
    mouvements = produit.mouvements.order_by('-date')[:20]
    return render(request, 'stock/produits/detail.html', {
        'produit':    produit,
        'lots':       lots,
        'mouvements': mouvements,
        'today':      timezone.now().date(),
    })


@login_required(login_url='login')
def produit_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    categories  = CategorieStock.objects.filter(actif=True).order_by('nom')
    fournisseurs = Fournisseur.objects.filter(actif=True).order_by('nom')
    unites_mesure = UniteMesure.objects.filter(actif=True).select_related('categorie').order_by('categorie__nom', 'nom')
    errors = {}

    if request.method == 'POST':
        nom  = request.POST.get('nom', '').strip()
        type_ = request.POST.get('type', 'medicament')
        cat_pk = request.POST.get('categorie', '')
        um_pk = request.POST.get('unite_mesure', '')
        if not nom:
            errors['nom'] = 'Le nom est obligatoire.'
        if not cat_pk:
            errors['categorie'] = 'La catégorie est obligatoire.'
        if not um_pk:
            errors['unite_mesure'] = "L'unité de mesure est obligatoire."
        if not errors:
            p = Produit(
                nom=nom, type=type_,
                dci=request.POST.get('dci', '').strip(),
                dosage=request.POST.get('dosage', '').strip(),
                forme=request.POST.get('forme', ''),
                description=request.POST.get('description', '').strip(),
                prescription_obligatoire=request.POST.get('prescription_obligatoire') == 'on',
            )
            p.unite_mesure = UniteMesure.objects.filter(pk=um_pk).first()
            try: p.stock_alerte  = float(request.POST.get('stock_alerte',  10) or 10)
            except ValueError: p.stock_alerte = 10
            try: p.stock_minimum = float(request.POST.get('stock_minimum', 5) or 5)
            except ValueError: p.stock_minimum = 5
            try: p.prix_achat    = float(request.POST.get('prix_achat', 0) or 0)
            except ValueError: p.prix_achat = 0
            try: p.prix_vente    = float(request.POST.get('prix_vente', 0) or 0)
            except ValueError: p.prix_vente = 0
            p.categorie = CategorieStock.objects.filter(pk=cat_pk).first()
            frn_pk = request.POST.get('fournisseur_principal', '')
            if frn_pk:
                p.fournisseur_principal = Fournisseur.objects.filter(pk=frn_pk).first()
            p.modifie_par = request.user
            p.modifie_le  = timezone.now()
            try:
                p.full_clean()
            except ValidationError as e:
                errors.update({k: ' '.join(v) for k, v in e.message_dict.items()})
            else:
                p.save()
                messages.success(request, f'Produit « {p.nom} » créé (code : {p.code}).')
                return redirect('stock_produit_detail', pk=p.pk)

    return render(request, 'stock/produits/form.html', {
        'mode':         'create',
        'categories':   categories,
        'fournisseurs': fournisseurs,
        'unites_mesure': unites_mesure,
        'errors':       errors,
        'post':         request.POST if request.method == 'POST' else None,
    })


@login_required(login_url='login')
def produit_edit(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    produit = get_object_or_404(Produit, pk=pk)
    categories   = CategorieStock.objects.filter(actif=True).order_by('nom')
    fournisseurs = Fournisseur.objects.filter(actif=True).order_by('nom')
    unites_mesure = UniteMesure.objects.filter(actif=True).select_related('categorie').order_by('categorie__nom', 'nom')
    errors = {}

    if request.method == 'POST':
        nom = request.POST.get('nom', '').strip()
        cat_pk = request.POST.get('categorie', '')
        um_pk = request.POST.get('unite_mesure', '')
        if not nom:
            errors['nom'] = 'Le nom est obligatoire.'
        if not cat_pk:
            errors['categorie'] = 'La catégorie est obligatoire.'
        if not um_pk:
            errors['unite_mesure'] = "L'unité de mesure est obligatoire."
        if not errors:
            produit.nom   = nom
            produit.type  = request.POST.get('type', produit.type)
            produit.dci   = request.POST.get('dci', '').strip()
            produit.dosage = request.POST.get('dosage', '').strip()
            produit.forme  = request.POST.get('forme', '')
            produit.unite_mesure = UniteMesure.objects.filter(pk=um_pk).first()
            produit.description  = request.POST.get('description', '').strip()
            produit.prescription_obligatoire = request.POST.get('prescription_obligatoire') == 'on'
            produit.actif = request.POST.get('actif') != 'off'
            try: produit.stock_alerte  = float(request.POST.get('stock_alerte',  produit.stock_alerte) or 10)
            except ValueError: pass
            try: produit.stock_minimum = float(request.POST.get('stock_minimum', produit.stock_minimum) or 5)
            except ValueError: pass
            try: produit.prix_achat    = float(request.POST.get('prix_achat', produit.prix_achat) or 0)
            except ValueError: pass
            try: produit.prix_vente    = float(request.POST.get('prix_vente', produit.prix_vente) or 0)
            except ValueError: pass
            produit.categorie = CategorieStock.objects.filter(pk=cat_pk).first()
            frn_pk = request.POST.get('fournisseur_principal', '')
            produit.fournisseur_principal = Fournisseur.objects.filter(pk=frn_pk).first() if frn_pk else None
            produit.modifie_par = request.user
            produit.modifie_le  = timezone.now()
            try:
                produit.full_clean()
            except ValidationError as e:
                errors.update({k: ' '.join(v) for k, v in e.message_dict.items()})
            else:
                produit.save()
                messages.success(request, f'Produit « {produit.nom} » mis à jour.')
                return redirect('stock_produit_detail', pk=produit.pk)

    return render(request, 'stock/produits/form.html', {
        'mode':         'edit',
        'produit':      produit,
        'categories':   categories,
        'fournisseurs': fournisseurs,
        'unites_mesure': unites_mesure,
        'errors':       errors,
    })


@login_required(login_url='login')
def mouvements_list(request):
    qs = MouvementStock.objects.select_related('produit', 'produit__unite_mesure').order_by('-date')
    type_filtre = request.GET.get('type', '')
    periode     = request.GET.get('periode', '')
    q           = request.GET.get('q', '').strip()

    if type_filtre:
        qs = qs.filter(type=type_filtre)
    if q:
        qs = qs.filter(Q(produit__nom__icontains=q) | Q(reference__icontains=q))

    today = timezone.now().date()
    if periode == 'today':
        qs = qs.filter(date__date=today)
    elif periode == 'week':
        debut_semaine = today - timezone.timedelta(days=today.weekday())
        qs = qs.filter(date__date__gte=debut_semaine)
    elif periode == 'month':
        qs = qs.filter(date__month=today.month, date__year=today.year)

    paginator = Paginator(qs, 40)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'stock/mouvements/list.html', {
        'page_obj':    page_obj,
        'type_filtre': type_filtre,
        'periode':     periode,
        'q':           q,
    })


@login_required(login_url='login')
def mouvement_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    produits = Produit.objects.filter(actif=True).order_by('nom')
    if request.method == 'POST':
        produit_pk = request.POST.get('produit', '')
        type_      = request.POST.get('type', '')
        motif      = request.POST.get('motif', '')
        try:
            quantite = float(request.POST.get('quantite', 0) or 0)
        except ValueError:
            quantite = 0
        notes    = request.POST.get('notes', '').strip()
        reference = request.POST.get('reference', '').strip()

        type_valide = type_ in dict(MouvementStock.TYPE_CHOICES)
        produit = None
        if produit_pk and quantite > 0 and type_valide:
            with transaction.atomic():
                produit = Produit.objects.select_for_update().filter(pk=produit_pk).first()
                if produit:
                    stock_avant = float(produit.stock_actuel)
                    if type_ == 'entree':
                        stock_apres = stock_avant + quantite
                    elif type_ in ('livraison', 'peremption', 'retour'):
                        stock_apres = max(0, stock_avant - quantite)
                    else:
                        stock_apres = quantite  # ajustement

                    MouvementStock.objects.create(
                        produit=produit, type=type_, motif=motif,
                        quantite=quantite, stock_avant=stock_avant,
                        stock_apres=stock_apres, notes=notes, reference=reference,
                        cree_par=request.user,
                    )
                    produit.stock_actuel = stock_apres
                    produit.save(update_fields=['stock_actuel'])
            if produit:
                messages.success(request, f'Mouvement enregistré pour « {produit.nom} ».')
                return redirect('stock_mouvements')
            messages.error(request, 'Produit introuvable.')
        else:
            messages.error(request, 'Produit, type et quantité valides requis.')

    return render(request, 'stock/mouvements/form.html', {'produits': produits})


@login_required(login_url='login')
def commandes_list(request):
    qs = CommandeStock.objects.select_related('fournisseur').order_by('-date_creation')
    statut_filtre = request.GET.get('statut', '')
    if statut_filtre:
        qs = qs.filter(statut=statut_filtre)
    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'stock/commandes/list.html', {
        'page_obj':     page_obj,
        'statut_filtre': statut_filtre,
    })


@login_required(login_url='login')
def commande_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    fournisseurs = Fournisseur.objects.filter(actif=True).order_by('nom')
    produits     = Produit.objects.filter(actif=True).order_by('nom')
    if request.method == 'POST':
        frn_pk = request.POST.get('fournisseur', '')
        fournisseur = Fournisseur.objects.filter(pk=frn_pk).first()
        if fournisseur:
            cmd = CommandeStock.objects.create(
                fournisseur=fournisseur,
                notes=request.POST.get('notes', '').strip(),
                cree_par=request.user,
            )
            messages.success(request, f'Commande {cmd.numero} créée.')
            return redirect('stock_commande_detail', pk=cmd.pk)
    return render(request, 'stock/commandes/form.html', {
        'fournisseurs': fournisseurs,
        'produits':     produits,
    })


@login_required(login_url='login')
@require_POST
def commande_ajouter_ligne(request, pk):
    if not can_manage_stock(request.user):
        return JsonResponse({'error': 'Permission refusée.'}, status=403)
    commande = get_object_or_404(CommandeStock, pk=pk)
    if commande.statut != 'brouillon':
        return JsonResponse({'error': 'La commande n\'est plus modifiable.'}, status=400)

    produit_pk  = request.POST.get('produit', '')
    try:
        quantite    = float(request.POST.get('quantite', 0) or 0)
        prix        = float(request.POST.get('prix_unitaire', 0) or 0)
    except ValueError:
        return JsonResponse({'error': 'Quantité ou prix invalide.'}, status=400)
    notes = request.POST.get('notes', '').strip()

    produit = Produit.objects.filter(pk=produit_pk).first()
    if not produit:
        return JsonResponse({'error': 'Produit introuvable.'}, status=400)
    if quantite <= 0:
        return JsonResponse({'error': 'La quantité doit être supérieure à 0.'}, status=400)

    ligne = LigneCommande.objects.create(
        commande=commande, produit=produit,
        quantite_commandee=quantite, prix_unitaire=prix, notes=notes,
    )

    # Recalculer le montant total
    total = sum(float(l.quantite_commandee) * float(l.prix_unitaire)
                for l in commande.lignes.all())
    commande.montant_total = total
    commande.save(update_fields=['montant_total'])

    return JsonResponse({
        'id':           ligne.pk,
        'produit_nom':  produit.nom,
        'produit_code': produit.code,
        'produit_pk':   produit.pk,
        'quantite':     float(ligne.quantite_commandee),
        'prix':         float(ligne.prix_unitaire),
        'montant':      float(ligne.montant),
        'total':        float(commande.montant_total),
        'nb_lignes':    commande.lignes.count(),
    })


@login_required(login_url='login')
@require_POST
def commande_ajouter_lot(request, pk):
    """Ajoute plusieurs produits à la commande en une seule soumission."""
    if not can_manage_stock(request.user):
        return JsonResponse({'error': 'Permission refusée.'}, status=403)
    commande = get_object_or_404(CommandeStock, pk=pk)
    if commande.statut != 'brouillon':
        return JsonResponse({'error': 'Non modifiable.'}, status=400)

    produits = Produit.objects.filter(actif=True)
    deja_liste = set(
        LigneCommande.objects.filter(commande=commande).values_list('produit_id', flat=True)
    )
    nouvelles_lignes = []
    ajouts = []
    deja_presents = []
    for p in produits:
        qte_str = request.POST.get(f'qte_{p.pk}', '').strip()
        prix_str = request.POST.get(f'prix_{p.pk}', '').strip()
        if not qte_str:
            continue
        try:
            qte = float(qte_str)
        except ValueError:
            continue
        if qte <= 0:
            continue
        try:
            prix = float(prix_str) if prix_str else float(p.prix_achat)
        except ValueError:
            prix = float(p.prix_achat)

        # Un produit ne peut apparaître qu'une seule fois par commande
        if p.pk in deja_liste:
            deja_presents.append(p.nom)
            continue
        nouvelles_lignes.append(LigneCommande(
            commande=commande, produit=p,
            quantite_commandee=qte, prix_unitaire=prix,
        ))
        ajouts.append(p.nom)

    if nouvelles_lignes:
        LigneCommande.objects.bulk_create(nouvelles_lignes)

    total = commande.lignes.aggregate(
        t=Sum(F('quantite_commandee') * F('prix_unitaire'))
    )['t'] or 0
    commande.montant_total = total
    commande.save(update_fields=['montant_total'])

    if ajouts:
        messages.success(request, f'{len(ajouts)} produit(s) ajouté(s) à la commande.')
    if deja_presents:
        noms = ', '.join(deja_presents)
        messages.warning(request, f'Déjà dans la commande (ignoré) : {noms}.')
    return redirect('stock_commande_detail', pk=pk)


@login_required(login_url='login')
@require_POST
def commande_supprimer_ligne(request, pk, ligne_pk):
    if not can_manage_stock(request.user):
        return JsonResponse({'error': 'Permission refusée.'}, status=403)
    commande = get_object_or_404(CommandeStock, pk=pk)
    if commande.statut != 'brouillon':
        return JsonResponse({'error': 'Non modifiable.'}, status=400)
    ligne = get_object_or_404(LigneCommande, pk=ligne_pk, commande=commande)
    ligne.delete()
    total = commande.lignes.aggregate(t=Sum(F('quantite_commandee') * F('prix_unitaire')))['t'] or 0
    commande.montant_total = total
    commande.save(update_fields=['montant_total'])
    return JsonResponse({'total': float(total), 'nb_lignes': commande.lignes.count()})


@login_required(login_url='login')
def commande_print(request, pk):
    commande = get_object_or_404(CommandeStock, pk=pk)
    lignes   = commande.lignes.select_related('produit').all()
    return render(request, 'stock/commandes/print.html', {
        'commande': commande,
        'lignes':   lignes,
        'today':    timezone.now().date(),
    })


@login_required(login_url='login')
def commande_detail(request, pk):
    commande = get_object_or_404(CommandeStock, pk=pk)
    lignes   = commande.lignes.select_related('produit').all()
    produits = Produit.objects.filter(actif=True).order_by('nom')
    return render(request, 'stock/commandes/detail.html', {
        'commande': commande,
        'lignes':   lignes,
        'produits': produits,
    })


@login_required(login_url='login')
@require_POST
def categorie_create_ajax(request):
    if not can_manage_stock(request.user):
        return JsonResponse({'error': 'Permission refusée.'}, status=403)
    nom  = request.POST.get('nom', '').strip()
    type_ = request.POST.get('type', 'medicament')
    if not nom:
        return JsonResponse({'error': 'Le nom est obligatoire.'}, status=400)
    cat, created = CategorieStock.objects.get_or_create(nom=nom, type=type_)
    return JsonResponse({'id': cat.pk, 'nom': cat.nom, 'type': cat.type, 'created': created})


# ── Configuration : Type de produit (lecture seule) / Catégorie de produit ──

@login_required(login_url='login')
def stock_types_produit(request):
    """Les types de produit sont des choix fixes (pas un modèle séparé) —
    page de référence en lecture seule, la vraie gestion se fait via les
    catégories (chacune rattachée à l'un de ces 3 types)."""
    types = [
        {
            'code': code, 'label': label,
            'nb_categories': CategorieStock.objects.filter(type=code).count(),
            'nb_produits':   Produit.objects.filter(type=code).count(),
        }
        for code, label in Produit.TYPE_CHOICES
    ]
    return render(request, 'stock/config/types_list.html', {'types': types})


def _categorie_stock_form_class():
    from django import forms

    class CategorieStockForm(forms.ModelForm):
        class Meta:
            model = CategorieStock
            fields = ['nom', 'type', 'description', 'actif']
            widgets = {
                'nom':         forms.TextInput(attrs={'placeholder': 'Ex : Antipaludéens'}),
                'type':        forms.Select(),
                'description': forms.Textarea(attrs={'rows': 3, 'placeholder': 'Description optionnelle'}),
            }
    return CategorieStockForm


@login_required(login_url='login')
def stock_categories_list(request):
    q = request.GET.get('q', '').strip()
    qs = CategorieStock.objects.annotate(nb_produits=Count('produit')).order_by('type', 'nom')
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(description__icontains=q))
    paginator = Paginator(qs, 15)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'stock/config/categories_list.html', {
        'page_obj': page_obj, 'q': q,
        'total': CategorieStock.objects.count(), 'total_filtre': qs.count(),
    })


@login_required(login_url='login')
def stock_categorie_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    Form = _categorie_stock_form_class()
    form = Form(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        obj = form.save()
        if is_ajax:
            return JsonResponse({'ok': True, 'message': f'Catégorie « {obj} » créée.'})
        messages.success(request, 'Catégorie créée.')
        return redirect('stock_categories_list')
    return render(request, 'stock/config/categorie_form_modal.html', {'form': form, 'titre': 'Nouvelle catégorie de produit'})


@login_required(login_url='login')
def stock_categorie_edit(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    obj = get_object_or_404(CategorieStock, pk=pk)
    Form = _categorie_stock_form_class()
    form = Form(request.POST or None, instance=obj)
    if request.method == 'POST' and form.is_valid():
        form.save()
        if is_ajax:
            return JsonResponse({'ok': True, 'message': f'Catégorie « {obj} » mise à jour.'})
        messages.success(request, 'Catégorie mise à jour.')
        return redirect('stock_categories_list')
    return render(request, 'stock/config/categorie_form_modal.html', {
        'form': form, 'titre': f'Modifier — {obj.nom}', 'obj': obj,
    })


@login_required(login_url='login')
@require_POST
def stock_categorie_delete(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    obj = get_object_or_404(CategorieStock, pk=pk)
    nom = obj.nom
    obj.delete()
    messages.success(request, f'Catégorie « {nom} » supprimée.')
    return redirect('stock_categories_list')


# ── Export / Import des catégories de produit ────────────────────────────────

_CATSTOCK_HDR = ['nom', 'type', 'description', 'actif']


def _catstock_row(c):
    return [c.nom, c.type, c.description, int(c.actif)]


@login_required(login_url='login')
def stock_export_categories_produit(request):
    fmt = request.GET.get('format', 'json')
    qs = CategorieStock.objects.order_by('type', 'nom')
    rows = [_catstock_row(c) for c in qs]
    return _export_file(fmt, 'categories_produit', _CATSTOCK_HDR, rows,
                        [dict(zip(_CATSTOCK_HDR, r)) for r in rows])


@login_required(login_url='login')
@require_POST
def stock_import_categories_produit(request):
    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('stock_categories_list')

    data, err = _parse_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('stock_categories_list')

    do_update = 'update' in request.POST
    created = updated = skipped = errors = 0

    for item in data:
        try:
            nom = _s(item.get('nom', ''))
            type_ = _s(item.get('type', 'medicament')) or 'medicament'
            if not nom:
                errors += 1
                continue
            defaults = {
                'nom': nom,
                'type': type_,
                'description': _s(item.get('description', '')),
                'actif': _b(item.get('actif', True)),
            }
            obj, was_created = CategorieStock.objects.get_or_create(nom__iexact=nom, type=type_, defaults=defaults)
            if was_created:
                created += 1
            elif do_update:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                updated += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    if errors:
        messages.warning(request, f'{created} créée(s), {updated} mise(s) à jour, {skipped} ignorée(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} catégorie(s) importée(s), {updated} mise(s) à jour, {skipped} ignorée(s).')
    return redirect('stock_categories_list')


@login_required(login_url='login')
def fournisseur_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    errors = {}
    if request.method == 'POST':
        nom       = request.POST.get('nom', '').strip()
        telephone = request.POST.get('telephone', '').strip()
        email     = request.POST.get('email', '').strip()
        adresse   = request.POST.get('adresse', '').strip()
        if not nom:
            errors['nom'] = 'Le nom est obligatoire.'
        if not telephone:
            errors['telephone'] = 'Le téléphone est obligatoire.'
        if not errors:
            f = Fournisseur.objects.create(
                nom=nom, telephone=telephone, email=email, adresse=adresse
            )
            messages.success(request, f'Fournisseur « {f.nom} » créé (code : {f.code}).')
            return redirect('stock_fournisseurs')
    return render(request, 'stock/fournisseurs/form.html', {
        'mode': 'create', 'errors': errors,
        'post': request.POST if request.method == 'POST' else None,
    })


@login_required(login_url='login')
def fournisseur_edit(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    fournisseur = get_object_or_404(Fournisseur, pk=pk)
    errors = {}
    if request.method == 'POST':
        nom       = request.POST.get('nom', '').strip()
        telephone = request.POST.get('telephone', '').strip()
        email     = request.POST.get('email', '').strip()
        adresse   = request.POST.get('adresse', '').strip()
        actif     = request.POST.get('actif') == 'on'
        if not nom:
            errors['nom'] = 'Le nom est obligatoire.'
        if not telephone:
            errors['telephone'] = 'Le téléphone est obligatoire.'
        if not errors:
            fournisseur.nom       = nom
            fournisseur.telephone = telephone
            fournisseur.email     = email
            fournisseur.adresse   = adresse
            fournisseur.actif     = actif
            fournisseur.save()
            messages.success(request, f'Fournisseur « {fournisseur.nom} » mis à jour.')
            return redirect('stock_fournisseurs')
    return render(request, 'stock/fournisseurs/form.html', {
        'mode': 'edit', 'fournisseur': fournisseur, 'errors': errors,
    })


@login_required(login_url='login')
def fournisseurs_list(request):
    qs = Fournisseur.objects.filter(actif=True).order_by('nom')
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(code__icontains=q))
    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'stock/fournisseurs/list.html', {
        'fournisseurs': page_obj,
        'page_obj':     page_obj,
        'q':            q,
    })


# ---------------------------------------------------------------------------
# Réception commande
# ---------------------------------------------------------------------------

@login_required(login_url='login')
@require_POST
def commande_envoyer(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    commande = get_object_or_404(CommandeStock, pk=pk)
    if commande.statut != 'brouillon':
        messages.error(request, 'Seule une commande en brouillon peut être envoyée.')
        return redirect('stock_commande_detail', pk=pk)
    if not commande.lignes.exists():
        messages.error(request, 'Ajoutez au moins un produit avant d\'envoyer la commande.')
        return redirect('stock_commande_detail', pk=pk)
    commande.statut = 'envoye'
    commande.save(update_fields=['statut'])
    messages.success(request, f'Commande {commande.numero} envoyée au fournisseur {commande.fournisseur.nom}.')
    return redirect('stock_commande_detail', pk=pk)


@login_required(login_url='login')
def commande_modifier(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    commande = get_object_or_404(CommandeStock, pk=pk)
    if commande.statut != 'brouillon':
        messages.error(request, 'Seule une commande en brouillon peut être modifiée.')
        return redirect('stock_commande_detail', pk=pk)
    fournisseurs = Fournisseur.objects.filter(actif=True).order_by('nom')
    errors = {}
    if request.method == 'POST':
        frn_pk = request.POST.get('fournisseur', '')
        fournisseur = Fournisseur.objects.filter(pk=frn_pk).first()
        if not fournisseur:
            errors['fournisseur'] = 'Sélectionnez un fournisseur.'
        if not errors:
            commande.fournisseur = fournisseur
            commande.notes = request.POST.get('notes', '').strip()
            date_liv = request.POST.get('date_livraison_prevue', '')
            commande.date_livraison_prevue = date_liv if date_liv else None
            commande.save(update_fields=['fournisseur', 'notes', 'date_livraison_prevue'])
            messages.success(request, f'Commande {commande.numero} mise à jour.')
            return redirect('stock_commande_detail', pk=pk)
    return render(request, 'stock/commandes/modifier.html', {
        'commande':    commande,
        'fournisseurs': fournisseurs,
        'errors':      errors,
    })


@login_required(login_url='login')
@require_POST
def commande_receptionner(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    from decimal import Decimal
    commande = get_object_or_404(CommandeStock, pk=pk)
    if commande.statut not in ('envoye', 'partiel'):
        messages.error(request, 'Cette commande ne peut pas être réceptionnée.')
        return redirect('stock_commande_detail', pk=pk)

    lignes = commande.lignes.select_related('produit').all()
    tout_recu = True

    with transaction.atomic():
        for ligne in lignes:
            key = f'recu_{ligne.pk}'
            try:
                qte_recue = float(request.POST.get(key, 0) or 0)
            except ValueError:
                qte_recue = 0
            if qte_recue <= 0:
                continue

            # Verrouille la ligne produit pour la durée de la transaction —
            # évite qu'une autre réception/mouvement concurrent sur le même
            # produit lise le même stock_actuel avant cette écriture.
            produit = Produit.objects.select_for_update().get(pk=ligne.produit_id)

            # Mettre à jour la ligne
            ligne.quantite_recue = min(ligne.quantite_recue + Decimal(str(qte_recue)), ligne.quantite_commandee)
            ligne.save(update_fields=['quantite_recue'])

            # Créer un lot
            num_lot = request.POST.get(f'lot_{ligne.pk}', '').strip() or f'LOT-{commande.numero}'
            date_peremption = request.POST.get(f'peremption_{ligne.pk}', '') or None
            lot = LotProduit.objects.create(
                produit=produit, numero_lot=num_lot,
                quantite_initiale=qte_recue, quantite_actuelle=qte_recue,
                fournisseur=commande.fournisseur, date_reception=timezone.now().date(),
                prix_achat_lot=ligne.prix_unitaire,
                date_peremption=date_peremption if date_peremption else None,
            )

            # Mouvement de stock
            stock_avant = float(produit.stock_actuel)
            stock_apres = stock_avant + qte_recue
            MouvementStock.objects.create(
                produit=produit, lot=lot, type='entree', motif='achat',
                quantite=qte_recue, stock_avant=stock_avant, stock_apres=stock_apres,
                reference=commande.numero, cree_par=request.user,
            )
            produit.stock_actuel = stock_apres
            produit.save(update_fields=['stock_actuel'])

            if ligne.quantite_recue < ligne.quantite_commandee:
                tout_recu = False

        # Mettre à jour le statut commande
        commande.statut = 'recu' if tout_recu else 'partiel'
        commande.date_reception = timezone.now().date()
        commande.save(update_fields=['statut', 'date_reception'])

    messages.success(request, f'Réception enregistrée pour la commande {commande.numero}.')
    return redirect('stock_commande_detail', pk=pk)


# ---------------------------------------------------------------------------
# Inventaire
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def inventaire_list(request):
    qs = Inventaire.objects.select_related('cree_par').annotate(
        nb_lignes=Count('lignes')
    ).order_by('-date_creation')
    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'stock/inventaire/list.html', {
        'inventaires': page_obj,
        'page_obj':    page_obj,
    })


@login_required(login_url='login')
def inventaire_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    produits = Produit.objects.filter(actif=True).order_by('type', 'nom')
    if request.method == 'POST':
        from decimal import Decimal
        inv = Inventaire.objects.create(
            notes=request.POST.get('notes', '').strip(),
            cree_par=request.user,
        )
        nouvelles_lignes = []
        for p in produits:
            val = request.POST.get(f'reel_{p.pk}', '')
            if val != '':
                try:
                    stock_reel = float(val)
                except ValueError:
                    stock_reel = float(p.stock_actuel)
                peremption = request.POST.get(f'peremption_{p.pk}', '').strip() or None
                # LigneInventaire.save() calcule normalement l'écart — bulk_create
                # n'appelle pas save(), donc on le calcule ici explicitement.
                ecart = Decimal(str(stock_reel)) - Decimal(str(p.stock_actuel))
                nouvelles_lignes.append(LigneInventaire(
                    inventaire=inv, produit=p,
                    stock_theorique=p.stock_actuel,
                    stock_reel=stock_reel,
                    ecart=ecart,
                    date_peremption=peremption,
                ))
        if nouvelles_lignes:
            LigneInventaire.objects.bulk_create(nouvelles_lignes)
        messages.success(request, f'Inventaire {inv.numero} créé.')
        return redirect('stock_inventaire_detail', pk=inv.pk)
    return render(request, 'stock/inventaire/form.html', {'produits': produits})


@login_required(login_url='login')
def inventaire_detail(request, pk):
    inv = get_object_or_404(Inventaire, pk=pk)
    lignes = inv.lignes.select_related('produit', 'produit__unite_mesure').all()

    if request.method == 'POST' and request.POST.get('action') == 'valider' and inv.statut == 'brouillon':
        if not can_manage_stock(request.user):
            raise PermissionDenied
        with transaction.atomic():
            for ligne in lignes:
                if ligne.ecart != 0:
                    produit = Produit.objects.select_for_update().get(pk=ligne.produit_id)
                    type_mvt = 'entree' if ligne.ecart > 0 else 'ajustement'
                    stock_avant = float(produit.stock_actuel)
                    stock_apres = float(ligne.stock_reel)
                    MouvementStock.objects.create(
                        produit=produit, type=type_mvt, motif='inventaire',
                        quantite=abs(float(ligne.ecart)),
                        stock_avant=stock_avant, stock_apres=stock_apres,
                        reference=inv.numero, cree_par=request.user,
                    )
                    produit.stock_actuel = stock_apres
                    produit.save(update_fields=['stock_actuel'])

                # Mettre à jour la date de péremption sur le lot principal si renseignée
                if ligne.date_peremption:
                    lot = ligne.produit.lots.order_by('-date_reception').first()
                    if lot:
                        lot.date_peremption = ligne.date_peremption
                        lot.save(update_fields=['date_peremption'])
            inv.statut = 'valide'
            inv.date_validation = timezone.now()
            inv.save(update_fields=['statut', 'date_validation'])
        messages.success(request, f'Inventaire {inv.numero} validé. Les stocks ont été mis à jour.')
        return redirect('stock_inventaire_detail', pk=pk)

    lignes_list = list(lignes)
    stats_ecarts = {
        'excedents': sum(1 for l in lignes_list if float(l.ecart) > 0),
        'manques':   sum(1 for l in lignes_list if float(l.ecart) < 0),
        'neutres':   sum(1 for l in lignes_list if float(l.ecart) == 0),
        'total':     len(lignes_list),
    }

    return render(request, 'stock/inventaire/detail.html', {
        'inv':          inv,
        'lignes':       lignes_list,
        'stats_ecarts': stats_ecarts,
    })


# ---------------------------------------------------------------------------
# Péremptions
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def peremptions_list(request):
    today = timezone.now().date()
    filtre = request.GET.get('filtre', '90j')

    if filtre == 'expires':
        lots = LotProduit.objects.filter(date_peremption__lt=today, quantite_actuelle__gt=0)
    elif filtre == '30j':
        lots = LotProduit.objects.filter(date_peremption__lte=today + timezone.timedelta(days=30), date_peremption__gte=today, quantite_actuelle__gt=0)
    elif filtre == '60j':
        lots = LotProduit.objects.filter(date_peremption__lte=today + timezone.timedelta(days=60), date_peremption__gte=today, quantite_actuelle__gt=0)
    else:
        lots = LotProduit.objects.filter(date_peremption__lte=today + timezone.timedelta(days=90), quantite_actuelle__gt=0)

    lots = lots.select_related('produit', 'produit__unite_mesure', 'fournisseur').order_by('date_peremption')
    stats = {
        'expires': LotProduit.objects.filter(date_peremption__lt=today, quantite_actuelle__gt=0).count(),
        'j30':     LotProduit.objects.filter(date_peremption__lte=today + timezone.timedelta(days=30), date_peremption__gte=today, quantite_actuelle__gt=0).count(),
        'j60':     LotProduit.objects.filter(date_peremption__lte=today + timezone.timedelta(days=60), date_peremption__gte=today, quantite_actuelle__gt=0).count(),
        'j90':     LotProduit.objects.filter(date_peremption__lte=today + timezone.timedelta(days=90), quantite_actuelle__gt=0).count(),
    }
    paginator = Paginator(lots, 30)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'stock/peremptions/list.html', {
        'lots': page_obj, 'page_obj': page_obj,
        'stats': stats, 'filtre': filtre, 'today': today,
    })


# ---------------------------------------------------------------------------
# Valorisation
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def valorisation(request):
    produits = Produit.objects.filter(actif=True, stock_actuel__gt=0).select_related('categorie')

    valeur_totale = sum(float(p.stock_actuel) * float(p.prix_achat) for p in produits)

    par_type = {
        'medicament':  {'valeur': 0, 'nb': 0},
        'consommable': {'valeur': 0, 'nb': 0},
        'equipement':  {'valeur': 0, 'nb': 0},
    }
    for p in produits:
        par_type[p.type]['valeur'] += float(p.stock_actuel) * float(p.prix_achat)
        par_type[p.type]['nb'] += 1

    paginator = Paginator(produits.order_by('-stock_actuel'), 30)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'stock/valorisation.html', {
        'produits':     page_obj,
        'page_obj':     page_obj,
        'valeur_totale': valeur_totale,
        'par_type':     par_type,
    })


# ---------------------------------------------------------------------------
# Rapports consommation
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def rapports_consommation(request):
    import datetime
    from collections import defaultdict
    today = timezone.now().date()
    mois  = int(request.GET.get('mois', today.month))
    annee = int(request.GET.get('annee', today.year))
    type_filtre = request.GET.get('type', '')

    # Livraisons = sorties vers pharmacies (type livraison)
    mouvements = MouvementStock.objects.filter(
        type='livraison',
        date__month=mois, date__year=annee,
    ).select_related('produit', 'produit__unite_mesure', 'lot')

    if type_filtre:
        mouvements = mouvements.filter(produit__type=type_filtre)

    # Agréger par produit (quantité + valorisation)
    consommation = defaultdict(lambda: {'produit': None, 'total': 0.0, 'nb_livraisons': 0, 'valeur': 0.0})
    for mv in mouvements:
        k = mv.produit.pk
        consommation[k]['produit']       = mv.produit
        consommation[k]['total']         += float(mv.quantite)
        consommation[k]['nb_livraisons'] += 1
        prix = float(mv.lot.prix_achat_lot) if mv.lot_id else float(mv.produit.prix_achat)
        consommation[k]['valeur']        += float(mv.quantite) * prix

    lignes = sorted(consommation.values(), key=lambda x: x['total'], reverse=True)

    total_quantite = sum(l['total'] for l in lignes)
    total_valeur   = sum(l['valeur'] for l in lignes)

    # Mois nav
    d_courante = datetime.date(annee, mois, 1)
    if mois == 1:
        prev_mois, prev_annee = 12, annee - 1
    else:
        prev_mois, prev_annee = mois - 1, annee
    if mois == 12:
        next_mois, next_annee = 1, annee + 1
    else:
        next_mois, next_annee = mois + 1, annee

    MOIS_NOMS = ['', 'Janvier','Février','Mars','Avril','Mai','Juin',
                 'Juillet','Août','Septembre','Octobre','Novembre','Décembre']
    annees_dispo = list(range(today.year, today.year - 3, -1))

    return render(request, 'stock/rapports/consommation.html', {
        'lignes':         lignes,
        'mois':           mois,
        'annee':          annee,
        'mois_nom':       MOIS_NOMS[mois],
        'mois_noms':      MOIS_NOMS,
        'annees_dispo':   annees_dispo,
        'type_filtre':    type_filtre,
        'total_quantite': total_quantite,
        'total_valeur':   total_valeur,
        'prev_mois':      prev_mois,
        'prev_annee':     prev_annee,
        'next_mois':      next_mois,
        'next_annee':     next_annee,
        'today':          today,
    })


@login_required(login_url='login')
def rapports_dotations(request):
    import json as _json, datetime
    from collections import defaultdict
    today   = timezone.now().date()
    annee   = int(request.GET.get('annee', today.year))
    pharmacie_filtre = request.GET.get('pharmacie', '')

    MOIS_NOMS = ['','Janv','Févr','Mars','Avr','Mai','Juin','Juil','Août','Sept','Oct','Nov','Déc']
    PHARMACIES_LABELS = {'wale_toumbokro': 'Walé Toumbokro', 'wale_yamoussoukro': 'Walé Yamoussoukro'}

    # ── 1. Tendance mensuelle des demandes par pharmacie ──
    tendance = {}
    for code, label in PHARMACIES_LABELS.items():
        tendance[code] = {'label': label, 'data': [0]*12}
    demandes = DemandePharmacie.objects.filter(date_demande__year=annee)
    if pharmacie_filtre:
        demandes = demandes.filter(pharmacie=pharmacie_filtre)
    for d in demandes:
        m = d.date_demande.month - 1
        if d.pharmacie in tendance:
            tendance[d.pharmacie]['data'][m] += 1

    chart_tendance = {
        'labels': MOIS_NOMS[1:],
        'datasets': [
            {'label': v['label'], 'data': v['data'],
             'borderColor': '#1a237e' if k == 'wale_yamoussoukro' else '#ef6c00',
             'backgroundColor': 'rgba(26,35,126,.1)' if k == 'wale_yamoussoukro' else 'rgba(239,108,0,.08)',
             'tension': 0.3, 'fill': True}
            for k, v in tendance.items()
        ]
    }

    # ── 2. Taux approbation par pharmacie ──
    taux = {}
    for code, label in PHARMACIES_LABELS.items():
        qs = DemandePharmacie.objects.filter(pharmacie=code, date_demande__year=annee)
        total     = qs.count()
        approuvee = qs.filter(statut__in=['approuvee', 'partielle', 'en_livraison']).count()
        taux[code] = {
            'label':     label,
            'total':     total,
            'approuvee': approuvee,
            'taux':      round(approuvee / total * 100) if total else 0,
        }

    # ── 3. Top 10 produits les plus demandés ──
    qs_lignes = LigneDemande.objects.filter(
        demande__date_demande__year=annee
    ).select_related('produit', 'demande')
    if pharmacie_filtre:
        qs_lignes = qs_lignes.filter(demande__pharmacie=pharmacie_filtre)

    prod_stats = defaultdict(lambda: {'produit': None, 'qte_demandee': 0.0, 'qte_approuvee': 0.0, 'nb': 0})
    for ligne in qs_lignes:
        if not ligne.produit:
            continue
        k = ligne.produit.pk
        prod_stats[k]['produit']      = ligne.produit
        prod_stats[k]['qte_demandee'] += float(ligne.quantite_demandee)
        prod_stats[k]['qte_approuvee'] += float(ligne.quantite_approuvee)
        prod_stats[k]['nb']           += 1

    top_produits = sorted(prod_stats.values(), key=lambda x: x['qte_demandee'], reverse=True)[:10]

    chart_top = {
        'labels':    [t['produit'].nom[:20] for t in top_produits],
        'demandee':  [t['qte_demandee']     for t in top_produits],
        'approuvee': [t['qte_approuvee']    for t in top_produits],
    }

    # ── 4. Comparaison volumes entre pharmacies ──
    comp = {}
    for code, label in PHARMACIES_LABELS.items():
        qte = sum(
            float(l.quantite_demandee)
            for l in LigneDemande.objects.filter(
                demande__pharmacie=code, demande__date_demande__year=annee
            )
        )
        comp[code] = {'label': label, 'qte': qte}

    chart_comp = {
        'labels': [v['label'] for v in comp.values()],
        'data':   [v['qte']   for v in comp.values()],
        'colors': ['#1a237e', '#ef6c00'],
    }

    # ── 5. Stats globales ──
    stats = {
        'total_demandes': DemandePharmacie.objects.filter(date_demande__year=annee).count(),
        'en_attente':     DemandePharmacie.objects.filter(date_demande__year=annee, statut='en_attente').count(),
        'approuvees':     DemandePharmacie.objects.filter(date_demande__year=annee, statut__in=['approuvee','en_livraison']).count(),
        'refusees':       DemandePharmacie.objects.filter(date_demande__year=annee, statut='refusee').count(),
    }

    annees_dispo = list(range(today.year, today.year - 3, -1))

    return render(request, 'stock/rapports/dotations.html', {
        'annee':           annee,
        'annees_dispo':    annees_dispo,
        'pharmacie_filtre': pharmacie_filtre,
        'pharmacies':      PHARMACIES_LABELS,
        'stats':           stats,
        'taux':            taux,
        'top_produits':    top_produits,
        'chart_tendance':  chart_tendance,
        'chart_top':       chart_top,
        'chart_comp':      chart_comp,
        'today':           today,
    })


# ---------------------------------------------------------------------------
@login_required(login_url='login')
def rapports_besoins(request):
    """Rapport besoins mensuels — style fiche commande pharmacie."""
    type_filtre = request.GET.get('type', '')
    today = timezone.now().date()

    qs = Produit.objects.filter(actif=True).select_related('unite_mesure').order_by('type', 'nom')
    if type_filtre:
        qs = qs.filter(type=type_filtre)

    # Calcul des indicateurs pour chaque produit
    produits_data = []
    for p in qs:
        cmm = p.cmm
        stock = float(p.stock_actuel)
        couverture = p.couverture_jours
        qte_cmd = p.qte_a_commander
        produits_data.append({
            'produit':    p,
            'cmm':        cmm,
            'stock':      stock,
            'couverture': couverture,
            'qte_cmd':    qte_cmd,
            'en_besoin':  qte_cmd > 0,
        })

    # Stats globales
    nb_rupture  = sum(1 for d in produits_data if d['stock'] <= 0)
    nb_critique = sum(1 for d in produits_data if 0 < d['stock'] <= float(d['produit'].stock_alerte))
    nb_besoin   = sum(1 for d in produits_data if d['qte_cmd'] > 0)

    ctx = {
        'produits_data': produits_data,
        'type_filtre':   type_filtre,
        'today':         today,
        'stats': {'rupture': nb_rupture, 'critique': nb_critique, 'besoin': nb_besoin, 'total': len(produits_data)},
    }
    return render(request, 'stock/rapports/besoins.html', ctx)


@login_required(login_url='login')
def rapports_besoins_print(request):
    """Version imprimable / PDF de la fiche de besoins."""
    type_filtre = request.GET.get('type', '')
    today = timezone.now().date()
    qs = Produit.objects.filter(actif=True).select_related('unite_mesure').order_by('type', 'nom')
    if type_filtre:
        qs = qs.filter(type=type_filtre)
    produits_data = []
    for p in qs:
        produits_data.append({
            'produit':    p,
            'cmm':        p.cmm,
            'stock':      float(p.stock_actuel),
            'couverture': p.couverture_jours,
            'qte_cmd':    p.qte_a_commander,
        })
    return render(request, 'stock/rapports/besoins_print.html', {
        'produits_data': produits_data,
        'type_filtre':   type_filtre,
        'today':         today,
        'type_label': {'medicament': 'Médicaments', 'consommable': 'Consommables', 'equipement': 'Équipements'}.get(type_filtre, 'Tous les produits'),
        'nb_a_commander': sum(1 for d in produits_data if d['qte_cmd'] > 0),
    })


@login_required(login_url='login')
def rapports_indicateurs(request):
    """Tableau de bord des indicateurs de performance du stock."""
    import json as _json
    today = timezone.now().date()

    produits = list(Produit.objects.filter(actif=True))

    # Taux de service = demandes satisfaites / demandes totales
    total_demandee  = sum(float(l.quantite_demandee)  for l in LigneDemande.objects.all())
    total_approuvee = sum(float(l.quantite_approuvee) for l in LigneDemande.objects.all())
    taux_service  = round(total_approuvee / total_demandee * 100, 1) if total_demandee > 0 else 0
    taux_rupture  = round((total_demandee - total_approuvee) / total_demandee * 100, 1) if total_demandee > 0 else 0

    # Valeur du stock
    valeur_stock = sum(float(p.stock_actuel) * float(p.prix_achat) for p in produits)

    # Taux de rotation = livraisons / stock moyen
    from django.db.models import Sum
    total_livraisons = float(MouvementStock.objects.filter(type='livraison').aggregate(
        t=Sum('quantite'))['t'] or 0)
    stock_total = sum(float(p.stock_actuel) for p in produits)
    taux_rotation = round(total_livraisons / stock_total, 2) if stock_total > 0 else 0

    # Couverture moyenne
    couvertures = [p.couverture_jours for p in produits if p.couverture_jours is not None]
    couverture_moy = round(sum(couvertures) / len(couvertures)) if couvertures else 0

    # Top 5 produits avec couverture critique (< 30 jours)
    critiques = sorted(
        [{'produit': p, 'couverture': p.couverture_jours, 'cmm': p.cmm}
         for p in produits if p.couverture_jours is not None and p.couverture_jours < 30],
        key=lambda x: x['couverture']
    )[:10]

    # Top 5 produits à commander
    a_commander = sorted(
        [{'produit': p, 'qte': p.qte_a_commander, 'cmm': p.cmm}
         for p in produits if p.qte_a_commander > 0],
        key=lambda x: x['qte'], reverse=True
    )[:10]

    # Prévision 3 mois (Feature 3)
    previsions = []
    for item in a_commander:
        p   = item['produit']
        cmm = item['cmm'] or 0
        s0  = float(p.stock_actuel)
        previsions.append({
            'produit':         p,
            'cmm':             cmm,
            'stock_actuel':    s0,
            'stock_m1':        max(0, s0 - cmm),
            'stock_m2':        max(0, s0 - 2 * cmm),
            'stock_m3':        max(0, s0 - 3 * cmm),
            'qte_a_commander': item['qte'],
        })

    chart_service_data = {
        'labels': ['Taux de service', 'Taux de rupture'],
        'data': [float(taux_service), float(taux_rupture)],
        'colors': ['#0d7a4c', '#e84545'],
    }
    chart_critiques_data = {
        'labels': [c['produit'].nom[:22] for c in critiques],
        'data': [int(c['couverture']) for c in critiques],
    }
    chart_commander_data = {
        'labels': [c['produit'].nom[:22] for c in a_commander],
        'data': [int(c['qte']) for c in a_commander],
    }

    return render(request, 'stock/rapports/indicateurs.html', {
        'today':           today,
        'taux_service':    taux_service,
        'taux_rupture':    taux_rupture,
        'taux_rotation':   taux_rotation,
        'couverture_moy':  couverture_moy,
        'valeur_stock':    valeur_stock,
        'critiques':       critiques,
        'a_commander':     a_commander,
        'previsions':      previsions,
        'total_produits':  len(produits),
        'chart_service':   chart_service_data,
        'chart_critiques': chart_critiques_data,
        'chart_commander': chart_commander_data,
    })


# ---------------------------------------------------------------------------
# Rapport de péremptions (Feature 1)
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def rapports_peremptions(request):
    from collections import defaultdict
    today = timezone.now().date()
    mois  = int(request.GET.get('mois', today.month))
    annee = int(request.GET.get('annee', today.year))

    mouvements = MouvementStock.objects.filter(
        type='peremption',
        date__month=mois, date__year=annee,
    ).select_related('produit', 'produit__unite_mesure', 'lot', 'cree_par').order_by('-date')

    total_quantite = sum(float(mv.quantite) for mv in mouvements)
    total_valeur   = sum(
        float(mv.quantite) * float(mv.lot.prix_achat_lot if mv.lot_id else mv.produit.prix_achat)
        for mv in mouvements
    )

    MOIS_NOMS = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    annees_dispo = list(range(today.year, today.year - 3, -1))
    prev_mois, prev_annee = (12, annee - 1) if mois == 1 else (mois - 1, annee)
    next_mois, next_annee = (1, annee + 1) if mois == 12 else (mois + 1, annee)

    return render(request, 'stock/rapports/peremptions.html', {
        'mouvements':     mouvements,
        'mois':           mois, 'annee': annee, 'mois_nom': MOIS_NOMS[mois],
        'annees_dispo':   annees_dispo,
        'total_quantite': total_quantite, 'total_valeur': total_valeur,
        'prev_mois': prev_mois, 'prev_annee': prev_annee,
        'next_mois': next_mois, 'next_annee': next_annee,
        'today': today,
    })


# ---------------------------------------------------------------------------
# Saisie des éliminations / destructions (Feature 2)
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def elimination_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    today = timezone.now().date()

    if request.method == 'POST':
        nb = 0
        notes_global = request.POST.get('notes', '').strip()
        with transaction.atomic():
            for key, val in request.POST.items():
                if not key.startswith('qte_'):
                    continue
                lot_pk = key[4:]
                try:
                    qte = float(val)
                except (ValueError, TypeError):
                    continue
                if qte <= 0:
                    continue
                try:
                    lot = LotProduit.objects.select_for_update().get(pk=lot_pk)
                except LotProduit.DoesNotExist:
                    continue
                qte = min(qte, float(lot.quantite_actuelle))
                if qte <= 0:
                    continue
                produit = Produit.objects.select_for_update().get(pk=lot.produit_id)
                sv = float(produit.stock_actuel)
                sa = max(0, sv - qte)
                MouvementStock.objects.create(
                    produit=produit, lot=lot,
                    type='peremption', motif='peremption',
                    quantite=qte, stock_avant=sv, stock_apres=sa,
                    notes=notes_global, cree_par=request.user,
                )
                lot.quantite_actuelle = max(0, float(lot.quantite_actuelle) - qte)
                lot.save(update_fields=['quantite_actuelle'])
                produit.stock_actuel = sa
                produit.save(update_fields=['stock_actuel'])
                nb += 1
        if nb:
            messages.success(request, f'{nb} lot(s) éliminé(s) et enregistré(s).')
        return redirect('stock_peremptions')

    # GET : lots périmés ou expirant dans 30 jours (la 1re condition est un
    # sous-ensemble de la 2e, donc un simple <= 30j suffit)
    lots = LotProduit.objects.filter(
        quantite_actuelle__gt=0,
        date_peremption__lte=today + timezone.timedelta(days=30),
    ).select_related('produit', 'produit__unite_mesure', 'fournisseur').order_by('date_peremption')

    return render(request, 'stock/peremptions/eliminer.html', {
        'lots': lots, 'today': today,
    })


# ---------------------------------------------------------------------------
# Génération automatique de fiche de besoins (Feature 4)
# ---------------------------------------------------------------------------

@login_required(login_url='login')
@require_POST
def besoins_generer_auto(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    import datetime as _dt
    today = timezone.now().date()
    debut = today.replace(day=1)
    if today.month == 12:
        fin = today.replace(day=31)
    else:
        fin = today.replace(month=today.month + 1, day=1) - _dt.timedelta(days=1)

    produits_all = list(Produit.objects.filter(actif=True))
    a_commander  = [p for p in produits_all if p.qte_a_commander > 0]

    if not a_commander:
        messages.warning(request, 'Aucun produit à commander actuellement.')
        return redirect('stock_rapports_indicateurs')

    fiche = FicheBesoins.objects.create(
        periode_debut=debut, periode_fin=fin,
        cree_par=request.user,
        notes='Généré automatiquement depuis les indicateurs de stock.',
    )
    LigneFicheBesoins.objects.bulk_create([
        LigneFicheBesoins(
            fiche=fiche, produit=p,
            stock_initial=p.stock_actuel,
            cmm=p.cmm,
            qte_commander=p.qte_a_commander,
        )
        for p in a_commander
    ])
    messages.success(request, f'Fiche {fiche.numero} créée avec {len(a_commander)} produit(s).')
    return redirect('stock_fiche_detail', pk=fiche.pk)


# ---------------------------------------------------------------------------
# Bilan mensuel (Feature 5)
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def rapports_bilan_mensuel(request):
    from collections import defaultdict
    today = timezone.now().date()
    mois  = int(request.GET.get('mois', today.month))
    annee = int(request.GET.get('annee', today.year))
    MOIS_NOMS = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    annees_dispo = list(range(today.year, today.year - 3, -1))
    prev_mois, prev_annee = (12, annee - 1) if mois == 1 else (mois - 1, annee)
    next_mois, next_annee = (1, annee + 1) if mois == 12 else (mois + 1, annee)

    mvts = list(MouvementStock.objects.filter(
        date__month=mois, date__year=annee,
    ).select_related('produit', 'lot').order_by('-date'))

    def _val(mv):
        prix = float(mv.lot.prix_achat_lot) if mv.lot_id else float(mv.produit.prix_achat)
        return float(mv.quantite) * prix

    entrees     = [mv for mv in mvts if mv.type == 'entree']
    sorties     = [mv for mv in mvts if mv.type == 'livraison']
    peremptes   = [mv for mv in mvts if mv.type == 'peremption']
    ajustements = [mv for mv in mvts if mv.type == 'ajustement']
    retours     = [mv for mv in mvts if mv.type == 'retour']

    conso = defaultdict(float)
    for mv in sorties:
        conso[mv.produit.nom[:30]] += float(mv.quantite)
    top_sorties = sorted(conso.items(), key=lambda x: x[1], reverse=True)[:5]

    return render(request, 'stock/rapports/bilan.html', {
        'mois': mois, 'annee': annee, 'mois_nom': MOIS_NOMS[mois],
        'annees_dispo': annees_dispo,
        'prev_mois': prev_mois, 'prev_annee': prev_annee,
        'next_mois': next_mois, 'next_annee': next_annee,
        'today': today,
        'entrees': entrees, 'sorties': sorties,
        'peremptes': peremptes, 'ajustements': ajustements, 'retours': retours,
        'qte_entrees':   sum(float(mv.quantite) for mv in entrees),
        'qte_sorties':   sum(float(mv.quantite) for mv in sorties),
        'qte_peremptes': sum(float(mv.quantite) for mv in peremptes),
        'val_entrees':   round(sum(_val(mv) for mv in entrees)),
        'val_sorties':   round(sum(_val(mv) for mv in sorties)),
        'val_peremptes': round(sum(_val(mv) for mv in peremptes)),
        'top_sorties':   top_sorties,
        'nb_ruptures':   Produit.objects.filter(actif=True, stock_actuel__lte=0).count(),
        'nb_alertes':    Produit.objects.filter(actif=True, stock_actuel__gt=0, stock_actuel__lte=F('stock_alerte')).count(),
    })


# ---------------------------------------------------------------------------
# Comparatif fournisseurs / historique des prix (Feature 6)
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def rapports_fournisseurs_prix(request):
    from collections import defaultdict
    fournisseur_pk = request.GET.get('fournisseur', '')
    produit_pk     = request.GET.get('produit', '')

    lots_qs = LotProduit.objects.filter(
        prix_achat_lot__gt=0,
    ).select_related('produit', 'fournisseur').order_by('produit__nom', 'date_reception')

    if fournisseur_pk:
        lots_qs = lots_qs.filter(fournisseur_id=fournisseur_pk)
    if produit_pk:
        lots_qs = lots_qs.filter(produit_id=produit_pk)

    par_produit = defaultdict(list)
    for lot in lots_qs:
        par_produit[lot.produit].append(lot)

    produits_data = []
    for prod, lots in par_produit.items():
        prices = [float(l.prix_achat_lot) for l in lots]
        pmin, pmax = min(prices), max(prices)
        produits_data.append({
            'produit':   prod,
            'lots':      lots,
            'prix_min':  pmin,
            'prix_max':  pmax,
            'prix_moy':  round(sum(prices) / len(prices)),
            'variation': round((pmax - pmin) / pmin * 100, 1) if pmin > 0 else 0,
        })
    produits_data.sort(key=lambda x: x['variation'], reverse=True)

    from achats.models import Fournisseur as FournisseurAchat
    fournisseurs = FournisseurAchat.objects.filter(actif=True).order_by('nom')
    produits_avec_lots = Produit.objects.filter(
        lots__prix_achat_lot__gt=0, actif=True
    ).distinct().order_by('nom')

    return render(request, 'stock/rapports/fournisseurs_prix.html', {
        'produits_data':      produits_data,
        'fournisseurs':       fournisseurs,
        'produits_avec_lots': produits_avec_lots,
        'fournisseur_pk':     fournisseur_pk,
        'produit_pk':         produit_pk,
        'today':              timezone.now().date(),
    })


# ---------------------------------------------------------------------------
# Retours pharmacie → stock central (Feature 8)
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def retour_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        produit_pk = request.POST.get('produit', '')
        lot_pk     = request.POST.get('lot', '')
        pharmacie  = request.POST.get('pharmacie', '')
        notes      = request.POST.get('notes', '').strip()
        try:
            qte = float(request.POST.get('quantite', 0) or 0)
        except ValueError:
            qte = 0

        if not produit_pk or qte <= 0:
            messages.error(request, 'Produit et quantité requis.')
            return redirect('stock_retour_create')

        with transaction.atomic():
            produit = get_object_or_404(Produit.objects.select_for_update(), pk=produit_pk)
            lot = LotProduit.objects.select_for_update().filter(pk=lot_pk, produit=produit).first() if lot_pk else None

            sv = float(produit.stock_actuel)
            sa = sv + qte
            MouvementStock.objects.create(
                produit=produit, lot=lot,
                type='retour', motif='retour',
                pharmacie=pharmacie, quantite=qte,
                stock_avant=sv, stock_apres=sa,
                notes=notes, cree_par=request.user,
            )
            produit.stock_actuel = sa
            produit.save(update_fields=['stock_actuel'])
            if lot:
                lot.quantite_actuelle = float(lot.quantite_actuelle) + qte
                lot.save(update_fields=['quantite_actuelle'])

        messages.success(request, f'Retour de {qte:.0f} {produit.unite_mesure.nom if produit.unite_mesure else ""} enregistré pour « {produit.nom} ».')
        return redirect('stock_mouvements')

    produits  = Produit.objects.filter(actif=True).order_by('nom')
    return render(request, 'stock/retours/form.html', {
        'produits':   produits,
        'pharmacies': PHARMACIES,
        'today':      timezone.now().date(),
    })


# ---------------------------------------------------------------------------
# Export Excel (CSV)
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def export_stock_excel(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    from core.utils import csv_response
    type_filtre = request.GET.get('type', '')
    qs = Produit.objects.filter(actif=True).select_related('categorie', 'unite_mesure')
    if type_filtre:
        qs = qs.filter(type=type_filtre)
    qs = qs.order_by('type', 'nom')

    headers = ['Code', 'Nom', 'Type', 'Catégorie', 'Unité', 'Stock actuel', 'Seuil alerte', 'Prix achat', 'Prix vente', 'État']
    rows = []
    for p in qs:
        etat = 'Rupture' if p.en_rupture else ('Alerte' if p.en_alerte else 'OK')
        rows.append([
            p.code, p.nom, p.get_type_display(),
            p.categorie.nom if p.categorie else '',
            p.unite_mesure.nom if p.unite_mesure else '', p.stock_actuel, p.stock_alerte,
            p.prix_achat, p.prix_vente, etat,
        ])
    return csv_response('stock', headers, rows)


_PRODUIT_HDR = [
    'code', 'nom', 'type', 'categorie', 'unite_mesure', 'dci', 'dosage', 'forme',
    'prescription_obligatoire', 'stock_actuel', 'stock_alerte', 'stock_minimum',
    'prix_achat', 'prix_vente', 'actif', 'numero_lot', 'date_peremption',
]


def _parse_date_cell(value):
    """Accepte aussi bien une vraie date Excel (openpyxl la lit comme
    datetime/date) qu'une chaîne 'AAAA-MM-JJ' (CSV/JSON)."""
    import datetime
    if not value:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    value = str(value).strip()
    if not value:
        return None
    try:
        return datetime.datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


# ── Export des produits (JSON/CSV/XLSX, comme les unités de mesure) ──

_PRD_HDR = [
    'code', 'nom', 'type', 'categorie', 'description', 'unite_mesure',
    'dci', 'dosage', 'forme', 'prescription_obligatoire',
    'stock_actuel', 'stock_alerte', 'stock_minimum',
    'prix_achat', 'prix_vente', 'actif',
]


def _prd_row(p):
    return [
        p.code, p.nom, p.type,
        p.categorie.nom if p.categorie else '',
        p.description,
        p.unite_mesure.nom if p.unite_mesure else '',
        p.dci, p.dosage, p.forme, int(p.prescription_obligatoire),
        float(p.stock_actuel), float(p.stock_alerte), float(p.stock_minimum),
        float(p.prix_achat), float(p.prix_vente), int(p.actif),
    ]


@login_required(login_url='login')
def export_produits(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    fmt = request.GET.get('format', 'json')
    qs = Produit.objects.select_related('categorie', 'unite_mesure').order_by('type', 'nom')
    rows = [_prd_row(p) for p in qs]
    return _export_file(fmt, 'produits', _PRD_HDR, rows,
                        [dict(zip(_PRD_HDR, r)) for r in rows])


@login_required(login_url='login')
@require_POST
def import_produits(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('stock_produits')

    data, err = _parse_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('stock_produits')

    do_update = 'update' in request.POST
    types_valides  = dict(Produit.TYPE_CHOICES)
    formes_valides = dict(Produit.FORME_CHOICES)
    created = updated = skipped = errors = 0
    unites_manquantes = set()

    for item in data:
        try:
            nom = _s(item.get('nom', ''))
            if not nom:
                errors += 1
                continue

            code = _s(item.get('code', ''))
            existing = Produit.objects.filter(code=code).first() if code else None
            if existing and not do_update:
                skipped += 1
                continue

            type_produit = _s(item.get('type', 'medicament')) or 'medicament'
            if type_produit not in types_valides:
                type_produit = 'medicament'
            forme = _s(item.get('forme', ''))
            if forme not in formes_valides:
                forme = ''

            cat_nom = _s(item.get('categorie', ''))
            categorie = None
            if cat_nom:
                categorie, _created_cat = CategorieStock.objects.get_or_create(
                    nom=cat_nom, defaults={'type': type_produit}
                )

            um_code = _s(item.get('unite_mesure', ''))
            unite = None
            if um_code:
                unite = UniteMesure.objects.filter(code=um_code).first()
                if not unite:
                    unites_manquantes.add(um_code)

            with transaction.atomic():
                obj = existing or Produit()
                if code:
                    obj.code = code
                obj.nom       = nom
                obj.type      = type_produit
                obj.categorie = categorie
                obj.unite_mesure = unite
                obj.dci    = _s(item.get('dci', ''))
                obj.dosage = _s(item.get('dosage', ''))
                obj.forme  = forme
                obj.prescription_obligatoire = _b(item.get('prescription_obligatoire', False))
                obj.stock_actuel  = item.get('stock_actuel') or 0
                obj.stock_alerte  = item.get('stock_alerte') or 10
                obj.stock_minimum = item.get('stock_minimum') or 5
                obj.prix_achat = item.get('prix_achat') or 0
                obj.prix_vente = item.get('prix_vente') or 0
                obj.actif = _b(item.get('actif', True))
                obj.modifie_par = request.user
                obj.modifie_le  = timezone.now()

                obj.full_clean()
                obj.save()

                # ── Lot + traçabilité (numero_lot / date_peremption facultatifs) ──
                # Une ligne = un produit = au plus un lot initial. La quantité du
                # lot suit le même principe « déclaratif » que stock_actuel :
                # elle est fixée à la valeur de la ligne, pas cumulée.
                numero_lot = _s(item.get('numero_lot', ''))
                if numero_lot:
                    date_peremption = _parse_date_cell(item.get('date_peremption'))
                    lot, lot_created = LotProduit.objects.get_or_create(
                        produit=obj, numero_lot=numero_lot,
                        defaults={
                            'date_peremption':   date_peremption,
                            'date_reception':    timezone.now().date(),
                            'quantite_initiale': obj.stock_actuel,
                            'quantite_actuelle': obj.stock_actuel,
                            'prix_achat_lot':    obj.prix_achat,
                        }
                    )
                    if not lot_created:
                        lot.quantite_initiale = obj.stock_actuel
                        lot.quantite_actuelle = obj.stock_actuel
                        if date_peremption:
                            lot.date_peremption = date_peremption
                        lot.save(update_fields=['quantite_initiale', 'quantite_actuelle', 'date_peremption'])
                    elif not existing:
                        # Nouveau produit + nouveau lot : trace le mouvement d'entrée
                        # correspondant, comme pour toute autre entrée en stock.
                        MouvementStock.objects.create(
                            produit=obj, lot=lot, type='entree', motif='inventaire',
                            quantite=obj.stock_actuel, stock_avant=0, stock_apres=obj.stock_actuel,
                            notes='Import de produits (fichier)', cree_par=request.user,
                        )

            if existing:
                updated += 1
            else:
                created += 1
        except ValidationError:
            errors += 1
        except Exception:
            errors += 1

    base = f'{created} créé(s), {updated} mis à jour, {skipped} ignoré(s)'
    if errors:
        base += f', {errors} erreur(s)'
    base += '.'
    suffix = _fk_warning([('Unité de mesure', unites_manquantes)])
    if errors or unites_manquantes:
        messages.warning(request, base + suffix)
    else:
        messages.success(request, base)
    return redirect('stock_produits')


@login_required(login_url='login')
def import_produits_modele(request):
    """Modèle Excel vierge pour l'import — mêmes colonnes que celles lues
    par import_produits (dans le même ordre)."""
    if not can_manage_stock(request.user):
        raise PermissionDenied
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"

    header_fill = PatternFill('solid', fgColor='2E7D32')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),  bottom=Side(style='thin', color='D0D0D0'),
    )

    headers = _PRODUIT_HDR
    exemple = [
        '', 'Paracétamol 500mg', 'medicament', 'Antalgiques', 'CP',
        'Paracétamol', '500mg', 'comprime', '0',
        '100', '20', '10', '150', '250', '1',
        'LOT2026-001', '2027-06-30',
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
    ws.row_dimensions[1].height = 20

    for col, val in enumerate(exemple, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.border = thin
        cell.font = Font(italic=True, color='888888')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Colonne date_peremption : format date pour guider la saisie
    date_peremption_col = openpyxl.utils.get_column_letter(headers.index('date_peremption') + 1)
    for row in range(2, 501):
        ws[f'{date_peremption_col}{row}'].number_format = 'YYYY-MM-DD'

    # ── Feuille cachée listant les valeurs valides, référencée par les
    # validations de données (listes déroulantes) ci-dessous.
    listes = wb.create_sheet("Listes")
    listes.sheet_state = 'hidden'

    types_vals      = [k for k, _ in Produit.TYPE_CHOICES]
    formes_vals     = [k for k, _ in Produit.FORME_CHOICES]
    bool_vals       = ['0', '1']
    categories_vals = list(CategorieStock.objects.order_by('nom').values_list('nom', flat=True)) or ['']
    unites_vals     = list(UniteMesure.objects.order_by('code').values_list('code', flat=True)) or ['']

    _listes_cols = [
        ('A', types_vals), ('B', formes_vals), ('C', bool_vals),
        ('D', categories_vals), ('E', unites_vals),
    ]
    for col_letter, values in _listes_cols:
        for i, v in enumerate(values, 1):
            listes[f'{col_letter}{i}'] = v

    NB_LIGNES = 500  # nombre de lignes du fichier couvertes par les listes déroulantes

    def _add_dropdown(col_letter, listes_col, count):
        dv = DataValidation(
            type='list',
            formula1=f"'Listes'!${listes_col}$1:${listes_col}${max(count, 1)}",
            allow_blank=True, showErrorMessage=False,
        )
        ws.add_data_validation(dv)
        dv.add(f'{col_letter}2:{col_letter}{NB_LIGNES}')

    # code, nom, dci, dosage : texte libre — pas de liste déroulante pertinente
    _add_dropdown('C', 'A', len(types_vals))       # type
    _add_dropdown('D', 'D', len(categories_vals))  # categorie (existantes — une nouvelle valeur reste acceptée)
    _add_dropdown('E', 'E', len(unites_vals))       # unite_mesure (existantes)
    _add_dropdown('H', 'B', len(formes_vals))      # forme
    _add_dropdown('I', 'C', len(bool_vals))        # prescription_obligatoire
    _add_dropdown('O', 'C', len(bool_vals))        # actif
    # stock_actuel, stock_alerte, stock_minimum, prix_achat, prix_vente : numériques — pas de liste déroulante

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="modele_import_produits.xlsx"'
    return resp


# ---------------------------------------------------------------------------
# Transfert interne
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def transfert_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    from .models import PHARMACIES
    produits = Produit.objects.filter(actif=True, stock_actuel__gt=0).order_by('nom')
    errors = {}
    if request.method == 'POST':
        produit_pk = request.POST.get('produit', '')
        pharmacie  = request.POST.get('pharmacie', '')
        notes      = request.POST.get('notes', '').strip()
        try:
            quantite = float(request.POST.get('quantite', 0) or 0)
        except ValueError:
            quantite = 0

        produit = Produit.objects.filter(pk=produit_pk).first()
        if not produit:
            errors['produit'] = 'Sélectionnez un produit.'
        if not pharmacie:
            errors['pharmacie'] = 'Sélectionnez la pharmacie destinataire.'
        if quantite <= 0:
            errors['quantite'] = 'La quantité doit être supérieure à 0.'
        elif produit and float(produit.stock_actuel) < quantite:
            errors['quantite'] = f'Stock insuffisant — disponible : {produit.stock_actuel} {produit.unite_mesure.nom if produit.unite_mesure else ""}.'

        if not errors:
            with transaction.atomic():
                produit = Produit.objects.select_for_update().get(pk=produit.pk)
                if float(produit.stock_actuel) < quantite:
                    errors['quantite'] = 'Stock insuffisant (modifié entre-temps).'
                else:
                    stock_avant = float(produit.stock_actuel)
                    stock_apres = stock_avant - quantite
                    pharma_label = dict(PHARMACIES).get(pharmacie, pharmacie)
                    MouvementStock.objects.create(
                        produit=produit, type='livraison', motif='livraison',
                        pharmacie=pharmacie,
                        quantite=quantite, stock_avant=stock_avant, stock_apres=stock_apres,
                        notes=notes, cree_par=request.user,
                        reference=f'Livraison → {pharma_label}',
                    )
                    produit.stock_actuel = stock_apres
                    produit.save(update_fields=['stock_actuel'])
            if not errors:
                messages.success(request, f'{quantite} {produit.unite_mesure.nom if produit.unite_mesure else ""} de « {produit.nom} » livrés à {pharma_label}.')
                return redirect('stock_mouvements')

    return render(request, 'stock/transfert/form.html', {
        'produits':   produits,
        'pharmacies': PHARMACIES,
        'errors':     errors,
        'post':       request.POST if request.method == 'POST' else None,
    })


# ── Dotation / Demandes pharmacies ─────────────────────────────────────────

@login_required(login_url='login')
def dotation_list(request):
    statut_filtre = request.GET.get('statut', 'en_attente')
    qs = DemandePharmacie.objects.annotate(nb_lignes=Count('lignes')).order_by('-date_demande')
    if statut_filtre and statut_filtre != 'tous':
        qs = qs.filter(statut=statut_filtre)
    stats = {
        'en_attente': DemandePharmacie.objects.filter(statut='en_attente').count(),
        'approuvee':  DemandePharmacie.objects.filter(statut='approuvee').count(),
        'partielle':  DemandePharmacie.objects.filter(statut='partielle').count(),
        'refusee':    DemandePharmacie.objects.filter(statut='refusee').count(),
    }
    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'stock/dotation/list.html', {
        'demandes':      page_obj,
        'page_obj':      page_obj,
        'statut_filtre': statut_filtre,
        'stats':         stats,
        'pharmacies':    PHARMACIES,
    })


@login_required(login_url='login')
def dotation_detail(request, pk):
    demande = get_object_or_404(DemandePharmacie, pk=pk)
    lignes  = demande.lignes.select_related('produit', 'produit__unite_mesure').all()
    lignes_enrichies = []
    for l in lignes:
        stock = float(l.produit.stock_actuel)
        demandee = float(l.quantite_demandee)
        lignes_enrichies.append({
            'ligne':        l,
            'produit':      l.produit,
            'stock':        stock,
            'suffisant':    stock >= demandee,
            'pct_dispo':    min(100, int(stock / demandee * 100)) if demandee > 0 else 0,
        })
    return render(request, 'stock/dotation/detail.html', {
        'demande':          demande,
        'lignes_enrichies': lignes_enrichies,
    })


@login_required(login_url='login')
@require_POST
def dotation_valider(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    demande = get_object_or_404(DemandePharmacie, pk=pk)
    if demande.statut != 'en_attente':
        messages.error(request, 'Cette demande a déjà été traitée.')
        return redirect('stock_dotation_detail', pk=pk)

    action = request.POST.get('action', 'approuver')

    if action == 'refuser':
        demande.statut          = 'refusee'
        demande.notes_stock     = request.POST.get('notes_stock', '').strip()
        demande.traite_par      = request.user
        demande.date_traitement = timezone.now()
        demande.save()
        messages.success(request, f'Demande {demande.numero} refusée.')
        return redirect('stock_dotation_list')

    today = timezone.now().date()
    lignes = demande.lignes.select_related('produit').all()
    tout_approuve = True
    manque_stock = []

    with transaction.atomic():
        for ligne in lignes:
            try:
                qte = float(request.POST.get(f'approuve_{ligne.pk}', 0) or 0)
            except ValueError:
                qte = 0
            qte = max(0, min(qte, float(ligne.quantite_demandee)))

            if qte > 0:
                produit = Produit.objects.select_for_update().get(pk=ligne.produit_id)
                # FEFO (First-Expired-First-Out) : déduit d'abord les lots qui
                # périment le plus tôt, jamais un lot déjà périmé — évite de
                # distribuer un produit périmé aux pharmacies alors que des
                # lots valides existent en stock.
                qte_restante = qte
                lots_dispo = LotProduit.objects.select_for_update().filter(
                    produit=produit, quantite_actuelle__gt=0,
                ).filter(
                    Q(date_peremption__isnull=True) | Q(date_peremption__gte=today)
                ).order_by(F('date_peremption').asc(nulls_last=True))
                for lot in lots_dispo:
                    if qte_restante <= 0:
                        break
                    qte_lot = min(float(lot.quantite_actuelle), qte_restante)
                    sv = float(produit.stock_actuel)
                    sa = max(0, sv - qte_lot)
                    MouvementStock.objects.create(
                        produit=produit, lot=lot,
                        type='livraison', motif='livraison',
                        pharmacie=demande.pharmacie, quantite=qte_lot,
                        stock_avant=sv, stock_apres=sa,
                        reference=demande.numero,
                        notes=f'Dotation {demande.get_pharmacie_display()} — Lot {lot.numero_lot}',
                        cree_par=request.user,
                    )
                    lot.quantite_actuelle = max(0, float(lot.quantite_actuelle) - qte_lot)
                    lot.save(update_fields=['quantite_actuelle'])
                    produit.stock_actuel = sa
                    produit.save(update_fields=['stock_actuel'])
                    qte_restante -= qte_lot

                # Repli si les lots ne couvrent pas la totalité : le mouvement
                # ne doit JAMAIS prétendre sortir plus que le stock réellement
                # disponible (sinon stock_avant-stock_apres ne correspond plus
                # à la quantité enregistrée — un mensonge dans le journal).
                if qte_restante > 0:
                    sv = float(produit.stock_actuel)
                    qte_reelle = min(qte_restante, sv)
                    if qte_reelle > 0:
                        sa = sv - qte_reelle
                        MouvementStock.objects.create(
                            produit=produit, type='livraison', motif='livraison',
                            pharmacie=demande.pharmacie, quantite=qte_reelle,
                            stock_avant=sv, stock_apres=sa,
                            reference=demande.numero,
                            notes=f'Dotation {demande.get_pharmacie_display()}',
                            cree_par=request.user,
                        )
                        produit.stock_actuel = sa
                        produit.save(update_fields=['stock_actuel'])
                    if qte_reelle < qte_restante:
                        # Stock réellement insuffisant : la quantité
                        # effectivement livrée est revue à la baisse pour que
                        # la ligne de demande ne mente jamais sur ce qui est parti.
                        qte = qte - (qte_restante - qte_reelle)
                        manque_stock.append(ligne.produit.nom)

            ligne.quantite_approuvee = qte
            ligne.save(update_fields=['quantite_approuvee'])
            if qte < float(ligne.quantite_demandee):
                tout_approuve = False

        # Le stock passe en "en_livraison" — la pharmacie doit confirmer la réception
        demande.statut          = 'en_livraison'
        demande.notes_stock     = request.POST.get('notes_stock', '').strip()
        demande.traite_par      = request.user
        demande.date_traitement = timezone.now()
        demande.save()

    if manque_stock:
        messages.warning(
            request,
            f"Stock insuffisant pour : {', '.join(manque_stock)} — quantité livrée réduite en conséquence."
        )
    messages.success(request, f'Demande {demande.numero} traitée — {demande.get_statut_display()}.')
    return redirect('stock_dotation_list')


@login_required(login_url='login')
def dotation_creer(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    produits = Produit.objects.filter(actif=True).order_by('type', 'nom')
    errors = {}
    if request.method == 'POST':
        pharmacie = request.POST.get('pharmacie', '')
        notes     = request.POST.get('notes', '').strip()
        if not pharmacie:
            errors['pharmacie'] = 'Sélectionnez une pharmacie.'
        if not errors:
            demande = DemandePharmacie.objects.create(
                pharmacie=pharmacie, notes=notes, cree_par=request.user,
            )
            nouvelles_lignes = []
            for p in produits:
                val = request.POST.get(f'qte_{p.pk}', '').strip()
                if val:
                    try:
                        qte = float(val)
                        if qte > 0:
                            nouvelles_lignes.append(LigneDemande(
                                demande=demande, produit=p, quantite_demandee=qte,
                            ))
                    except ValueError:
                        pass
            if nouvelles_lignes:
                LigneDemande.objects.bulk_create(nouvelles_lignes)
            if demande.lignes.exists():
                messages.success(request, f'Demande {demande.numero} créée.')
                return redirect('stock_dotation_detail', pk=demande.pk)
            demande.delete()
            errors['lignes'] = 'Ajoutez au moins un produit avec une quantité.'
    return render(request, 'stock/dotation/form.html', {
        'produits': produits, 'pharmacies': PHARMACIES, 'errors': errors,
        'post': request.POST if request.method == 'POST' else None,
    })


# ── Fiches de besoins ──────────────────────────────────────────────────────

@login_required(login_url='login')
def fiche_list(request):
    qs = FicheBesoins.objects.select_related('cree_par').prefetch_related(
        'besoins_achats'
    ).annotate(nb_lignes=Count('lignes')).order_by('-date_creation')
    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'stock/fiches/list.html', {'fiches': page_obj, 'page_obj': page_obj})


@login_required(login_url='login')
def fiche_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    produits = Produit.objects.filter(actif=True).select_related('categorie').order_by('categorie__type', 'categorie__nom', 'nom')
    if request.method == 'POST':
        periode_debut = request.POST.get('periode_debut', '')
        periode_fin   = request.POST.get('periode_fin', '')
        notes = request.POST.get('notes', '').strip()

        lignes_data = [(p, request.POST.get(f'cmd_{p.pk}', '').strip()) for p in produits]
        lignes_data = [(p, q) for p, q in lignes_data if q]

        if not lignes_data:
            messages.error(request, "Sélectionnez au moins un produit et renseignez les quantités.")
        elif periode_debut and periode_fin:
            from decimal import Decimal

            fiche = FicheBesoins.objects.create(
                periode_debut=periode_debut, periode_fin=periode_fin,
                notes=notes, cree_par=request.user, statut='brouillon',
            )
            LigneFicheBesoins.objects.bulk_create([
                LigneFicheBesoins(
                    fiche=fiche, produit=p,
                    stock_initial=Decimal(str(p.stock_actuel or 0)),
                    cmm=Decimal(str(p.cmm or 0)),
                    qte_commander=Decimal(qte_cmd), qte_accordee=Decimal(qte_cmd),
                )
                for p, qte_cmd in lignes_data
            ])

            messages.success(request, f'Fiche {fiche.numero} créée. Vérifiez et envoyez aux achats.')
            return redirect('stock_fiche_detail', pk=fiche.pk)

    categories = CategorieStock.objects.filter(actif=True).order_by('type', 'nom')
    return render(request, 'stock/fiches/form.html', {
        'mode': 'create', 'produits': produits,
        'categories': categories,
        'today': timezone.now().date(),
    })


@login_required(login_url='login')
def fiche_detail(request, pk):
    fiche  = get_object_or_404(FicheBesoins, pk=pk)
    lignes = fiche.lignes.select_related('produit__categorie', 'produit__unite_mesure').order_by('produit__categorie__type', 'produit__categorie__nom', 'produit__nom')
    return render(request, 'stock/fiches/detail.html', {
        'fiche': fiche, 'lignes': lignes,
    })


@login_required(login_url='login')
def fiche_edit(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    fiche  = get_object_or_404(FicheBesoins, pk=pk)
    if fiche.statut not in ('brouillon',):
        messages.error(request, 'Seule une fiche en brouillon est modifiable.')
        return redirect('stock_fiche_detail', pk=pk)
    lignes = fiche.lignes.select_related('produit__categorie', 'produit__unite_mesure').order_by('produit__categorie__type', 'produit__categorie__nom', 'produit__nom')
    if request.method == 'POST':
        from decimal import Decimal
        for ligne in lignes:
            ligne.qte_commander = Decimal(request.POST.get(f'cmd_{ligne.pk}', 0) or 0)
            ligne.notes         = request.POST.get(f'notes_{ligne.pk}', '').strip()
            ligne.save()
        fiche.notes = request.POST.get('notes', '').strip()
        fiche.save(update_fields=['notes'])
        messages.success(request, 'Fiche mise à jour.')
        return redirect('stock_fiche_detail', pk=pk)
    return render(request, 'stock/fiches/edit.html', {
        'fiche': fiche, 'lignes': lignes,
    })


@login_required(login_url='login')
@require_POST
def fiche_soumettre(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    fiche = get_object_or_404(FicheBesoins, pk=pk)
    if fiche.statut == 'brouillon':
        fiche.statut = 'soumis'
        fiche.save(update_fields=['statut'])
        messages.success(request, f'Fiche {fiche.numero} soumise pour validation.')
    return redirect('stock_fiche_detail', pk=pk)


@login_required(login_url='login')
def fiche_valider(request, pk):
    fiche  = get_object_or_404(FicheBesoins, pk=pk)
    lignes = fiche.lignes.select_related('produit', 'produit__unite_mesure').all()
    if request.method == 'POST':
        if not can_manage_stock(request.user):
            raise PermissionDenied
        from decimal import Decimal
        action = request.POST.get('action', 'valider')
        for ligne in lignes:
            ligne.qte_accordee = Decimal(request.POST.get(f'accorde_{ligne.pk}', ligne.qte_commander) or ligne.qte_commander)
            ligne.save(update_fields=['qte_accordee'])
        fiche.notes_direction = request.POST.get('notes_direction', '').strip()
        fiche.valide_par      = request.user
        fiche.date_validation = timezone.now()
        fiche.statut          = 'valide' if action == 'valider' else 'rejete'
        fiche.save()
        messages.success(request, f'Fiche {fiche.numero} {"validée" if action == "valider" else "rejetée"}.')
        return redirect('stock_fiche_detail', pk=pk)
    return render(request, 'stock/fiches/valider.html', {
        'fiche': fiche, 'lignes': lignes,
    })


@login_required(login_url='login')
@require_POST
def fiche_envoyer_achats(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    from achats.models import BesoinAchat, LigneBesoin
    fiche = get_object_or_404(FicheBesoins, pk=pk)

    besoin_existant = fiche.besoins_achats.first()
    if besoin_existant:
        messages.warning(request, f"Cette fiche a déjà été transmise ({besoin_existant.numero}).")
        return redirect('achats:besoin_detail', pk=besoin_existant.pk)

    if fiche.statut == 'rejete':
        messages.error(request, "Cette fiche a été rejetée et ne peut pas être transmise aux achats.")
        return redirect('stock_fiche_detail', pk=pk)

    if not fiche.lignes.exists():
        messages.error(request, "Cette fiche ne contient aucun produit à transmettre.")
        return redirect('stock_fiche_detail', pk=pk)

    besoin = BesoinAchat.objects.create(
        titre=f"Fiche besoins {fiche.numero} — {fiche.periode_debut.strftime('%d/%m/%Y')} au {fiche.periode_fin.strftime('%d/%m/%Y')}",
        fiche_besoins=fiche, statut='soumis', cree_par=request.user,
        notes=f"Généré depuis la fiche {fiche.numero}",
    )
    lignes_creees = 0
    for ligne in fiche.lignes.select_related('produit', 'produit__unite_mesure').all():
        qte = ligne.qte_accordee if ligne.qte_accordee else ligne.qte_commander
        LigneBesoin.objects.create(
            besoin=besoin, produit=ligne.produit,
            quantite=qte,
            unite=ligne.produit.unite_mesure.nom if ligne.produit.unite_mesure else 'unité',
        )
        lignes_creees += 1

    fiche.statut          = 'valide'
    fiche.valide_par       = request.user
    fiche.date_validation  = timezone.now()
    fiche.save(update_fields=['statut', 'valide_par', 'date_validation'])

    messages.success(request, f"Besoin {besoin.numero} transmis aux achats ({lignes_creees} produit(s)).")
    return redirect('achats:besoin_detail', pk=besoin.pk)


@login_required(login_url='login')
def fiche_print(request, pk):
    fiche  = get_object_or_404(FicheBesoins, pk=pk)
    lignes = fiche.lignes.select_related('produit', 'produit__unite_mesure').all()
    return render(request, 'stock/fiches/print.html', {
        'fiche': fiche, 'lignes': lignes, 'today': timezone.now().date(),
    })


# ──────────────────────────────────────────────────────────────
# INTÉGRATION DES RÉCEPTIONS D'ACHATS
# ──────────────────────────────────────────────────────────────

@login_required(login_url='login')
def receptions_a_integrer(request):
    from achats.models import ReceptionAchat
    receptions = ReceptionAchat.objects.filter(
        integre_en_stock=False
    ).select_related(
        'commande__fournisseur',
        'commande__proforma__besoin__fiche_besoins',
        'receptionne_par',
    ).order_by('-date_creation')
    return render(request, 'stock/receptions_a_integrer.html', {
        'receptions': receptions,
        'nb': receptions.count(),
    })


@login_required(login_url='login')
@require_POST
def integrer_reception(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    from achats.models import ReceptionAchat, LigneReceptionAchat

    with transaction.atomic():
        # select_for_update() + le filtre integre_en_stock=False dans la même transaction
        # empêchent un double clic/double soumission d'intégrer deux fois la même réception.
        reception = get_object_or_404(
            ReceptionAchat.objects.select_for_update(), pk=pk, integre_en_stock=False
        )
        commande = reception.commande
        lignes_ignorees = []

        for lr in reception.lignes.select_related(
            'ligne_commande__ligne_proforma__ligne_besoin__produit'
        ).all():
            if lr.quantite_recue <= 0:
                continue
            lc = lr.ligne_commande
            produit = None
            if (lc.ligne_proforma and
                    lc.ligne_proforma.ligne_besoin and
                    lc.ligne_proforma.ligne_besoin.produit):
                produit = lc.ligne_proforma.ligne_besoin.produit
            if produit is None:
                lignes_ignorees.append(lc.designation)
                continue

            numero_lot = lr.numero_lot or reception.numero
            lot = LotProduit.objects.filter(produit=produit, numero_lot=numero_lot).first()
            if lot:
                lot.quantite_actuelle += lr.quantite_recue
                lot.quantite_initiale += lr.quantite_recue
                update_lot_fields = ['quantite_actuelle', 'quantite_initiale']
                if not lot.date_peremption and lr.date_peremption:
                    lot.date_peremption = lr.date_peremption
                    update_lot_fields.append('date_peremption')
                lot.save(update_fields=update_lot_fields)
            else:
                lot = LotProduit.objects.create(
                    produit=produit,
                    numero_lot=numero_lot,
                    date_reception=reception.date_reception,
                    date_peremption=lr.date_peremption,
                    quantite_initiale=lr.quantite_recue,
                    quantite_actuelle=lr.quantite_recue,
                    fournisseur=commande.fournisseur,
                    prix_achat_lot=lc.prix_unitaire,
                )
            stock_avant = produit.stock_actuel
            produit.stock_actuel = produit.stock_actuel + lr.quantite_recue
            produit.save(update_fields=['stock_actuel'])
            MouvementStock.objects.create(
                produit=produit, lot=lot,
                type='entree', motif='achat',
                quantite=lr.quantite_recue,
                stock_avant=stock_avant,
                stock_apres=produit.stock_actuel,
                reference=reception.numero,
                notes=f'Réception {reception.numero} — Commande {commande.numero}',
                cree_par=request.user,
            )

        reception.integre_en_stock = True
        reception.date_integration = timezone.now()
        reception.integre_par = request.user
        if lignes_ignorees:
            note = f"Non intégré en stock (produit non résolu) : {', '.join(lignes_ignorees)}"
            reception.notes = f"{reception.notes}\n{note}".strip()
            reception.save(update_fields=['integre_en_stock', 'date_integration', 'integre_par', 'notes'])
        else:
            reception.save(update_fields=['integre_en_stock', 'date_integration', 'integre_par'])

    if lignes_ignorees:
        messages.warning(
            request,
            f"Réception {reception.numero} intégrée, mais {len(lignes_ignorees)} ligne(s) ignorée(s) "
            f"faute de produit identifiable : {', '.join(lignes_ignorees)}. Vérifiez manuellement le stock pour ces articles."
        )
    else:
        messages.success(request, f'Réception {reception.numero} intégrée dans le stock.')
    return redirect('stock_receptions_a_integrer')


# ── Unités de mesure ───────────────────────────────────────────────────────

from .models import CategorieUniteMesure
from .forms import UniteMesureForm, CategorieUniteMesureForm
from services.views import _export_file, _parse_upload, _s, _b, _fk_warning


@login_required(login_url='login')
def unites_list(request):
    q = request.GET.get('q', '').strip()
    categorie_id = request.GET.get('categorie', '')
    vue = request.GET.get('vue', 'liste')
    qs = UniteMesure.objects.select_related('categorie').all()
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(code__icontains=q))
    if categorie_id:
        qs = qs.filter(categorie_id=categorie_id)
    total_all = UniteMesure.objects.count()
    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'stock/unites/list.html', {
        'page_obj': page_obj,
        'categories_um': CategorieUniteMesure.objects.all(),
        'q': q,
        'categorie_id': categorie_id,
        'vue': vue,
        'total': total_all,
        'total_filtre': qs.count(),
    })


@login_required(login_url='login')
def unite_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        form = UniteMesureForm(request.POST)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f'Unité « {obj.nom} » créée.')
            return redirect('stock_unites')
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        form = UniteMesureForm()
    return render(request, 'stock/unites/form.html', {
        'form': form,
        'titre': 'Nouvelle unité de mesure',
        'edit': False,
    })


@login_required(login_url='login')
def unite_edit(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    obj = get_object_or_404(UniteMesure, pk=pk)
    if request.method == 'POST':
        form = UniteMesureForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, f'Unité « {obj.nom} » mise à jour.')
            return redirect('stock_unite_detail', pk=obj.pk)
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        form = UniteMesureForm(instance=obj)
    return render(request, 'stock/unites/form.html', {
        'form': form,
        'obj': obj,
        'titre': f'Modifier — {obj.nom}',
        'edit': True,
    })


@login_required(login_url='login')
def unite_detail(request, pk):
    obj = get_object_or_404(UniteMesure, pk=pk)
    return render(request, 'stock/unites/detail.html', {'obj': obj})


@login_required(login_url='login')
def unite_delete(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    obj = get_object_or_404(UniteMesure, pk=pk)
    if request.method == 'POST':
        nom = obj.nom
        obj.delete()
        messages.success(request, f'Unité « {nom} » supprimée.')
    return redirect('stock_unites')


@login_required(login_url='login')
@require_POST
def unite_bulk_delete(request):
    if not can_manage_stock(request.user):
        return JsonResponse({'error': 'Permission refusée.'}, status=403)
    ids = request.POST.getlist('ids[]')
    if ids:
        count, _ = UniteMesure.objects.filter(pk__in=ids).delete()
        return JsonResponse({'ok': True, 'count': count})
    return JsonResponse({'ok': False}, status=400)


@login_required(login_url='login')
def categories_unites_list(request):
    q = request.GET.get('q', '').strip()
    vue = request.GET.get('vue', 'liste')
    qs = CategorieUniteMesure.objects.all()
    if q:
        qs = qs.filter(nom__icontains=q)
    total_all = CategorieUniteMesure.objects.count()
    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'stock/unites/categories/list.html', {
        'page_obj': page_obj,
        'q': q,
        'vue': vue,
        'total': total_all,
        'total_filtre': qs.count(),
    })


@login_required(login_url='login')
def categorie_unite_create(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        form = CategorieUniteMesureForm(request.POST)
        if form.is_valid():
            obj = form.save()
            messages.success(request, f'Catégorie « {obj.nom} » créée.')
            return redirect('stock_categories_unites')
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        form = CategorieUniteMesureForm()
    return render(request, 'stock/unites/categories/form.html', {
        'form': form,
        'titre': "Nouvelle catégorie d'unité",
        'edit': False,
    })


@login_required(login_url='login')
def categorie_unite_detail(request, pk):
    obj = get_object_or_404(CategorieUniteMesure, pk=pk)
    return render(request, 'stock/unites/categories/detail.html', {'obj': obj})


@login_required(login_url='login')
def categorie_unite_edit(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    obj = get_object_or_404(CategorieUniteMesure, pk=pk)
    if request.method == 'POST':
        form = CategorieUniteMesureForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, f'Catégorie « {obj.nom} » mise à jour.')
            return redirect('stock_categorie_unite_detail', pk=obj.pk)
        else:
            messages.error(request, 'Veuillez corriger les erreurs du formulaire.')
    else:
        form = CategorieUniteMesureForm(instance=obj)
    return render(request, 'stock/unites/categories/form.html', {
        'form': form,
        'obj': obj,
        'titre': f'Modifier — {obj.nom}',
        'edit': True,
    })


@login_required(login_url='login')
def categorie_unite_delete(request, pk):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    obj = get_object_or_404(CategorieUniteMesure, pk=pk)
    if request.method == 'POST':
        nom = obj.nom
        obj.delete()
        messages.success(request, f'Catégorie « {nom} » supprimée.')
    return redirect('stock_categories_unites')


@login_required(login_url='login')
@require_POST
def categorie_unite_bulk_delete(request):
    if not can_manage_stock(request.user):
        return JsonResponse({'error': 'Permission refusée.'}, status=403)
    ids = request.POST.getlist('ids[]')
    if ids:
        count, _ = CategorieUniteMesure.objects.filter(pk__in=ids).delete()
        return JsonResponse({'ok': True, 'count': count})
    return JsonResponse({'ok': False}, status=400)


_UM_HDR = ['code', 'nom', 'categorie', 'type_unite', 'ratio', 'precision_arrondi', 'actif']


def _um_row(u):
    return [
        u.code, u.nom,
        u.categorie.nom if u.categorie else '',
        u.type_unite, float(u.ratio), float(u.precision_arrondi), int(u.actif),
    ]


@login_required(login_url='login')
def export_unites(request):
    fmt = request.GET.get('format', 'json')
    qs = UniteMesure.objects.select_related('categorie')
    rows = [_um_row(u) for u in qs]
    return _export_file(fmt, 'unites_mesure', _UM_HDR, rows,
                        [dict(zip(_UM_HDR, r)) for r in rows])


_CU_HDR = ['nom']


@login_required(login_url='login')
def export_categories_unites(request):
    fmt = request.GET.get('format', 'json')
    qs = CategorieUniteMesure.objects.all()
    rows = [[c.nom] for c in qs]
    return _export_file(fmt, 'categories_unites', _CU_HDR, rows,
                        [{'nom': c.nom} for c in qs])


@login_required(login_url='login')
@require_POST
def import_unites(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('stock_unites')

    data, err = _parse_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('stock_unites')

    do_update = 'update' in request.POST
    created = updated = skipped = errors = 0

    for item in data:
        try:
            code = _s(item.get('code', ''))
            if not code:
                errors += 1
                continue
            cat_nom = _s(item.get('categorie', ''))
            cat = None
            if cat_nom:
                cat, _ = CategorieUniteMesure.objects.get_or_create(nom=cat_nom)
            defaults = {
                'nom': _s(item.get('nom', code)),
                'categorie': cat,
                'type_unite': _s(item.get('type_unite', 'umrc')),
                'ratio': item.get('ratio') or 1,
                'precision_arrondi': item.get('precision_arrondi') or 0.01,
                'actif': _b(item.get('actif', True)),
            }
            obj, was_created = UniteMesure.objects.get_or_create(code=code, defaults=defaults)
            if was_created:
                created += 1
            elif do_update:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                updated += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    if errors:
        messages.warning(request, f'{created} créée(s), {updated} mise(s) à jour, {skipped} ignorée(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} unité(s) importée(s), {updated} mise(s) à jour, {skipped} ignorée(s).')
    return redirect('stock_unites')


@login_required(login_url='login')
@require_POST
def import_categories_unites(request):
    if not can_manage_stock(request.user):
        raise PermissionDenied
    upload = request.FILES.get('fichier')
    if not upload:
        messages.error(request, 'Aucun fichier sélectionné.')
        return redirect('stock_categories_unites')

    data, err = _parse_upload(upload)
    if err:
        messages.error(request, err)
        return redirect('stock_categories_unites')

    created = skipped = errors = 0
    for item in data:
        try:
            nom = _s(item.get('nom', ''))
            if not nom:
                errors += 1
                continue
            _, was_created = CategorieUniteMesure.objects.get_or_create(nom=nom)
            if was_created:
                created += 1
            else:
                skipped += 1
        except Exception:
            errors += 1

    if errors:
        messages.warning(request, f'{created} créée(s), {skipped} ignorée(s), {errors} erreur(s).')
    else:
        messages.success(request, f'{created} catégorie(s) importée(s), {skipped} ignorée(s).')
    return redirect('stock_categories_unites')
